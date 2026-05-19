"""
API de Conciliacao Bancaria.

Endpoints:
- GET  /              - info do servico
- GET  /health        - healthcheck
- POST /conciliar/ofx - recebe 1 ou 2 arquivos OFX e devolve relatorio em Markdown
- POST /conciliar/csv - recebe CSV extrato + CSV razao contabil e devolve relatorio

Execucao:
    uvicorn api.main:app --reload --port 8765
"""

from __future__ import annotations

import io
import json
import logging
import os
import re
import secrets
import xml.etree.ElementTree as ET
from pathlib import Path
from typing import List, Optional

import pdfplumber
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from slowapi import Limiter, _rate_limit_exceeded_handler
from slowapi.errors import RateLimitExceeded
from slowapi.util import get_remote_address

from anthropic import Anthropic, APIStatusError
from fastapi import Depends, FastAPI, File, Header, HTTPException, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, Response
from fastapi.staticfiles import StaticFiles
from markdown import markdown as md_to_html
import contextlib
from pydantic import BaseModel
from sqlalchemy import text as sql_text

# --- Config via .env (deve rodar ANTES dos imports de banco) ---
_env_path = Path(__file__).resolve().parent.parent / ".env"
if _env_path.exists():
    load_dotenv(_env_path, override=True)

# Imports de banco (opcionais — ativos somente se DATABASE_URL estiver configurado)
_DB_IMPORTS_OK = False
try:
    from api.db.client import SessionLocal, engine
    from api.db import models
    from api.db import clientes as crud_clientes
    _DB_IMPORTS_OK = True
except Exception:
    pass

AUTH_TOKEN = os.environ.get("ORGCONC_AUTH_TOKEN", "").strip()
CORS_ORIGINS = [
    o.strip() for o in os.environ.get(
        "ORGCONC_CORS_ORIGINS",
        "http://127.0.0.1:8765,http://localhost:8765",
    ).split(",") if o.strip()
]
MAX_UPLOAD_MB = int(os.environ.get("ORGCONC_MAX_UPLOAD_MB", "10"))
MAX_UPLOAD_BYTES = MAX_UPLOAD_MB * 1024 * 1024
DATA_DIR = Path(os.environ.get("ORGCONC_DATA_DIR", "./data")).resolve()
DATA_DIR.mkdir(parents=True, exist_ok=True)

_DB_URL = os.environ.get("DATABASE_URL", "").strip()
DB_DISPONIVEL = _DB_IMPORTS_OK and bool(_DB_URL) and not re.search(r"\[.+?\]", _DB_URL)

# --- Logging ---
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
)
log = logging.getLogger("orgconc")

# --- Rate limiter ---
limiter = Limiter(key_func=get_remote_address, default_limits=["120/minute"])

SYSTEM_PROMPT = (
    "Voce e um agente especializado em conciliacao bancaria para escritorios "
    "contabeis brasileiros. Recebe extratos (OFX/CSV) e/ou razao contabil, "
    "cruza por data/valor/descricao, identifica conciliados, divergencias, "
    "duplicidades e nao conciliados, e gera relatorio em portugues com resumo "
    "executivo, achados criticos, classificacao contabil e plano de acao."
)

@contextlib.asynccontextmanager
async def _lifespan(app: FastAPI):
    if not os.environ.get("ANTHROPIC_API_KEY"):
        log.warning("ANTHROPIC_API_KEY nao configurada — /conciliar/ofx (sem simular) nao funcionara")
    if DB_DISPONIVEL:
        log.info("Banco configurado: %s", _DB_URL.split("@")[-1] if "@" in _DB_URL else "ok")
    else:
        log.info("Banco nao configurado — persistencia JSON local ativa")
    yield
    if DB_DISPONIVEL and _DB_IMPORTS_OK:
        await engine.dispose()


app = FastAPI(
    title="ORGATEC · Conciliacao Bancaria API",
    description="Cruza extratos OFX/PDF/XML contra razao contabil. Gera HTML/XLSX/PDF.",
    version="0.3.0",
    lifespan=_lifespan,
)
app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)
app.add_middleware(
    CORSMiddleware,
    allow_origins=CORS_ORIGINS if CORS_ORIGINS else ["*"],
    allow_methods=["GET", "POST", "OPTIONS"],
    allow_headers=["*"],
)

from starlette.middleware.base import BaseHTTPMiddleware
from starlette.requests import Request as StarletteRequest
from starlette.responses import Response as StarletteResponse

_CSP = (
    "default-src 'self'; "
    "script-src 'self' 'unsafe-inline' cdn.jsdelivr.net; "
    "style-src 'self' 'unsafe-inline' fonts.googleapis.com; "
    "font-src 'self' fonts.gstatic.com; "
    "img-src 'self' data:; "
    "object-src 'none'; "
    "frame-ancestors 'none'; "
    "base-uri 'self'"
)

class _SecurityHeadersMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request: StarletteRequest, call_next) -> StarletteResponse:
        response = await call_next(request)
        response.headers["Content-Security-Policy"] = _CSP
        response.headers["X-Content-Type-Options"] = "nosniff"
        response.headers["X-Frame-Options"] = "DENY"
        response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
        response.headers["Permissions-Policy"] = "geolocation=(), microphone=(), camera=()"
        return response

app.add_middleware(_SecurityHeadersMiddleware)


def auth(authorization: Optional[str] = Header(None)) -> None:
    """Auth opcional via Bearer token (so valida se ORGCONC_AUTH_TOKEN estiver definido)."""
    if not AUTH_TOKEN:
        return
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Token Bearer ausente")
    token = authorization.split(" ", 1)[1].strip()
    if not secrets.compare_digest(token, AUTH_TOKEN):
        raise HTTPException(status_code=401, detail="Token invalido")


# ── Modelos Pydantic ──────────────────────────────────────────────────────

def _validar_cnpj(cnpj: str) -> bool:
    digits = re.sub(r"\D", "", cnpj)
    if len(digits) != 14 or len(set(digits)) == 1:
        return False
    def _calc(d, pesos):
        s = sum(int(d[i]) * pesos[i] for i in range(len(pesos)))
        r = s % 11
        return 0 if r < 2 else 11 - r
    p1 = [5,4,3,2,9,8,7,6,5,4,3,2]
    p2 = [6,5,4,3,2,9,8,7,6,5,4,3,2]
    return int(digits[12]) == _calc(digits, p1) and int(digits[13]) == _calc(digits, p2)


class ClienteCreate(BaseModel):
    nome: str
    cnpj: Optional[str] = None
    email: Optional[str] = None
    telefone: Optional[str] = None
    plano: str = "basico"

    def model_post_init(self, __context) -> None:
        if self.cnpj and not _validar_cnpj(self.cnpj):
            raise ValueError(f"CNPJ inválido: {self.cnpj}")

class ClienteUpdate(BaseModel):
    nome: Optional[str] = None
    email: Optional[str] = None
    telefone: Optional[str] = None
    plano: Optional[str] = None
    ativo: Optional[bool] = None


# ── Persistência no banco ─────────────────────────────────────────────────

async def _salvar_no_banco(
    report_id: str,
    extratos: list[dict],
    anomalias: list[dict],
    modo: str,
    cliente_id: Optional[str] = None,
) -> None:
    if not DB_DISPONIVEL:
        return
    import uuid as _uuid
    from datetime import date
    try:
        total_cred = sum(t["valor"] for e in extratos for t in e["transacoes"] if t["valor"] > 0)
        total_deb  = sum(t["valor"] for e in extratos for t in e["transacoes"] if t["valor"] < 0)
        datas = sorted({t["data"] for e in extratos for t in e["transacoes"] if t.get("data")})
        cid = _uuid.UUID(cliente_id) if cliente_id else None
        async with SessionLocal() as db:
            async with db.begin():
                conc = models.Conciliacao(
                    cliente_id=cid,
                    report_id=report_id,
                    modo=modo,
                    total_transacoes=sum(e["qtd"] for e in extratos),
                    total_anomalias=len(anomalias),
                    valor_total_credito=total_cred,
                    valor_total_debito=total_deb,
                    periodo_inicio=date.fromisoformat(datas[0]) if datas else None,
                    periodo_fim=date.fromisoformat(datas[-1]) if datas else None,
                )
                db.add(conc)
                await db.flush()
                anomalias_set = {
                    (a.get("conta", ""), round(abs(a.get("valor", 0)), 2))
                    for a in anomalias
                }
                txs = [
                    models.Transacao(
                        conciliacao_id=conc.id,
                        cliente_id=cid,
                        data_lancamento=date.fromisoformat(t["data"]) if t.get("data") else date.today(),
                        valor=t["valor"],
                        memo=t.get("memo"),
                        categoria=_classificar(t.get("memo", ""), t.get("nome", "")),
                        banco=e.get("conta"),
                        tipo=t.get("tipo"),
                        eh_anomalia=(e.get("conta", ""), round(abs(t["valor"]), 2)) in anomalias_set,
                    )
                    for e in extratos for t in e["transacoes"]
                ]
                db.add_all(txs)
            log.info("Conciliacao %s salva no banco (%d transacoes)", report_id, len(txs))
    except Exception:
        log.exception("Falha ao salvar no banco (conciliacao %s) — JSON preservado", report_id)


async def read_limited(up: UploadFile, max_bytes: int = MAX_UPLOAD_BYTES) -> bytes:
    """Le upload com limite de tamanho. Aborta se exceder."""
    chunks: list[bytes] = []
    total = 0
    while True:
        chunk = await up.read(1024 * 256)
        if not chunk:
            break
        total += len(chunk)
        if total > max_bytes:
            raise HTTPException(
                status_code=413,
                detail=f"Arquivo {up.filename} excede {MAX_UPLOAD_MB} MB",
            )
        chunks.append(chunk)
    return b"".join(chunks)

_static_dir = Path(__file__).resolve().parent.parent / "static"
if _static_dir.exists():
    app.mount("/ui", StaticFiles(directory=str(_static_dir), html=True), name="ui")

# Logo embutida (base64) para HTML standalone e PDF
import base64
_LOGO_PATH = _static_dir / "logo.png"
_LOGO_B64 = ""
_LOGO_DATA_URI = ""
if _LOGO_PATH.exists():
    _LOGO_B64 = base64.b64encode(_LOGO_PATH.read_bytes()).decode()
    _LOGO_DATA_URI = f"data:image/png;base64,{_LOGO_B64}"


def _get_client() -> Anthropic:
    key = os.environ.get("ANTHROPIC_API_KEY")
    if not key:
        raise HTTPException(
            status_code=500,
            detail="ANTHROPIC_API_KEY nao configurada no servidor",
        )
    return Anthropic(api_key=key)


def _salvar_dataset(extratos: list[dict], anomalias: list[dict], relatorio: str) -> str:
    """Persiste o dataset em disco (./data/{rid}.json) e retorna o ID."""
    import uuid
    rid = uuid.uuid4().hex[:12]
    path = DATA_DIR / f"{rid}.json"
    payload = {
        "extratos": extratos,
        "anomalias": anomalias,
        "relatorio": relatorio,
    }
    path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
    # Limpa datasets antigos: mantem so os 50 mais recentes
    existing = sorted(DATA_DIR.glob("*.json"), key=lambda p: p.stat().st_mtime, reverse=True)
    for old in existing[50:]:
        try: old.unlink()
        except OSError: pass
    log.info("Dataset salvo: %s (%d transacoes, %d anomalias)", rid,
             sum(e["qtd"] for e in extratos), len(anomalias))
    return rid


def _carregar_dataset(rid: str) -> dict:
    """Carrega dataset persistido."""
    # Sanitiza o ID para evitar path traversal
    if not re.fullmatch(r"[a-f0-9]{12}", rid):
        raise HTTPException(status_code=400, detail="ID invalido")
    path = DATA_DIR / f"{rid}.json"
    if not path.exists():
        raise HTTPException(status_code=404, detail="Relatorio nao encontrado ou expirado")
    return json.loads(path.read_text(encoding="utf-8"))


def _render_html(relatorio_md: str) -> str:
    """Renderiza relatorio Markdown como HTML completo standalone (fundo branco, logo embutida)."""
    body = md_to_html(relatorio_md, extensions=["tables", "fenced_code"])
    from datetime import datetime
    agora = datetime.now().strftime("%d/%m/%Y %H:%M")
    logo_html = (
        f'<img src="{_LOGO_DATA_URI}" alt="ORGATEC" style="width:64px;height:64px;'
        f'filter:drop-shadow(0 0 14px rgba(255,255,255,0.4));">'
        if _LOGO_DATA_URI else ""
    )
    return f"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8">
<title>ORGATEC - Relatorio de Conciliacao Bancaria</title>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&family=JetBrains+Mono:wght@400;500&display=swap" rel="stylesheet">
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  html, body {{ background: #ffffff; }}
  body {{
    font-family: "Inter", -apple-system, "Segoe UI", Roboto, sans-serif;
    color: #1a202c;
    line-height: 1.65;
    font-size: 14px;
    -webkit-font-smoothing: antialiased;
  }}
  .wrap {{ max-width: 980px; margin: 0 auto; background: #ffffff; }}

  .hd {{
    background: linear-gradient(135deg, #0a3a7a 0%, #1e6fd9 60%, #4dc8ff 100%);
    color: white;
    padding: 32px 44px;
    display: flex;
    align-items: center;
    gap: 22px;
    position: relative;
    overflow: hidden;
  }}
  .hd::after {{
    content: "";
    position: absolute;
    right: -120px; top: -120px;
    width: 360px; height: 360px;
    background: radial-gradient(circle, rgba(255,255,255,0.15), transparent 65%);
    pointer-events: none;
  }}
  .hd .brand .nm {{
    font-size: 28px;
    font-weight: 800;
    letter-spacing: 0.5px;
    text-shadow: 0 2px 8px rgba(0,0,0,0.15);
  }}
  .hd .brand .tg {{
    font-size: 11px;
    letter-spacing: 3px;
    text-transform: uppercase;
    opacity: 0.92;
    margin-top: 3px;
  }}
  .hd .spacer {{ flex: 1; }}
  .hd .meta {{
    text-align: right;
    font-size: 11px;
    background: rgba(255,255,255,0.12);
    border: 1px solid rgba(255,255,255,0.25);
    border-radius: 10px;
    padding: 10px 16px;
    backdrop-filter: blur(8px);
    position: relative; z-index: 1;
  }}
  .hd .meta .ttl {{
    font-weight: 700;
    letter-spacing: 1.8px;
    font-size: 10.5px;
    text-transform: uppercase;
    opacity: 0.95;
  }}
  .hd .meta .dt {{ font-size: 12.5px; margin-top: 3px; font-weight: 600; }}

  .content {{ padding: 36px 44px 16px; background: #ffffff; }}

  h1 {{
    font-size: 24px;
    color: #0a3a7a;
    border-bottom: 3px solid #1e6fd9;
    padding-bottom: 10px;
    margin-bottom: 18px;
    font-weight: 800;
    letter-spacing: -0.3px;
  }}
  h2 {{
    font-size: 17px;
    color: #1e6fd9;
    margin-top: 28px;
    margin-bottom: 10px;
    font-weight: 700;
    padding-left: 12px;
    border-left: 3px solid #4dc8ff;
  }}
  h3 {{
    font-size: 14px;
    color: #0a3a7a;
    margin-top: 18px;
    margin-bottom: 8px;
    font-weight: 700;
  }}
  p {{ margin-bottom: 10px; color: #2d3748; }}
  ul, ol {{ padding-left: 22px; margin-bottom: 12px; }}
  li {{ margin-bottom: 5px; color: #2d3748; }}

  table {{
    border-collapse: collapse;
    width: 100%;
    margin: 14px 0;
    font-size: 12.5px;
    box-shadow: 0 1px 3px rgba(0,0,0,0.05);
    border-radius: 8px;
    overflow: hidden;
  }}
  th, td {{ border: 1px solid #e2e8f0; padding: 9px 12px; text-align: left; }}
  th {{
    background: linear-gradient(180deg, #0a3a7a, #1e6fd9);
    color: white;
    font-weight: 600;
    font-size: 12px;
    letter-spacing: 0.3px;
  }}
  td {{ color: #2d3748; background: #ffffff; }}
  tr:nth-child(even) td {{ background: #f8fafc; }}
  tr:hover td {{ background: #eff6ff; }}

  code {{
    background: #eff6ff;
    color: #1e6fd9;
    padding: 2px 7px;
    border-radius: 4px;
    font-size: 12px;
    font-family: "JetBrains Mono", "Consolas", monospace;
    font-weight: 500;
  }}
  strong {{ color: #0a3a7a; font-weight: 700; }}
  em {{ color: #4a5568; }}

  .ft {{
    background: #f7fafc;
    border-top: 1px solid #e2e8f0;
    padding: 16px 44px;
    color: #718096;
    font-size: 11px;
    display: flex;
    justify-content: space-between;
    flex-wrap: wrap;
    gap: 8px;
  }}
  .ft .pill {{
    background: #eff6ff;
    color: #1e6fd9;
    padding: 3px 10px;
    border-radius: 999px;
    font-weight: 600;
    font-size: 10.5px;
    letter-spacing: 0.5px;
  }}
  @media print {{ body {{ background: white; }} .wrap {{ box-shadow: none; }} .shimmer {{ display: none; }} }}

  /* Camada branca oscilante */
  @keyframes oscila {{
    0%   {{ transform: translateX(-120%) skewX(-18deg); opacity: 0; }}
    15%  {{ opacity: 1; }}
    85%  {{ opacity: 1; }}
    100% {{ transform: translateX(220%) skewX(-18deg); opacity: 0; }}
  }}
  .shimmer {{
    position: fixed;
    top: 0; left: 0;
    width: 35%; height: 100%;
    background: linear-gradient(
      90deg,
      transparent 0%,
      rgba(255,255,255,0.055) 40%,
      rgba(255,255,255,0.11) 50%,
      rgba(255,255,255,0.055) 60%,
      transparent 100%
    );
    animation: oscila 7s cubic-bezier(0.4, 0, 0.6, 1) infinite;
    pointer-events: none;
    z-index: 9999;
  }}
  @media (prefers-reduced-motion: reduce) {{
    .shimmer {{ display: none; }}
  }}
</style>
</head>
<body>
  <div class="shimmer" aria-hidden="true"></div>
  <div class="wrap">
    <div class="hd">
      {logo_html}
      <div class="brand">
        <div class="nm">ORGATEC</div>
        <div class="tg">Contabilidade &amp; Auditoria</div>
      </div>
      <div class="spacer"></div>
      <div class="meta">
        <div class="ttl">Relatório de Conciliação</div>
        <div class="dt">{agora}</div>
      </div>
    </div>
    <div class="content">{body}</div>
    <div class="ft">
      <div>© ORGATEC Contabilidade e Auditoria · orgatec.cloud@gmail.com</div>
      <div><span class="pill">OrgAudi 1.0</span></div>
    </div>
  </div>
</body>
</html>"""


def _xlsx_estilos() -> dict:
    """Paleta de cores, fontes e bordas compartilhados entre as abas XLSX."""
    BLUE_DARK = "0A3A7A"; BLUE = "1E6FD9"; WHITE = "FFFFFF"
    GRAY_BORDER = "E2E8F0"; GRAY_LIGHT = "F7FAFC"; GRAY_HOVER = "EFF6FF"
    RED = "DC2626"; RED_BG = "FEE2E2"
    ORANGE = "EA580C"; ORANGE_BG = "FFEDD5"
    YELLOW = "CA8A04"; YELLOW_BG = "FEF9C3"
    GREEN = "16A34A"
    side_thin = Side(border_style="thin", color=GRAY_BORDER)
    return dict(
        BLUE_DARK=BLUE_DARK, BLUE=BLUE, WHITE=WHITE,
        RED=RED, ORANGE=ORANGE, YELLOW=YELLOW, GREEN=GREEN,
        fill_blue_dark=PatternFill("solid", fgColor=BLUE_DARK),
        fill_blue=PatternFill("solid", fgColor=BLUE),
        fill_zebra=PatternFill("solid", fgColor=GRAY_LIGHT),
        fill_kpi_blue=PatternFill("solid", fgColor=GRAY_HOVER),
        fill_critico=PatternFill("solid", fgColor=RED_BG),
        fill_alerta=PatternFill("solid", fgColor=ORANGE_BG),
        fill_atencao=PatternFill("solid", fgColor=YELLOW_BG),
        font_h_white=Font(bold=True, color=WHITE, size=11, name="Calibri"),
        font_brand=Font(bold=True, size=24, color=BLUE_DARK, name="Calibri"),
        font_brand_sub=Font(color=BLUE, size=10, italic=True, name="Calibri"),
        font_section=Font(bold=True, size=13, color=BLUE_DARK, name="Calibri"),
        font_kpi_lbl=Font(bold=True, size=9, color="64748B", name="Calibri"),
        font_kpi_val_red=Font(bold=True, size=22, color=RED, name="Calibri"),
        font_kpi_val_orange=Font(bold=True, size=22, color=ORANGE, name="Calibri"),
        font_kpi_val_yellow=Font(bold=True, size=22, color=YELLOW, name="Calibri"),
        font_kpi_val_blue=Font(bold=True, size=22, color=BLUE_DARK, name="Calibri"),
        side_thin=side_thin,
        border_all=Border(left=side_thin, right=side_thin, top=side_thin, bottom=side_thin),
        border_kpi=Border(left=side_thin, right=side_thin,
                          top=Side(border_style="medium", color=BLUE), bottom=side_thin),
        FMT_BRL='R$ #,##0.00;[Red]-R$ #,##0.00',
        FMT_BRL_POS='R$ #,##0.00',
    )


def _xlsx_aba_resumo(ws, extratos: list[dict], anomalias: list[dict], e: dict) -> None:
    """Preenche a aba Resumo com cabecalho, KPIs e tabelas."""
    from openpyxl.drawing.image import Image as XLImage
    from datetime import datetime

    ws.title = "Resumo"
    ws.sheet_view.showGridLines = False

    def estilo_header(cells, fill=None, font=None):
        fill = fill or e["fill_blue_dark"]
        font = font or e["font_h_white"]
        for c in cells:
            c.fill = fill; c.font = font
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border = e["border_all"]

    def linha_borda(ws_inner, row, cols):
        for c in range(1, cols + 1):
            ws_inner.cell(row=row, column=c).border = e["border_all"]

    # Cabecalho ORGATEC (linhas 1-3)
    if _LOGO_PATH.exists():
        try:
            img = XLImage(str(_LOGO_PATH))
            img.width = 64; img.height = 64
            ws.add_image(img, "A1")
        except Exception:
            pass
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 24
    ws.row_dimensions[3].height = 10

    ws["B1"] = "ORGATEC"
    ws["B1"].font = e["font_brand"]
    ws["B2"] = "Contabilidade & Auditoria"
    ws["B2"].font = e["font_brand_sub"]
    ws.merge_cells("B1:E1"); ws.merge_cells("B2:E2")

    ws["F1"] = "RELATÓRIO DE CONCILIAÇÃO"
    ws["F1"].font = Font(bold=True, color=e["BLUE_DARK"], size=11, name="Calibri")
    ws["F1"].alignment = Alignment(horizontal="right", vertical="bottom")
    ws.merge_cells("F1:H1")
    ws["F2"] = f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["F2"].font = Font(italic=True, color="64748B", size=10, name="Calibri")
    ws["F2"].alignment = Alignment(horizontal="right", vertical="top")
    ws.merge_cells("F2:H2")

    for col in range(1, 9):
        ws.cell(row=3, column=col).fill = e["fill_blue"]
    ws.row_dimensions[3].height = 4

    # KPIs (linhas 5-10)
    ws.cell(row=5, column=1, value="VISÃO GERAL").font = e["font_section"]
    ws.merge_cells("A5:H5")

    total_tx   = sum(ex["qtd"] for ex in extratos)
    total_cred = sum(t["valor"] for ex in extratos for t in ex["transacoes"] if t["valor"] > 0)
    total_deb  = sum(t["valor"] for ex in extratos for t in ex["transacoes"] if t["valor"] < 0)
    saldo = total_cred + total_deb

    sev_count = {"critico": 0, "alerta": 0, "atencao": 0}
    for a in anomalias:
        sev_count[a["severidade"]] = sev_count.get(a["severidade"], 0) + 1

    kpis = [
        ("TRANSAÇÕES",  total_tx,   e["font_kpi_val_blue"],
         e["fill_kpi_blue"], None),
        ("CRÉDITOS",    total_cred,
         Font(bold=True, size=18, color=e["GREEN"], name="Calibri"),
         e["fill_kpi_blue"], e["FMT_BRL_POS"]),
        ("DÉBITOS",     total_deb,
         Font(bold=True, size=18, color=e["RED"], name="Calibri"),
         e["fill_kpi_blue"], e["FMT_BRL"]),
        ("SALDO",       saldo,
         Font(bold=True, size=18, color=e["BLUE_DARK"], name="Calibri"),
         e["fill_kpi_blue"], e["FMT_BRL"]),
    ]
    sev_kpis = [
        ("🔴 CRÍTICAS", sev_count["critico"], e["font_kpi_val_red"],    e["fill_critico"]),
        ("🟠 ALERTAS",  sev_count["alerta"],  e["font_kpi_val_orange"], e["fill_alerta"]),
        ("🟡 ATENÇÃO",  sev_count["atencao"], e["font_kpi_val_yellow"], e["fill_atencao"]),
        ("✅ TOTAL",    len(anomalias),        e["font_kpi_val_blue"],   e["fill_kpi_blue"]),
    ]

    border_kpi_bottom = Border(left=e["side_thin"], right=e["side_thin"], bottom=e["side_thin"])

    def aplicar_kpi(ws_inner, row_lbl, row_val, col, label, val, font_val, fill, fmt=None):
        lbl = ws_inner.cell(row=row_lbl, column=col, value=label)
        lbl.font = e["font_kpi_lbl"]
        lbl.alignment = Alignment(horizontal="left", vertical="bottom")
        lbl.fill = fill; lbl.border = e["border_kpi"]
        v = ws_inner.cell(row=row_val, column=col, value=val)
        v.font = font_val
        v.alignment = Alignment(horizontal="left", vertical="center")
        v.fill = fill; v.border = border_kpi_bottom
        if fmt:
            v.number_format = fmt
        lbl2 = ws_inner.cell(row=row_lbl, column=col + 1)
        lbl2.fill = fill; lbl2.border = e["border_kpi"]
        v2 = ws_inner.cell(row=row_val, column=col + 1)
        v2.fill = fill; v2.border = border_kpi_bottom
        ws_inner.merge_cells(start_row=row_lbl, start_column=col, end_row=row_lbl, end_column=col + 1)
        ws_inner.merge_cells(start_row=row_val, start_column=col, end_row=row_val, end_column=col + 1)

    for i, (label, val, font_val, fill, fmt) in enumerate(kpis):
        aplicar_kpi(ws, 6, 7, 1 + i * 2, label, val, font_val, fill, fmt)
    ws.row_dimensions[6].height = 18
    ws.row_dimensions[7].height = 30

    for i, (label, val, font_val, fill) in enumerate(sev_kpis):
        aplicar_kpi(ws, 9, 10, 1 + i * 2, label, val, font_val, fill)
    ws.row_dimensions[9].height = 18
    ws.row_dimensions[10].height = 30

    # Tabela por conta
    r = 12
    ws.cell(row=r, column=1, value="MOVIMENTAÇÃO POR CONTA").font = e["font_section"]
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    r += 1
    headers_conta = ["Conta", "Arquivo", "Transações", "Créditos", "Débitos", "Saldo", "% do Total"]
    for col, txt in enumerate(headers_conta, 1):
        ws.cell(row=r, column=col, value=txt)
    estilo_header([ws.cell(row=r, column=c) for c in range(1, len(headers_conta) + 1)])
    ws.row_dimensions[r].height = 24
    r += 1
    for i, ex in enumerate(extratos):
        cred = sum(t["valor"] for t in ex["transacoes"] if t["valor"] > 0)
        deb  = sum(t["valor"] for t in ex["transacoes"] if t["valor"] < 0)
        sld  = cred + deb
        pct  = (ex["qtd"] / total_tx) if total_tx else 0
        ws.cell(row=r, column=1, value=ex["conta"])
        ws.cell(row=r, column=2, value=ex["arquivo"])
        ws.cell(row=r, column=3, value=ex["qtd"])
        c = ws.cell(row=r, column=4, value=cred)
        c.number_format = e["FMT_BRL_POS"]; c.font = Font(color=e["GREEN"], name="Calibri")
        c = ws.cell(row=r, column=5, value=deb)
        c.number_format = e["FMT_BRL"];     c.font = Font(color=e["RED"], name="Calibri")
        c = ws.cell(row=r, column=6, value=sld)
        c.number_format = e["FMT_BRL"];     c.font = Font(bold=True, name="Calibri")
        c = ws.cell(row=r, column=7, value=pct); c.number_format = '0.0%'
        if i % 2 == 1:
            for col in range(1, len(headers_conta) + 1):
                ws.cell(row=r, column=col).fill = e["fill_zebra"]
        linha_borda(ws, r, len(headers_conta))
        ws.row_dimensions[r].height = 22
        r += 1

    # Top categorias contabeis
    stats = _top_categorias_e_contrapartes(extratos)
    cats = stats["cats"]
    if cats:
        r += 2
        ws.cell(row=r, column=1, value="🏷 DISTRIBUIÇÃO POR CATEGORIA").font = e["font_section"]
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        r += 1
        for col, txt in enumerate(["Categoria", "Qtd", "Valor Total", "Ticket Médio", "% Volume", "", "", ""], 1):
            if txt:
                ws.cell(row=r, column=col, value=txt)
        estilo_header([ws.cell(row=r, column=c) for c in range(1, 6)],
                      fill=e["fill_blue"], font=e["font_h_white"])
        ws.row_dimensions[r].height = 22
        r += 1
        vol_total = sum(abs(d["valor"]) for d in cats.values()) or 1
        for i, cat in enumerate(sorted(cats, key=lambda k: -abs(cats[k]["valor"]))):
            d = cats[cat]
            tk = d["valor"] / d["qtd"] if d["qtd"] else 0
            pct = abs(d["valor"]) / vol_total
            ws.cell(row=r, column=1, value=cat)
            ws.cell(row=r, column=2, value=d["qtd"])
            c = ws.cell(row=r, column=3, value=d["valor"]); c.number_format = e["FMT_BRL"]
            c.font = Font(color=e["GREEN"] if d["valor"] > 0 else e["RED"], name="Calibri")
            c = ws.cell(row=r, column=4, value=tk); c.number_format = e["FMT_BRL"]
            c = ws.cell(row=r, column=5, value=pct); c.number_format = '0.0%'
            if i % 2 == 1:
                for col in range(1, 6):
                    ws.cell(row=r, column=col).fill = e["fill_zebra"]
            for col in range(1, 6):
                ws.cell(row=r, column=col).border = e["border_all"]
            r += 1

    # Top contrapartes
    top_cps = sorted(stats["contrapartes"].items(), key=lambda x: -abs(x[1]["valor"]))[:10]
    if top_cps:
        r += 2
        ws.cell(row=r, column=1, value="🏆 TOP 10 CONTRAPARTES").font = e["font_section"]
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        r += 1
        for col, txt in enumerate(["#", "Contraparte (CNPJ/CPF/Nome)", "Transações", "Volume", "Tipo"], 1):
            ws.cell(row=r, column=col, value=txt)
        estilo_header([ws.cell(row=r, column=c) for c in range(1, 6)],
                      fill=e["fill_blue"], font=e["font_h_white"])
        ws.row_dimensions[r].height = 22
        r += 1
        for i, (chave, d) in enumerate(top_cps, 1):
            ws.cell(row=r, column=1, value=i)
            ws.cell(row=r, column=2, value=chave)
            ws.cell(row=r, column=3, value=d["qtd"])
            c = ws.cell(row=r, column=4, value=d["valor"]); c.number_format = e["FMT_BRL"]
            c.font = Font(color=e["GREEN"] if d["valor"] > 0 else e["RED"], bold=True, name="Calibri")
            ws.cell(row=r, column=5, value="Recebimento" if d["valor"] > 0 else "Pagamento")
            if i % 2 == 0:
                for col in range(1, 6):
                    ws.cell(row=r, column=col).fill = e["fill_zebra"]
            for col in range(1, 6):
                ws.cell(row=r, column=col).border = e["border_all"]
            r += 1

    # Top anomalias criticas
    crit_lista = [a for a in anomalias if a["severidade"] == "critico"][:10]
    if crit_lista:
        r += 2
        ws.cell(row=r, column=1, value="🔴 ACHADOS CRÍTICOS").font = e["font_section"]
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        r += 1
        for col, txt in enumerate(["Tipo", "Título", "Conta", "Valor", "Detalhe"], 1):
            ws.cell(row=r, column=col, value=txt)
        estilo_header([ws.cell(row=r, column=c) for c in range(1, 6)],
                      fill=e["fill_blue"], font=e["font_h_white"])
        ws.row_dimensions[r].height = 22
        r += 1
        for a in crit_lista:
            ws.cell(row=r, column=1, value=a["tipo"])
            ws.cell(row=r, column=2, value=a["titulo"])
            ws.cell(row=r, column=3, value=a["conta"])
            c = ws.cell(row=r, column=4, value=a.get("valor", 0)); c.number_format = e["FMT_BRL"]
            ws.cell(row=r, column=5, value=a["detalhe"])
            for col in range(1, 6):
                ws.cell(row=r, column=col).fill = e["fill_critico"]
                ws.cell(row=r, column=col).border = e["border_all"]
            r += 1

    for col, w in zip("ABCDEFGH", [26, 32, 13, 16, 16, 16, 13, 8]):
        ws.column_dimensions[col].width = w


def _xlsx_aba_transacoes(wb, extratos: list[dict], e: dict) -> None:
    """Cria e preenche a aba Transacoes."""
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet("Transações")
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 30
    cabec = ["Conta", "Data", "Tipo", "Valor", "Memo", "Nome", "Doc"]
    for col, txt in enumerate(cabec, 1):
        cell = ws.cell(row=1, column=col, value=txt)
        cell.fill = e["fill_blue_dark"]
        cell.font = e["font_h_white"]
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = e["border_all"]
    r = 2
    for ex in extratos:
        for t in ex["transacoes"]:
            ws.cell(row=r, column=1, value=ex["conta"])
            ws.cell(row=r, column=2, value=t["data"])
            tipo_cell = ws.cell(row=r, column=3, value=t["tipo"])
            if t["tipo"] == "CREDIT":
                tipo_cell.font = Font(color=e["GREEN"], bold=True, name="Calibri", size=10)
            elif t["tipo"] == "DEBIT":
                tipo_cell.font = Font(color=e["RED"], bold=True, name="Calibri", size=10)
            c = ws.cell(row=r, column=4, value=t["valor"])
            c.number_format = e["FMT_BRL"]
            c.font = Font(color=e["GREEN"] if t["valor"] > 0 else e["RED"],
                          name="Calibri", size=10, bold=True)
            ws.cell(row=r, column=5, value=t["memo"])
            ws.cell(row=r, column=6, value=t["nome"])
            ws.cell(row=r, column=7, value=t["checknum"])
            if r % 2 == 0:
                for col in range(1, len(cabec) + 1):
                    if col != 3 and col != 4:
                        ws.cell(row=r, column=col).fill = e["fill_zebra"]
            for col in range(1, len(cabec) + 1):
                ws.cell(row=r, column=col).border = e["border_all"]
            r += 1
    for col, w in zip("ABCDEFG", [26, 12, 10, 16, 52, 30, 14]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cabec))}{r-1}"


def _xlsx_aba_anomalias(wb, anomalias: list[dict], e: dict) -> None:
    """Cria e preenche a aba Anomalias."""
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet("Anomalias")
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 30
    cabec = ["Severidade", "Tipo", "Título", "Conta", "Valor", "Detalhe"]
    for col, txt in enumerate(cabec, 1):
        cell = ws.cell(row=1, column=col, value=txt)
        cell.fill = e["fill_blue_dark"]
        cell.font = e["font_h_white"]
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = e["border_all"]
    r = 2
    sev_meta = {
        "critico": ("🔴 CRÍTICO", e["fill_critico"],
                    Font(bold=True, color=e["RED"],    name="Calibri", size=10)),
        "alerta":  ("🟠 ALERTA",  e["fill_alerta"],
                    Font(bold=True, color=e["ORANGE"], name="Calibri", size=10)),
        "atencao": ("🟡 ATENÇÃO", e["fill_atencao"],
                    Font(bold=True, color=e["YELLOW"], name="Calibri", size=10)),
    }
    for a in anomalias:
        label, fill, font_sev = sev_meta.get(a["severidade"], ("?", None, None))
        sev_cell = ws.cell(row=r, column=1, value=label)
        if font_sev:
            sev_cell.font = font_sev
        ws.cell(row=r, column=2, value=a["tipo"])
        ws.cell(row=r, column=3, value=a["titulo"])
        ws.cell(row=r, column=4, value=a["conta"])
        c = ws.cell(row=r, column=5, value=a.get("valor", 0))
        c.number_format = e["FMT_BRL"]
        ws.cell(row=r, column=6, value=a["detalhe"])
        if fill:
            for col in range(1, len(cabec) + 1):
                ws.cell(row=r, column=col).fill = fill
        for col in range(1, len(cabec) + 1):
            ws.cell(row=r, column=col).border = e["border_all"]
        r += 1
    for col, w in zip("ABCDEF", [15, 22, 42, 28, 16, 60]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    if r > 2:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(cabec))}{r-1}"


def _gerar_xlsx(extratos: list[dict], anomalias: list[dict]) -> bytes:
    """Gera planilha XLSX com 3 abas estilizadas: Resumo, Transacoes, Anomalias."""
    wb = Workbook()
    e = _xlsx_estilos()
    _xlsx_aba_resumo(wb.active, extratos, anomalias, e)
    _xlsx_aba_transacoes(wb, extratos, e)
    _xlsx_aba_anomalias(wb, anomalias, e)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()



def _parse_pdf(content: bytes, filename: str) -> list[dict]:
    """Extrai transacoes de PDF de extrato bancario com 3 estrategias em fallback."""
    transacoes: list[dict] = []
    conta_default = f"PDF ({filename})"

    # Tenta extrair conta/agencia do header do PDF
    conta_detectada = None
    rx_conta = re.compile(
        r"(?:AG[EÊE]?N?CIA|AG[ÊE]?)\s*:?\s*(\d{3,5}[-\d]?)\s+"
        r"(?:CONTA|C\.?C\.?|CC)\s*:?\s*(\d{4,10}[-\d]?)",
        re.IGNORECASE,
    )

    # Regex para sinal explicito ao final (D / C)
    rx_sinal_dc = re.compile(
        r"(\d{2}/\d{2}/\d{4})\s+(.{5,80}?)\s+([\d.]+,\d{2})\s*([CD])\b",
        re.IGNORECASE,
    )
    # Regex padrao com sinal +/- ou R$
    rx_padrao = re.compile(
        r"(\d{2}/\d{2}/\d{4})\s+(.{5,80}?)\s+([+\-]?\s*R?\$?\s*[\d.]+,\d{2})"
    )
    # Regex compacta DD/MM com valor entre parenteses (negativo) ou nao
    rx_compacta = re.compile(
        r"(\d{2}/\d{2}/\d{2,4})\s+(.{3,80}?)\s+(\(?\s*[+\-]?\s*[\d.]+,\d{2}\s*\)?)"
    )

    keywords_debito = ("PAGTO", "DEBITO", "DÉBITO", "DEB ", "PIX EMITIDO", "PIX ENVIADO",
                       "SAQUE", "COMPRA", "TARIFA", "JUROS", "IOF", "BOLETO", "TED ENVIADA",
                       "DOC ENVIADO", "PAGAMENTO", "ESTORNO DEB", "RETIRADA")

    def parse_valor(s: str) -> Optional[float]:
        """Converte string '1.234,56' ou '(1.234,56)' ou '-1.234,56' em float."""
        s = s.strip()
        neg = s.startswith("(") and s.endswith(")") or s.startswith("-")
        s = s.strip("()").replace("R", "").replace("$", "").replace(" ", "")
        s = s.replace(".", "").replace(",", ".").lstrip("+-")
        try:
            v = float(s)
            return -v if neg else v
        except ValueError:
            return None

    def parse_data(s: str) -> Optional[str]:
        partes = s.split("/")
        if len(partes) != 3: return None
        dia, mes, ano = partes
        if len(ano) == 2: ano = "20" + ano
        if len(dia) != 2 or len(mes) != 2 or len(ano) != 4: return None
        return f"{ano}-{mes}-{dia}"

    try:
        with pdfplumber.open(io.BytesIO(content)) as pdf:
            for page in pdf.pages:
                text = page.extract_text() or ""
                if conta_detectada is None:
                    m_conta = rx_conta.search(text)
                    if m_conta:
                        ag, cc = m_conta.groups()
                        conta_detectada = f"AG {ag} / CC {cc}"

                vistos = set()  # evitar duplicar mesma transacao detectada por mais de um regex

                # Estrategia 1: data + desc + valor + sinal C/D explicito
                for m in rx_sinal_dc.finditer(text):
                    data_br, desc, valor_s, sinal = m.groups()
                    data_iso = parse_data(data_br)
                    valor = parse_valor(valor_s)
                    if not data_iso or valor is None:
                        continue
                    if sinal.upper() == "D":
                        valor = -abs(valor)
                    else:
                        valor = abs(valor)
                    chave = (data_iso, round(valor, 2), desc.strip()[:40])
                    if chave in vistos: continue
                    vistos.add(chave)
                    transacoes.append({
                        "conta": conta_detectada or conta_default,
                        "data": data_iso, "tipo": "CREDIT" if valor > 0 else "DEBIT",
                        "valor": valor, "memo": desc.strip(),
                        "nome": "", "checknum": "",
                    })

                # Estrategia 2: data + desc + valor (sem sinal explicito, heuristica por keyword)
                for m in rx_padrao.finditer(text):
                    data_br, desc, valor_s = m.groups()
                    data_iso = parse_data(data_br)
                    valor = parse_valor(valor_s)
                    if not data_iso or valor is None:
                        continue
                    desc_up = desc.upper()
                    # So aplica heuristica de sinal se nao tinha sinal explicito
                    if "+" not in valor_s and "-" not in valor_s and "(" not in valor_s:
                        if any(k in desc_up for k in keywords_debito):
                            valor = -abs(valor)
                    chave = (data_iso, round(valor, 2), desc.strip()[:40])
                    if chave in vistos: continue
                    vistos.add(chave)
                    transacoes.append({
                        "conta": conta_detectada or conta_default,
                        "data": data_iso, "tipo": "CREDIT" if valor > 0 else "DEBIT",
                        "valor": valor, "memo": desc.strip(),
                        "nome": "", "checknum": "",
                    })

                # Estrategia 3: fallback compacta (DD/MM/YY)
                if not transacoes:
                    for m in rx_compacta.finditer(text):
                        data_br, desc, valor_s = m.groups()
                        data_iso = parse_data(data_br)
                        valor = parse_valor(valor_s)
                        if not data_iso or valor is None:
                            continue
                        chave = (data_iso, round(valor, 2), desc.strip()[:40])
                        if chave in vistos: continue
                        vistos.add(chave)
                        transacoes.append({
                            "conta": conta_detectada or conta_default,
                            "data": data_iso, "tipo": "CREDIT" if valor > 0 else "DEBIT",
                            "valor": valor, "memo": desc.strip(),
                            "nome": "", "checknum": "",
                        })
    except Exception as e:
        log.exception("Erro parseando PDF %s", filename)
        raise HTTPException(status_code=400, detail=f"PDF invalido ou corrompido: {e}")

    log.info("PDF %s: %d transacoes extraidas", filename, len(transacoes))
    return transacoes


def _parse_xml(text: str, filename: str) -> list[dict]:
    """Extrai transacoes de XML (CAMT.053, padrao bancario brasileiro, ou OFX em XML)."""
    transacoes = []
    conta_default = f"XML ({filename})"
    try:
        root = ET.fromstring(text)
    except ET.ParseError:
        return []

    # Strip namespaces de toda a arvore (cria nova arvore sem namespaces)
    def _strip_ns(el):
        if "}" in el.tag:
            el.tag = el.tag.split("}", 1)[1]
        for child in el:
            _strip_ns(child)
    _strip_ns(root)

    conta = conta_default
    # Tenta achar identificacao da conta (preferindo Acct/Id/Othr/Id de CAMT)
    acct = None
    for xpath in (".//Acct/Id/Othr/Id", ".//Acct/Id", ".//ACCTID", ".//Id"):
        acct = root.find(xpath)
        if acct is not None:
            break
    if acct is not None and acct.text:
        conta = f"Conta {acct.text.strip()}"

    # CAMT.053: <Ntry> blocks
    for ntry in root.iter("Ntry"):
        amt = ntry.find("Amt")
        cdtdbt = ntry.find("CdtDbtInd")
        dt_el = ntry.find("BookgDt/Dt")
        if dt_el is None:
            dt_el = ntry.find("ValDt/Dt")
        dt = dt_el
        info = ntry.find(".//AddtlNtryInf")
        if info is None:
            info = ntry.find(".//RmtInf/Ustrd")
        if amt is None or cdtdbt is None or dt is None:
            continue
        try:
            valor = float(amt.text)
        except (TypeError, ValueError):
            continue
        if cdtdbt.text == "DBIT":
            valor = -abs(valor)
        transacoes.append({
            "conta": conta,
            "data": dt.text[:10],
            "tipo": "CREDIT" if valor > 0 else "DEBIT",
            "valor": valor,
            "memo": (info.text.strip() if info is not None and info.text else ""),
            "nome": "",
            "checknum": "",
        })

    # Se nao encontrou CAMT, tenta STMTTRN (OFX em XML)
    if not transacoes:
        for tr in root.iter("STMTTRN"):
            data = (tr.findtext("DTPOSTED") or "")[:8]
            data_iso = f"{data[:4]}-{data[4:6]}-{data[6:8]}" if len(data) == 8 else data
            try:
                valor = float(tr.findtext("TRNAMT") or 0)
            except ValueError:
                continue
            transacoes.append({
                "conta": conta,
                "data": data_iso,
                "tipo": tr.findtext("TRNTYPE") or "",
                "valor": valor,
                "memo": (tr.findtext("MEMO") or "").strip(),
                "nome": (tr.findtext("NAME") or "").strip(),
                "checknum": (tr.findtext("CHECKNUM") or "").strip(),
            })

    return transacoes


def _parse_arquivo(content: bytes, filename: str) -> list[dict]:
    """Detecta tipo do arquivo e roteia para o parser correto."""
    ext = Path(filename).suffix.lower()
    if ext == ".ofx":
        return _parse_ofx(content.decode("latin-1", errors="ignore"))
    if ext == ".pdf":
        return _parse_pdf(content, filename)
    if ext == ".xml":
        return _parse_xml(content.decode("utf-8", errors="ignore"), filename)
    raise HTTPException(
        status_code=400,
        detail=f"Extensao nao suportada: {ext}. Use .ofx, .pdf ou .xml",
    )


def _parse_ofx(text: str) -> list[dict]:
    """Parser OFX minimalista (SGML)."""
    branch_m = re.search(r"<BRANCHID>([^<\n]+)", text)
    acct_m = re.search(r"<ACCTID>([^<\n]+)", text)
    conta = f"AG {branch_m.group(1).strip() if branch_m else '?'} / CC {acct_m.group(1).strip() if acct_m else '?'}"
    transacoes = []
    for bloco in re.findall(r"<STMTTRN>(.*?)</STMTTRN>", text, flags=re.DOTALL):
        def fld(tag: str) -> str:
            m = re.search(rf"<{tag}>([^<\n]*)", bloco)
            return m.group(1).strip() if m else ""

        data_raw = fld("DTPOSTED")[:8]
        data = (
            f"{data_raw[:4]}-{data_raw[4:6]}-{data_raw[6:8]}"
            if len(data_raw) == 8 else data_raw
        )
        transacoes.append({
            "conta": conta,
            "data": data,
            "tipo": fld("TRNTYPE"),
            "valor": float(fld("TRNAMT") or 0),
            "memo": fld("MEMO"),
            "nome": fld("NAME"),
            "checknum": fld("CHECKNUM"),
        })
    return transacoes


def _fmt_csv(transacoes: list[dict]) -> str:
    linhas = ["data,tipo,valor,memo,nome,checknum"]
    for t in transacoes:
        memo = t["memo"].replace(",", " ").replace("\n", " ")
        nome = t["nome"].replace(",", " ").replace("\n", " ")
        linhas.append(
            f"{t['data']},{t['tipo']},{t['valor']:.2f},{memo},{nome},{t['checknum']}"
        )
    return "\n".join(linhas)


_REGRAS_ANTES_PIX: list[tuple[tuple[str, ...], str]] = [
    (("INTERCREDIS", "TRANSF.CONTAS", "TRANSF. CONTAS", "TRANSF MESMA TIT",
      "TRANSFERENCIA MESMA TITULARIDADE", "TRANSFERENCIA ENTRE CONTAS PROPRIAS"),
     "Transferencia entre contas proprias"),
    (("DAS ", "DARF", "RFB", "INSS", "FGTS", "DAE", "GPS", "GNRE", "DAR ",
      "IRRF", "IRPJ", "CSLL", "ICMS", "ISS", "GUIA"),
     "Tributo"),
    (("IOF",), "Despesa Financeira - IOF"),
    (("JUROS", "MORA"), "Despesa Financeira - Juros"),
    (("MULTA",), "Despesa Financeira - Multa"),
    (("PAGAMENTO TD", "LIBERACAO TD", "LIBERAÇÃO TD", "CRED.LIBERA",
      "DESCONTO TITULO", "CREDITO ROTATIVO", "ANTECIPACAO RECEBIVEL"),
     "Operacao de Credito - TD"),
    (("EMPRESTIMO", "EMPRÉSTIMO", "FINANCIAMENTO", "CDC", "PARCELA EMP"),
     "Pagamento de Emprestimo"),
    (("CHEQUE ESPECIAL", "LIMITE CONTA"), "Despesa Financeira - Cheque Especial"),
    (("SEGURO", "PRESTAMISTA", "PROTECAO", "PROTEÇÃO"), "Despesa - Seguro"),
]
_REGRAS_APOS_PIX: list[tuple[tuple[str, ...], str]] = [
    (("COMPRA MASTERCARD", "COMPRA VISA", "COMPRA CARTAO", "COMPRA ELO",
      "COMPRA HIPERCARD", "COMPRA AMEX", "COMPRA DEBITO", "DEBITO COMPRA"),
     "Compra Cartao"),
    (("FATURA CARTAO", "PAGTO FATURA", "PAGAMENTO CARTAO CRED"), "Pagamento Fatura Cartao"),
    (("PEDAGIO", "PEDÁGIO", "SICOOB TAG", "SEM PARAR", "MOVE MAIS", "CONECTCAR"),
     "Despesa - Pedagio"),
    (("POSTO ", "COMBUSTIVEL", "GASOLINA", "ETANOL", "DIESEL", "SHELL", "IPIRANGA"),
     "Despesa - Combustivel"),
    (("TARIFA", "MENSALIDADE", "ANUIDADE", "CESTA ", "PACOTE ", "MANUTENCAO",
      "MANUTENÇÃO CONTA"),
     "Despesa Bancaria - Tarifa"),
    (("BOLETO", "COBRAN", "COMPE", "COMPENSADO", "TITULO PAGO"), "Pagamento Boleto"),
    (("SALARIO", "SALÁRIO", "FOLHA PGTO", "PAGAMENTO FOLHA", "PROVENTO", "ADIANTAMENTO SAL"),
     "Folha de Pagamento"),
    (("PRO LABORE", "PRÓ-LABORE", "PRO-LABORE", "RETIRADA SOCIO"),
     "Pro-Labore / Retirada Socio"),
    (("ALUGUEL", "CONDOMINIO", "CONDOMÍNIO"), "Despesa - Aluguel/Condominio"),
    (("ENERGIA ELETRICA", "ENERGIA ELÉTRICA", "ENEL", "CEMIG", "COELBA", "COPEL",
      "CELPE", "CELESC", "ELEKTRO", "LIGHT", "EQUATORIAL"),
     "Despesa - Energia Eletrica"),
    (("AGUA", "ÁGUA", "SABESP", "CEDAE", "COPASA", "EMBASA", "SANEPAR"), "Despesa - Agua"),
    (("TELEFON", "VIVO", "CLARO", "OI ", "TIM ", "INTERNET", "OPERADORA"),
     "Despesa - Telecom"),
    (("SAQUE", "RETIRADA"), "Saque"),
    (("DEPOSITO", "DEPÓSITO"), "Deposito em Dinheiro"),
    (("ESTORNO", "DEVOLUC"), "Estorno"),
]


def _classificar(memo: str, nome: str) -> str:
    """Classificacao contabil heuristica multi-banco (Sicoob, BB, Itau, Bradesco, Santander, Caixa, Inter, Nubank, C6)."""
    s = f"{memo} {nome}".upper()
    match = lambda *t: any(x in s for x in t)

    for termos, cat in _REGRAS_ANTES_PIX:
        if any(t in s for t in termos):
            return cat

    if "PIX" in s:
        if match("EMITIDO", "ENVIADO", "PAGAMENTO PIX", "PIX SAIDA", "DEBITO PIX"):
            return "Pagamento PIX - Fornecedor/Despesa"
        if match("RECEB", "CREDITO PIX", "CRÉDITO PIX", "PIX ENTRADA", "PIX RECEBIDO"):
            return "Receita PIX"
        return "PIX - A classificar"

    if match("TED ", "DOC "):
        if match("RECEB", "CREDITO", "CRÉDITO"):
            return "Receita TED/DOC"
        return "Pagamento TED/DOC"

    for termos, cat in _REGRAS_APOS_PIX:
        if any(t in s for t in termos):
            return cat

    return "A classificar"



def _detectar_anomalias(extratos: list[dict]) -> list[dict]:
    """Identifica anomalias com severidade (critico/alerta/atencao)."""
    from collections import Counter

    anomalias = []

    # Duplicidades
    for e in extratos:
        contagem = Counter(
            (t["data"], round(t["valor"], 2), t["memo"][:40]) for t in e["transacoes"]
        )
        for (data, valor, memo), n in contagem.items():
            if n < 2:
                continue
            sev = "critico" if n >= 3 else "alerta"
            anomalias.append({
                "severidade": sev,
                "tipo": "Duplicidade",
                "titulo": f"{n}x lançamento idêntico em {data}",
                "conta": e["conta"],
                "valor": valor,
                "detalhe": f"R$ {valor:,.2f} | {memo} | {n} ocorrências",
            })

    # Transacoes atipicas (>R$ 50k = critico, >R$ 10k = atencao)
    for e in extratos:
        for t in e["transacoes"]:
            v = abs(t["valor"])
            memo = (t["memo"] or t["nome"])[:60]
            if v > 50000:
                anomalias.append({
                    "severidade": "alerta",
                    "tipo": "Valor alto",
                    "titulo": f"Transação de R$ {v:,.2f}",
                    "conta": e["conta"],
                    "valor": t["valor"],
                    "detalhe": f"{t['data']} | {memo}",
                })
            elif v > 10000:
                anomalias.append({
                    "severidade": "atencao",
                    "tipo": "Valor alto",
                    "titulo": f"Transação de R$ {v:,.2f}",
                    "conta": e["conta"],
                    "valor": t["valor"],
                    "detalhe": f"{t['data']} | {memo}",
                })

    # Estornos
    for e in extratos:
        for t in e["transacoes"]:
            s = (t["memo"] + t["nome"]).upper()
            if "ESTORNO" in s:
                anomalias.append({
                    "severidade": "critico",
                    "tipo": "Estorno",
                    "titulo": "Operação estornada",
                    "conta": e["conta"],
                    "valor": t["valor"],
                    "detalhe": f"{t['data']} | R$ {t['valor']:,.2f} | {t['memo'][:60]}",
                })

    # Transferências internas sem par — verifica todos os pares de contas
    _KEYWORDS_TRANSF = ("INTERCREDIS", "TRANSF.CONTAS", "TRANSF MESMA TIT", "TRANSFERENCIA ENTRE CONTAS")
    if len(extratos) >= 2:
        from itertools import combinations
        for c1, c2 in combinations(extratos, 2):
            def _eh_transf(t):
                s = (t["memo"] + t["nome"]).upper()
                return any(k in s for k in _KEYWORDS_TRANSF)
            tx1 = [t for t in c1["transacoes"] if _eh_transf(t)]
            tx2 = [t for t in c2["transacoes"] if _eh_transf(t)]
            usados = set()
            casados = 0
            for t1 in tx1:
                for j, t2 in enumerate(tx2):
                    if j in usados:
                        continue
                    if abs(abs(t1["valor"]) - abs(t2["valor"])) < 0.01 and t1["valor"] * t2["valor"] < 0:
                        usados.add(j); casados += 1; break
            sem_par = (len(tx1) - casados) + (len(tx2) - casados)
            if sem_par > 0:
                anomalias.append({
                    "severidade": "alerta",
                    "tipo": "Transferencia sem par",
                    "titulo": f"{sem_par} transferência(s) interna(s) sem par",
                    "conta": f"{c1['conta']} ↔ {c2['conta']}",
                    "valor": 0,
                    "detalhe": (
                        f"{c1['conta']}: {len(tx1) - casados} sem par | "
                        f"{c2['conta']}: {len(tx2) - casados} sem par"
                    ),
                })

    # Ordena: critico > alerta > atencao, depois por |valor|
    ordem = {"critico": 0, "alerta": 1, "atencao": 2}
    anomalias.sort(key=lambda a: (ordem[a["severidade"]], -abs(a.get("valor", 0))))
    return anomalias


def _top_categorias_e_contrapartes(extratos: list[dict]) -> dict:
    """Retorna estatisticas: categorias, top contrapartes, evolucao diaria."""
    from collections import defaultdict, Counter
    import re as _re

    cats = defaultdict(lambda: {"qtd": 0, "valor": 0.0, "transacoes": []})
    contrapartes = defaultdict(lambda: {"qtd": 0, "valor": 0.0})
    diario = defaultdict(lambda: {"cred": 0.0, "deb": 0.0})

    rx_cnpj = _re.compile(r"(\d{2}\.\d{3}\.\d{3}[/ ]?\d{4}[- ]?\d{2})")
    rx_cpf = _re.compile(r"(\*{3}\.\d{3}\.\d{3}-\*{2})")

    for e in extratos:
        for t in e["transacoes"]:
            c = _classificar(t["memo"], t["nome"])
            cats[c]["qtd"] += 1
            cats[c]["valor"] += t["valor"]
            cats[c]["transacoes"].append(t)

            # Detectar contraparte (CNPJ/CPF no memo+nome)
            texto = f"{t['memo']} {t['nome']}"
            m_cnpj = rx_cnpj.search(texto)
            m_cpf = rx_cpf.search(texto)
            chave = None
            if m_cnpj:
                chave = m_cnpj.group(1).strip()
            elif m_cpf:
                chave = m_cpf.group(1).strip()
            elif t["nome"]:
                chave = t["nome"][:30].strip()
            if chave:
                contrapartes[chave]["qtd"] += 1
                contrapartes[chave]["valor"] += t["valor"]

            # Evolucao diaria
            if t["valor"] > 0:
                diario[t["data"]]["cred"] += t["valor"]
            else:
                diario[t["data"]]["deb"] += t["valor"]

    return {"cats": dict(cats), "contrapartes": dict(contrapartes), "diario": dict(diario)}


def _conciliacao_local(extratos: list[dict], anomalias: list[dict]) -> str:
    """Gera relatorio de conciliacao deterministicamente (sem LLM) — versao enriquecida."""
    from collections import Counter, defaultdict
    from datetime import datetime

    out = ["# Relatório de Conciliação Bancária\n"]
    out.append(f"**Gerado em:** {datetime.now().strftime('%d/%m/%Y %H:%M')}  \n")
    out.append(f"**Modo:** Simulação local (sem LLM)  \n")
    out.append(f"**Extratos analisados:** {len(extratos)}\n\n")

    stats = _top_categorias_e_contrapartes(extratos)

    # Achados criticos no topo
    crit = [a for a in anomalias if a["severidade"] == "critico"]
    alerta = [a for a in anomalias if a["severidade"] == "alerta"]
    atencao = [a for a in anomalias if a["severidade"] == "atencao"]
    out.append("## ⚠️ Achados de Anomalias\n\n")
    out.append(f"- 🔴 **Críticos:** {len(crit)}\n")
    out.append(f"- 🟠 **Alertas:** {len(alerta)}\n")
    out.append(f"- 🟡 **Atenção:** {len(atencao)}\n\n")
    if crit:
        out.append("### 🔴 Críticos\n\n")
        for a in crit:
            out.append(f"- **[{a['tipo']}]** {a['titulo']} — {a['conta']}\n  {a['detalhe']}\n")
        out.append("\n")
    if alerta:
        out.append("### 🟠 Alertas\n\n")
        for a in alerta[:15]:
            out.append(f"- **[{a['tipo']}]** {a['titulo']} — {a['conta']}\n  {a['detalhe']}\n")
        if len(alerta) > 15:
            out.append(f"- _...e mais {len(alerta) - 15} alerta(s)_\n")
        out.append("\n")

    # === 1. RESUMO EXECUTIVO ===
    out.append("## 1. Resumo Executivo\n\n")
    total_tx = sum(e["qtd"] for e in extratos)
    total_cred = sum(t["valor"] for e in extratos for t in e["transacoes"] if t["valor"] > 0)
    total_deb = sum(t["valor"] for e in extratos for t in e["transacoes"] if t["valor"] < 0)
    total_liq = total_cred + total_deb
    ticket_medio = (total_cred + abs(total_deb)) / max(total_tx, 1)
    # Periodo
    datas = sorted({t["data"] for e in extratos for t in e["transacoes"] if t["data"]})
    periodo = f"{datas[0]} a {datas[-1]}" if datas else "indefinido"
    out.append(f"**Período analisado:** {periodo}  \n")
    out.append(f"**Total de transações:** {total_tx}  \n")
    out.append(f"**Ticket médio:** R$ {ticket_medio:,.2f}  \n")
    out.append(f"**Volume bruto movimentado:** R$ {total_cred + abs(total_deb):,.2f}\n\n")

    out.append("| Conta | Transações | Créditos | Débitos | Saldo Líquido | % Volume |\n")
    out.append("|---|---:|---:|---:|---:|---:|\n")
    vol_total = total_cred + abs(total_deb)
    for e in extratos:
        cred = sum(t["valor"] for t in e["transacoes"] if t["valor"] > 0)
        deb = sum(t["valor"] for t in e["transacoes"] if t["valor"] < 0)
        liq = cred + deb
        vol_e = cred + abs(deb)
        pct = (vol_e / vol_total * 100) if vol_total else 0
        out.append(
            f"| {e['conta']} | {e['qtd']} | R$ {cred:,.2f} | R$ {deb:,.2f} | "
            f"**R$ {liq:,.2f}** | {pct:.1f}% |\n"
        )
    out.append(f"| **CONSOLIDADO** | **{total_tx}** | **R$ {total_cred:,.2f}** | "
               f"**R$ {total_deb:,.2f}** | **R$ {total_liq:,.2f}** | 100,0% |\n\n")

    # === 2. INDICADORES OPERACIONAIS ===
    out.append("## 2. Indicadores Operacionais\n\n")

    sev_count = {"critico": len(crit), "alerta": len(alerta), "atencao": len(atencao)}
    saude = "🟢 Boa" if sev_count["critico"] == 0 else (
        "🟡 Atenção" if sev_count["critico"] <= 2 else "🔴 Crítica"
    )

    n_dias = len(datas) or 1
    media_diaria_tx = total_tx / n_dias
    cats_count = sum(1 for k in stats["cats"] if k != "A classificar")
    pct_classif = (
        (sum(d["qtd"] for k, d in stats["cats"].items() if k != "A classificar") / total_tx * 100)
        if total_tx else 0
    )

    out.append("| Indicador | Valor |\n|---|---:|\n")
    out.append(f"| Saúde da conciliação | {saude} |\n")
    out.append(f"| Dias com movimento | {n_dias} |\n")
    out.append(f"| Média de transações/dia | {media_diaria_tx:.1f} |\n")
    out.append(f"| Categorias contábeis detectadas | {cats_count} |\n")
    out.append(f"| Cobertura de classificação | {pct_classif:.1f}% |\n")
    out.append(f"| Total de anomalias | {len(anomalias)} ({sev_count['critico']} críticas) |\n\n")

    # === 3. TRANSFERENCIAS ENTRE CONTAS ===
    out.append("## 3. Transferências entre Contas Próprias\n\n")
    _KEYWORDS_TRANSF = ("INTERCREDIS", "TRANSF.CONTAS", "TRANSF MESMA TIT", "TRANSFERENCIA ENTRE CONTAS")
    def _eh_transf(t):
        s = (t["memo"] + t["nome"]).upper()
        return any(k in s for k in _KEYWORDS_TRANSF)

    if len(extratos) >= 2:
        from itertools import combinations as _combis
        total_pares_encontrados = 0
        total_volume = 0.0
        out.append("| Data | Origem | Destino | Valor | Status |\n|---|---|---|---:|:-:|\n")
        for c1, c2 in _combis(extratos, 2):
            tx1 = [t for t in c1["transacoes"] if _eh_transf(t)]
            tx2 = [t for t in c2["transacoes"] if _eh_transf(t)]
            usados = set()
            for t1 in tx1:
                for j, t2 in enumerate(tx2):
                    if j in usados:
                        continue
                    if abs(abs(t1["valor"]) - abs(t2["valor"])) < 0.01 and t1["valor"] * t2["valor"] < 0:
                        origem = c1["conta"] if t1["valor"] < 0 else c2["conta"]
                        destino = c2["conta"] if t1["valor"] < 0 else c1["conta"]
                        v = abs(t1["valor"])
                        out.append(f"| {t1['data']} | {origem} | {destino} | R$ {v:,.2f} | ✅ CASADO |\n")
                        total_volume += v
                        total_pares_encontrados += 1
                        usados.add(j)
                        break
        if total_pares_encontrados:
            out.append(f"\n**Resumo:** {total_pares_encontrados} par(es) conciliado(s) · "
                       f"Volume total R$ {total_volume:,.2f}\n\n")
        else:
            out.append("\n_Nenhuma transferência entre contas detectada._\n\n")
    else:
        out.append("_Apenas 1 extrato enviado — cruzamento entre contas não aplicável._\n\n")

    # === 4. TOP CATEGORIAS CONTÁBEIS ===
    out.append("## 4. Distribuição por Categoria Contábil\n\n")
    cats = stats["cats"]
    out.append("| Categoria | Qtd | Valor Total | Ticket Médio | % do Volume |\n|---|---:|---:|---:|---:|\n")
    for cat in sorted(cats, key=lambda k: -abs(cats[k]["valor"])):
        d = cats[cat]
        tk = d["valor"] / d["qtd"] if d["qtd"] else 0
        pct = (abs(d["valor"]) / vol_total * 100) if vol_total else 0
        out.append(f"| {cat} | {d['qtd']} | R$ {d['valor']:,.2f} | R$ {tk:,.2f} | {pct:.1f}% |\n")
    out.append("\n")

    # === 5. TOP CONTRAPARTES ===
    out.append("## 5. Top Contrapartes (Pareto)\n\n")
    cps = stats["contrapartes"]
    top_cps = sorted(cps.items(), key=lambda x: -abs(x[1]["valor"]))[:12]
    if top_cps:
        out.append("| # | Contraparte | Transações | Volume | Tipo |\n|---:|---|---:|---:|:-:|\n")
        for i, (chave, d) in enumerate(top_cps, 1):
            tipo = "💚 Recebimento" if d["valor"] > 0 else "🔻 Pagamento"
            out.append(f"| {i} | `{chave}` | {d['qtd']} | R$ {d['valor']:,.2f} | {tipo} |\n")
        out.append("\n")
    else:
        out.append("_Sem contrapartes identificáveis por CNPJ/CPF._\n\n")

    # === 6. DUPLICIDADES DETALHADAS ===
    out.append("## 6. Duplicidades Detectadas\n\n")
    achou_dup = False
    for e in extratos:
        contagem = Counter(
            (t["data"], round(t["valor"], 2), t["memo"][:40]) for t in e["transacoes"]
        )
        dups = [k for k, n in contagem.items() if n > 1]
        if dups:
            achou_dup = True
            out.append(f"### {e['conta']}\n\n")
            out.append("| Data | Valor | Memo | Ocorrências | Impacto Total |\n|---|---:|---|:-:|---:|\n")
            for data, valor, memo in sorted(dups, key=lambda k: -contagem[k]):
                n = contagem[(data, valor, memo)]
                impacto = valor * n
                emoji = "🔴" if n >= 3 else "🟠"
                out.append(f"| {data} | R$ {valor:,.2f} | `{memo}` | {emoji} **{n}x** | R$ {impacto:,.2f} |\n")
            out.append("\n")
    if not achou_dup:
        out.append("✅ _Nenhuma duplicidade detectada._\n\n")

    # === 7. TRANSACOES ATIPICAS ===
    out.append("## 7. Transações Atípicas (> R$ 10.000)\n\n")
    atipicas = []
    for e in extratos:
        for t in e["transacoes"]:
            if abs(t["valor"]) > 10000:
                atipicas.append((e["conta"], t))
    if atipicas:
        atipicas.sort(key=lambda x: -abs(x[1]["valor"]))
        out.append("| Conta | Data | Valor | Memo | Classificação |\n|---|---|---:|---|---|\n")
        for conta, t in atipicas[:20]:
            memo = (t["memo"] or t["nome"])[:55]
            cat = _classificar(t["memo"], t["nome"])
            out.append(f"| {conta} | {t['data']} | **R$ {t['valor']:,.2f}** | {memo} | {cat} |\n")
        if len(atipicas) > 20:
            out.append(f"\n_...e mais {len(atipicas) - 20} transação(ões) atípica(s)._\n")
        out.append("\n")
    else:
        out.append("✅ _Nenhuma transação acima de R$ 10.000._\n\n")

    # === 8. EVOLUCAO DIARIA ===
    diario = stats["diario"]
    if len(diario) > 1:
        out.append("## 8. Evolução Diária do Fluxo\n\n")
        out.append("| Data | Créditos | Débitos | Saldo do Dia |\n|---|---:|---:|---:|\n")
        for data in sorted(diario.keys()):
            d = diario[data]
            sld = d["cred"] + d["deb"]
            out.append(f"| {data} | R$ {d['cred']:,.2f} | R$ {d['deb']:,.2f} | R$ {sld:,.2f} |\n")
        out.append("\n")

    # === 9. PLANO DE AÇÃO ===
    out.append("## 9. Plano de Ação Recomendado\n\n")
    if crit:
        out.append("### 🔴 Imediato (24-48h)\n\n")
        for a in crit[:5]:
            out.append(f"1. **{a['titulo']}** — {a['conta']}  \n   _{a['detalhe']}_\n")
        out.append("\n")
    if alerta:
        out.append("### 🟠 Curto prazo (esta semana)\n\n")
        for a in alerta[:5]:
            out.append(f"1. Investigar: **{a['titulo']}** — {a['conta']}\n")
        out.append("\n")

    a_classificar = cats.get("A classificar", {}).get("qtd", 0)
    if a_classificar > 0:
        out.append(f"### 🟡 Médio prazo\n\n")
        out.append(f"- Classificar manualmente **{a_classificar} transação(ões)** sem regra automática\n")
        out.append(f"- Refinar regras de classificação para reduzir cobertura abaixo de 100%\n\n")

    out.append("### ✅ Boas práticas\n\n")
    out.append("- Implementar conciliação diária (vs. mensal) para detectar duplicidades cedo\n")
    out.append("- Confirmar comprovantes de PIX acima de R$ 1.000\n")
    out.append("- Revisar estornos com o banco antes do fechamento contábil\n")
    out.append("- Documentar transferências entre contas próprias com referência cruzada\n")

    return "".join(out)


@app.get("/")
def root():
    return {
        "service": "Conciliacao Bancaria API",
        "version": "0.1.0",
        "endpoints": ["/health", "/conciliar/ofx", "/conciliar/csv"],
    }


@app.get("/health")
async def health():
    db_status = "nao_configurado"
    if DB_DISPONIVEL:
        try:
            async with SessionLocal() as db:
                await db.execute(sql_text("SELECT 1"))
            db_status = "ok"
        except Exception:
            db_status = "erro"
    return {
        "status": "ok",
        "versao": "0.3.0",
        "api_key_configured": bool(os.environ.get("ANTHROPIC_API_KEY")),
        "banco_dados": db_status,
    }


# ── Clientes ──────────────────────────────────────────────────────────────

@app.post("/clientes", dependencies=[Depends(auth)], status_code=201, tags=["clientes"])
@limiter.limit("20/minute")
async def criar_cliente(request: Request, payload: ClienteCreate):
    """Cadastra um novo cliente."""
    if not DB_DISPONIVEL:
        raise HTTPException(503, "Banco de dados nao configurado — adicione DATABASE_URL ao .env")
    async with SessionLocal() as db:
        cliente = await crud_clientes.criar_cliente(
            db, nome=payload.nome, cnpj=payload.cnpj,
            email=payload.email, telefone=payload.telefone, plano=payload.plano,
        )
    return {
        "id": str(cliente.id), "nome": cliente.nome, "cnpj": cliente.cnpj,
        "email": cliente.email, "plano": cliente.plano,
        "criado_em": cliente.criado_em.isoformat(),
    }


@app.get("/clientes", dependencies=[Depends(auth)], tags=["clientes"])
@limiter.limit("30/minute")
async def listar_clientes(request: Request, apenas_ativos: bool = True):
    """Lista todos os clientes."""
    if not DB_DISPONIVEL:
        raise HTTPException(503, "Banco de dados nao configurado")
    async with SessionLocal() as db:
        clientes = await crud_clientes.listar_clientes(db, apenas_ativos=apenas_ativos)
    return [
        {"id": str(c.id), "nome": c.nome, "cnpj": c.cnpj, "email": c.email,
         "plano": c.plano, "ativo": c.ativo}
        for c in clientes
    ]


@app.get("/clientes/{cliente_id}", dependencies=[Depends(auth)], tags=["clientes"])
@limiter.limit("30/minute")
async def buscar_cliente(request: Request, cliente_id: str):
    """Busca um cliente pelo ID."""
    if not DB_DISPONIVEL:
        raise HTTPException(503, "Banco de dados nao configurado")
    import uuid as _uuid
    try:
        cid = _uuid.UUID(cliente_id)
    except ValueError:
        raise HTTPException(400, "ID invalido")
    async with SessionLocal() as db:
        cliente = await crud_clientes.buscar_cliente(db, cid)
    if not cliente:
        raise HTTPException(404, "Cliente nao encontrado")
    return {
        "id": str(cliente.id), "nome": cliente.nome, "cnpj": cliente.cnpj,
        "email": cliente.email, "telefone": cliente.telefone,
        "plano": cliente.plano, "ativo": cliente.ativo,
        "criado_em": cliente.criado_em.isoformat(),
    }


@app.patch("/clientes/{cliente_id}", dependencies=[Depends(auth)], tags=["clientes"])
@limiter.limit("20/minute")
async def atualizar_cliente(request: Request, cliente_id: str, payload: ClienteUpdate):
    """Atualiza dados de um cliente."""
    if not DB_DISPONIVEL:
        raise HTTPException(503, "Banco de dados nao configurado")
    import uuid as _uuid
    try:
        cid = _uuid.UUID(cliente_id)
    except ValueError:
        raise HTTPException(400, "ID invalido")
    campos = {k: v for k, v in payload.model_dump().items() if v is not None}
    async with SessionLocal() as db:
        cliente = await crud_clientes.atualizar_cliente(db, cid, **campos)
    if not cliente:
        raise HTTPException(404, "Cliente nao encontrado")
    return {"id": str(cliente.id), "nome": cliente.nome, "plano": cliente.plano, "ativo": cliente.ativo}


# ── Conciliação ───────────────────────────────────────────────────────────

@app.post("/conciliar/ofx", dependencies=[Depends(auth)])
@limiter.limit("20/minute")
async def conciliar_ofx(
    request: Request,
    arquivos: List[UploadFile] = File(..., description="1 a 50 arquivos (.ofx, .pdf ou .xml)"),
    max_tokens: int = 16000,
    simular: bool = False,
    cliente_id: Optional[str] = None,
):
    """Cruza ate 50 extratos bancarios (OFX, PDF ou XML).

    Aceita arquivos de multiplos bancos e periodos simultaneamente.
    Use simular=true para gerar relatorio local (sem chamar a API).
    """
    if not (1 <= len(arquivos) <= 50):
        raise HTTPException(status_code=400, detail="Envie entre 1 e 50 arquivos")

    extratos_parsed = []
    for up in arquivos:
        content = await read_limited(up)
        try:
            txs = _parse_arquivo(content, up.filename)
        except HTTPException:
            raise
        except Exception as e:
            log.exception("Falha parseando %s", up.filename)
            raise HTTPException(
                status_code=400,
                detail=f"Falha ao parsear {up.filename}: {type(e).__name__}",
            )
        if not txs:
            raise HTTPException(
                status_code=400,
                detail=f"Nao foi possivel extrair transacoes de {up.filename}",
            )
        extratos_parsed.append({
            "arquivo": up.filename,
            "conta": txs[0]["conta"],
            "qtd": len(txs),
            "transacoes": txs,
        })

    if simular:
        anomalias = _detectar_anomalias(extratos_parsed)
        relatorio = _conciliacao_local(extratos_parsed, anomalias)
        rid = _salvar_dataset(extratos_parsed, anomalias, relatorio)
        await _salvar_no_banco(rid, extratos_parsed, anomalias, "simulacao", cliente_id)
        return JSONResponse({
            "modo": "simulacao_local",
            "report_id": rid,
            "extratos": [
                {"arquivo": e["arquivo"], "conta": e["conta"], "qtd": e["qtd"]}
                for e in extratos_parsed
            ],
            "anomalias": anomalias,
            "relatorio_md": relatorio,
            "relatorio_html": _render_html(relatorio),
        })

    blocos = []
    for e in extratos_parsed:
        blocos.append(
            f"=== {e['conta']} ({e['arquivo']}) ===\n"
            f"Total: {e['qtd']} transacoes\n{_fmt_csv(e['transacoes'])}"
        )

    n_contas = len(extratos_parsed)
    prompt = (
        f"Analise os {n_contas} extrato(s) bancario(s) abaixo. "
        "Identifique transferencias entre contas proprias (INTERCREDIS/TED entre as mesmas contas), "
        "duplicidades, transacoes atipicas e pre-classifique para lancamento contabil. "
        "Consolide o fluxo de caixa considerando todas as contas em conjunto. "
        "Gere relatorio em portugues em Markdown.\n\n"
        + "\n\n".join(blocos)
    )

    client = _get_client()
    try:
        resp = client.messages.create(
            model="claude-sonnet-4-5",
            max_tokens=max_tokens,
            system=SYSTEM_PROMPT,
            messages=[{"role": "user", "content": prompt}],
        )
    except APIStatusError as e:
        body = getattr(e, "body", None) or {}
        msg = (body.get("error") or {}).get("message") or str(e)
        # Mensagens amigaveis em portugues para erros comuns
        if "credit balance" in msg.lower():
            friendly = ("Saldo de creditos Anthropic esgotado. "
                        "Recarregue em https://platform.claude.com/settings/billing "
                        "ou use ?simular=true para gerar relatorio local.")
        elif "rate" in msg.lower() and "limit" in msg.lower():
            friendly = "Rate limit da Anthropic atingido. Aguarde alguns segundos."
        else:
            friendly = msg
        log.warning("Anthropic API error %s: %s", e.status_code, msg)
        raise HTTPException(status_code=e.status_code, detail={"anthropic_error": friendly})
    relatorio = "\n".join(b.text for b in resp.content if b.type == "text")
    anomalias = _detectar_anomalias(extratos_parsed)
    rid = _salvar_dataset(extratos_parsed, anomalias, relatorio)
    await _salvar_no_banco(rid, extratos_parsed, anomalias, "llm", cliente_id)

    return JSONResponse({
        "modo": "claude_llm",
        "report_id": rid,
        "extratos": [
            {"arquivo": e["arquivo"], "conta": e["conta"], "qtd": e["qtd"]}
            for e in extratos_parsed
        ],
        "anomalias": anomalias,
        "usage": {
            "input_tokens": resp.usage.input_tokens,
            "output_tokens": resp.usage.output_tokens,
        },
        "stop_reason": resp.stop_reason,
        "relatorio_md": relatorio,
        "relatorio_html": _render_html(relatorio),
    })


@app.get("/logo-base64")
def logo_base64():
    """Devolve a logo como data URI (usado pelo frontend para PDF)."""
    return {"data_uri": _LOGO_DATA_URI}


@app.get("/export/html/{rid}", dependencies=[Depends(auth)])
def export_html(rid: str):
    """Baixa o relatorio renderizado em HTML standalone."""
    ds = _carregar_dataset(rid)
    html = _render_html(ds["relatorio"])
    return Response(
        content=html,
        media_type="text/html",
        headers={"Content-Disposition": f'attachment; filename="conciliacao_{rid}.html"'},
    )


@app.get("/export/xlsx/{rid}", dependencies=[Depends(auth)])
def export_xlsx(rid: str):
    """Baixa o dataset em XLSX (3 abas: Resumo, Transacoes, Anomalias)."""
    ds = _carregar_dataset(rid)
    blob = _gerar_xlsx(ds["extratos"], ds["anomalias"])
    return Response(
        content=blob,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="conciliacao_{rid}.xlsx"'},
    )


@app.post("/conciliar/csv")
async def conciliar_csv(
    extrato: UploadFile = File(..., description="CSV do extrato bancario"),
    razao: UploadFile = File(..., description="CSV do razao contabil"),
    max_tokens: int = 16000,
):
    """Cruza extrato bancario CSV contra razao contabil CSV."""
    extrato_text = (await extrato.read()).decode("utf-8", errors="ignore")
    razao_text = (await razao.read()).decode("utf-8", errors="ignore")

    prompt = (
        "Realize a conciliacao bancaria entre o extrato e o razao contabil "
        "abaixo. Liste conciliados, divergencias, duplicidades e pendencias.\n\n"
        f"=== EXTRATO BANCARIO ({extrato.filename}) ===\n{extrato_text}\n\n"
        f"=== RAZAO CONTABIL ({razao.filename}) ===\n{razao_text}"
    )

    client = _get_client()
    try:
        resp = client.messages.create(
            model="claude-sonnet-4-5",
            max_tokens=max_tokens,
            system=SYSTEM_PROMPT,
            messages=[{"role": "user", "content": prompt}],
        )
    except APIStatusError as e:
        body = getattr(e, "body", None) or {}
        msg = (body.get("error") or {}).get("message") or str(e)
        raise HTTPException(status_code=e.status_code, detail={"anthropic_error": msg})
    relatorio = "\n".join(b.text for b in resp.content if b.type == "text")

    return JSONResponse({
        "extrato": extrato.filename,
        "razao": razao.filename,
        "usage": {
            "input_tokens": resp.usage.input_tokens,
            "output_tokens": resp.usage.output_tokens,
        },
        "stop_reason": resp.stop_reason,
        "relatorio_md": relatorio,
    })
