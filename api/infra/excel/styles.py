"""Paleta de cores, fontes e bordas compartilhados entre as abas XLSX.

Extraido de api/services/excel.py (item 28). Para retrocompat, o arquivo
original re-exporta `_xlsx_estilos` daqui.
"""
from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl.styles import Border, Font, PatternFill, Side


LOGO_PATH = Path(__file__).resolve().parent.parent.parent.parent / "static" / "logo.png"


def estilos_xlsx() -> dict[str, Any]:
    """Retorna dict com paleta + fonts + borders + format codes."""
    BLUE_DARK = "0A3A7A"
    BLUE = "1E6FD9"
    WHITE = "FFFFFF"
    GRAY_BORDER = "E2E8F0"
    GRAY_LIGHT = "F7FAFC"
    GRAY_HOVER = "EFF6FF"
    RED = "DC2626"; RED_BG = "FEE2E2"
    ORANGE = "EA580C"; ORANGE_BG = "FFEDD5"
    YELLOW = "CA8A04"; YELLOW_BG = "FEF9C3"
    GREEN = "16A34A"

    side_thin = Side(border_style="thin", color=GRAY_BORDER)

    return dict(
        # Cores soltas (para uso ad-hoc)
        BLUE_DARK=BLUE_DARK, BLUE=BLUE, WHITE=WHITE,
        RED=RED, ORANGE=ORANGE, YELLOW=YELLOW, GREEN=GREEN,

        # Fills
        fill_blue_dark=PatternFill("solid", fgColor=BLUE_DARK),
        fill_blue=PatternFill("solid", fgColor=BLUE),
        fill_zebra=PatternFill("solid", fgColor=GRAY_LIGHT),
        fill_kpi_blue=PatternFill("solid", fgColor=GRAY_HOVER),
        fill_critico=PatternFill("solid", fgColor=RED_BG),
        fill_alerta=PatternFill("solid", fgColor=ORANGE_BG),
        fill_atencao=PatternFill("solid", fgColor=YELLOW_BG),

        # Fonts
        font_h_white=Font(bold=True, color=WHITE, size=11, name="Calibri"),
        font_brand=Font(bold=True, size=24, color=BLUE_DARK, name="Calibri"),
        font_brand_sub=Font(color=BLUE, size=10, italic=True, name="Calibri"),
        font_section=Font(bold=True, size=13, color=BLUE_DARK, name="Calibri"),
        font_kpi_lbl=Font(bold=True, size=9, color="64748B", name="Calibri"),
        font_kpi_val_red=Font(bold=True, size=22, color=RED, name="Calibri"),
        font_kpi_val_orange=Font(bold=True, size=22, color=ORANGE, name="Calibri"),
        font_kpi_val_yellow=Font(bold=True, size=22, color=YELLOW, name="Calibri"),
        font_kpi_val_blue=Font(bold=True, size=22, color=BLUE_DARK, name="Calibri"),

        # Borders
        side_thin=side_thin,
        border_all=Border(left=side_thin, right=side_thin, top=side_thin, bottom=side_thin),
        border_kpi=Border(
            left=side_thin, right=side_thin,
            top=Side(border_style="medium", color=BLUE),
            bottom=side_thin,
        ),

        # Format codes BR
        FMT_BRL='R$ #,##0.00;[Red]-R$ #,##0.00',
        FMT_BRL_POS='R$ #,##0.00',
    )
