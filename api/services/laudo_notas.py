"""Laudo de Documentos Fiscais (NF-e/CT-e/NFS-e) — relatório baseado SOMENTE nos
documentos fiscais, sem extrato bancário (OFX).

Útil quando se tem os XMLs mas não o extrato, ou para auditar a base documental
isoladamente. Gera um XLSX com:
  1. Resumo        — KPIs + alertas (chaves inválidas, emitentes não-ativos)
  2. Por Fornecedor— emitentes por volume + situação cadastral
  3. Natureza/CFOP — distribuição por natureza de operação e por CFOP
  4. Por Dia       — distribuição temporal
  5. Notas         — detalhe documento a documento
  6. Alertas       — chaves inválidas + canceladas/denegadas + emitentes não-ativos

`situacao_por_cnpj` é opcional (dict CNPJ->situação cadastral) — o router pode
derivá-lo do cache de CNPJ; sem ele, a coluna cadastral fica vazia (sem rede).
"""
from __future__ import annotations

import html as _html
from collections import Counter, defaultdict
from typing import Iterable, Optional

from openpyxl import Workbook
from openpyxl.styles import Border, Font, PatternFill, Side

from api.matchers.xml_fiscal import DocumentoFiscalLido

_NAVY = "0F172A"
_HDR_FILL = PatternFill("solid", fgColor=_NAVY)
_HDR_FONT = Font(color="FFFFFF", bold=True)
_TITULO = Font(bold=True, size=14, color=_NAVY)
_BOLD = Font(bold=True)
_RED = PatternFill("solid", fgColor="FDE2E2")
_AMBER = PatternFill("solid", fgColor="FEF3C7")
_THIN = Border(*[Side(style="thin", color="E2E8F0")] * 4)
_MONEY = "#,##0.00"

ABAS_NOTAS = [
    "1. Capa", "2. Por Fornecedor", "3. Natureza CFOP",
    "4. Por Dia", "5. Notas", "6. Alertas",
]


def _hdr(ws, row: int, cols: list[str]) -> None:
    for c, t in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=t)
        cell.fill = _HDR_FILL
        cell.font = _HDR_FONT
        cell.border = _THIN


def _larguras(ws, larguras: dict[str, int]) -> None:
    for col, w in larguras.items():
        ws.column_dimensions[col].width = w


def _situacao_nao_ativa(sit: str) -> bool:
    return bool(sit) and "ATIVA" not in sit.upper()


def gerar_laudo_notas_workbook(
    documentos: Iterable[DocumentoFiscalLido],
    situacao_por_cnpj: Optional[dict[str, str]] = None,
) -> tuple[Workbook, dict]:
    """Gera o XLSX do laudo de notas. Retorna (workbook, stats)."""
    docs = [d for d in documentos if d.chave]
    sit_map = situacao_por_cnpj or {}

    def sit_cad(cnpj: str) -> str:
        return sit_map.get(cnpj) or sit_map.get(cnpj[:8], "") if cnpj else ""

    total = sum(d.valor_total for d in docs)
    icms = sum(d.valor_icms for d in docs)
    pis = sum(d.valor_pis for d in docs)
    cofins = sum(d.valor_cofins for d in docs)
    iss = sum(d.valor_iss for d in docs)
    por_situacao = Counter(d.situacao for d in docs)
    invalidas = [d for d in docs if d.chave_valida is False]
    canceladas = [d for d in docs if d.situacao in ("CANCELADA", "DENEGADA")]
    nao_ativos = [d for d in docs if _situacao_nao_ativa(sit_cad(d.emit_cnpj))]
    datas = sorted(d.data_emissao for d in docs if d.data_emissao)

    # Estilo padrão ORGATEC (mesmo do laudo forense) — import lazy p/ não pesar o load.
    from api.services.laudo_forense import (
        ALERT_FILL, ALERT_FILL_MEDIO, SUBTITLE_FONT, ZEBRA_FILL,
        aba_capa, cabecalho_padrao, style_header,
    )

    periodo_str = (f"{datas[0]} a {datas[-1]}") if datas else "-"
    linha2 = f"Base documental fiscal · Periodo {periodo_str} · {len(docs)} documentos"
    n_alertas_tot = len(invalidas) + len(canceladas) + len(nao_ativos)

    def _cab(ws, ncols, secao):
        return cabecalho_padrao(ws, ncols, titulo="Laudo de Documentos Fiscais",
                                linha2=linha2, secao=secao)

    def _zebra(ws, row, ncols):
        if row % 2 == 0:
            for c in range(1, ncols + 1):
                ws.cell(row=row, column=c).fill = ZEBRA_FILL

    wb = Workbook()

    # ── Aba 1: Capa (índice + sumário executivo) ─────────────────────────
    aba_capa(
        wb,
        titulo_relatorio="Laudo de Documentos Fiscais",
        linha2=linha2,
        subtitulo="Sistema OrgConc · Auditoria Fiscal",
        objeto="Análise da base documental (NF-e/CT-e/NFS-e): volume por fornecedor, natureza de "
               "operação, CFOP, situação cadastral e alertas.",
        secoes=[
            ("2", "Por Fornecedor", "Emitentes por volume + situacao cadastral", "2. Por Fornecedor"),
            ("3", "Natureza / CFOP", "Distribuicao por natureza de operacao e CFOP", "3. Natureza CFOP"),
            ("4", "Por Dia", "Distribuicao temporal", "4. Por Dia"),
            ("5", "Notas", "Detalhe documento-a-documento", "5. Notas"),
            ("6", "Alertas", "Chaves invalidas, canceladas, emitentes nao-ativos", "6. Alertas"),
        ],
        sumario=[
            ("Periodo analisado", periodo_str),
            ("Total de documentos", f"{len(docs):,}"),
            ("Volume total documentado", _money(total)),
            ("ICMS total", _money(icms)),
            ("Fornecedores distintos", str(len({d.emit_cnpj for d in docs if d.emit_cnpj}))),
            ("Canceladas / denegadas", str(por_situacao.get("CANCELADA", 0) + por_situacao.get("DENEGADA", 0))),
            ("Chaves invalidas (mod-11)", str(len(invalidas))),
            ("Alertas no total", str(n_alertas_tot)),
        ],
    )

    # ── Aba 2: Por Fornecedor ────────────────────────────────────────────
    ws = wb.create_sheet("2. Por Fornecedor")
    agg: dict[str, dict] = defaultdict(lambda: {"n": 0, "vol": 0.0, "icms": 0.0, "nome": ""})
    for d in docs:
        a = agg[d.emit_cnpj]
        a["n"] += 1
        a["vol"] += d.valor_total
        a["icms"] += d.valor_icms
        a["nome"] = d.emit_nome or a["nome"]
    r0 = _cab(ws, 6, "Fornecedores por volume + situacao cadastral")
    cols = ["CNPJ", "Razao Social", "Qtd", "Volume (R$)", "ICMS (R$)", "Situacao Cadastral"]
    for c, t in enumerate(cols, 1):
        ws.cell(row=r0, column=c, value=t)
    style_header(ws, r0, 6)
    r = r0 + 1
    for cnpj, a in sorted(agg.items(), key=lambda kv: -kv[1]["vol"]):
        sc = sit_cad(cnpj)
        ws.cell(row=r, column=1, value=cnpj)
        ws.cell(row=r, column=2, value=(a["nome"] or "")[:40])
        ws.cell(row=r, column=3, value=a["n"])
        ws.cell(row=r, column=4, value=round(a["vol"], 2)).number_format = _MONEY
        ws.cell(row=r, column=5, value=round(a["icms"], 2)).number_format = _MONEY
        ws.cell(row=r, column=6, value=sc or "(sem cadastro)")
        _zebra(ws, r, 6)
        if _situacao_nao_ativa(sc):
            for col in range(1, 7):
                ws.cell(row=r, column=col).fill = ALERT_FILL_MEDIO
        r += 1
    _larguras(ws, {"A": 18, "B": 42, "C": 8, "D": 16, "E": 14, "F": 20})
    ws.freeze_panes = ws.cell(row=r0 + 1, column=1).coordinate

    # ── Aba 3: Natureza de Operacao + CFOP ───────────────────────────────
    ws = wb.create_sheet("3. Natureza CFOP")
    nat: dict[str, dict] = defaultdict(lambda: {"n": 0, "vol": 0.0})
    for d in docs:
        key = (d.natureza_operacao or "(nao informado)").upper()[:60]
        nat[key]["n"] += 1
        nat[key]["vol"] += d.valor_total
    r0 = _cab(ws, 4, "Distribuicao por natureza de operacao e CFOP")
    for c, t in enumerate(["Natureza de Operacao", "Qtd", "Volume (R$)", "% Volume"], 1):
        ws.cell(row=r0, column=c, value=t)
    style_header(ws, r0, 4)
    r = r0 + 1
    for k, a in sorted(nat.items(), key=lambda kv: -kv[1]["vol"]):
        ws.cell(row=r, column=1, value=k)
        ws.cell(row=r, column=2, value=a["n"])
        ws.cell(row=r, column=3, value=round(a["vol"], 2)).number_format = _MONEY
        ws.cell(row=r, column=4, value=round(a["vol"] / total * 100, 1) if total else 0).number_format = "0.0"
        _zebra(ws, r, 4)
        r += 1
    cfop_count: Counter = Counter()
    for d in docs:
        for c in (d.cfops or ([d.cfop] if d.cfop else [])):
            cfop_count[c] += 1
    r += 2
    ws.cell(row=r, column=1, value="DISTRIBUICAO POR CFOP (itens)").font = SUBTITLE_FONT
    r += 1
    for c, t in enumerate(["CFOP", "Qtd documentos"], 1):
        ws.cell(row=r, column=c, value=t)
    style_header(ws, r, 2)
    r += 1
    for cfop, n in cfop_count.most_common():
        ws.cell(row=r, column=1, value=cfop)
        ws.cell(row=r, column=2, value=n)
        r += 1
    _larguras(ws, {"A": 62, "B": 16, "C": 16, "D": 10})

    # ── Aba 4: Por Dia ───────────────────────────────────────────────────
    ws = wb.create_sheet("4. Por Dia")
    dia: dict[str, dict] = defaultdict(lambda: {"n": 0, "vol": 0.0})
    for d in docs:
        k = d.data_emissao or "-"
        dia[k]["n"] += 1
        dia[k]["vol"] += d.valor_total
    r0 = _cab(ws, 3, "Distribuicao temporal")
    for c, t in enumerate(["Data", "Qtd", "Volume (R$)"], 1):
        ws.cell(row=r0, column=c, value=t)
    style_header(ws, r0, 3)
    r = r0 + 1
    for k in sorted(dia):
        ws.cell(row=r, column=1, value=k)
        ws.cell(row=r, column=2, value=dia[k]["n"])
        ws.cell(row=r, column=3, value=round(dia[k]["vol"], 2)).number_format = _MONEY
        _zebra(ws, r, 3)
        r += 1
    _larguras(ws, {"A": 14, "B": 10, "C": 16})

    # ── Aba 5: Notas (detalhe) ───────────────────────────────────────────
    ws = wb.create_sheet("5. Notas")
    r0 = _cab(ws, 14, "Detalhe documento-a-documento")
    cols = ["Chave", "Tipo", "Num", "Serie", "Data", "Emit CNPJ", "Emit Nome",
            "Valor", "ICMS", "PIS", "COFINS", "CFOP", "Situacao", "Chave OK"]
    for c, t in enumerate(cols, 1):
        ws.cell(row=r0, column=c, value=t)
    style_header(ws, r0, 14)
    r = r0 + 1
    for d in sorted(docs, key=lambda d: (d.data_emissao, d.emit_nome or "")):
        chave_ok = "OK" if d.chave_valida else ("X" if d.chave_valida is False else "-")
        vals = [d.chave, d.tipo, d.numero, d.serie, d.data_emissao, d.emit_cnpj,
                (d.emit_nome or "")[:30], d.valor_total, d.valor_icms, d.valor_pis,
                d.valor_cofins, d.cfop, d.situacao, chave_ok]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            if c in (8, 9, 10, 11):
                cell.number_format = _MONEY
        if d.chave_valida is False or d.situacao in ("CANCELADA", "DENEGADA"):
            for c in range(1, 15):
                ws.cell(row=r, column=c).fill = ALERT_FILL
        else:
            _zebra(ws, r, 14)
        r += 1
    _larguras(ws, dict(zip("ABCDEFGHIJKLMN",
              [46, 7, 8, 6, 12, 16, 32, 14, 12, 10, 12, 12, 12, 9])))
    ws.freeze_panes = ws.cell(row=r0 + 1, column=1).coordinate

    # ── Aba 6: Alertas ───────────────────────────────────────────────────
    ws = wb.create_sheet("6. Alertas")
    r0 = _cab(ws, 5, "Documentos que exigem atencao")
    for c, t in enumerate(["Tipo Alerta", "Chave", "Emitente", "Valor (R$)", "Detalhe"], 1):
        ws.cell(row=r0, column=c, value=t)
    style_header(ws, r0, 5)
    r = r0 + 1

    def _linha_alerta(tipo, d, detalhe, fill):
        nonlocal r
        ws.cell(row=r, column=1, value=tipo).font = _BOLD
        ws.cell(row=r, column=2, value=d.chave)
        ws.cell(row=r, column=3, value=(d.emit_nome or "")[:30])
        ws.cell(row=r, column=4, value=round(d.valor_total, 2)).number_format = _MONEY
        ws.cell(row=r, column=5, value=detalhe)
        for c in range(1, 6):
            ws.cell(row=r, column=c).fill = fill
        r += 1

    for d in invalidas:
        _linha_alerta("CHAVE INVALIDA", d, "DV mod-11 nao confere", ALERT_FILL)
    for d in canceladas:
        _linha_alerta(d.situacao, d, "documento sem validade fiscal", ALERT_FILL)
    for d in nao_ativos:
        _linha_alerta("EMITENTE NAO-ATIVO", d, sit_cad(d.emit_cnpj), ALERT_FILL_MEDIO)
    if r == r0 + 1:
        ws.cell(row=r, column=1,
                value="Nenhum alerta — todas as notas validas, ativas e nao-canceladas.").font = _BOLD
    _larguras(ws, {"A": 20, "B": 46, "C": 32, "D": 14, "E": 30})

    stats = {
        "total_documentos": len(docs),
        "volume_total": round(total, 2),
        "icms": round(icms, 2), "pis": round(pis, 2), "cofins": round(cofins, 2),
        "por_situacao": dict(por_situacao),
        "chaves_invalidas": len(invalidas),
        "canceladas": len(canceladas),
        "emitentes_nao_ativos": len(nao_ativos),
        "fornecedores": len({d.emit_cnpj for d in docs if d.emit_cnpj}),
        "periodo": (datas[0], datas[-1]) if datas else None,
    }
    return wb, stats


# ─────────────────────────────────────────────────────────────────────────────
# Versão HTML (e, via Playwright, PDF) — mesmos dados do XLSX
# ─────────────────────────────────────────────────────────────────────────────

def _money(x: float) -> str:
    return "R$ " + format(x or 0.0, ",.2f")


def _mdcell(s) -> str:
    """Escapa uma célula de tabela Markdown (markdown deixa HTML cru passar)."""
    return _html.escape(str(s)).replace("|", "/").replace("\n", " ")


def gerar_laudo_notas_html(
    documentos: Iterable[DocumentoFiscalLido],
    situacao_por_cnpj: Optional[dict[str, str]] = None,
) -> tuple[str, dict]:
    """Versão HTML do laudo de notas, no MESMO visual do laudo forense
    (capa + CSS Playfair/Source Sans + assinatura). Gera o conteúdo em Markdown
    e envelopa via laudo_forense.gerar_html. Retorna (html, stats)."""
    from api.services.laudo_forense import gerar_html  # evita import pesado no load

    docs = [d for d in documentos if d.chave]
    sit_map = situacao_por_cnpj or {}

    def sit_cad(cnpj: str) -> str:
        return (sit_map.get(cnpj) or sit_map.get(cnpj[:8], "")) if cnpj else ""

    total = sum(d.valor_total for d in docs)
    icms = sum(d.valor_icms for d in docs)
    por_situacao = Counter(d.situacao for d in docs)
    invalidas = [d for d in docs if d.chave_valida is False]
    canceladas = [d for d in docs if d.situacao in ("CANCELADA", "DENEGADA")]
    nao_ativos = [d for d in docs if _situacao_nao_ativa(sit_cad(d.emit_cnpj))]
    datas = sorted(d.data_emissao for d in docs if d.data_emissao)
    tipos = Counter(d.tipo for d in docs)
    n_alertas = len(invalidas) + len(canceladas) + len(nao_ativos)
    periodo = (datas[0] + " a " + datas[-1]) if datas else "-"

    agg: dict[str, dict] = defaultdict(lambda: {"n": 0, "vol": 0.0, "nome": ""})
    for d in docs:
        a = agg[d.emit_cnpj]
        a["n"] += 1
        a["vol"] += d.valor_total
        a["nome"] = d.emit_nome or a["nome"]
    top_forn = sorted(agg.items(), key=lambda kv: -kv[1]["vol"])[:25]

    nat: dict[str, dict] = defaultdict(lambda: {"n": 0, "vol": 0.0})
    for d in docs:
        k = (d.natureza_operacao or "(nao informado)").upper()[:50]
        nat[k]["n"] += 1
        nat[k]["vol"] += d.valor_total
    top_nat = sorted(nat.items(), key=lambda kv: -kv[1]["vol"])[:10]

    cfop_count: Counter = Counter()
    for d in docs:
        for c in (d.cfops or ([d.cfop] if d.cfop else [])):
            cfop_count[c] += 1

    # ---- corpo em Markdown (envelopado pelo template do forense) ----
    L = []
    L.append("## 1. Resumo Executivo")
    L.append("")
    L.append("| Indicador | Valor |")
    L.append("|---|---:|")
    L.append("| Documentos fiscais | " + format(len(docs), ",") + " |")
    L.append("| Tipos | " + _mdcell(", ".join(t + " " + str(n) for t, n in tipos.items())) + " |")
    L.append("| Volume documentado | " + _money(total) + " |")
    L.append("| ICMS | " + _money(icms) + " |")
    L.append("| Fornecedores | " + str(len({d.emit_cnpj for d in docs if d.emit_cnpj})) + " |")
    L.append("| Canceladas/denegadas | " + str(por_situacao.get("CANCELADA", 0) + por_situacao.get("DENEGADA", 0)) + " |")
    L.append("| Chaves invalidas (mod-11) | " + str(len(invalidas)) + " |")
    L.append("| **Alertas** | **" + str(n_alertas) + "** |")
    L.append("")
    L.append("## 2. Top Fornecedores (por volume)")
    L.append("")
    L.append("| CNPJ | Razao Social | Qtd | Volume | Situacao |")
    L.append("|---|---|---:|---:|---|")
    for cnpj, a in top_forn:
        L.append("| " + _mdcell(cnpj) + " | " + _mdcell((a["nome"] or "")[:38]) + " | " + str(a["n"])
                 + " | " + _money(a["vol"]) + " | " + _mdcell(sit_cad(cnpj) or "(sem cadastro)") + " |")
    L.append("")
    L.append("## 3. Natureza de Operacao")
    L.append("")
    L.append("| Natureza | Qtd | Volume | % |")
    L.append("|---|---:|---:|---:|")
    for k, a in top_nat:
        pct = (a["vol"] / total * 100) if total else 0
        L.append("| " + _mdcell(k) + " | " + str(a["n"]) + " | " + _money(a["vol"]) + " | "
                 + format(pct, ".1f") + "% |")
    L.append("")
    L.append("## 4. Distribuicao por CFOP")
    L.append("")
    L.append("| CFOP | Qtd documentos |")
    L.append("|---|---:|")
    for c, n in cfop_count.most_common(12):
        L.append("| " + _mdcell(c) + " | " + str(n) + " |")
    L.append("")
    L.append("## 5. Alertas")
    L.append("")
    if n_alertas == 0:
        L.append("*Nenhum alerta — todas as notas validas, ativas e nao-canceladas.*")
    else:
        L.append("| Alerta | Chave | Emitente | Valor |")
        L.append("|---|---|---|---:|")
        for d in invalidas[:50]:
            L.append("| CHAVE INVALIDA | " + _mdcell(d.chave) + " | " + _mdcell((d.emit_nome or "")[:30])
                     + " | " + _money(d.valor_total) + " |")
        for d in canceladas[:50]:
            L.append("| " + _mdcell(d.situacao) + " | " + _mdcell(d.chave) + " | "
                     + _mdcell((d.emit_nome or "")[:30]) + " | " + _money(d.valor_total) + " |")
        for d in nao_ativos[:50]:
            L.append("| EMITENTE NAO-ATIVO | " + _mdcell(d.chave) + " | " + _mdcell((d.emit_nome or "")[:30])
                     + " | " + _money(d.valor_total) + " |")
    L.append("")
    L.append("---")
    L.append("*Validacao de chave estrutural (mod-11) — nao prova autenticidade (exige SEFAZ). "
             "Documentos cancelados/denegados excluidos da cobertura. Detalhe documento-a-documento no XLSX.*")
    md = "\n".join(L)

    html = gerar_html(
        md, periodo=periodo,
        titulo="Laudo de<br>Documentos Fiscais",
        subtitulo="Sistema OrgConc · Auditoria Fiscal",
        objeto="Análise da base documental (NF-e/CT-e/NFS-e): volume por fornecedor, natureza de "
               "operação, CFOP, situação cadastral e alertas (chaves inválidas, canceladas, emitentes não-ativos).",
        razao="Base documental fiscal",
        cnpj="—",
    )

    stats = {
        "total_documentos": len(docs),
        "volume_total": round(total, 2),
        "fornecedores": len({d.emit_cnpj for d in docs if d.emit_cnpj}),
        "chaves_invalidas": len(invalidas),
        "canceladas": len(canceladas),
    }
    return html, stats
