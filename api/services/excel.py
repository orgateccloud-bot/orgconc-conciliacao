"""Geracao de planilha XLSX com 3 abas (Resumo, Transacoes, Anomalias).

NOTE (item 28 do roadmap, em andamento): Modularizacao incremental.
- `_xlsx_estilos`  -> movido para `api/infra/excel/styles.py` (re-exportado abaixo).
- (futuro) abas    -> a quebrar em `api/infra/excel/aba_*.py`.

Por ora, este arquivo continua sendo o ponto de entrada via `_gerar_xlsx`.
"""
from __future__ import annotations

import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Re-exporta para nao quebrar imports existentes (`from api.services.excel import _xlsx_estilos`)
from api.infra.excel.styles import LOGO_PATH as _LOGO_PATH  # noqa: F401
from api.infra.excel.styles import estilos_xlsx as _xlsx_estilos  # noqa: F401

from api.parsers import _top_categorias_e_contrapartes


def _xlsx_aba_resumo(ws, extratos: list[dict], anomalias: list[dict], e: dict) -> None:
    """Preenche a aba Resumo com cabecalho, KPIs e tabelas."""
    ws.title = "Resumo"
    ws.sheet_view.showGridLines = False

    def estilo_header(cells, fill=None, font=None):
        fill = fill or e["fill_blue_dark"]
        font = font or e["font_h_white"]
        for c in cells:
            c.fill = fill; c.font = font
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border = e["border_all"]

    def linha_borda(ws_inner, row, cols):
        for c in range(1, cols + 1):
            ws_inner.cell(row=row, column=c).border = e["border_all"]

    if _LOGO_PATH.exists():
        try:
            img = XLImage(str(_LOGO_PATH))
            img.width = 64; img.height = 64
            ws.add_image(img, "A1")
        except Exception:
            pass
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 24
    ws.row_dimensions[3].height = 10

    ws["B1"] = "ORGATEC"
    ws["B1"].font = e["font_brand"]
    ws["B2"] = "Contabilidade & Auditoria"
    ws["B2"].font = e["font_brand_sub"]
    ws.merge_cells("B1:E1"); ws.merge_cells("B2:E2")

    ws["F1"] = "RELATÓRIO DE CONCILIAÇÃO"
    ws["F1"].font = Font(bold=True, color=e["BLUE_DARK"], size=11, name="Calibri")
    ws["F1"].alignment = Alignment(horizontal="right", vertical="bottom")
    ws.merge_cells("F1:H1")
    ws["F2"] = f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["F2"].font = Font(italic=True, color="64748B", size=10, name="Calibri")
    ws["F2"].alignment = Alignment(horizontal="right", vertical="top")
    ws.merge_cells("F2:H2")

    for col in range(1, 9):
        ws.cell(row=3, column=col).fill = e["fill_blue"]
    ws.row_dimensions[3].height = 4

    ws.cell(row=5, column=1, value="VISÃO GERAL").font = e["font_section"]
    ws.merge_cells("A5:H5")

    total_tx   = sum(ex["qtd"] for ex in extratos)
    total_cred = sum(t["valor"] for ex in extratos for t in ex["transacoes"] if t["valor"] > 0)
    total_deb  = sum(t["valor"] for ex in extratos for t in ex["transacoes"] if t["valor"] < 0)
    saldo = total_cred + total_deb

    sev_count = {"critico": 0, "alerta": 0, "atencao": 0}
    for a in anomalias:
        sev_count[a["severidade"]] = sev_count.get(a["severidade"], 0) + 1

    kpis = [
        ("TRANSAÇÕES",  total_tx,   e["font_kpi_val_blue"], e["fill_kpi_blue"], None),
        ("CRÉDITOS",    total_cred,
         Font(bold=True, size=18, color=e["GREEN"], name="Calibri"),
         e["fill_kpi_blue"], e["FMT_BRL_POS"]),
        ("DÉBITOS",     total_deb,
         Font(bold=True, size=18, color=e["RED"], name="Calibri"),
         e["fill_kpi_blue"], e["FMT_BRL"]),
        ("SALDO",       saldo,
         Font(bold=True, size=18, color=e["BLUE_DARK"], name="Calibri"),
         e["fill_kpi_blue"], e["FMT_BRL"]),
    ]
    sev_kpis = [
        ("🔴 CRÍTICAS", sev_count["critico"], e["font_kpi_val_red"],    e["fill_critico"]),
        ("🟠 ALERTAS",  sev_count["alerta"],  e["font_kpi_val_orange"], e["fill_alerta"]),
        ("🟡 ATENÇÃO",  sev_count["atencao"], e["font_kpi_val_yellow"], e["fill_atencao"]),
        ("✅ TOTAL",    len(anomalias),        e["font_kpi_val_blue"],   e["fill_kpi_blue"]),
    ]

    border_kpi_bottom = Border(left=e["side_thin"], right=e["side_thin"], bottom=e["side_thin"])

    def aplicar_kpi(ws_inner, row_lbl, row_val, col, label, val, font_val, fill, fmt=None):
        lbl = ws_inner.cell(row=row_lbl, column=col, value=label)
        lbl.font = e["font_kpi_lbl"]
        lbl.alignment = Alignment(horizontal="left", vertical="bottom")
        lbl.fill = fill; lbl.border = e["border_kpi"]
        v = ws_inner.cell(row=row_val, column=col, value=val)
        v.font = font_val
        v.alignment = Alignment(horizontal="left", vertical="center")
        v.fill = fill; v.border = border_kpi_bottom
        if fmt:
            v.number_format = fmt
        lbl2 = ws_inner.cell(row=row_lbl, column=col + 1)
        lbl2.fill = fill; lbl2.border = e["border_kpi"]
        v2 = ws_inner.cell(row=row_val, column=col + 1)
        v2.fill = fill; v2.border = border_kpi_bottom
        ws_inner.merge_cells(start_row=row_lbl, start_column=col, end_row=row_lbl, end_column=col + 1)
        ws_inner.merge_cells(start_row=row_val, start_column=col, end_row=row_val, end_column=col + 1)

    for i, (label, val, font_val, fill, fmt) in enumerate(kpis):
        aplicar_kpi(ws, 6, 7, 1 + i * 2, label, val, font_val, fill, fmt)
    ws.row_dimensions[6].height = 18
    ws.row_dimensions[7].height = 30

    for i, (label, val, font_val, fill) in enumerate(sev_kpis):
        aplicar_kpi(ws, 9, 10, 1 + i * 2, label, val, font_val, fill)
    ws.row_dimensions[9].height = 18
    ws.row_dimensions[10].height = 30

    r = 12
    ws.cell(row=r, column=1, value="MOVIMENTAÇÃO POR CONTA").font = e["font_section"]
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    r += 1
    headers_conta = ["Conta", "Arquivo", "Transações", "Créditos", "Débitos", "Saldo", "% do Total"]
    for col, txt in enumerate(headers_conta, 1):
        ws.cell(row=r, column=col, value=txt)
    estilo_header([ws.cell(row=r, column=c) for c in range(1, len(headers_conta) + 1)])
    ws.row_dimensions[r].height = 24
    r += 1
    for i, ex in enumerate(extratos):
        cred = sum(t["valor"] for t in ex["transacoes"] if t["valor"] > 0)
        deb  = sum(t["valor"] for t in ex["transacoes"] if t["valor"] < 0)
        sld  = cred + deb
        vol_e = cred + abs(deb)
        vol_total_contas = sum(
            sum(t["valor"] for t in ex2["transacoes"] if t["valor"] > 0)
            + abs(sum(t["valor"] for t in ex2["transacoes"] if t["valor"] < 0))
            for ex2 in extratos
        ) or 1
        pct = vol_e / vol_total_contas
        ws.cell(row=r, column=1, value=ex["conta"])
        ws.cell(row=r, column=2, value=ex["arquivo"])
        ws.cell(row=r, column=3, value=ex["qtd"])
        c = ws.cell(row=r, column=4, value=cred)
        c.number_format = e["FMT_BRL_POS"]; c.font = Font(color=e["GREEN"], name="Calibri")
        c = ws.cell(row=r, column=5, value=deb)
        c.number_format = e["FMT_BRL"];     c.font = Font(color=e["RED"], name="Calibri")
        c = ws.cell(row=r, column=6, value=sld)
        c.number_format = e["FMT_BRL"];     c.font = Font(bold=True, name="Calibri")
        c = ws.cell(row=r, column=7, value=pct); c.number_format = '0.0%'
        if i % 2 == 1:
            for col in range(1, len(headers_conta) + 1):
                ws.cell(row=r, column=col).fill = e["fill_zebra"]
        linha_borda(ws, r, len(headers_conta))
        ws.row_dimensions[r].height = 22
        r += 1

    stats = _top_categorias_e_contrapartes(extratos)
    cats = stats["cats"]
    if cats:
        r += 2
        ws.cell(row=r, column=1, value="🏷 DISTRIBUIÇÃO POR CATEGORIA").font = e["font_section"]
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        r += 1
        for col, txt in enumerate(["Categoria", "Qtd", "Valor Total", "Ticket Médio", "% Volume", "", "", ""], 1):
            if txt:
                ws.cell(row=r, column=col, value=txt)
        estilo_header([ws.cell(row=r, column=c) for c in range(1, 6)],
                      fill=e["fill_blue"], font=e["font_h_white"])
        ws.row_dimensions[r].height = 22
        r += 1
        vol_total = sum(abs(d["valor"]) for d in cats.values()) or 1
        for i, cat in enumerate(sorted(cats, key=lambda k: -abs(cats[k]["valor"]))):
            d = cats[cat]
            tk = d["valor"] / d["qtd"] if d["qtd"] else 0
            pct = abs(d["valor"]) / vol_total
            ws.cell(row=r, column=1, value=cat)
            ws.cell(row=r, column=2, value=d["qtd"])
            c = ws.cell(row=r, column=3, value=d["valor"]); c.number_format = e["FMT_BRL"]
            c.font = Font(color=e["GREEN"] if d["valor"] > 0 else e["RED"], name="Calibri")
            c = ws.cell(row=r, column=4, value=tk); c.number_format = e["FMT_BRL"]
            c = ws.cell(row=r, column=5, value=pct); c.number_format = '0.0%'
            if i % 2 == 1:
                for col in range(1, 6):
                    ws.cell(row=r, column=col).fill = e["fill_zebra"]
            for col in range(1, 6):
                ws.cell(row=r, column=col).border = e["border_all"]
            r += 1

    top_cps = sorted(stats["contrapartes"].items(), key=lambda x: -abs(x[1]["valor"]))[:10]
    if top_cps:
        r += 2
        ws.cell(row=r, column=1, value="🏆 TOP 10 CONTRAPARTES").font = e["font_section"]
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        r += 1
        for col, txt in enumerate(["#", "Contraparte (CNPJ/CPF/Nome)", "Transações", "Volume", "Tipo"], 1):
            ws.cell(row=r, column=col, value=txt)
        estilo_header([ws.cell(row=r, column=c) for c in range(1, 6)],
                      fill=e["fill_blue"], font=e["font_h_white"])
        ws.row_dimensions[r].height = 22
        r += 1
        for i, (chave, d) in enumerate(top_cps, 1):
            ws.cell(row=r, column=1, value=i)
            ws.cell(row=r, column=2, value=chave)
            ws.cell(row=r, column=3, value=d["qtd"])
            c = ws.cell(row=r, column=4, value=d["valor"]); c.number_format = e["FMT_BRL"]
            c.font = Font(color=e["GREEN"] if d["valor"] > 0 else e["RED"], bold=True, name="Calibri")
            ws.cell(row=r, column=5, value="Recebimento" if d["valor"] > 0 else "Pagamento")
            if i % 2 == 0:
                for col in range(1, 6):
                    ws.cell(row=r, column=col).fill = e["fill_zebra"]
            for col in range(1, 6):
                ws.cell(row=r, column=col).border = e["border_all"]
            r += 1

    crit_lista = [a for a in anomalias if a["severidade"] == "critico"][:10]
    if crit_lista:
        r += 2
        ws.cell(row=r, column=1, value="🔴 ACHADOS CRÍTICOS").font = e["font_section"]
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
        r += 1
        for col, txt in enumerate(["Tipo", "Título", "Conta", "Valor", "Detalhe"], 1):
            ws.cell(row=r, column=col, value=txt)
        estilo_header([ws.cell(row=r, column=c) for c in range(1, 6)],
                      fill=e["fill_blue"], font=e["font_h_white"])
        ws.row_dimensions[r].height = 22
        r += 1
        for a in crit_lista:
            ws.cell(row=r, column=1, value=a["tipo"])
            ws.cell(row=r, column=2, value=a["titulo"])
            ws.cell(row=r, column=3, value=a["conta"])
            c = ws.cell(row=r, column=4, value=a.get("valor", 0)); c.number_format = e["FMT_BRL"]
            ws.cell(row=r, column=5, value=a["detalhe"])
            for col in range(1, 6):
                ws.cell(row=r, column=col).fill = e["fill_critico"]
                ws.cell(row=r, column=col).border = e["border_all"]
            r += 1

    for col, w in zip("ABCDEFGH", [26, 32, 13, 16, 16, 16, 13, 8]):
        ws.column_dimensions[col].width = w


def _xlsx_aba_transacoes(wb, extratos: list[dict], e: dict) -> None:
    """Cria e preenche a aba Transacoes."""
    ws = wb.create_sheet("Transações")
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 30
    cabec = ["Conta", "Data", "Tipo", "Valor", "Memo", "Nome", "Doc"]
    for col, txt in enumerate(cabec, 1):
        cell = ws.cell(row=1, column=col, value=txt)
        cell.fill = e["fill_blue_dark"]
        cell.font = e["font_h_white"]
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = e["border_all"]
    r = 2
    for ex in extratos:
        for t in ex["transacoes"]:
            ws.cell(row=r, column=1, value=ex["conta"])
            ws.cell(row=r, column=2, value=t["data"])
            tipo_cell = ws.cell(row=r, column=3, value=t["tipo"])
            if t["tipo"] == "CREDIT":
                tipo_cell.font = Font(color=e["GREEN"], bold=True, name="Calibri", size=10)
            elif t["tipo"] == "DEBIT":
                tipo_cell.font = Font(color=e["RED"], bold=True, name="Calibri", size=10)
            c = ws.cell(row=r, column=4, value=t["valor"])
            c.number_format = e["FMT_BRL"]
            c.font = Font(color=e["GREEN"] if t["valor"] > 0 else e["RED"],
                          name="Calibri", size=10, bold=True)
            ws.cell(row=r, column=5, value=t["memo"])
            ws.cell(row=r, column=6, value=t["nome"])
            ws.cell(row=r, column=7, value=t["checknum"])
            if r % 2 == 0:
                for col in range(1, len(cabec) + 1):
                    if col != 3 and col != 4:
                        ws.cell(row=r, column=col).fill = e["fill_zebra"]
            for col in range(1, len(cabec) + 1):
                ws.cell(row=r, column=col).border = e["border_all"]
            r += 1
    for col, w in zip("ABCDEFG", [26, 12, 10, 16, 52, 30, 14]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cabec))}{r-1}"


def _xlsx_aba_anomalias(wb, anomalias: list[dict], e: dict) -> None:
    """Cria e preenche a aba Anomalias."""
    ws = wb.create_sheet("Anomalias")
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 30
    cabec = ["Severidade", "Tipo", "Título", "Conta", "Valor", "Detalhe"]
    for col, txt in enumerate(cabec, 1):
        cell = ws.cell(row=1, column=col, value=txt)
        cell.fill = e["fill_blue_dark"]
        cell.font = e["font_h_white"]
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = e["border_all"]
    r = 2
    sev_meta = {
        "critico": ("🔴 CRÍTICO", e["fill_critico"],
                    Font(bold=True, color=e["RED"],    name="Calibri", size=10)),
        "alerta":  ("🟠 ALERTA",  e["fill_alerta"],
                    Font(bold=True, color=e["ORANGE"], name="Calibri", size=10)),
        "atencao": ("🟡 ATENÇÃO", e["fill_atencao"],
                    Font(bold=True, color=e["YELLOW"], name="Calibri", size=10)),
    }
    for a in anomalias:
        label, fill, font_sev = sev_meta.get(a["severidade"], ("?", None, None))
        sev_cell = ws.cell(row=r, column=1, value=label)
        if font_sev:
            sev_cell.font = font_sev
        ws.cell(row=r, column=2, value=a["tipo"])
        ws.cell(row=r, column=3, value=a["titulo"])
        ws.cell(row=r, column=4, value=a["conta"])
        c = ws.cell(row=r, column=5, value=a.get("valor", 0))
        c.number_format = e["FMT_BRL"]
        ws.cell(row=r, column=6, value=a["detalhe"])
        if fill:
            for col in range(1, len(cabec) + 1):
                ws.cell(row=r, column=col).fill = fill
        for col in range(1, len(cabec) + 1):
            ws.cell(row=r, column=col).border = e["border_all"]
        r += 1
    for col, w in zip("ABCDEF", [15, 22, 42, 28, 16, 60]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    if r > 2:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(cabec))}{r-1}"


def _gerar_xlsx(extratos: list[dict], anomalias: list[dict]) -> bytes:
    """Gera planilha XLSX com 3 abas estilizadas: Resumo, Transacoes, Anomalias."""
    wb = Workbook()
    e = _xlsx_estilos()
    _xlsx_aba_resumo(wb.active, extratos, anomalias, e)
    _xlsx_aba_transacoes(wb, extratos, e)
    _xlsx_aba_anomalias(wb, anomalias, e)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
