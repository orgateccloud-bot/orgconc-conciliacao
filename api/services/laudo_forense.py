"""Laudo Integrado de Auditoria Bancaria — modelo PRINCIPAL (OrgConc).

Gera um relatorio forense de 11 abas (Capa, Identificacao, Resumo Executivo,
Transacoes, Disposicoes 27-col, Risk Heatmap, CNPJs, Partes Relacionadas,
MEIs Teto, Status Tributario, Pos-Baixa) + MD + HTML + PDF, a partir de uma
pasta de extratos OFX + enriquecimento cadastral (RFB/BrasilAPI).

Parametrizado (sem dados de cliente no codigo):
    python scripts/relatorio_integrado.py --pasta <dir> --conta <id> \\
        --empresa-cnpj <14d> --tag <nome> [--enrich-all]

Nucleo reutilizavel pela API em api/services/laudo_forense.py (Fase 2).
"""
from __future__ import annotations

import asyncio
import base64
import contextvars
import os
import re
import xml.etree.ElementTree as ET
# F-Sec: parse de XML de upload via defusedxml (anti-XXE/billion-laughs).
# Mantém ET para ET.ParseError e tipos; só a leitura usa o parser seguro.
from defusedxml.ElementTree import fromstring as _safe_fromstring
import zipfile
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from api.services.report_utils import format_brl, format_num, format_pct, font_faces_css

_LOGO_PATH = Path(__file__).resolve().parents[2] / "assets" / "orgatec_logo.png"


def _logo_data_uri() -> str:
    if not _LOGO_PATH.exists():
        return ""
    return f"data:image/png;base64,{base64.b64encode(_LOGO_PATH.read_bytes()).decode('ascii')}"


def html_logo_inline() -> str:
    uri = _logo_data_uri()
    return (f'<img src="{uri}" alt="ORGATEC" style="width:64px;height:64px;'
            f'vertical-align:middle;margin-right:18px;"/>') if uri else ""


def inserir_logo_xlsx(ws, anchor: str = "A1", largura_px: int = 70, altura_px: int = 70):
    if not _LOGO_PATH.exists():
        return False
    try:
        from openpyxl.drawing.image import Image as XLImage
        img = XLImage(str(_LOGO_PATH))
        img.width, img.height, img.anchor = largura_px, altura_px, anchor
        ws.add_image(img)
        return True
    except Exception:  # noqa: BLE001
        return False
from api.matchers.cascata import classificar, ler_ofx
from api.matchers.cnpj_enricher import _carregar_cache, enriquecer_lote
from api.matchers.auditoria_forense import _meses_observados
from api.matchers.regime_fiscal import TETO_SIMPLES_EPP, analisar_regime
from api.matchers.forensics import (
    calcular_agregados,
    calcular_risk_score,
    classificar_tributario,
    detectar_carrossel,
    detectar_meio,
    detectar_primeira_vez,
    detectar_smurfing,
    detectar_valor_redondo,
    hash_linha,
    periodo_fiscal,
)

# ════════════════════════════════════════════════════════════════════════
# Configuração parametrizável — SEM dados de cliente hardcoded.
# EMPRESA é montado em runtime de --empresa-cnpj + enriquecimento RFB/BrasilAPI.
# ════════════════════════════════════════════════════════════════════════

MESES_PT = ["JAN", "FEV", "MAR", "ABR", "MAI", "JUN", "JUL", "AGO", "SET", "OUT", "NOV", "DEZ"]
ENRICH_LIMITE_PADRAO = 300

# Empresa auditada do laudo atual. Guardada em contextvars (não em global mutável)
# para isolar gerações concorrentes — sem race entre clientes e sem lock. Definida
# por set_empresa() e lida por _emp() em toda a cadeia de geração.
_EMPRESA_VAR: contextvars.ContextVar[dict] = contextvars.ContextVar("empresa_laudo", default={})


def set_empresa(empresa: dict) -> dict:
    """Define a empresa auditada no contexto atual (request/thread) e a retorna."""
    _EMPRESA_VAR.set(empresa)
    return empresa


def _emp() -> dict:
    """Empresa auditada do contexto atual (vazio se não definida)."""
    return _EMPRESA_VAR.get()

PASTA_DEFAULT = os.environ.get("ORGCONC_OFX_DIR", "")
OUT_DIR = Path(os.environ.get("ORGCONC_OUT_DIR", "."))
# O serviço NÃO faz I/O de arquivo nem mantém estado de saída — o chamador (CLI/API)
# salva o Workbook retornado por gerar_laudo_workbook().


def _mes_label(data_iso: str) -> str:
    """'2026-01-15' -> 'JAN/2026'. Vazio/inválido -> '??/????'."""
    try:
        return f"{MESES_PT[int(data_iso[5:7]) - 1]}/{data_iso[:4]}"
    except (ValueError, IndexError):
        return "??/????"


def construir_empresa(cnpj: str, cache: dict) -> dict:
    """Monta EMPRESA do cache de CNPJ (RFB/BrasilAPI). Campos fora da base
    pública (sócio, contrato) ficam '—' — preenchíveis manualmente depois."""
    c = re.sub(r"\D", "", cnpj or "")
    info = cache.get(c, {}) if c else {}
    cnpj_fmt = f"{c[:2]}.{c[2:5]}.{c[5:8]}/{c[8:12]}-{c[12:]}" if len(c) == 14 else (cnpj or "—")
    cnae = info.get("cnae_principal") or ""
    if cnae and info.get("cnae_descricao"):
        cnae = f"{cnae} {info['cnae_descricao']}"
    return {
        "razao_social": info.get("razao_social") or "—",
        "razao_anterior": "—",
        "nome_fantasia": info.get("nome_fantasia") or "—",
        "cnpj": cnpj_fmt,
        "cnpj_basico": c,
        "data_abertura": "—",
        "situacao": info.get("situacao") or "—",
        "porte_declarado": info.get("porte") or "—",
        "natureza_juridica": "—",
        "capital_social": float(info.get("capital_social") or 0.0),
        "cnae_principal": cnae or "—",
        "cnae_secundario": "—",
        "endereco_sede": " - ".join(x for x in [info.get("municipio"), info.get("uf")] if x) or "—",
        "endereco_admin": "—",
        "email": "—",
        "telefones": "—",
        "socio_nome": "—",
        "socio_cpf": "—",
        "socio_quotas": "—",
        "socio_nascimento": "—",
        "socio_endereco": "—",
        "ultima_alteracao": info.get("data_situacao") or "—",
    }

RX_CNPJ = re.compile(r"(\d{2})[.](\d{3})[.](\d{3})[ /](\d{4})[-](\d{2})")
LIMITE_MEI_PADRAO = 81_000.0
LIMITE_MEI_TAC = 251_600.0   # MEI Transportador Autonomo de Cargas (LC 188/2021)
CNAES_TRANSPORTE = ("4930", "5320", "4911")  # Transporte rodoviario carga, courier, ferroviario


def _limite_mei_por_cnae(cnae: str) -> float:
    """Retorna teto MEI conforme CNAE: TAC para transporte, padrao p/ outros."""
    cnae_clean = (cnae or "").replace(".", "").replace("-", "").replace("/", "")
    if any(cnae_clean.startswith(c) for c in CNAES_TRANSPORTE):
        return LIMITE_MEI_TAC
    return LIMITE_MEI_PADRAO

NAVY = "12345E"          # títulos/cabeçalhos (paleta AuditTax/OrgAudi)
ACCENT = "1F7FB8"        # azul de destaque (alinhado ao laudo HTML/PDF)
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
HEADER_FONT = Font(bold=True, color="EEF4FB", size=11)
TOTAL_FILL = PatternFill("solid", fgColor=ACCENT)
TOTAL_FONT = Font(bold=True, color="FFFFFF", size=11)
ZEBRA_FILL = PatternFill("solid", fgColor="EFF4F9")
ALERT_FILL = PatternFill("solid", fgColor="FBE6E4")        # risco alto/crítico (vermelho suave)
ALERT_FILL_MEDIO = PatternFill("solid", fgColor="FBF2DD")  # risco médio (âmbar)
INFO_FILL = PatternFill("solid", fgColor="EEF6FB")         # info (azul claro)
SUCCESS_FILL = PatternFill("solid", fgColor="E6F0E8")      # baixo (verde)
TITLE_FONT = Font(bold=True, size=14, color=NAVY)
SUBTITLE_FONT = Font(bold=True, size=11, color=NAVY)
BORDER = Side(border_style="thin", color="D4DDE6")
THIN_BORDER = Border(top=BORDER, left=BORDER, right=BORDER, bottom=BORDER)


def _extrair_cnpj(t):
    for fonte in (t.nome or "", t.memo or ""):
        m = RX_CNPJ.search(fonte)
        if m:
            return "".join(m.groups())
    return None


def style_header(ws, row, n_cols):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = THIN_BORDER


def cabecalho(ws, ultima_col, secao):
    c1 = ws.cell(row=1, column=1, value=f"    ORGATEC · Relatorio Integrado de Auditoria · {_emp().get('razao_social', '')}")
    c1.font = Font(bold=True, size=14, color="FFFFFF")
    c1.fill = PatternFill("solid", fgColor=NAVY)
    c1.alignment = Alignment(horizontal="center", vertical="center", indent=2)
    ws.merge_cells(f"A1:{get_column_letter(ultima_col)}1")
    ws.row_dimensions[1].height = 60
    if ws.column_dimensions["A"].width is None or ws.column_dimensions["A"].width < 12:
        ws.column_dimensions["A"].width = 12
    inserir_logo_xlsx(ws, "A1", largura_px=60, altura_px=60)

    c2 = ws.cell(row=2, column=1,
        value=f"Empresa: {_emp()['razao_social']} | CNPJ: {_emp()['cnpj']} | Socio: {_emp()['socio_nome']} (CPF {_emp()['socio_cpf']})")
    c2.font = Font(bold=True, size=10, color="FFFFFF")
    c2.fill = PatternFill("solid", fgColor="1F7FB8")
    c2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(f"A2:{get_column_letter(ultima_col)}2")

    c3 = ws.cell(row=3, column=1,
        value=f"{_emp().get('razao_social', '')} · CNPJ {_emp().get('cnpj', '—')} · Secao: {secao}")
    c3.font = Font(size=9, color="12345E")
    c3.fill = INFO_FILL
    c3.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(f"A3:{get_column_letter(ultima_col)}3")
    return 5


def cabecalho_padrao(ws, ultima_col, *, titulo, linha2="", secao="", com_logo=True):
    """Cabeçalho XLSX padrão ORGATEC reutilizável por qualquer laudo (não depende
    da empresa do contexto). 3 linhas: banda navy (título + logo), banda azul (linha2),
    banda info (seção). Retorna a primeira linha de conteúdo (5)."""
    c1 = ws.cell(row=1, column=1, value="    ORGATEC · " + titulo)
    c1.font = Font(bold=True, size=14, color="FFFFFF")
    c1.fill = PatternFill("solid", fgColor=NAVY)
    c1.alignment = Alignment(horizontal="center", vertical="center", indent=2)
    ws.merge_cells(f"A1:{get_column_letter(ultima_col)}1")
    ws.row_dimensions[1].height = 60
    if ws.column_dimensions["A"].width is None or ws.column_dimensions["A"].width < 12:
        ws.column_dimensions["A"].width = 12
    if com_logo:
        inserir_logo_xlsx(ws, "A1", largura_px=60, altura_px=60)
    if linha2:
        c2 = ws.cell(row=2, column=1, value=linha2)
        c2.font = Font(bold=True, size=10, color="FFFFFF")
        c2.fill = PatternFill("solid", fgColor="1F7FB8")
        c2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.merge_cells(f"A2:{get_column_letter(ultima_col)}2")
    if secao:
        c3 = ws.cell(row=3, column=1, value=secao)
        c3.font = Font(size=9, color="12345E")
        c3.fill = INFO_FILL
        c3.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.merge_cells(f"A3:{get_column_letter(ultima_col)}3")
    return 5


def aba_capa(wb, *, titulo_relatorio, linha2, objeto, secoes, sumario, subtitulo="Sistema OrgConc"):
    """Cria a primeira aba 'Capa' no padrão do laudo forense (reutilizável):
    cabeçalho + título + data + ÍNDICE DE SEÇÕES (hyperlinks) + SUMÁRIO EXECUTIVO.

    secoes: list[(num, nome, descricao, sheet_name)] — links para as demais abas.
    sumario: list[(rotulo, valor)] — KPIs principais.
    Usa a aba ativa (renomeia para '1. Capa'); demais abas são criadas pelo caller.
    """
    ws = wb.active
    ws.title = "1. Capa"
    start = cabecalho_padrao(ws, 6, titulo=titulo_relatorio, linha2=linha2, secao=subtitulo)

    ws.cell(row=start, column=1, value=titulo_relatorio.upper()).font = TITLE_FONT
    ws.merge_cells(f"A{start}:F{start}")
    ws.cell(row=start + 1, column=1, value=objeto).font = Font(size=9, color="51616F")
    ws.merge_cells(f"A{start + 1}:F{start + 1}")
    ws.cell(row=start + 2, column=1,
            value="Gerado em " + datetime.now().strftime("%d/%m/%Y %H:%M")).font = Font(size=9, italic=True, color="65778A")

    r = start + 4
    ws.cell(row=r, column=1, value="INDICE DE SECOES").font = SUBTITLE_FONT
    ws.merge_cells(f"A{r}:F{r}")
    r += 1
    ws.cell(row=r, column=1, value="#")
    ws.cell(row=r, column=2, value="Secao")
    ws.cell(row=r, column=3, value="Conteudo")
    style_header(ws, r, 3)
    ws.merge_cells(f"C{r}:F{r}")
    r += 1
    for num, nome, desc, sheet in secoes:
        ws.cell(row=r, column=1, value=num)
        c = ws.cell(row=r, column=2, value=nome)
        c.font = Font(bold=True, color="1F7FB8", underline="single")
        c.hyperlink = f"#'{sheet}'!A1"
        c.style = "Hyperlink"
        ws.cell(row=r, column=3, value=desc)
        ws.merge_cells(f"C{r}:F{r}")
        for cc in range(1, 7):
            ws.cell(row=r, column=cc).border = THIN_BORDER
            if r % 2 == 0:
                ws.cell(row=r, column=cc).fill = ZEBRA_FILL
        r += 1

    r += 2
    ws.cell(row=r, column=1, value="SUMARIO EXECUTIVO").font = SUBTITLE_FONT
    ws.merge_cells(f"A{r}:F{r}")
    r += 1
    for rotulo, valor in sumario:
        ws.cell(row=r, column=1, value=rotulo).font = Font(bold=True)
        cv = ws.cell(row=r, column=2, value=valor)
        cv.alignment = Alignment(horizontal="left")
        ws.merge_cells(f"B{r}:F{r}")
        for cc in range(1, 7):
            ws.cell(row=r, column=cc).border = THIN_BORDER
        r += 1

    ws.column_dimensions["A"].width = 16
    for col in "BCDEF":
        ws.column_dimensions[col].width = 22
    return ws


# ════════════════════════════════════════════════════════════════════════
# Coleta de dados (carrega todos os 5 OFXs)
# ════════════════════════════════════════════════════════════════════════


def montar_dados(transacoes):
    """Lista de Transacao (deduped) -> (todos[(mes, t, r)], saldos{mes: {...}}).

    Reusado pela API (que já tem as transações do upload) e pelo CLI."""
    todos, saldos, saldo_corr = [], {}, 0.0
    for t in sorted(transacoes, key=lambda x: x.data or ""):
        mes = _mes_label(t.data)
        todos.append((mes, t, classificar(t)))
        s = saldos.setdefault(mes, {"saldo_final": 0.0, "n": 0, "cred": 0.0, "deb": 0.0})
        s["n"] += 1
        if t.valor > 0:
            s["cred"] += t.valor
        else:
            s["deb"] += t.valor
        saldo_corr += t.valor
        s["saldo_final"] = round(saldo_corr, 2)
    return todos, saldos


async def coletar_dados(pasta: str, conta_filtro: str, empresa_cnpj: str, enrich_all: bool):
    print(f"Coletando OFX de {pasta} ...")
    cache = _carregar_cache()
    arquivos = sorted(Path(pasta).glob("*.ofx"))
    brutas = []
    for p in arquivos:
        try:
            brutas.extend(ler_ofx(str(p)))
        except Exception as e:  # noqa: BLE001
            print(f"  AVISO OFX {p.name}: {e}")

    # dedup por (conta, fitid) — corrige downloads sobrepostos
    vistos, dedup = set(), []
    for t in brutas:
        k = (t.conta, t.fitid) if t.fitid else (t.conta, t.data, round(t.valor, 2), t.memo, t.nome)
        if k in vistos:
            continue
        vistos.add(k)
        dedup.append(t)
    if conta_filtro:
        dedup = [t for t in dedup if conta_filtro in (t.conta or "")]
    dedup.sort(key=lambda t: t.data or "")
    print(f"  {len(arquivos)} arquivos, {format_num(len(brutas))} linhas -> {format_num(len(dedup))} após dedup"
          + (f" (conta '{conta_filtro}')" if conta_filtro else ""))

    # bucket por mês via montar_dados (mesma lógica reusada pela API)
    todos, saldos = montar_dados(dedup)

    # enriquecer CNPJs faltantes (BrasilAPI → cache)
    cnpjs = {_extrair_cnpj(t) for _, t, _ in todos if _extrair_cnpj(t)}
    cnpjs.discard(None)
    falt = [c for c in cnpjs if c not in cache]
    if not enrich_all:
        falt = falt[:ENRICH_LIMITE_PADRAO]
    if falt:
        print(f"  Enriquecendo {len(falt)} CNPJs via BrasilAPI...")
        await enriquecer_lote(falt, db=None)
        cache = _carregar_cache()

    empresa = set_empresa(construir_empresa(empresa_cnpj, cache))
    print(f"  {format_num(len(todos))} transacoes | {len(saldos)} meses | {len(cnpjs)} CNPJs"
          f" | empresa: {empresa['razao_social']}")
    return todos, saldos, cache


# ════════════════════════════════════════════════════════════════════════
# Gerar XLSX integrado (11 abas)
# ════════════════════════════════════════════════════════════════════════


# ════════════════════════════════════════════════════════════════════════
# Documentos fiscais (NF-e / CT-e) — parsers por conteúdo + cruzamento OFX
# ════════════════════════════════════════════════════════════════════════
def _local(tag):
    return tag.split("}")[-1]


def _filho(elem, nome):
    if elem is None:
        return None
    for f in elem:
        if _local(f.tag) == nome:
            return f
    return None


def _texto(elem, *caminho):
    cur = elem
    for nome in caminho:
        cur = _filho(cur, nome)
        if cur is None:
            return ""
    return cur.text.strip() if cur is not None and cur.text else ""


def parse_nfe(conteudo):
    try:
        root = _safe_fromstring(conteudo)
    except ET.ParseError:
        return None
    inf = next((el for el in root.iter() if _local(el.tag) == "infNFe"), None)
    if inf is None:
        return None
    ide, emit = _filho(inf, "ide"), _filho(inf, "emit")
    total = _filho(inf, "total")
    icms_tot = _filho(total, "ICMSTot") if total is not None else None
    return {
        "chave": (inf.get("Id") or "").lstrip("NFe"),
        "data": (_texto(ide, "dhEmi") or _texto(ide, "dEmi"))[:10],
        "emit_cnpj": _texto(emit, "CNPJ"), "emit_nome": _texto(emit, "xNome"),
        "valor": float(_texto(icms_tot, "vNF") or 0) if icms_tot is not None else 0.0,
        "modelo": _texto(ide, "mod") or "55", "uf": _texto(emit, "enderEmit", "UF"),
    }


def parse_cte(conteudo):
    try:
        root = _safe_fromstring(conteudo)
    except ET.ParseError:
        return None
    inf = next((el for el in root.iter() if _local(el.tag) in ("infCte", "infCTe")), None)
    if inf is None:
        return None
    ide, dest = _filho(inf, "ide"), _filho(inf, "dest")
    vprest = _filho(inf, "vPrest")
    return {
        "chave": (inf.get("Id") or "").lstrip("CTe"),
        "data": (_texto(ide, "dhEmi") or _texto(ide, "dEmi"))[:10],
        "dest_cnpj": _texto(dest, "CNPJ") if dest is not None else "",
        "dest_nome": _texto(dest, "xNome") if dest is not None else "",
        "valor": float(_texto(vprest, "vTPrest") or 0) if vprest is not None else 0.0,
        "modelo": "57", "uf": _texto(ide, "UFIni"),
    }


def carregar_docs(pasta: str):
    """Glob de *.zip; parse de TODO XML por conteúdo (infNFe/infCte); dedup por chave."""
    nfes: dict[str, dict] = {}
    ctes: dict[str, dict] = {}
    n_xml = 0
    for z in sorted(Path(pasta).glob("*.zip")):
        try:
            zf = zipfile.ZipFile(z)
        except Exception:  # noqa: BLE001
            continue
        with zf:
            for m in zf.namelist():
                if not m.lower().endswith(".xml"):
                    continue
                n_xml += 1
                try:
                    raw = zf.read(m)
                except Exception:  # noqa: BLE001
                    continue
                doc = parse_nfe(raw)
                if doc and doc.get("chave"):
                    nfes.setdefault(doc["chave"], doc)
                    continue
                cte = parse_cte(raw)
                if cte and cte.get("chave"):
                    ctes.setdefault(cte["chave"], cte)
    return list(nfes.values()), list(ctes.values()), n_xml


def carregar_docs_xmls(xmls):
    """Variante de carregar_docs para uploads em memória: aceita lista de
    (nome, bytes) de XML e/ou ZIP (expande ZIPs). Retorna (nfes, ctes, n_xml).
    Mesma engine — usada pelo endpoint /fiscal/laudo p/ alimentar as abas fiscais."""
    import io as _io

    nfes: dict[str, dict] = {}
    ctes: dict[str, dict] = {}
    n_xml = 0

    def _proc(raw):
        nonlocal n_xml
        n_xml += 1
        doc = parse_nfe(raw)
        if doc and doc.get("chave"):
            nfes.setdefault(doc["chave"], doc)
            return
        cte = parse_cte(raw)
        if cte and cte.get("chave"):
            ctes.setdefault(cte["chave"], cte)

    for nome, conteudo in xmls:
        low = (nome or "").lower()
        if low.endswith(".zip"):
            try:
                with zipfile.ZipFile(_io.BytesIO(conteudo)) as zf:
                    for m in zf.namelist():
                        if m.lower().endswith(".xml"):
                            _proc(zf.read(m))
            except zipfile.BadZipFile:
                continue
        elif low.endswith(".xml"):
            _proc(conteudo)
    return list(nfes.values()), list(ctes.values()), n_xml


def _norm_nome(s: str) -> str:
    return re.sub(r"\s+", " ", re.sub(r"[^A-Z0-9 ]", " ", (s or "").upper())).strip()


def aba_documentos_fiscais(wb, nfes, ctes):
    ws = wb.create_sheet("12. Documentos Fiscais")
    start = cabecalho(ws, 6, "Documentos Fiscais")
    nfe_vol = sum(n["valor"] for n in nfes)
    cte_vol = sum(c["valor"] for c in ctes)
    ws.cell(row=start, column=1, value=f"DOCUMENTOS FISCAIS PROCESSADOS - {format_num(len(nfes) + len(ctes))} XMLs").font = TITLE_FONT
    ws.merge_cells(f"A{start}:F{start}")

    def _per(docs):
        ds = [d["data"] for d in docs if d.get("data")]
        return f"{min(ds)} a {max(ds)}" if ds else "-"

    r = start + 2
    for c, h in enumerate(["Modelo", "Tipo", "Qtde", "Valor Total (R$)", "Periodo", "Fonte"], 1):
        ws.cell(row=r, column=c, value=h)
    style_header(ws, r, 6)
    r += 1
    for mod, tipo, qtd, val, per in [
        ("55", "NF-e (Nota Fiscal Eletronica)", len(nfes), nfe_vol, _per(nfes)),
        ("57", "CT-e (Conhecimento de Transporte)", len(ctes), cte_vol, _per(ctes)),
    ]:
        ws.cell(row=r, column=1, value=mod).font = Font(bold=True)
        ws.cell(row=r, column=2, value=tipo)
        ws.cell(row=r, column=3, value=qtd).number_format = "#,##0"
        ws.cell(row=r, column=4, value=round(val, 2)).number_format = "#,##0.00"
        ws.cell(row=r, column=5, value=per)
        ws.cell(row=r, column=6, value="ZIPs")
        for c in range(1, 7):
            ws.cell(row=r, column=c).border = THIN_BORDER
        r += 1
    ws.cell(row=r, column=2, value="TOTAL").font = TOTAL_FONT
    ws.cell(row=r, column=3, value=len(nfes) + len(ctes)).number_format = "#,##0"
    ws.cell(row=r, column=4, value=round(nfe_vol + cte_vol, 2)).number_format = "#,##0.00"
    for c in range(1, 7):
        ws.cell(row=r, column=c).fill = TOTAL_FILL
        ws.cell(row=r, column=c).font = TOTAL_FONT
    r += 3

    ws.cell(row=r, column=1, value="TOP 10 EMITENTES DE NF-E (FORNECEDORES)").font = SUBTITLE_FONT
    ws.merge_cells(f"A{r}:F{r}")
    r += 1
    for c, h in enumerate(["#", "Emitente", "CNPJ", "UF", "Qtd", "Valor Total"], 1):
        ws.cell(row=r, column=c, value=h)
    style_header(ws, r, 6)
    r += 1
    emit = defaultdict(lambda: {"qtd": 0, "vol": 0.0, "cnpj": "", "uf": ""})
    for n in nfes:
        k = (n.get("emit_nome") or "")[:50]
        emit[k]["qtd"] += 1
        emit[k]["vol"] += n["valor"]
        emit[k]["cnpj"] = n.get("emit_cnpj", "")
        emit[k]["uf"] = n.get("uf", "")
    for i, (nome, info) in enumerate(sorted(emit.items(), key=lambda x: -x[1]["vol"])[:10], 1):
        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=2, value=nome)
        ws.cell(row=r, column=3, value=info["cnpj"])
        ws.cell(row=r, column=4, value=info["uf"])
        ws.cell(row=r, column=5, value=info["qtd"]).number_format = "#,##0"
        ws.cell(row=r, column=6, value=round(info["vol"], 2)).number_format = "#,##0.00"
        for c in range(1, 7):
            ws.cell(row=r, column=c).border = THIN_BORDER
            if r % 2 == 0:
                ws.cell(row=r, column=c).fill = ZEBRA_FILL
        r += 1
    for col, w in {1: 5, 2: 46, 3: 20, 4: 6, 5: 10, 6: 18}.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = f"A{start + 2}"


# CNPJ embutido no memo bancário: aceita 'XX.XXX.XXX/XXXX-XX', 'XX.XXX.XXX XXXX-XX',
# 'XX XXX XXX XXXX XX' e 'XXXXXXXXXXXXXX' (separadores opcionais).
_RX_CNPJ_FLEX = re.compile(r"\b(\d{2})[.\s]?(\d{3})[.\s]?(\d{3})[/\s]?(\d{4})[-\s]?(\d{2})\b")


def _fmt_cnpj(c: str) -> str:
    c = re.sub(r"\D", "", c or "")
    return f"{c[:2]}.{c[2:5]}.{c[5:8]}/{c[8:12]}-{c[12:]}" if len(c) == 14 else (c or "")


def _cnpj_do_texto(texto: str) -> str:
    """Extrai um CNPJ (14 dígitos) do texto bancário (ex.: 'Pagamento Pix 12.345.678 0001-90')."""
    m = _RX_CNPJ_FLEX.search(texto or "")
    return "".join(m.groups()) if m else ""


def _conformidade_lista(nfes, ctes, transacoes, top=30):
    """Cruza pagamentos OFX com NF-e (emitente) e CT-e (dest) por fornecedor.

    RECONHECE o CNPJ embutido no memo bancário e agrupa/casa por CNPJ (preciso);
    cai para nome quando não há CNPJ. Exibe a razão social (cache) no lugar do memo
    cru. Exclui auto-movimentação (próprio CNPJ / MESMA TIT / própria razão)."""
    cache = _carregar_cache()
    cb = re.sub(r"\D", "", _emp().get("cnpj_basico", "") or "")
    razao = _norm_nome(_emp().get("razao_social", ""))
    # NF-e indexada por CNPJ emitente (match preciso) e por nome (fallback)
    nfe_cnpj = defaultdict(float)
    nfe_nome = defaultdict(float)
    for n in nfes:
        c = re.sub(r"\D", "", n.get("emit_cnpj", "") or "")
        if len(c) == 14:
            nfe_cnpj[c] += n["valor"]
        nm = _norm_nome(n.get("emit_nome"))[:60]
        if nm:
            nfe_nome[nm] += n["valor"]
    cte_nome = defaultdict(float)
    for c in ctes:
        nm = _norm_nome(c.get("dest_nome"))[:60]
        if nm:
            cte_nome[nm] += c["valor"]
    # Agrupa pagamentos por CNPJ reconhecido (senão por nome)
    pag = defaultdict(lambda: {"vol": 0.0, "n": 0, "cnpj": "", "label": ""})
    for t in transacoes:
        if t.valor >= 0:
            continue
        full = (t.nome or "") + " " + (t.memo or "")
        if "MESMA TIT" in full.upper():
            continue
        cnpj = _cnpj_do_texto(full)
        if cnpj and cnpj == cb:
            continue
        nm = _norm_nome(t.nome)[:60]
        if not nm:
            continue
        if len(razao) >= 8 and razao[:20] in nm:
            continue
        key = cnpj or nm
        d = pag[key]
        d["vol"] += abs(t.valor)
        d["n"] += 1
        if cnpj and not d["cnpj"]:
            d["cnpj"] = cnpj
            d["label"] = cache.get(cnpj, {}).get("razao_social") or _fmt_cnpj(cnpj)
        if not d["label"]:
            d["label"] = nm.title()
    out = []
    for key, d in sorted(pag.items(), key=lambda x: -x[1]["vol"])[:top]:
        vp, cnpj = d["vol"], d["cnpj"]
        if cnpj and cnpj in nfe_cnpj:                       # match preciso por CNPJ
            vn = nfe_cnpj[cnpj]
        else:                                               # fallback fuzzy por nome
            nmk = _norm_nome(d["label"])[:60]
            vn = next((v for k, v in nfe_nome.items()
                       if k[:30] == nmk[:30] or (nmk[:20] and nmk[:20] in k) or (k[:20] and k[:20] in nmk)), 0.0)
        vc = cte_nome.get(_norm_nome(d["label"])[:60], 0.0)
        conf = ((vn + vc) / vp * 100) if vp else 0
        classe = "CONFORME" if conf >= 80 else "MEDIO" if conf >= 50 else "ALTO" if conf >= 20 else "CRITICO"
        out.append({"nome": d["label"], "cnpj": _fmt_cnpj(cnpj) if cnpj else "",
                    "vp": vp, "vn": vn, "vc": vc, "conf": conf, "classe": classe})
    return out


def aba_conformidade_fiscal(wb, conf):
    ws = wb.create_sheet("13. Conformidade Fiscal")
    start = cabecalho(ws, 8, "Conformidade Fiscal")
    ws.cell(row=start, column=1, value="CRUZAMENTO OFX x NF-e x CT-e - SCORE POR FORNECEDOR").font = TITLE_FONT
    ws.merge_cells(f"A{start}:H{start}")
    r = start + 2
    for col, h in enumerate(["#", "Fornecedor", "CNPJ", "Vol Pago OFX", "Vol NF-e", "Vol CT-e", "Conf %", "Classe"], 1):
        ws.cell(row=r, column=col, value=h)
    style_header(ws, r, 8)
    r += 1
    for i, c in enumerate(conf, 1):
        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=2, value=c["nome"][:48].title())
        ws.cell(row=r, column=3, value=c["cnpj"])
        ws.cell(row=r, column=4, value=round(c["vp"], 2)).number_format = "#,##0.00"
        ws.cell(row=r, column=5, value=round(c["vn"], 2)).number_format = "#,##0.00"
        ws.cell(row=r, column=6, value=round(c["vc"], 2)).number_format = "#,##0.00"
        ws.cell(row=r, column=7, value=c["conf"] / 100).number_format = "0.0%"
        cc = ws.cell(row=r, column=8, value=c["classe"])
        if c["classe"] == "CRITICO":
            cc.fill = ALERT_FILL
            cc.font = Font(bold=True, color="B33A3A")
        elif c["classe"] == "ALTO":
            cc.fill = ALERT_FILL_MEDIO
        elif c["classe"] == "CONFORME":
            cc.fill = SUCCESS_FILL
        for cx in range(1, 9):
            ws.cell(row=r, column=cx).border = THIN_BORDER
        r += 1
    for col, w in {1: 4, 2: 42, 3: 18, 4: 16, 5: 16, 6: 14, 7: 10, 8: 12}.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = f"A{start + 2}"


def _parse_data_xlsx(data_str: str | None):
    """Converte string ISO 'YYYY-MM-DD' em datetime.date para o Excel tratar como data real."""
    if not data_str:
        return None
    try:
        return date.fromisoformat(str(data_str)[:10])
    except (ValueError, TypeError):
        return None


def gerar_laudo_workbook(todos, saldos, cache, nfes=None, ctes=None):
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # ── Metadados do workbook ─────────────────────────────────────────────
    emp = _emp()
    razao_wb = emp.get("razao_social", "ORGATEC")
    wb.properties.title = f"Laudo de Auditoria — {razao_wb}"
    wb.properties.subject = "Auditoria Bancária Forense"
    wb.properties.creator = "ORGATEC — OrgAudi"
    wb.properties.company = "ORGATEC Contabilidade e Auditoria"
    wb.properties.keywords = "auditoria; forense; bancário; OFX; ORGATEC"

    # ── Pre-calculos ────────────────────────────────────────────────────
    n_total = len(todos)
    meses = list(saldos.keys())              # ordem cronológica (derivada dos dados)
    n_meses = len(meses)
    meses_curto = [m.split("/")[0] for m in meses]
    ncol_trib = 4 + n_meses
    _datas = sorted(t.data for _, t, _ in todos if t.data)

    def _br(d):
        return f"{d[8:10]}/{d[5:7]}/{d[:4]}" if d and len(d) >= 10 else (d or "—")

    periodo_str = f"{_br(_datas[0])} a {_br(_datas[-1])}" if _datas else "—"
    cred_total = sum(s["cred"] for s in saldos.values())
    deb_total = sum(s["deb"] for s in saldos.values())
    _m0 = saldos[meses[0]] if meses else {"saldo_final": 0.0, "cred": 0.0, "deb": 0.0}
    saldo_ini_jan = _m0["saldo_final"] - (_m0["cred"] + _m0["deb"])
    saldo_fim_mai = saldos[meses[-1]]["saldo_final"] if meses else 0.0
    volume_bruto = abs(cred_total) + abs(deb_total)
    # Anualização e múltiplo do teto via MOTOR validado (dias corridos) — laudo == motor.
    meses_obs = _meses_observados([t for _, t, _ in todos])
    _regime = analisar_regime(cred_total, deb_total, meses_obs, TETO_SIMPLES_EPP)
    anualizado = _regime.volume_anualizado
    multiplo = _regime.multiplo_do_teto

    # Disposicoes com classificacao forense
    todas_disps = []
    fake_disps_para_agg = []
    for mes, t, r in todos:
        cnpj = _extrair_cnpj(t)
        info = cache.get(cnpj, {}) if cnpj else {}
        sit = info.get("situacao", "")
        porte = info.get("porte", "")
        meio = detectar_meio(t.memo or "", t.nome or "")
        razao = info.get("razao_social", "")

        # Detecta pos-baixa
        disp_name = "NORMAL"
        flag = ""
        if ("BAIXADA" in sit or "INAPTA" in sit) and info.get("data_situacao"):
            try:
                db_d = date.fromisoformat(info["data_situacao"][:10])
                dt_d = date.fromisoformat(t.data[:10])
                if dt_d > db_d:
                    disp_name = "ALERTA_POS_BAIXA"
                    flag = f"Pos-baixa: {(dt_d - db_d).days}d apos {info['data_situacao']}"
            except (ValueError, TypeError):
                pass

        class _D:
            pass
        d = _D()
        d.transacao = t
        d.estagio = r.estagio
        d.disposicao = disp_name
        d.contraparte = razao
        d.flag = flag
        d.meio = meio
        d.cnpj = cnpj
        d.info_cnpj = info
        d.mes = mes
        todas_disps.append(d)
        fake_disps_para_agg.append(d)

    agg = calcular_agregados(fake_disps_para_agg)

    # ════════════════════════════════════════════════════════════════════
    # Aba 1: Capa / Indice
    # ════════════════════════════════════════════════════════════════════
    ws = wb.create_sheet("1. Capa")
    start = cabecalho(ws, 6, "Capa e Indice")
    ws.cell(row=start, column=1, value="RELATORIO INTEGRADO DE AUDITORIA BANCARIA").font = TITLE_FONT
    ws.merge_cells(f"A{start}:F{start}")
    ws.cell(row=start+1, column=1, value=f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}").font = Font(italic=True, color="64748B", size=9)
    ws.merge_cells(f"A{start+1}:F{start+1}")

    r = start + 3
    ws.cell(row=r, column=1, value="INDICE DE SECOES").font = SUBTITLE_FONT
    ws.merge_cells(f"A{r}:F{r}")
    r += 1
    ws.cell(row=r, column=1, value="#")
    ws.cell(row=r, column=2, value="Secao")
    ws.cell(row=r, column=3, value="Conteudo")
    style_header(ws, r, 3)
    r += 1

    indice = [
        ("1", "Capa", "Indice, sumario executivo, totais", "1. Capa"),
        ("2", "Identificacao Cadastral", "Dados RFB + contrato social + quadro societario", "2. Identificacao"),
        ("3", "Resumo Executivo", "KPIs principais e evolucao mensal", "3. Resumo Executivo"),
        ("4", "Transacoes", "Lancamentos com saldo acumulado e contraparte", "4. Transacoes"),
        ("5", "Disposicoes Forenses", "Classificacao em 27 colunas + Risk Score", "5. Disposicoes"),
        ("6", "Risk Heatmap", "Distribuicao por classe (CRITICO/ALTO/MEDIO/BAIXO)", "6. Risk Heatmap"),
        ("7", "CNPJs Enriquecidos", "Contrapartes identificadas via RFB / BrasilAPI", "7. CNPJs"),
        ("8", "Partes Relacionadas", "Auto-movimentacao (proprio CNPJ) + mesma titularidade", "8. Partes Relacionadas"),
        ("9", "MEIs Estourando Teto", "Fornecedores PJ acima do teto MEI", "9. MEIs Teto"),
        ("10", "Status Tributario", "Categorias fiscais + retencoes estimadas", "10. Status Tributario"),
        ("11", "Pagamentos Pos-Baixa", "Transacoes a CNPJs ja baixados", "11. Pos-Baixa"),
    ]
    if nfes or ctes:
        indice += [
            ("12", "Documentos Fiscais", "NF-e + CT-e processados (XMLs) + top emitentes", "12. Documentos Fiscais"),
            ("13", "Conformidade Fiscal", "Cruzamento OFX x NF-e x CT-e por fornecedor", "13. Conformidade Fiscal"),
        ]
    for num, sec, desc, sheet_name in indice:
        ws.cell(row=r, column=1, value=num)
        c_sec = ws.cell(row=r, column=2, value=sec)
        c_sec.font = Font(bold=True, color="1F7FB8", underline="single")
        # Hyperlink para a aba correspondente
        c_sec.hyperlink = f"#'{sheet_name}'!A1"
        c_sec.style = "Hyperlink"
        ws.cell(row=r, column=3, value=desc)
        ws.merge_cells(f"C{r}:F{r}")
        for c in range(1, 7):
            ws.cell(row=r, column=c).border = THIN_BORDER
            if r % 2 == 0:
                ws.cell(row=r, column=c).fill = ZEBRA_FILL
        r += 1

    # Volume liquido = bruto menos transferencias internas (auto-movimentacao
    # por proprio CNPJ + mesma titularidade entre contas proprias), que NAO
    # representam movimentacao economica real. Mesma deteccao da aba 8.
    _cnpj_basico = _emp().get("cnpj_basico", "")
    vol_transf_interna = 0.0
    for _d in todas_disps:
        _txt = ((_d.transacao.nome or "") + " " + (_d.transacao.memo or "")).upper()
        if (_cnpj_basico and (_d.cnpj == _cnpj_basico or _cnpj_basico in re.sub(r"\D", "", _txt))) \
           or "MESMA TIT" in _txt or "MESMA TITULAR" in _txt:
            vol_transf_interna += abs(_d.transacao.valor)
    volume_liquido = volume_bruto - vol_transf_interna

    # Sumario rapido
    r += 2
    ws.cell(row=r, column=1, value="SUMARIO EXECUTIVO").font = SUBTITLE_FONT
    ws.merge_cells(f"A{r}:F{r}")
    r += 1
    resumo = [
        ("Periodo analisado", f"{periodo_str} ({n_meses} meses)"),
        ("Total de transacoes", f"{format_num(n_total)}"),
        ("Volume bruto movimentado", f"R$ {format_brl(volume_bruto)}"),
        ("Volume liquido (excl. transf. internas)", f"R$ {format_brl(volume_liquido)}"),
        ("Volume anualizado projetado", f"R$ {format_brl(anualizado)}"),
        ("Saldo inicial do periodo", f"R$ {format_brl(saldo_ini_jan)}"),
        ("Saldo final do periodo", f"R$ {format_brl(saldo_fim_mai)}"),
        ("Variacao do periodo", f"R$ {format_brl(saldo_fim_mai - saldo_ini_jan)}"),
        ("CNPJs identificados", f"{sum(1 for d in todas_disps if d.cnpj)}"),
        ("Alertas pos-baixa", f"{sum(1 for d in todas_disps if d.disposicao == 'ALERTA_POS_BAIXA')}"),
    ]
    for k, v in resumo:
        ws.cell(row=r, column=1, value=k).font = Font(bold=True)
        c = ws.cell(row=r, column=2, value=v)
        if "anualizado" in k or "Volume bruto" in k:
            c.font = Font(bold=True, color="B33A3A")
        ws.merge_cells(f"B{r}:F{r}")
        for col in range(1, 7):
            ws.cell(row=r, column=col).border = THIN_BORDER
            if r % 2 == 0:
                ws.cell(row=r, column=col).fill = ZEBRA_FILL
        r += 1

    for col, w in {1: 6, 2: 26, 3: 50, 4: 12, 5: 12, 6: 12}.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = f"A{start + 2}"

    # ════════════════════════════════════════════════════════════════════
    # Aba 2: Identificacao Cadastral
    # ════════════════════════════════════════════════════════════════════
    ws = wb.create_sheet("2. Identificação")
    start = cabecalho(ws, 5, "Identificacao Cadastral")
    ws.cell(row=start, column=1, value="DADOS CADASTRAIS DA EMPRESA AUDITADA").font = TITLE_FONT
    ws.merge_cells(f"A{start}:E{start}")
    ws.cell(row=start+1, column=1, value="Fontes: Contrato Social + Cartao CNPJ (RFB / BrasilAPI)").font = Font(italic=True, color="64748B", size=9)
    ws.merge_cells(f"A{start+1}:E{start+1}")

    r = start + 3
    ws.cell(row=r, column=1, value="DADOS DA PESSOA JURIDICA").font = SUBTITLE_FONT
    ws.merge_cells(f"A{r}:E{r}")
    r += 1
    dados_pj = [
        ("Razao Social Atual", _emp()["razao_social"]),
        ("Razao Social Anterior", _emp()["razao_anterior"]),
        ("Nome Fantasia", _emp()["nome_fantasia"]),
        ("CNPJ", _emp()["cnpj"]),
        ("Situacao Cadastral", _emp()["situacao"]),
        ("Data de Abertura", _emp()["data_abertura"]),
        ("Porte Declarado", _emp()["porte_declarado"]),
        ("Natureza Jurídica", _emp()["natureza_juridica"]),
        ("Capital Social", f"R$ {format_brl(_emp()['capital_social'])}"),
        ("CNAE Principal", _emp()["cnae_principal"]),
        ("CNAE Secundário", _emp()["cnae_secundario"]),
        ("Endereço Sede", _emp()["endereco_sede"]),
        ("Escritorio Administrativo", _emp()["endereco_admin"]),
        ("Email", _emp()["email"]),
        ("Telefones", _emp()["telefones"]),
        ("Ultima Alteracao Contratual", _emp()["ultima_alteracao"]),
    ]
    for k, v in dados_pj:
        ws.cell(row=r, column=1, value=k).font = Font(bold=True)
        ws.cell(row=r, column=2, value=str(v))
        ws.merge_cells(f"B{r}:E{r}")
        for c in range(1, 6):
            ws.cell(row=r, column=c).border = THIN_BORDER
            if r % 2 == 0:
                ws.cell(row=r, column=c).fill = ZEBRA_FILL
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="QUADRO SOCIETARIO").font = SUBTITLE_FONT
    ws.merge_cells(f"A{r}:E{r}")
    r += 1
    socio = [
        ("Socio Unico (100%)", _emp()["socio_nome"]),
        ("CPF", _emp()["socio_cpf"]),
        ("Participacao", _emp()["socio_quotas"]),
        ("Data de Nascimento", _emp()["socio_nascimento"]),
        ("Endereco Residencial", _emp()["socio_endereco"]),
        ("Funcao", "—"),
    ]
    for k, v in socio:
        ws.cell(row=r, column=1, value=k).font = Font(bold=True)
        ws.cell(row=r, column=2, value=str(v))
        ws.merge_cells(f"B{r}:E{r}")
        for c in range(1, 6):
            ws.cell(row=r, column=c).border = THIN_BORDER
            if r % 2 == 0:
                ws.cell(row=r, column=c).fill = ZEBRA_FILL
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="DIVERGENCIAS IDENTIFICADAS").font = Font(bold=True, size=11, color="B33A3A")
    ws.merge_cells(f"A{r}:E{r}")
    r += 1
    # Divergências DATA-DRIVEN — calculadas dos dados, não hardcoded.
    _anual = anualizado
    _mult = _anual / 4_800_000
    _cap = _emp().get("capital_social", 0) or 0
    divergencias = []
    if _mult > 1:
        divergencias.append(("[!] Porte EPP vs Movimentacao",
                             f"Teto EPP: R$ 4.800.000/ano | Anualizado: R$ {format_brl(_anual)} ({_mult:.1f}x o teto)"))
    if _cap > 0 and _anual > 0:
        divergencias.append(("[!] Capital vs Giro",
                             f"Capital R$ {format_brl(_cap)} | Giro anualizado R$ {format_brl(_anual)} (razao 1:{_anual / _cap:.0f})"))
    if _mult > 1:
        divergencias.append(("[!] Regime tributario",
                             "Volume anualizado pode exceder o sublimite Simples/EPP — verificar enquadramento"))
    if _emp().get("razao_anterior", "—") not in ("—", "", None):
        divergencias.append(("[!] Mudanca de Razao Social", f"Anterior: {_emp()['razao_anterior']}"))
    if not divergencias:
        ws.cell(row=r, column=1, value="Nenhuma divergencia identificada nos testes deterministicos.").font = Font(italic=True, color="2F7D4F")
        ws.merge_cells(f"A{r}:E{r}")
        r += 1
    for k, v in divergencias:
        ws.cell(row=r, column=1, value=k).font = Font(bold=True, color="B33A3A")
        ws.cell(row=r, column=1).fill = ALERT_FILL
        ws.cell(row=r, column=2, value=v).fill = ALERT_FILL
        ws.merge_cells(f"B{r}:E{r}")
        for c in range(1, 6):
            ws.cell(row=r, column=c).border = THIN_BORDER
        r += 1

    for col, w in {1: 32, 2: 60, 3: 10, 4: 10, 5: 10}.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = f"A{start + 2}"

    # ════════════════════════════════════════════════════════════════════
    # Aba 3: Resumo Executivo (KPIs + Evolucao Mensal)
    # ════════════════════════════════════════════════════════════════════
    ws = wb.create_sheet("3. Resumo Executivo")
    start = cabecalho(ws, 7, "Resumo Executivo")
    ws.cell(row=start, column=1, value="INDICADORES PRINCIPAIS").font = TITLE_FONT
    ws.merge_cells(f"A{start}:G{start}")

    r = start + 2
    ws.cell(row=r, column=1, value="Indicador")
    ws.cell(row=r, column=2, value="Valor")
    style_header(ws, r, 2)
    r += 1
    kpis = [
        (f"Total de transacoes ({n_meses} meses)", n_total),
        ("Volume de creditos", cred_total),
        ("Volume de debitos", deb_total),
        ("Volume bruto movimentado", volume_bruto),
        ("Volume liquido (excl. transf. internas)", volume_liquido),
        ("Saldo inicial do periodo", saldo_ini_jan),
        ("Saldo final do periodo", saldo_fim_mai),
        ("Variacao do periodo", saldo_fim_mai - saldo_ini_jan),
        ("Volume anualizado projetado", anualizado),
        ("Limite EPP (referencia)", 4_800_000),
        ("Multiplo do teto EPP", multiplo),
    ]
    for k, v in kpis:
        ws.cell(row=r, column=1, value=k).font = Font(bold=True)
        c = ws.cell(row=r, column=2, value=v)
        if isinstance(v, (int, float)) and "Volume" in k or "Saldo" in k or "Variacao" in k or "Limite" in k:
            c.number_format = "#,##0.00"
        elif "Multiplo" in k:
            c.number_format = "0.0\\x"
            c.font = Font(bold=True, color="B33A3A")
        for col in range(1, 3):
            ws.cell(row=r, column=col).border = THIN_BORDER
            if r % 2 == 0:
                ws.cell(row=r, column=col).fill = ZEBRA_FILL
        r += 1

    # Evolucao mensal
    r += 2
    ws.cell(row=r, column=1, value="EVOLUCAO MENSAL").font = SUBTITLE_FONT
    ws.merge_cells(f"A{r}:G{r}")
    r += 1
    headers_m = ["Mes", "Transacoes", "Creditos (R$)", "Debitos (R$)", "Fluxo Liquido", "Saldo Final", "Var. Mes Ant."]
    for c, h in enumerate(headers_m, start=1):
        ws.cell(row=r, column=c, value=h)
    style_header(ws, r, 7)
    r += 1

    saldo_anterior = saldo_ini_jan
    for mes in meses:
        s = saldos[mes]
        var = s["saldo_final"] - saldo_anterior
        ws.cell(row=r, column=1, value=mes).font = Font(bold=True)
        ws.cell(row=r, column=2, value=s["n"]).number_format = "#,##0"
        ws.cell(row=r, column=3, value=round(s["cred"], 2)).number_format = "#,##0.00"
        ws.cell(row=r, column=3).font = Font(color="2F7D4F")
        ws.cell(row=r, column=4, value=round(s["deb"], 2)).number_format = "#,##0.00"
        ws.cell(row=r, column=4).font = Font(color="B33A3A")
        fl = s["cred"] + s["deb"]
        ws.cell(row=r, column=5, value=round(fl, 2)).number_format = "#,##0.00"
        ws.cell(row=r, column=5).font = Font(bold=True, color="B33A3A" if fl < 0 else "2F7D4F")
        ws.cell(row=r, column=6, value=round(s["saldo_final"], 2)).number_format = "#,##0.00"
        ws.cell(row=r, column=6).font = Font(bold=True, color="B33A3A" if s["saldo_final"] < 0 else "12345E")
        ws.cell(row=r, column=7, value=round(var, 2)).number_format = "#,##0.00"
        ws.cell(row=r, column=7).font = Font(color="B33A3A" if var < 0 else "2F7D4F")
        for c in range(1, 8):
            ws.cell(row=r, column=c).border = THIN_BORDER
            if r % 2 == 0:
                ws.cell(row=r, column=c).fill = ZEBRA_FILL
        saldo_anterior = s["saldo_final"]
        r += 1

    # Linha total
    ws.cell(row=r, column=1, value="TOTAL").font = TOTAL_FONT
    ws.cell(row=r, column=2, value=n_total).number_format = "#,##0"
    ws.cell(row=r, column=3, value=round(cred_total, 2)).number_format = "#,##0.00"
    ws.cell(row=r, column=4, value=round(deb_total, 2)).number_format = "#,##0.00"
    ws.cell(row=r, column=5, value=round(cred_total + deb_total, 2)).number_format = "#,##0.00"
    ws.cell(row=r, column=6, value=round(saldo_fim_mai, 2)).number_format = "#,##0.00"
    for c in range(1, 8):
        ws.cell(row=r, column=c).fill = TOTAL_FILL
        ws.cell(row=r, column=c).font = TOTAL_FONT

    for col, w in {1: 28, 2: 14, 3: 18, 4: 18, 5: 18, 6: 18, 7: 18}.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = f"A{start + 2}"

    # ════════════════════════════════════════════════════════════════════
    # Aba 4: Transacoes (extrato com saldo acumulado)
    # ════════════════════════════════════════════════════════════════════
    ws = wb.create_sheet("4. Transações")
    start = cabecalho(ws, 9, "Transacoes")
    ws.cell(row=start, column=1, value=f"EXTRATO DETALHADO - {format_num(n_total)} TRANSACOES").font = TITLE_FONT
    ws.merge_cells(f"A{start}:I{start}")
    ws.cell(row=start+1, column=1, value=f"Saldo inicial: R$ {format_brl(saldo_ini_jan)} | Saldo final: R$ {format_brl(saldo_fim_mai)}").font = Font(italic=True, color="64748B", size=9)
    ws.merge_cells(f"A{start+1}:I{start+1}")

    headers_t = ["#", "Mes", "Data", "Tipo", "Valor (R$)", "Memo", "Nome", "Contraparte (RFB)", "Saldo Acumulado (R$)"]
    r = start + 3
    for c, h in enumerate(headers_t, start=1):
        ws.cell(row=r, column=c, value=h)
    style_header(ws, r, 9)
    r += 1

    txs_ord = sorted(todas_disps, key=lambda x: x.transacao.data)
    saldo_corrente = saldo_ini_jan
    for i, d in enumerate(txs_ord, start=1):
        t = d.transacao
        saldo_corrente += t.valor
        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=2, value=d.mes)
        dc = ws.cell(row=r, column=3, value=_parse_data_xlsx(t.data))
        dc.number_format = "DD/MM/YYYY"
        ws.cell(row=r, column=4, value=t.tipo)
        cv = ws.cell(row=r, column=5, value=round(t.valor, 2))
        cv.number_format = "#,##0.00"
        cv.font = Font(color=("B33A3A" if t.valor < 0 else "2F7D4F"))
        ws.cell(row=r, column=6, value=t.memo or "")
        ws.cell(row=r, column=7, value=t.nome or "")
        cc = ws.cell(row=r, column=8, value=d.contraparte or "")
        if d.disposicao == "ALERTA_POS_BAIXA":
            cc.font = Font(bold=True, color="B33A3A")
        cs = ws.cell(row=r, column=9, value=round(saldo_corrente, 2))
        cs.number_format = "#,##0.00"
        cs.font = Font(bold=True, color=("B33A3A" if saldo_corrente < 0 else "12345E"))
        for c in range(1, 10):
            ws.cell(row=r, column=c).border = THIN_BORDER
            if d.disposicao == "ALERTA_POS_BAIXA":
                ws.cell(row=r, column=c).fill = ALERT_FILL
        r += 1

    for col, w in {1: 5, 2: 10, 3: 12, 4: 8, 5: 14, 6: 32, 7: 30, 8: 35, 9: 20}.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = f"A{start + 4}"
    ws.auto_filter.ref = f"A{start + 3}:I{r - 1}"

    # ════════════════════════════════════════════════════════════════════
    # Aba 5: Disposicoes Forenses (27 colunas)
    # ════════════════════════════════════════════════════════════════════
    ws = wb.create_sheet("5. Disposições")
    start = cabecalho(ws, 27, "Disposicoes Forenses")
    ws.cell(row=start, column=1, value="DISPOSICOES POR TRANSACAO - Auditoria Forense").font = TITLE_FONT
    ws.merge_cells(f"A{start}:AA{start}")
    ws.cell(row=start+1, column=1, value="Eixos: A=Compliance | B=Identificacao | C=Padroes | D=Risk Score | E=Rastreabilidade").font = Font(italic=True, color="64748B", size=9)
    ws.merge_cells(f"A{start+1}:AA{start+1}")

    headers = [
        "Data", "Tipo", "Valor (R$)", "Memo", "Nome (banco)",
        "FITID", "CheckNum", "Meio",
        "CNPJ", "Contraparte (RFB)",
        "CNAE", "UF", "Municipio", "Porte",
        "Acumulado Mes (R$)", "1a Vez?", "Valor Redondo", "Smurfing", "Carrossel",
        "Disposicao", "Flag",
        "Risk Score", "Risk Class",
        "Periodo Fiscal", "Hash Linha", "Status Revisao", "Comentario Revisor",
    ]
    r = start + 3
    for c, h in enumerate(headers, start=1):
        ws.cell(row=r, column=c, value=h)
    style_header(ws, r, len(headers))
    r += 1

    for d in sorted(todas_disps, key=lambda x: x.transacao.data):
        t = d.transacao
        cnpj = d.cnpj
        cnpj_fmt = ""
        if cnpj:
            cnpj_fmt = f"{cnpj[:2]}.{cnpj[2:5]}.{cnpj[5:8]}/{cnpj[8:12]}-{cnpj[12:14]}"
        info = d.info_cnpj
        cnae = info.get("cnae_descricao", "")
        uf = info.get("uf", "")
        municipio = info.get("municipio", "")
        porte = info.get("porte", "")
        sit = info.get("situacao", "")
        meio = d.meio
        mes = t.data[:7]
        acum = agg.acumulado_mes.get((cnpj, mes), 0.0) if cnpj else 0.0
        pv = detectar_primeira_vez(cnpj, t.data, agg)
        vr = detectar_valor_redondo(t.valor)
        sm = detectar_smurfing(cnpj, t.data, agg)
        car = detectar_carrossel(cnpj, agg)
        score, classe = calcular_risk_score(t.valor, d.disposicao, sit, porte, meio, vr, sm, car, pv, acum)
        pf = periodo_fiscal(t.data)
        h_linha = hash_linha(t.data, t.valor, t.memo or "", t.fitid or "")
        is_alerta = d.disposicao == "ALERTA_POS_BAIXA"
        is_critico = classe == "CRITICO"

        # Preenche linha
        dc2 = ws.cell(row=r, column=1, value=_parse_data_xlsx(t.data))
        dc2.number_format = "DD/MM/YYYY"
        ws.cell(row=r, column=2, value=t.tipo)
        cv = ws.cell(row=r, column=3, value=round(t.valor, 2))
        cv.number_format = "#,##0.00"
        cv.font = Font(color=("B33A3A" if t.valor < 0 else "2F7D4F"), bold=is_alerta or is_critico)
        ws.cell(row=r, column=4, value=t.memo or "")
        ws.cell(row=r, column=5, value=t.nome or "")
        ws.cell(row=r, column=6, value=t.fitid or "").font = Font(name="Consolas", size=9, color="64748B")
        ws.cell(row=r, column=7, value=t.checknum or "")
        ws.cell(row=r, column=8, value=meio)
        ws.cell(row=r, column=9, value=cnpj_fmt).font = Font(name="Consolas", size=10)
        ws.cell(row=r, column=10, value=d.contraparte or "")
        ws.cell(row=r, column=11, value=cnae[:60])
        ws.cell(row=r, column=12, value=uf)
        ws.cell(row=r, column=13, value=municipio)
        ws.cell(row=r, column=14, value=porte)
        ws.cell(row=r, column=15, value=round(acum, 2)).number_format = "#,##0.00"
        ws.cell(row=r, column=16, value=pv)
        ws.cell(row=r, column=17, value=vr)
        ws.cell(row=r, column=18, value=sm)
        ws.cell(row=r, column=19, value=car)
        ws.cell(row=r, column=20, value=d.disposicao)
        ws.cell(row=r, column=21, value=d.flag or "")
        c_score = ws.cell(row=r, column=22, value=score)
        c_score.number_format = "0"
        c_classe = ws.cell(row=r, column=23, value=classe)
        c_classe.font = Font(bold=True)
        fill_classe = {"CRITICO": "FEE2E2", "ALTO": "FEF3C7", "MEDIO": "DBEAFE", "BAIXO": "DCFCE7"}.get(classe)
        if fill_classe:
            c_classe.fill = PatternFill("solid", fgColor=fill_classe)
        ws.cell(row=r, column=24, value=pf)
        ws.cell(row=r, column=25, value=h_linha).font = Font(name="Consolas", size=8, color="94A3B8")
        ws.cell(row=r, column=26, value="PENDENTE").font = Font(italic=True, color="64748B", size=9)
        ws.cell(row=r, column=27, value="")

        for c in range(1, 28):
            ws.cell(row=r, column=c).border = THIN_BORDER
            if is_alerta and c != 23:
                ws.cell(row=r, column=c).fill = ALERT_FILL
        r += 1

    larguras = {1: 11, 2: 7, 3: 13, 4: 28, 5: 28, 6: 18, 7: 10, 8: 11, 9: 19, 10: 32,
                11: 38, 12: 5, 13: 22, 14: 22, 15: 16, 16: 8, 17: 13, 18: 22, 19: 11,
                20: 22, 21: 32, 22: 10, 23: 10, 24: 12, 25: 17, 26: 13, 27: 30}
    for col, w in larguras.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = f"F{start + 4}"
    ws.auto_filter.ref = f"A{start + 3}:AA{r - 1}"

    # ════════════════════════════════════════════════════════════════════
    # Aba 6: Risk Heatmap
    # ════════════════════════════════════════════════════════════════════
    ws = wb.create_sheet("6. Risk Heatmap")
    start = cabecalho(ws, 6, "Risk Heatmap")
    ws.cell(row=start, column=1, value="DISTRIBUICAO POR CLASSE DE RISCO").font = TITLE_FONT
    ws.merge_cells(f"A{start}:F{start}")

    classe_counts = {"CRITICO": [0, 0.0], "ALTO": [0, 0.0], "MEDIO": [0, 0.0], "BAIXO": [0, 0.0]}
    for d in todas_disps:
        t = d.transacao
        cnpj = d.cnpj
        info = d.info_cnpj
        sit = info.get("situacao", "")
        porte = info.get("porte", "")
        meio = d.meio
        mes = t.data[:7]
        acum = agg.acumulado_mes.get((cnpj, mes), 0.0) if cnpj else 0.0
        pv = detectar_primeira_vez(cnpj, t.data, agg)
        vr = detectar_valor_redondo(t.valor)
        sm = detectar_smurfing(cnpj, t.data, agg)
        car = detectar_carrossel(cnpj, agg)
        _, classe = calcular_risk_score(t.valor, d.disposicao, sit, porte, meio, vr, sm, car, pv, acum)
        classe_counts[classe][0] += 1
        classe_counts[classe][1] += abs(t.valor)

    r = start + 2
    headers_h = ["Classe", "Qtd Transacoes", "% do Total", "Volume (R$)", "% Volume", "Acao Sugerida"]
    for c, h in enumerate(headers_h, start=1):
        ws.cell(row=r, column=c, value=h)
    style_header(ws, r, 6)
    r += 1

    cores = {
        "CRITICO": ("B33A3A", "FEE2E2", "Auditoria imediata - investigar"),
        "ALTO":    ("D97706", "FEF3C7", "Revisão prioritária"),
        "MEDIO":   ("1F7FB8", "DBEAFE", "Conferir em lote"),
        "BAIXO":   ("2F7D4F", "DCFCE7", "Auto-aprovar apos confirmacao"),
    }
    total_qtd = sum(v[0] for v in classe_counts.values())
    total_vol = sum(v[1] for v in classe_counts.values())
    for classe in ("CRITICO", "ALTO", "MEDIO", "BAIXO"):
        qtd, vol = classe_counts[classe]
        cor, fill_cor, acao = cores[classe]
        c1 = ws.cell(row=r, column=1, value=classe)
        c1.font = Font(bold=True, color=cor)
        c1.fill = PatternFill("solid", fgColor=fill_cor)
        ws.cell(row=r, column=2, value=qtd).number_format = "#,##0"
        ws.cell(row=r, column=3, value=qtd / max(total_qtd, 1)).number_format = "0.0%"
        ws.cell(row=r, column=4, value=round(vol, 2)).number_format = "#,##0.00"
        ws.cell(row=r, column=5, value=vol / max(total_vol, 1)).number_format = "0.0%"
        ws.cell(row=r, column=6, value=acao)
        for c in range(1, 7):
            ws.cell(row=r, column=c).border = THIN_BORDER
        r += 1

    ws.cell(row=r, column=1, value="TOTAL").font = TOTAL_FONT
    ws.cell(row=r, column=2, value=total_qtd).number_format = "#,##0"
    ws.cell(row=r, column=4, value=round(total_vol, 2)).number_format = "#,##0.00"
    for c in range(1, 7):
        ws.cell(row=r, column=c).fill = TOTAL_FILL
        ws.cell(row=r, column=c).font = TOTAL_FONT

    for col, w in {1: 12, 2: 16, 3: 12, 4: 18, 5: 11, 6: 36}.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = f"A{start + 2}"

    # ════════════════════════════════════════════════════════════════════
    # Aba 7: CNPJs Enriquecidos
    # ════════════════════════════════════════════════════════════════════
    ws = wb.create_sheet("7. CNPJs")
    start = cabecalho(ws, 8, "CNPJs Enriquecidos")
    ws.cell(row=start, column=1, value="CONTRAPARTES IDENTIFICADAS VIA BASE RFB").font = TITLE_FONT
    ws.merge_cells(f"A{start}:H{start}")

    cnpjs_unicos_usados = {d.cnpj for d in todas_disps if d.cnpj}
    cnpjs_unicos_usados.discard(None)

    headers_c = ["CNPJ", "Razao Social", "Situacao", "Data Situacao", "UF",
                 "Municipio", "CNAE", "Porte"]
    r = start + 2
    for c, h in enumerate(headers_c, start=1):
        ws.cell(row=r, column=c, value=h)
    style_header(ws, r, 8)
    r += 1

    # Agrega por aparicoes
    aparicoes = Counter(d.cnpj for d in todas_disps if d.cnpj)
    for cnpj in sorted(cnpjs_unicos_usados, key=lambda c: -aparicoes[c]):
        info = cache.get(cnpj, {})
        fmt = f"{cnpj[:2]}.{cnpj[2:5]}.{cnpj[5:8]}/{cnpj[8:12]}-{cnpj[12:14]}"
        is_baixada = "BAIXADA" in info.get("situacao", "") or "INAPTA" in info.get("situacao", "")
        ws.cell(row=r, column=1, value=fmt).font = Font(name="Consolas", size=10)
        ws.cell(row=r, column=2, value=info.get("razao_social", "(nao enriquecido)"))
        c_sit = ws.cell(row=r, column=3, value=info.get("situacao", ""))
        if is_baixada:
            c_sit.font = Font(bold=True, color="B33A3A")
        dc3 = ws.cell(row=r, column=4, value=_parse_data_xlsx(info.get("data_situacao", "")))
        dc3.number_format = "DD/MM/YYYY"
        ws.cell(row=r, column=5, value=info.get("uf", ""))
        ws.cell(row=r, column=6, value=info.get("municipio", ""))
        ws.cell(row=r, column=7, value=info.get("cnae_descricao", "")[:55])
        ws.cell(row=r, column=8, value=info.get("porte", ""))
        for c in range(1, 9):
            ws.cell(row=r, column=c).border = THIN_BORDER
            if is_baixada:
                ws.cell(row=r, column=c).fill = ALERT_FILL
            elif r % 2 == 0:
                ws.cell(row=r, column=c).fill = ZEBRA_FILL
        r += 1

    for col, w in {1: 20, 2: 42, 3: 14, 4: 13, 5: 5, 6: 22, 7: 38, 8: 24}.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = f"A{start + 3}"
    ws.auto_filter.ref = f"A{start + 2}:H{r - 1}"

    # ════════════════════════════════════════════════════════════════════
    # Aba 8: Partes Relacionadas
    # ════════════════════════════════════════════════════════════════════
    ws = wb.create_sheet("8. Partes Relacionadas")
    start = cabecalho(ws, 5, "Partes Relacionadas")
    razao = _emp().get("razao_social", "a entidade")
    ws.cell(row=start, column=1, value=f"MOVIMENTACAO COM PARTES RELACIONADAS — {razao}").font = TITLE_FONT
    ws.merge_cells(f"A{start}:E{start}")

    # Fluxos data-driven: auto-movimentacao (proprio CNPJ) + mesma titularidade.
    # Partes relacionadas nominais (coligadas/socios) exigem cadastro manual.
    cnpj_basico = _emp().get("cnpj_basico", "")
    fluxos = {
        "Auto-movimentacao (proprio CNPJ)": {"n": 0, "cred": 0.0, "deb": 0.0},
        "Mesma titularidade (transf. entre contas proprias)": {"n": 0, "cred": 0.0, "deb": 0.0},
    }
    for d in todas_disps:
        texto_up = ((d.transacao.nome or "") + " " + (d.transacao.memo or "")).upper()
        if cnpj_basico and (d.cnpj == cnpj_basico or cnpj_basico in re.sub(r"\D", "", texto_up)):
            target = fluxos["Auto-movimentacao (proprio CNPJ)"]
        elif "MESMA TIT" in texto_up or "MESMA TITULAR" in texto_up:
            target = fluxos["Mesma titularidade (transf. entre contas proprias)"]
        else:
            continue
        target["n"] += 1
        if d.transacao.valor > 0:
            target["cred"] += d.transacao.valor
        else:
            target["deb"] += d.transacao.valor

    r = start + 2
    h = ["Entidade", "Qtd", "Creditos (R$)", "Debitos (R$)", "Volume Total (R$)"]
    for c, hd in enumerate(h, start=1):
        ws.cell(row=r, column=c, value=hd)
    style_header(ws, r, 5)
    r += 1
    total_pr = 0.0
    for nome, dados in fluxos.items():
        vol = dados["cred"] + abs(dados["deb"])
        ws.cell(row=r, column=1, value=nome)
        ws.cell(row=r, column=2, value=dados["n"]).number_format = "#,##0"
        ws.cell(row=r, column=3, value=round(dados["cred"], 2)).number_format = "#,##0.00"
        ws.cell(row=r, column=3).font = Font(color="2F7D4F")
        ws.cell(row=r, column=4, value=round(dados["deb"], 2)).number_format = "#,##0.00"
        ws.cell(row=r, column=4).font = Font(color="B33A3A")
        ws.cell(row=r, column=5, value=round(vol, 2)).number_format = "#,##0.00"
        ws.cell(row=r, column=5).font = Font(bold=True)
        for c in range(1, 6):
            ws.cell(row=r, column=c).border = THIN_BORDER
            if r % 2 == 0:
                ws.cell(row=r, column=c).fill = ZEBRA_FILL
        total_pr += vol
        r += 1

    ws.cell(row=r, column=1, value="TOTAL").font = TOTAL_FONT
    ws.cell(row=r, column=5, value=round(total_pr, 2)).number_format = "#,##0.00"
    for c in range(1, 6):
        ws.cell(row=r, column=c).fill = TOTAL_FILL
        ws.cell(row=r, column=c).font = TOTAL_FONT

    for col, w in {1: 50, 2: 8, 3: 22, 4: 22, 5: 22}.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = f"A{start + 2}"

    # ════════════════════════════════════════════════════════════════════
    # Aba 9: MEIs Teto (com diferenciacao MEI-TAC vs MEI Padrao)
    # ════════════════════════════════════════════════════════════════════
    ws = wb.create_sheet("9. MEIs Teto")
    start = cabecalho(ws, 10, "MEIs Estourando Teto (MEI-TAC vs Padrao)")
    ws.cell(row=start, column=1, value="ANALISE DE MEIs - MEI-TAC (R$ 251.600/ano) vs MEI Padrao (R$ 81.000/ano)").font = TITLE_FONT
    ws.merge_cells(f"A{start}:J{start}")
    ws.cell(row=start+1, column=1, value="Lei Complementar 188/2021: MEI-TAC (caminhoneiros, CNAEs 4930-*, 5320-*, 4911-*) tem teto de R$ 251.600/ano. MEI padrao mantem R$ 81.000/ano.").font = Font(italic=True, color="64748B", size=9)
    ws.merge_cells(f"A{start+1}:J{start+1}")

    # Agrega todos os MEIs (TAC e padrao)
    por_cnpj_mei = defaultdict(lambda: {"n": 0, "deb": 0.0})
    for d in todas_disps:
        if d.transacao.valor >= 0 or not d.cnpj:
            continue
        if d.info_cnpj.get("porte") != "MICRO EMPRESA":
            continue
        por_cnpj_mei[d.cnpj]["n"] += 1
        por_cnpj_mei[d.cnpj]["deb"] += abs(d.transacao.valor)

    # Classifica cada MEI conforme CNAE
    meis_tac_ok, meis_tac_estourados = [], []
    meis_padrao_ok, meis_padrao_estourados = [], []
    for cnpj, dd in por_cnpj_mei.items():
        info = cache.get(cnpj, {})
        cnae = info.get("cnae_principal", "")
        anualizado = dd["deb"] * 12 / max(meses_obs, 1)
        teto = _limite_mei_por_cnae(cnae)
        eh_tac = teto == LIMITE_MEI_TAC
        excesso = anualizado - teto if anualizado > teto else 0.0
        item = {
            "cnpj": cnpj, "razao": info.get("razao_social", ""),
            "cnae": cnae, "cnae_desc": info.get("cnae_descricao", ""),
            "uf": info.get("uf", ""), "n": dd["n"],
            "deb_5m": dd["deb"], "anualizado": anualizado,
            "teto": teto, "excesso": excesso, "eh_tac": eh_tac,
        }
        if eh_tac and excesso > 0:
            meis_tac_estourados.append(item)
        elif eh_tac:
            meis_tac_ok.append(item)
        elif excesso > 0:
            meis_padrao_estourados.append(item)
        else:
            meis_padrao_ok.append(item)

    # Sumario
    r = start + 3
    ws.cell(row=r, column=1, value="SUMARIO DA RECLASSIFICACAO").font = SUBTITLE_FONT
    ws.merge_cells(f"A{r}:J{r}")
    r += 1
    hdr_sum = ["Categoria", "Teto Legal", "Total MEIs", "Dentro do Teto", "Acima do Teto", "Status"]
    for c, h in enumerate(hdr_sum, start=1):
        ws.cell(row=r, column=c, value=h)
    style_header(ws, r, 6)
    r += 1
    linhas_sum = [
        ("MEI-TAC (caminhoneiros)", f"R$ {format_brl(LIMITE_MEI_TAC)}/ano",
         len(meis_tac_ok) + len(meis_tac_estourados),
         len(meis_tac_ok), len(meis_tac_estourados),
         "OK" if not meis_tac_estourados else "ATENCAO"),
        ("MEI Padrao (outros CNAEs)", f"R$ {format_brl(LIMITE_MEI_PADRAO)}/ano",
         len(meis_padrao_ok) + len(meis_padrao_estourados),
         len(meis_padrao_ok), len(meis_padrao_estourados),
         "OK" if not meis_padrao_estourados else "ATENCAO"),
        ("TOTAL", "—",
         len(por_cnpj_mei), len(meis_tac_ok) + len(meis_padrao_ok),
         len(meis_tac_estourados) + len(meis_padrao_estourados), "—"),
    ]
    for cat, teto, total, dentro, acima, status in linhas_sum:
        c1 = ws.cell(row=r, column=1, value=cat)
        c1.font = Font(bold=True)
        ws.cell(row=r, column=2, value=teto)
        ws.cell(row=r, column=3, value=total).number_format = "#,##0"
        ws.cell(row=r, column=4, value=dentro).number_format = "#,##0"
        ws.cell(row=r, column=4).font = Font(color="2F7D4F", bold=True)
        ws.cell(row=r, column=5, value=acima).number_format = "#,##0"
        if acima > 0:
            ws.cell(row=r, column=5).font = Font(color="B33A3A", bold=True)
        c_status = ws.cell(row=r, column=6, value=status)
        if status == "OK":
            c_status.fill = SUCCESS_FILL
            c_status.font = Font(bold=True, color="2F7D4F")
        elif status == "ATENCAO":
            c_status.fill = ALERT_FILL_MEDIO
            c_status.font = Font(bold=True, color="D97706")
        for c in range(1, 7):
            ws.cell(row=r, column=c).border = THIN_BORDER
        if cat == "TOTAL":
            for c in range(1, 7):
                ws.cell(row=r, column=c).fill = TOTAL_FILL
                ws.cell(row=r, column=c).font = TOTAL_FONT
        r += 1

    # Detalhamento dos casos acima do teto
    r += 1
    ws.cell(row=r, column=1, value="DETALHAMENTO - CASOS ACIMA DO TETO").font = SUBTITLE_FONT
    ws.merge_cells(f"A{r}:J{r}")
    r += 1

    headers = ["#", "Tipo", "CNPJ", "Razao Social", "CNAE", "UF",
               "Trans.", "Pago 5m (R$)", "Anualizado (R$)", "Excesso (R$)"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=r, column=c, value=h)
    style_header(ws, r, 10)
    r += 1

    todos_estourados = (
        [(m, "MEI-TAC") for m in meis_tac_estourados] +
        [(m, "MEI Padrao") for m in meis_padrao_estourados]
    )
    todos_estourados.sort(key=lambda x: -x[0]["anualizado"])
    total_excesso = 0.0

    if not todos_estourados:
        ws.cell(row=r, column=1, value="(nenhum MEI excede o teto legal correspondente)").font = Font(italic=True, color="2F7D4F")
    else:
        for i, (m, tipo) in enumerate(todos_estourados, start=1):
            cnpj_fmt = f"{m['cnpj'][:2]}.{m['cnpj'][2:5]}.{m['cnpj'][5:8]}/{m['cnpj'][8:12]}-{m['cnpj'][12:14]}"
            ws.cell(row=r, column=1, value=i)
            c_tipo = ws.cell(row=r, column=2, value=tipo)
            c_tipo.font = Font(bold=True)
            ws.cell(row=r, column=3, value=cnpj_fmt).font = Font(name="Consolas", size=10)
            ws.cell(row=r, column=4, value=m["razao"][:45])
            ws.cell(row=r, column=5, value=m["cnae_desc"][:35] if m["cnae_desc"] else m["cnae"])
            ws.cell(row=r, column=6, value=m["uf"])
            ws.cell(row=r, column=7, value=m["n"]).number_format = "#,##0"
            ws.cell(row=r, column=8, value=round(m["deb_5m"], 2)).number_format = "#,##0.00"
            c9 = ws.cell(row=r, column=9, value=round(m["anualizado"], 2))
            c9.number_format = "#,##0.00"
            c9.font = Font(bold=True, color="B33A3A")
            c10 = ws.cell(row=r, column=10, value=round(m["excesso"], 2))
            c10.number_format = "#,##0.00"
            c10.font = Font(bold=True, color="B33A3A")
            for c in range(1, 11):
                ws.cell(row=r, column=c).border = THIN_BORDER
                if m["excesso"] > 100_000:
                    ws.cell(row=r, column=c).fill = ALERT_FILL
                elif r % 2 == 0:
                    ws.cell(row=r, column=c).fill = ZEBRA_FILL
            total_excesso += m["excesso"]
            r += 1

        ws.cell(row=r, column=1, value=f"TOTAL ({len(todos_estourados)} MEIs)").font = TOTAL_FONT
        ws.cell(row=r, column=10, value=round(total_excesso, 2)).number_format = "#,##0.00"
        for c in range(1, 11):
            ws.cell(row=r, column=c).fill = TOTAL_FILL
            ws.cell(row=r, column=c).font = TOTAL_FONT

    for col, w in {1: 4, 2: 12, 3: 20, 4: 38, 5: 32, 6: 5, 7: 8, 8: 16, 9: 17, 10: 17}.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = f"A{start + 4}"
    # Mantem para compatibilidade com codigo posterior
    meis = [m for m, _ in todos_estourados]

    # ════════════════════════════════════════════════════════════════════
    # Aba 10: Status Tributario
    # ════════════════════════════════════════════════════════════════════
    ws = wb.create_sheet("10. Status Tributário")
    start = cabecalho(ws, ncol_trib, "Status Tributario")
    ws.cell(row=start, column=1, value=f"STATUS TRIBUTARIO CONSOLIDADO - {n_meses} MESES").font = TITLE_FONT
    ws.merge_cells(f"A{start}:{get_column_letter(ncol_trib)}{start}")

    cat_count = Counter()
    cat_volume = defaultdict(float)
    cat_retencao = defaultdict(float)
    cat_por_mes = defaultdict(lambda: defaultdict(float))

    for d in todas_disps:
        t = d.transacao
        porte = d.info_cnpj.get("porte", "")
        trib = classificar_tributario(t.memo or "", t.nome or "", t.valor, d.cnpj or "", porte)
        cat_count[trib["categoria"]] += 1
        cat_volume[trib["categoria"]] += abs(t.valor)
        cat_retencao[trib["categoria"]] += trib["valor_retencao"]
        cat_por_mes[d.mes][trib["categoria"]] += trib["valor_retencao"]

    r = start + 2
    headers = ["Categoria", "Qtd", "Volume (R$)", "Retencao (R$)"] + meses_curto
    for c, h in enumerate(headers, start=1):
        ws.cell(row=r, column=c, value=h)
    style_header(ws, r, ncol_trib)
    r += 1

    CATS = ["RETENCAO_PJ", "RETENCAO_PF", "OPERACAO_CREDITO", "IOF", "JUROS",
            "PAGAMENTO_TRIBUTO", "TARIFA", "PIX_RECEBIDO", "BOLETO",
            "COMPRA_CARTAO", "NAO_TRIBUTAVEL", "OUTRO"]
    total_ret_5m = 0.0
    for cat in CATS:
        qtd = cat_count.get(cat, 0)
        if qtd == 0:
            continue
        c1 = ws.cell(row=r, column=1, value=cat)
        c1.font = Font(bold=True, color="B33A3A" if cat.startswith("RETENCAO") else "12345E")
        ws.cell(row=r, column=2, value=qtd).number_format = "#,##0"
        ws.cell(row=r, column=3, value=round(cat_volume[cat], 2)).number_format = "#,##0.00"
        cret = ws.cell(row=r, column=4, value=round(cat_retencao[cat], 2))
        cret.number_format = "#,##0.00"
        if cat_retencao[cat] > 0:
            cret.font = Font(bold=True, color="D97706")
            total_ret_5m += cat_retencao[cat]
        for i, mes in enumerate(meses, start=5):
            val = cat_por_mes[mes].get(cat, 0)
            cm = ws.cell(row=r, column=i, value=round(val, 2) if val else "")
            if val:
                cm.number_format = "#,##0.00"
        for c in range(1, ncol_trib + 1):
            ws.cell(row=r, column=c).border = THIN_BORDER
            if r % 2 == 0:
                ws.cell(row=r, column=c).fill = ZEBRA_FILL
        r += 1

    ws.cell(row=r, column=1, value="TOTAL").font = TOTAL_FONT
    ws.cell(row=r, column=2, value=sum(cat_count.values())).number_format = "#,##0"
    ws.cell(row=r, column=3, value=round(sum(cat_volume.values()), 2)).number_format = "#,##0.00"
    ws.cell(row=r, column=4, value=round(total_ret_5m, 2)).number_format = "#,##0.00"
    for i, mes in enumerate(meses, start=5):
        ws.cell(row=r, column=i, value=round(sum(cat_por_mes[mes].values()), 2)).number_format = "#,##0.00"
    for c in range(1, ncol_trib + 1):
        ws.cell(row=r, column=c).fill = TOTAL_FILL
        ws.cell(row=r, column=c).font = TOTAL_FONT

    larguras_trib = {1: 22, 2: 8, 3: 17, 4: 16}
    larguras_trib.update({c: 13 for c in range(5, ncol_trib + 1)})
    for col, w in larguras_trib.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = f"A{start + 2}"

    # ════════════════════════════════════════════════════════════════════
    # Aba 11: Pagamentos Pos-Baixa
    # ════════════════════════════════════════════════════════════════════
    ws = wb.create_sheet("11. Pós-Baixa")
    start = cabecalho(ws, 7, "Pagamentos Pos-Baixa")
    ws.cell(row=start, column=1, value="PAGAMENTOS APOS BAIXA DO CNPJ - CRITICO").font = TITLE_FONT
    ws.merge_cells(f"A{start}:G{start}")
    ws.cell(row=start+1, column=1, value="Transacoes a CNPJs ja BAIXADOS na data do pagamento - red flag forense").font = Font(italic=True, color="64748B", size=9)
    ws.merge_cells(f"A{start+1}:G{start+1}")

    pos_baixa = []
    for d in todas_disps:
        if d.disposicao == "ALERTA_POS_BAIXA":
            try:
                db_d = date.fromisoformat(d.info_cnpj["data_situacao"][:10])
                dt_d = date.fromisoformat(d.transacao.data[:10])
                pos_baixa.append({
                    "mes": d.mes, "t": d.transacao, "cnpj": d.cnpj,
                    "info": d.info_cnpj, "dias": (dt_d - db_d).days,
                })
            except (ValueError, TypeError, KeyError):
                pass

    pos_baixa.sort(key=lambda x: -x["dias"])

    headers = ["#", "Mes", "Data", "Valor (R$)", "Razao Social", "Data Baixa", "Dias Apos"]
    r = start + 3
    for c, h in enumerate(headers, start=1):
        ws.cell(row=r, column=c, value=h)
    style_header(ws, r, 7)
    r += 1
    total_pb = 0.0
    for i, p in enumerate(pos_baixa, start=1):
        cnpj_fmt = f"{p['cnpj'][:2]}.{p['cnpj'][2:5]}.{p['cnpj'][5:8]}/{p['cnpj'][8:12]}-{p['cnpj'][12:14]}"
        razao = p["info"].get("razao_social", "")[:35]
        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=2, value=p["mes"])
        dpb = ws.cell(row=r, column=3, value=_parse_data_xlsx(p["t"].data))
        dpb.number_format = "DD/MM/YYYY"
        cv = ws.cell(row=r, column=4, value=round(p["t"].valor, 2))
        cv.number_format = "#,##0.00"
        cv.font = Font(bold=True, color="B33A3A")
        ws.cell(row=r, column=5, value=f"{cnpj_fmt} - {razao}")
        db2 = ws.cell(row=r, column=6, value=_parse_data_xlsx(p["info"].get("data_situacao", "")))
        db2.number_format = "DD/MM/YYYY"
        db2.font = Font(bold=True, color="B33A3A")
        ws.cell(row=r, column=7, value=p["dias"]).font = Font(bold=True, color="B33A3A")
        ws.cell(row=r, column=7).number_format = "#,##0"
        for c in range(1, 8):
            ws.cell(row=r, column=c).border = THIN_BORDER
            ws.cell(row=r, column=c).fill = ALERT_FILL
        total_pb += abs(p["t"].valor)
        r += 1

    ws.cell(row=r, column=1, value=f"TOTAL ({len(pos_baixa)} alertas)").font = TOTAL_FONT
    ws.cell(row=r, column=4, value=round(total_pb, 2)).number_format = "#,##0.00"
    for c in range(1, 8):
        ws.cell(row=r, column=c).fill = TOTAL_FILL
        ws.cell(row=r, column=c).font = TOTAL_FONT

    for col, w in {1: 4, 2: 10, 3: 12, 4: 14, 5: 60, 6: 13, 7: 12}.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = f"A{start + 4}"

    # Abas fiscais (12, 13) — só quando há documentos NF-e/CT-e
    fiscal = None
    if nfes or ctes:
        aba_documentos_fiscais(wb, nfes or [], ctes or [])
        _cb = re.sub(r"\D", "", _emp().get("cnpj_basico", "") or "")
        nfes_receb = [n for n in (nfes or []) if re.sub(r"\D", "", n.get("emit_cnpj", "")) != _cb]
        conf = _conformidade_lista(nfes_receb, ctes or [], [t for _, t, _ in todos])
        aba_conformidade_fiscal(wb, conf)
        fiscal = {
            "n_nfe": len(nfes or []), "n_cte": len(ctes or []),
            "vol_nfe": round(sum(n["valor"] for n in (nfes or [])), 2),
            "vol_cte": round(sum(c["valor"] for c in (ctes or [])), 2),
            "criticos": [c for c in conf if c["classe"] in ("CRITICO", "ALTO")][:10],
        }

    # ── Print setup em todas as abas (Fase 3) ─────────────────────────────
    for _ws in wb.worksheets:
        _ws.page_setup.orientation = "landscape"
        _ws.page_setup.fitToWidth = 1
        _ws.page_setup.fitToHeight = 0
        _ws.sheet_properties.pageSetUpPr.fitToPage = True
        if _ws.title not in ("1. Capa",):
            _ws.print_title_rows = "1:2"

    return wb, {
        "fiscal": fiscal,
        "anualizado": anualizado, "multiplo": multiplo, "meses_obs": meses_obs,
        "n_total": n_total, "volume_bruto": volume_bruto,
        "volume_liquido": volume_liquido, "vol_transf_interna": vol_transf_interna,
        "cred_total": cred_total, "deb_total": deb_total,
        "saldo_ini": saldo_ini_jan, "saldo_fim": saldo_fim_mai,
        "saldos": saldos, "todas_disps": todas_disps,
        "classe_counts": classe_counts, "cat_count": cat_count,
        "cat_volume": cat_volume, "cat_retencao": cat_retencao,
        "fluxos": fluxos, "meis": meis, "pos_baixa": pos_baixa,
        "total_ret_5m": total_ret_5m,
        "periodo_str": periodo_str, "n_meses": n_meses, "meses": meses,
    }


# ════════════════════════════════════════════════════════════════════════
# Geracao MD / HTML / PDF
# ════════════════════════════════════════════════════════════════════════


def gerar_md(stats):
    n = stats["n_total"]
    cred = stats["cred_total"]
    deb = stats["deb_total"]
    vol = stats["volume_bruto"]
    saldo_ini = stats["saldo_ini"]
    saldo_fim = stats["saldo_fim"]

    lines = [
        f"# RELATÓRIO INTEGRADO DE AUDITORIA — {_emp()['razao_social']}",
        "",
        "**ORGATEC · Contabilidade · Auditoria · Compliance**",
        "",
        f"**Gerado em:** {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        f"**Periodo:** {stats['periodo_str']} ({stats['n_meses']} meses, {format_num(n)} transacoes)",
        f"**CNPJ:** {_emp()['cnpj']}",
        "",
        "---",
        "",
        "## 1. Sumário Executivo",
        "",
        "| Indicador | Valor |",
        "|---|---:|",
        f"| Total de transacoes | {format_num(n)} |",
        f"| Volume de creditos | R$ {format_brl(cred)} |",
        f"| Volume de debitos | R$ {format_brl(deb)} |",
        f"| Volume bruto movimentado | **R$ {format_brl(vol)}** |",
        f"| Volume liquido (excl. transf. internas) | **R$ {format_brl(stats.get('volume_liquido', vol))}** |",
        f"| Saldo inicial do periodo | R$ {format_brl(saldo_ini)} |",
        f"| Saldo final do periodo | R$ {format_brl(saldo_fim)} |",
        f"| Variacao do periodo | R$ {format_brl(saldo_fim - saldo_ini)} |",
        f"| **Volume anualizado projetado** | **R$ {format_brl(stats['anualizado'])}** |",
        "| Limite EPP (referencia) | R$ 4.800.000,00 |",
        f"| **Multiplo do teto EPP** | **{stats['multiplo']:.1f}x** |",
        f"| Alertas pos-baixa | {len(stats['pos_baixa'])} |",
        f"| Retencao estimada no periodo | R$ {format_brl(stats['total_ret_5m'])} |",
        "",
        "## 2. Identificação Cadastral",
        "",
        "### Dados da Pessoa Jurídica (Contrato Social + Cartão CNPJ)",
        "",
        "| Campo | Valor |",
        "|---|---|",
    ]
    dados_pj = [
        ("Razão Social", _emp()["razao_social"]),
        ("Razão Anterior", _emp()["razao_anterior"]),
        ("Nome Fantasia", _emp()["nome_fantasia"]),
        ("CNPJ", _emp()["cnpj"]),
        ("Situação", _emp()["situacao"]),
        ("Data Abertura", _emp()["data_abertura"]),
        ("Porte Declarado", _emp()["porte_declarado"]),
        ("Natureza Jurídica", _emp()["natureza_juridica"]),
        ("Capital Social", f"R$ {format_brl(_emp()['capital_social'])}"),
        ("CNAE Principal", _emp()["cnae_principal"]),
        ("CNAE Secundário", _emp()["cnae_secundario"]),
        ("Endereço Sede", _emp()["endereco_sede"]),
        ("Escritório Admin", _emp()["endereco_admin"]),
        ("Email", _emp()["email"]),
        ("Telefones", _emp()["telefones"]),
        ("Última Alteração", _emp()["ultima_alteracao"]),
    ]
    for k, v in dados_pj:
        lines.append(f"| **{k}** | {v} |")

    lines += [
        "",
        "### Quadro Societário",
        "",
        "| Socio | CPF | Quotas | % |",
        "|---|---|---:|---:|",
        f"| **{_emp()['socio_nome']}** | {_emp()['socio_cpf']} | {_emp()['socio_quotas']} | — |",
        "",
        f"- **Nascimento:** {_emp()['socio_nascimento']}",
        f"- **Endereco:** {_emp()['socio_endereco']}",
        "- **Função:** Administrador único por prazo indeterminado",
        "",
        "## 3. Evolução Mensal",
        "",
        "| Mês | Transações | Créditos (R$) | Débitos (R$) | Saldo Final |",
        "|---|---:|---:|---:|---:|",
    ]
    for mes in stats["meses"]:
        s = stats["saldos"][mes]
        lines.append(f"| {mes} | {format_num(s['n'])} | {format_brl(s['cred'])} | {format_brl(s['deb'])} | {format_brl(s['saldo_final'])} |")
    lines.append(f"| **TOTAL** | **{format_num(n)}** | **{format_brl(cred)}** | **{format_brl(deb)}** | **{format_brl(saldo_fim)}** |")

    # Risk Heatmap
    lines += [
        "",
        "## 4. Risk Heatmap (Distribuição por Classe)",
        "",
        "| Classe | Transações | % | Volume (R$) | Ação |",
        "|---|---:|---:|---:|---|",
    ]
    cores_md = {"CRITICO": "🔴", "ALTO": "🟠", "MEDIO": "🔵", "BAIXO": "🟢"}
    acoes_md = {"CRITICO": "Auditoria imediata", "ALTO": "Revisão prioritária",
                "MEDIO": "Conferir em lote", "BAIXO": "Auto-aprovar"}
    total_qtd = sum(v[0] for v in stats["classe_counts"].values())
    for classe in ("CRITICO", "ALTO", "MEDIO", "BAIXO"):
        qtd, vol_c = stats["classe_counts"][classe]
        pct = 100 * qtd / max(total_qtd, 1)
        lines.append(f"| {cores_md[classe]} {classe} | {format_num(qtd)} | {pct:.1f}% | {format_brl(vol_c)} | {acoes_md[classe]} |")

    # Partes Relacionadas
    lines += [
        "",
        "## 5. Partes Relacionadas",
        "",
        "| Entidade | Trans | Créditos (R$) | Débitos (R$) | Volume (R$) |",
        "|---|---:|---:|---:|---:|",
    ]
    total_pr = 0.0
    for nome, dados in stats["fluxos"].items():
        v = dados["cred"] + abs(dados["deb"])
        lines.append(f"| {nome} | {dados['n']} | {format_brl(dados['cred'])} | {format_brl(dados['deb'])} | {format_brl(v)} |")
        total_pr += v
    lines.append(f"| **TOTAL** | | | | **R$ {format_brl(total_pr)}** |")

    # MEIs (com reclassificacao MEI-TAC vs Padrao)
    lines += [
        "",
        "## 6. MEIs Fornecedores — Reclassificação MEI-TAC vs Padrão",
        "",
        "**Aplicação do limite correto** após confirmação do cliente que muitos são caminhoneiros:",
        "",
        "- **MEI-TAC** (caminhoneiros, CNAEs 4930-*, 5320-*, 4911-*): teto **R$ 251.600/ano** (LC 188/2021)",
        "- **MEI Padrão** (outros CNAEs): teto **R$ 81.000/ano** (LC 123/2006)",
        "",
        f"### Casos acima do teto correspondente ({len(stats['meis'])} fornecedores)",
        "",
    ]
    total_exc = 0.0
    if stats["meis"]:
        lines += [
            "| # | CNPJ | Razão Social | CNAE | Anualizado | Excesso |",
            "|---|---|---|---|---:|---:|",
        ]
        for i, m in enumerate(stats["meis"][:15], start=1):
            fmt = f"{m['cnpj'][:2]}.{m['cnpj'][2:5]}.{m['cnpj'][5:8]}/{m['cnpj'][8:12]}-{m['cnpj'][12:14]}"
            cnae_desc = m.get("cnae_desc", "") or m.get("cnae", "")
            lines.append(f"| {i} | {fmt} | {m['razao'][:30]} | {cnae_desc[:25]} | **R$ {format_brl(m['anualizado'])}** | R$ {format_brl(m['excesso'])} |")
            total_exc += m["excesso"]
        lines.append(f"| | | | **TOTAL EXCESSO:** | | **R$ {format_brl(total_exc)}** |")
    else:
        lines.append("**Nenhum MEI excede o teto legal correspondente.**")

    # Status Tributario
    lines += [
        "",
        "## 7. Status Tributário Consolidado",
        "",
        "| Categoria | Qtd | Volume (R$) | Retenção (R$) |",
        "|---|---:|---:|---:|",
    ]
    CATS_ORD = ["RETENCAO_PJ", "RETENCAO_PF", "OPERACAO_CREDITO", "IOF", "JUROS",
                "PAGAMENTO_TRIBUTO", "TARIFA", "PIX_RECEBIDO", "BOLETO",
                "COMPRA_CARTAO", "NAO_TRIBUTAVEL", "OUTRO"]
    for cat in CATS_ORD:
        qtd = stats["cat_count"].get(cat, 0)
        if qtd == 0:
            continue
        vol_c = stats["cat_volume"][cat]
        ret = stats["cat_retencao"][cat]
        ret_fmt = f"**{format_brl(ret)}**" if ret > 0 else "—"
        lines.append(f"| {cat} | {format_num(qtd)} | {format_brl(vol_c)} | {ret_fmt} |")
    lines.append(f"| **TOTAL** | **{format_num(sum(stats['cat_count'].values()))}** | | **R$ {format_brl(stats['total_ret_5m'])}** |")

    # Pos-Baixa
    lines += [
        "",
        f"## 8. Pagamentos Pós-Baixa ({len(stats['pos_baixa'])} alertas)",
        "",
        "| Mês | Data | Valor (R$) | Razão Social | Data Baixa | Dias Após |",
        "|---|---|---:|---|---|---:|",
    ]
    total_pb = 0.0
    for p in stats["pos_baixa"]:
        cnpj_fmt = f"{p['cnpj'][:2]}.{p['cnpj'][2:5]}.{p['cnpj'][5:8]}/{p['cnpj'][8:12]}-{p['cnpj'][12:14]}"
        razao = p["info"].get("razao_social", "")[:30]
        lines.append(f"| {p['mes']} | {p['t'].data} | **{format_brl(p['t'].valor)}** | {cnpj_fmt} - {razao} | {p['info'].get('data_situacao','')} | **{p['dias']}** |")
        total_pb += abs(p["t"].valor)
    lines.append(f"| | | **R$ {format_brl(total_pb)}** | TOTAL | | |")

    # Conclusao
    # Conclusao DATA-DRIVEN — só afirma achados que existem (honesto p/ caso limpo/modelo).
    mult = stats.get("multiplo", 0.0)
    achados = []
    if mult > 1:
        achados.append(f"**Regime x teto** — volume anualizado ≈ **{mult:.1f}x** o teto EPP (R$ 4,8M); "
                       "indicador de porte/incompatibilidade a verificar")
    if stats["total_ret_5m"] > 0:
        achados.append(f"**Retenções estimadas na fonte** — R$ {format_brl(stats['total_ret_5m'])} em "
                       f"{stats['n_meses']} meses (PIS/COFINS/CSLL/IRRF sobre pagamentos a PJ de servico)")
    if stats["meis"]:
        achados.append(f"**{len(stats['meis'])} fornecedores PJ acima do teto MEI** — risco de pejotização")
    if stats["pos_baixa"]:
        achados.append(f"**{len(stats['pos_baixa'])} pagamentos pós-baixa** — R$ {format_brl(total_pb)} a CNPJ já baixado")
    if total_pr > 0:
        achados.append(f"**R$ {format_brl(total_pr)} com partes relacionadas** — necessita lastro contratual")
    fisc = stats.get("fiscal")
    if fisc:
        lines += ["", "## Conformidade Fiscal (OFX × NF-e × CT-e)", "",
                  f"**Documentos:** {format_num(fisc['n_nfe'])} NF-e (R$ {format_brl(fisc['vol_nfe'])}) + "
                  f"{format_num(fisc['n_cte'])} CT-e (R$ {format_brl(fisc['vol_cte'])}).", ""]
        if fisc["criticos"]:
            lines += ["Fornecedores com **gap de conformidade** (pago sem cobertura fiscal proporcional):", "",
                      "| Fornecedor | Pago (OFX) | NF-e+CT-e | Conf % | Classe |", "|---|---:|---:|---:|:--:|"]
            for c in fisc["criticos"]:
                lines.append(f"| {c['nome'][:34].title()} | R$ {format_brl(c['vp'])} | "
                             f"R$ {format_brl(c['vn'] + c['vc'])} | {c['conf']:.0f}% | {c['classe']} |")
        else:
            lines.append("Sem fornecedores em faixa crítica de conformidade no recorte.")

    lines += ["", "## 9. Conclusão", ""]
    if achados:
        lines.append("Os testes determinísticos aplicados apontam os seguintes **pontos de atenção**, que "
                     "demandam verificação documental e, se confirmados, regularização:")
        lines.append("")
        lines += [f"{i}. {a}" for i, a in enumerate(achados, 1)]
        lines += ["", "Recomenda-se a verificação dos itens acima e, quando aplicável, o acionamento das "
                  "medidas formais cabíveis (ex.: denúncia espontânea — CTN art. 138)."]
    else:
        lines.append("Nos testes determinísticos aplicados a este recorte **não foram identificados achados "
                     "materiais** (regime compatível com o teto, sem pagamentos pós-baixa, sem retenções "
                     "estimadas e sem MEIs acima do teto). Ver ressalvas de escopo e enriquecimento cadastral.")
    lines += ["", "---", "",
              "*Documento gerado pelo OrgConc — Sistema OrgAudi. Indicadores determinísticos; "
              "NÃO constituem conclusão de auditoria sem verificação documental.*"]

    return "\n".join(lines), {"total_pr": total_pr, "total_exc": total_exc, "total_pb": total_pb}


def gerar_html(md_text, periodo="", titulo=None, objeto=None, razao=None, cnpj=None, subtitulo=None, laudo_id=None):
    """HTML do laudo com fontes embarcadas (Manrope/Instrument Serif/JetBrains Mono),
    tarja CONFIDENCIAL em todas as páginas, capa modernizada e responsável técnico.
    Reutilizável por outros laudos — basta passar titulo/objeto.
    """
    import markdown as mdlib
    body = mdlib.markdown(md_text, extensions=["tables", "fenced_code"])
    # Remove cabeçalho redundante do MD (antes do primeiro ## N.) — a capa já tem o título.
    if "<h2" in body:
        body = body[body.index("<h2"):]
    agora = datetime.now().strftime("%d/%m/%Y %H:%M")
    razao = razao or (_emp().get("razao_social", "") or "—")
    cnpj = cnpj or _emp().get("cnpj", "—")
    titulo_display = titulo or "Laudo de Auditoria<br><em>Bancária Forense</em>"
    subtitulo = subtitulo or "Sistema OrgAudi · Auditoria Bancária Forense"
    objeto = objeto or (
        razao + " — análise forense de extratos bancários (OFX): regime × teto, "
        "retenções na fonte, tipologias (smurfing, carrossel, pós-baixa) e cruzamento cadastral RFB/BrasilAPI."
    )
    laudo_id = laudo_id or f"ORG-{datetime.now().strftime('%Y%m%d-%H%M')}"

    # Logo SVG inline (~500B) — sem dependência de arquivo PNG externo; sem problema de alt
    logo_svg = (
        '<svg width="52" height="52" viewBox="0 0 52 52" xmlns="http://www.w3.org/2000/svg"'
        ' role="img" aria-label="ORGATEC">'
        '<rect width="52" height="52" rx="7" fill="#1A3A6B"/>'
        '<text x="26" y="18" text-anchor="middle" font-family="Consolas,monospace"'
        ' font-size="9" font-weight="700" fill="#5BA9D6" letter-spacing="1">ORG</text>'
        '<text x="26" y="31" text-anchor="middle" font-family="Consolas,monospace"'
        ' font-size="9" font-weight="700" fill="#B8DDEE" letter-spacing="1">AUD</text>'
        '<text x="26" y="44" text-anchor="middle" font-family="Consolas,monospace"'
        ' font-size="6.5" fill="#5BA9D6" letter-spacing="0.5">ORGATEC</text>'
        '</svg>'
    )

    font_css = font_faces_css()  # @font-face data-URIs; vazio se fontes ausentes (fallback seguro)

    css = f"""
{font_css}

/* ── Página CSS Paged Media Level 3 (WeasyPrint) ── */
@page {{
  size: A4 landscape;
  margin: 14mm 16mm 18mm 16mm;
  @top-right {{
    content: "CONFIDENCIAL \2014 USO RESTRITO";
    font-family: 'JetBrains Mono', Consolas, monospace;
    font-size: 7pt; color: #991B1B; letter-spacing: 0.08em;
    border-bottom: 0.5pt solid #B91C1C; padding-bottom: 1mm;
  }}
  @top-left {{
    content: string(seccao-atual);
    font-family: 'Manrope', 'Segoe UI', Arial, sans-serif;
    font-size: 7pt; color: #3F5A78;
  }}
  @bottom-center {{
    content: "ORGATEC \00B7 OrgAudi \00B7 Auditoria Banc\00e1ria Forense  \2014  p\00e1gina " counter(page) " de " counter(pages);
    font-family: 'Manrope', 'Segoe UI', Arial, sans-serif;
    font-size: 7pt; color: #94A3B8;
  }}
}}
@page :first {{ @top-left {{ content: ""; }} @top-right {{ content: ""; }} }}

* {{ box-sizing: border-box; margin: 0; padding: 0; }}
body {{
  font-family: 'Manrope', 'Segoe UI', Arial, sans-serif;
  font-size: 10pt; color: #0E2A47; line-height: 1.6;
}}

/* ── Capa ── */
.capa {{
  height: 176mm; padding: 14mm 24mm;
  display: flex; flex-direction: column; justify-content: space-between;
  page-break-after: always; position: relative;
  background:
    radial-gradient(ellipse 60% 45% at 12% 8%, rgba(91,169,214,.28), transparent 60%),
    radial-gradient(ellipse 55% 42% at 90% 16%, rgba(184,221,238,.35), transparent 63%),
    linear-gradient(180deg,#d8eef7 0%,#e6f3f9 30%,#f0f7fb 56%,#f8fbfc 78%,#fff 100%);
}}
.capa::before {{
  content:""; position:absolute; top:0; left:0; right:0; height:5mm;
  background:linear-gradient(90deg,#1A3A6B 0%,#5BA9D6 60%,#B8DDEE 100%);
}}
.capa-brand {{ display:flex; align-items:center; gap:14px; }}
.capa-wm .nome {{
  font-size:20pt; font-weight:700; letter-spacing:.18em; color:#1A3A6B; line-height:1;
}}
.capa-wm .desc {{
  font-family:'JetBrains Mono',Consolas,monospace;
  font-size:7pt; letter-spacing:.24em; text-transform:uppercase; color:#3F5A78; margin-top:1.5mm;
}}
.capa-eyebrow {{
  font-family:'JetBrains Mono',Consolas,monospace;
  font-size:7.5pt; letter-spacing:.14em; text-transform:uppercase; color:#5BA9D6; margin-top:6mm;
}}
.capa-mid {{ display:flex; flex-direction:column; }}
.capa-titulo {{
  font-size:28pt; font-weight:300; line-height:1.16; color:#1A3A6B;
}}
.capa-titulo em {{
  font-family:'Instrument Serif',Georgia,serif;
  font-style:italic; font-weight:400; color:#5BA9D6;
}}
.capa-rule {{
  width:42mm; height:2.4pt; margin:7mm 0;
  background:linear-gradient(90deg,#1A3A6B,#5BA9D6,#B8DDEE);
}}
.capa-objeto {{ font-size:10.5pt; color:#3F5A78; max-width:170mm; line-height:1.5; }}
.capa-selo {{
  margin-top:7mm; display:inline-block; align-self:flex-start;
  border:0.7pt solid #5BA9D6; color:#1A3A6B;
  font-family:'JetBrains Mono',Consolas,monospace;
  font-size:7pt; letter-spacing:.12em; text-transform:uppercase; padding:1.5mm 3.5mm;
}}
.capa-meta {{
  border-top:0.5pt solid #B8DDEE; padding-top:5mm;
  display:flex; gap:18mm; font-size:8.5pt; color:#3F5A78;
}}
.capa-meta strong {{
  font-family:'JetBrains Mono',Consolas,monospace;
  color:#1A3A6B; font-weight:600; display:block; font-size:7pt;
  letter-spacing:.1em; text-transform:uppercase; margin-bottom:1mm;
}}

/* ── Conteúdo ── */
.conteudo {{ padding-top:2mm; }}
h1 {{
  font-size:14pt; font-weight:700; color:#1A3A6B;
  margin:20px 0 8px; padding-bottom:4px;
  border-bottom:1.4pt solid #1A3A6B; page-break-after:avoid;
  string-set:seccao-atual content();
}}
h2 {{
  font-size:11.5pt; font-weight:600; color:#1A3A6B;
  margin:16px 0 6px; padding:3px 0 3px 10px;
  border-left:3pt solid #5BA9D6; page-break-after:avoid;
  string-set:seccao-atual content();
}}
h3 {{ font-size:10pt; color:#5BA9D6; margin:10px 0 5px; font-weight:600; }}
p {{ margin-bottom:6px; }}
table {{
  width:100%; border-collapse:collapse; margin:6px 0 14px;
  font-size:8.8pt; border:0.6pt solid #DDE1E5;
}}
thead {{ display:table-header-group; }}
th {{ background:#1A3A6B; color:#F4F9FC; padding:5px 9px; text-align:left; font-weight:600; font-size:8.4pt; }}
td {{ padding:4px 9px; border:0.5pt solid #DDE1E5; vertical-align:top; }}
tr:nth-child(even) td {{ background:#F4F9FC; }}
td.num,th.num {{ text-align:right; font-variant-numeric:tabular-nums; }}
strong {{ color:#1A3A6B; font-weight:700; }}
ul,ol {{ padding-left:22px; margin-bottom:8px; }}
li {{ margin-bottom:3px; page-break-inside:avoid; }}
em {{ color:#3F5A78; font-style:italic; }}
hr {{ border:none; border-top:0.5pt solid #DDE1E5; margin:14px 0; }}
blockquote {{
  border-left:3pt solid #5BA9D6; background:#F4F9FC;
  padding:3mm 4mm; margin:4mm 0; color:#0E2A47; font-size:9.6pt;
}}

/* ── Chips de severidade (cor + rótulo — imprimíveis sem cor) ── */
.sev-critico {{ background:#FEE2E2; color:#991B1B; padding:1px 5px; border-radius:3px; font-weight:700; }}
.sev-alto    {{ background:#FFEDD5; color:#9A3412; padding:1px 5px; border-radius:3px; font-weight:700; }}
.sev-medio   {{ background:#FEF9C3; color:#854D0E; padding:1px 5px; border-radius:3px; font-weight:700; }}
.sev-baixo   {{ background:#DCFCE7; color:#166534; padding:1px 5px; border-radius:3px; font-weight:700; }}

/* ── Responsável técnico / Assinatura ── */
.assinatura {{
  margin-top:14mm; page-break-inside:avoid;
  border-top:0.5pt solid #DDE1E5; padding-top:6mm;
}}
.assinatura .linha {{ border-top:0.6pt solid #1A3A6B; width:76mm; margin-bottom:1.5mm; }}
.assinatura .nome {{ font-weight:700; color:#1A3A6B; font-size:9pt; }}
.assinatura .cargo {{ color:#3F5A78; margin-top:1mm; font-size:8.6pt; }}
.assinatura .id-laudo {{
  font-family:'JetBrains Mono',Consolas,monospace;
  font-size:7pt; color:#94A3B8; margin-top:3mm; letter-spacing:.05em;
}}
"""

    capa_html = f"""<section class="capa">
  <div>
    <div class="capa-brand">{logo_svg}<div class="capa-wm">
      <div class="nome">ORGATEC</div>
      <div class="desc">Contabilidade &middot; Auditoria &middot; Compliance</div>
    </div></div>
    <div class="capa-eyebrow">{subtitulo}</div>
  </div>
  <div class="capa-mid">
    <div class="capa-titulo">{titulo_display}</div>
    <div class="capa-rule"></div>
    <div class="capa-objeto">{objeto}</div>
    <div class="capa-selo">Parecer t&eacute;cnico &middot; assessoria &middot; car&aacute;ter indicativo</div>
  </div>
  <div class="capa-meta">
    <div><strong>Entidade auditada</strong>{razao}<br>CNPJ {cnpj}</div>
    <div><strong>Per&iacute;odo</strong>{periodo or '&mdash;'}</div>
    <div><strong>Emiss&atilde;o</strong>{agora}</div>
    <div><strong>ID do laudo</strong><span style="font-family:'JetBrains Mono',Consolas,monospace;font-size:8pt">{laudo_id}</span></div>
  </div>
</section>"""

    assinatura_html = f"""<div class="assinatura">
  <div class="linha"></div>
  <div class="nome">ORGATEC &mdash; Contabilidade e Auditoria</div>
  <div class="cargo">Sistema OrgAudi &middot; Auditoria Banc&aacute;ria Forense assistida por IA</div>
  <div class="id-laudo">Laudo {laudo_id} &middot; Gerado em {agora}</div>
</div>"""

    return f"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8">
<meta name="description" content="Laudo de Auditoria Banc&aacute;ria Forense &mdash; {razao}">
<meta name="generator" content="OrgConc OrgAudi">
<title>Laudo de Auditoria &middot; {razao}</title>
<style>{css}</style>
</head>
<body>
{capa_html}
<section class="conteudo">
{body}
{assinatura_html}
</section>
</body>
</html>"""


def _block_url_fetcher(url, **_kwargs):
    """Bloqueia fetch de URLs externas pelo WeasyPrint (anti-SSRF)."""
    return {"string": b"", "mime_type": "text/plain", "encoding": "utf-8"}


async def html_para_pdf_bytes(html_text, landscape=True):
    """Renderiza HTML em PDF via WeasyPrint e retorna os bytes.

    A orientacao e o rodape vem do `@page` do proprio HTML (size A4 landscape +
    contador de paginas); `landscape` permanece por compat de assinatura. WeasyPrint
    ja e dep de producao (requirements-prod) e o Dockerfile instala as libs nativas
    (libpango/cairo). write_pdf() e CPU-bound -> roda em thread pool; url_fetcher
    bloqueia fetch externo (anti-SSRF). Retorna None em falha (chamador decide o
    fallback / erro HTTP).
    """
    try:
        import weasyprint
    except ImportError:
        return None
    try:
        return await asyncio.to_thread(
            lambda: weasyprint.HTML(
                string=html_text, base_url=None, url_fetcher=_block_url_fetcher
            ).write_pdf()
        )
    except (OSError, RuntimeError):
        return None


async def gerar_pdf(html_text, out_pdf):
    """Escreve o PDF em `out_pdf` (compat: CLI/laudo). Usa html_para_pdf_bytes."""
    try:
        blob = await html_para_pdf_bytes(html_text, landscape=True)
        if not blob:
            return False
        with open(out_pdf, "wb") as fh:
            fh.write(blob)
        return True
    except Exception as exc:  # noqa: BLE001
        print(f"PDF failed: {exc}")
        return False


# CLI: scripts/relatorio_integrado.py (wrapper fino que importa este módulo).
# Núcleo reusável: montar_dados() + gerar_laudo_workbook() + coletar_dados().
