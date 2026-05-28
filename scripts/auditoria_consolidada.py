"""Auditoria consolidada cross-mensal - conta 158083-3 (jan-mai/2026).

Pega os 5 OFXs e gera 1 relatorio unificado com:
  - Evolucao mensal de KPIs
  - Top contrapartes acumuladas
  - Alertas cross-mes (mesmas contrapartes baixadas em multiplos meses)
  - Risk Score consolidado por contraparte
  - Status tributario agregado (retencao total estimada)
"""
from __future__ import annotations

import asyncio
import re
import sys
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
sys.path.insert(0, str(Path(__file__).resolve().parent))
from _logo_helper import html_logo_inline, inserir_logo_xlsx
from api.matchers.cascata import classificar, ler_ofx
from api.matchers.cnpj_enricher import _carregar_cache
from api.matchers.forensics import (
    calcular_agregados,
    calcular_risk_score,
    classificar_tributario,
    detectar_carrossel,
    detectar_meio,
    detectar_primeira_vez,
    detectar_smurfing,
    detectar_valor_redondo,
)

# Mesma estilizacao do conciliar_ofx_unico
NAVY = "0F172A"
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TOTAL_FILL = PatternFill("solid", fgColor="1E3A8A")
TOTAL_FONT = Font(bold=True, color="FFFFFF", size=11)
ZEBRA_FILL = PatternFill("solid", fgColor="F8FAFC")
ALERT_FILL = PatternFill("solid", fgColor="FEE2E2")
TITLE_FONT = Font(bold=True, size=14, color=NAVY)
BORDER = Side(border_style="thin", color="E2E8F0")
THIN_BORDER = Border(top=BORDER, left=BORDER, right=BORDER, bottom=BORDER)

OFX_LIST = [
    (r"C:\Users\Veloso\Downloads\extrato-conta-corrente-ofx-unix_202605_20260514110822.ofx", "JAN/2026"),
    (r"C:\Users\Veloso\Downloads\extrato-conta-corrente-ofx-unix_202605_20260514110841.ofx", "FEV/2026"),
    (r"C:\Users\Veloso\Downloads\extrato-conta-corrente-ofx-unix_202605_20260514110900.ofx", "MAR/2026"),
    (r"C:\Users\Veloso\Downloads\extrato-conta-corrente-ofx-unix_202605_20260514110917.ofx", "ABR/2026"),
    (r"C:\Users\Veloso\Downloads\extrato-conta-corrente-ofx-unix_202605_20260514110938.ofx", "MAI/2026"),
]
CONTA = "158083-3"
AGENCIA = "3333-2"
BANCO = "SICOOB - Banco Cooperativo do Brasil (756)"

OUT_BASE = r"C:\Users\Veloso\Downloads\AUDITORIA_CONSOLIDADA_158083-3_5MESES"
OUT_XLSX = Path(f"{OUT_BASE}.xlsx")
OUT_MD = Path(f"{OUT_BASE}.md")
OUT_HTML = Path(f"{OUT_BASE}.html")
OUT_PDF = Path(f"{OUT_BASE}.pdf")

RX_CNPJ = re.compile(r"(\d{2})[.](\d{3})[.](\d{3})[ /](\d{4})[-](\d{2})")


def _extrair_cnpj(t):
    for fonte in (t.nome or "", t.memo or ""):
        m = RX_CNPJ.search(fonte)
        if m:
            return "".join(m.groups())
    return None


def style_header(ws, row, n):
    for c in range(1, n + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = THIN_BORDER


def cabecalho(ws, ultima_col=8):
    c1 = ws.cell(row=1, column=1, value="    ORGATEC · Auditoria Consolidada 5 Meses · Conta 158083-3")
    c1.font = Font(bold=True, size=14, color="FFFFFF")
    c1.fill = PatternFill("solid", fgColor=NAVY)
    c1.alignment = Alignment(horizontal="center", vertical="center", indent=2)
    ws.merge_cells(f"A1:{get_column_letter(ultima_col)}1")
    ws.row_dimensions[1].height = 60
    if ws.column_dimensions[get_column_letter(1)].width is None or ws.column_dimensions[get_column_letter(1)].width < 12:
        ws.column_dimensions[get_column_letter(1)].width = 12
    inserir_logo_xlsx(ws, "A1", largura_px=60, altura_px=60)

    c2 = ws.cell(row=2, column=1,
        value=f"Empresa: [NAO CADASTRADO] | CNPJ: [PENDENTE] | Socios: [PENDENTE]")
    c2.font = Font(bold=True, size=10, color="FFFFFF")
    c2.fill = PatternFill("solid", fgColor="1E3A8A")
    c2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(f"A2:{get_column_letter(ultima_col)}2")

    c3 = ws.cell(row=3, column=1,
        value=f"Correntista: [PENDENTE] > Agencia: {AGENCIA} > Conta: {CONTA} > Banco: {BANCO} > Periodo: 01/01/2026 a 14/05/2026")
    c3.font = Font(size=9, color="0F172A")
    c3.fill = PatternFill("solid", fgColor="DBEAFE")
    c3.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(f"A3:{get_column_letter(ultima_col)}3")
    return 5


def coletar_dados():
    """Le todos os 5 OFXs e retorna dados consolidados."""
    cache = _carregar_cache()
    todos = []  # (mes_label, transacao, classificacao)
    saldos = {}  # mes -> (saldo_final, n_trans, cred, deb)

    for path, mes in OFX_LIST:
        txs = ler_ofx(path)
        cred = sum(t.valor for t in txs if t.valor > 0)
        deb = sum(t.valor for t in txs if t.valor < 0)
        # Le saldo final do OFX
        raw = Path(path).read_text(encoding="latin-1", errors="ignore")
        bal_m = re.search(r"<BALAMT>([\d.\-]+)", raw)
        saldo_final = float(bal_m.group(1)) if bal_m else 0.0
        saldos[mes] = {
            "saldo_final": saldo_final, "n": len(txs),
            "cred": cred, "deb": deb,
        }
        for t in txs:
            r = classificar(t)
            todos.append((mes, t, r))

    return todos, saldos, cache


def gerar_xlsx(todos, saldos, cache):
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # ────────────────────────────────────────────────────────────────
    # Aba 1: Resumo Executivo Consolidado
    # ────────────────────────────────────────────────────────────────
    ws = wb.create_sheet("Resumo Executivo")
    start = cabecalho(ws, 6)
    ws.cell(row=start, column=1, value="RESUMO EXECUTIVO - 5 MESES").font = TITLE_FONT
    ws.merge_cells(f"A{start}:F{start}")

    n_total = len(todos)
    cred_total = sum(s["cred"] for s in saldos.values())
    deb_total = sum(s["deb"] for s in saldos.values())
    saldo_inicial_jan = saldos["JAN/2026"]["saldo_final"] - (saldos["JAN/2026"]["cred"] + saldos["JAN/2026"]["deb"])
    saldo_final_mai = saldos["MAI/2026"]["saldo_final"]

    r = start + 2
    indicadores = [
        ("Total de transacoes (5 meses)", n_total),
        ("Volume de creditos", cred_total),
        ("Volume de debitos", deb_total),
        ("Fluxo liquido acumulado", cred_total + deb_total),
        ("Saldo inicial (01/01)", saldo_inicial_jan),
        ("Saldo final (14/05)", saldo_final_mai),
        ("Variacao do periodo", saldo_final_mai - saldo_inicial_jan),
        ("Volume movimentado bruto", abs(cred_total) + abs(deb_total)),
        ("Volume anualizado projetado", (abs(cred_total) + abs(deb_total)) * 12 / 4.5),
    ]
    ws.cell(row=r, column=1, value="INDICADOR")
    ws.cell(row=r, column=2, value="VALOR")
    style_header(ws, r, 2)
    r += 1
    for k, v in indicadores:
        ws.cell(row=r, column=1, value=k).font = Font(bold=True)
        c = ws.cell(row=r, column=2, value=v)
        if isinstance(v, (int, float)) and ("Volume" in k or "Saldo" in k or "Fluxo" in k or "Variacao" in k):
            c.number_format = "#,##0.00"
            c.font = Font(bold=True, color="DC2626" if v < 0 else "16A34A" if v > 0 else "0F172A")
        if r % 2 == 0:
            for col in (1, 2):
                ws.cell(row=r, column=col).fill = ZEBRA_FILL
        r += 1

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 22

    # ────────────────────────────────────────────────────────────────
    # Aba 2: Evolucao Mensal
    # ────────────────────────────────────────────────────────────────
    ws = wb.create_sheet("Evolucao Mensal")
    start = cabecalho(ws, 7)
    ws.cell(row=start, column=1, value="EVOLUCAO MENSAL").font = TITLE_FONT
    ws.merge_cells(f"A{start}:G{start}")

    headers_m = ["Mes", "Transacoes", "Creditos (R$)", "Debitos (R$)",
                 "Fluxo Liquido", "Saldo Final", "Var. Mes Ant."]
    r = start + 2
    for c, h in enumerate(headers_m, start=1):
        ws.cell(row=r, column=c, value=h)
    style_header(ws, r, 7)
    r += 1

    saldo_anterior = saldo_inicial_jan
    for mes in ["JAN/2026", "FEV/2026", "MAR/2026", "ABR/2026", "MAI/2026"]:
        s = saldos[mes]
        var = s["saldo_final"] - saldo_anterior
        ws.cell(row=r, column=1, value=mes).font = Font(bold=True)
        ws.cell(row=r, column=2, value=s["n"]).number_format = "#,##0"
        ws.cell(row=r, column=3, value=round(s["cred"], 2)).number_format = "#,##0.00"
        ws.cell(row=r, column=3).font = Font(color="16A34A")
        ws.cell(row=r, column=4, value=round(s["deb"], 2)).number_format = "#,##0.00"
        ws.cell(row=r, column=4).font = Font(color="DC2626")
        fl = s["cred"] + s["deb"]
        ws.cell(row=r, column=5, value=round(fl, 2)).number_format = "#,##0.00"
        ws.cell(row=r, column=5).font = Font(bold=True, color="DC2626" if fl < 0 else "16A34A")
        ws.cell(row=r, column=6, value=round(s["saldo_final"], 2)).number_format = "#,##0.00"
        ws.cell(row=r, column=6).font = Font(bold=True, color="DC2626" if s["saldo_final"] < 0 else "0F172A")
        ws.cell(row=r, column=7, value=round(var, 2)).number_format = "#,##0.00"
        ws.cell(row=r, column=7).font = Font(color="DC2626" if var < 0 else "16A34A")
        for c in range(1, 8):
            ws.cell(row=r, column=c).border = THIN_BORDER
            if r % 2 == 0:
                ws.cell(row=r, column=c).fill = ZEBRA_FILL
        saldo_anterior = s["saldo_final"]
        r += 1

    # Totais
    ws.cell(row=r, column=1, value="TOTAL").font = TOTAL_FONT
    ws.cell(row=r, column=2, value=n_total).number_format = "#,##0"
    ws.cell(row=r, column=3, value=round(cred_total, 2)).number_format = "#,##0.00"
    ws.cell(row=r, column=4, value=round(deb_total, 2)).number_format = "#,##0.00"
    ws.cell(row=r, column=5, value=round(cred_total + deb_total, 2)).number_format = "#,##0.00"
    ws.cell(row=r, column=6, value=round(saldo_final_mai, 2)).number_format = "#,##0.00"
    for c in range(1, 8):
        ws.cell(row=r, column=c).fill = TOTAL_FILL
        ws.cell(row=r, column=c).font = TOTAL_FONT

    for col, w in {1: 12, 2: 12, 3: 18, 4: 18, 5: 18, 6: 18, 7: 18}.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = f"A{start + 3}"

    # ────────────────────────────────────────────────────────────────
    # Aba 3: Top Contrapartes Cross-Mensal
    # ────────────────────────────────────────────────────────────────
    ws = wb.create_sheet("Top Contrapartes")
    start = cabecalho(ws, 10)
    ws.cell(row=start, column=1, value="TOP 50 CONTRAPARTES CONSOLIDADAS").font = TITLE_FONT
    ws.merge_cells(f"A{start}:J{start}")

    # Agrega por CNPJ
    por_cnpj = defaultdict(lambda: {
        "n": 0, "cred": 0.0, "deb": 0.0, "meses": set(),
    })
    for mes, t, _ in todos:
        cnpj = _extrair_cnpj(t)
        if cnpj:
            d = por_cnpj[cnpj]
            d["n"] += 1
            d["meses"].add(mes)
            if t.valor > 0:
                d["cred"] += t.valor
            else:
                d["deb"] += t.valor

    headers = ["CNPJ", "Razao Social", "Situacao", "UF", "Trans.",
               "Volume Total (R$)", "Creditos", "Debitos", "Meses Atuantes", "Alerta"]
    r = start + 2
    for c, h in enumerate(headers, start=1):
        ws.cell(row=r, column=c, value=h)
    style_header(ws, r, 10)
    r += 1

    top_cnpjs = sorted(
        por_cnpj.items(),
        key=lambda x: -(x[1]["cred"] + abs(x[1]["deb"]))
    )[:50]
    for cnpj, dados in top_cnpjs:
        info = cache.get(cnpj, {})
        razao = info.get("razao_social", "(nao enriquecido)")[:50]
        sit = info.get("situacao", "?")
        uf = info.get("uf", "")
        is_baixada = "BAIXADA" in sit or "INAPTA" in sit
        volume = dados["cred"] + abs(dados["deb"])
        n_meses = len(dados["meses"])

        # Alerta consolidado
        alerta = ""
        if is_baixada:
            alerta = "BAIXADA/INAPTA"
        elif n_meses >= 5:
            alerta = "RECORRENTE 5 MESES"
        elif n_meses == 1 and volume > 50_000:
            alerta = "PONTUAL ALTO VALOR"

        fmt = f"{cnpj[:2]}.{cnpj[2:5]}.{cnpj[5:8]}/{cnpj[8:12]}-{cnpj[12:14]}"
        ws.cell(row=r, column=1, value=fmt).font = Font(name="Consolas", size=10)
        ws.cell(row=r, column=2, value=razao)
        c_sit = ws.cell(row=r, column=3, value=sit)
        if is_baixada:
            c_sit.font = Font(bold=True, color="DC2626")
        ws.cell(row=r, column=4, value=uf)
        ws.cell(row=r, column=5, value=dados["n"]).number_format = "#,##0"
        cv = ws.cell(row=r, column=6, value=round(volume, 2))
        cv.number_format = "#,##0.00"
        cv.font = Font(bold=True)
        ws.cell(row=r, column=7, value=round(dados["cred"], 2)).number_format = "#,##0.00"
        ws.cell(row=r, column=7).font = Font(color="16A34A")
        ws.cell(row=r, column=8, value=round(dados["deb"], 2)).number_format = "#,##0.00"
        ws.cell(row=r, column=8).font = Font(color="DC2626")
        ws.cell(row=r, column=9, value=", ".join(sorted(dados["meses"], key=lambda x: ["JAN", "FEV", "MAR", "ABR", "MAI"].index(x[:3]))))
        c_alerta = ws.cell(row=r, column=10, value=alerta)
        if alerta == "BAIXADA/INAPTA":
            c_alerta.font = Font(bold=True, color="DC2626")
            for c in range(1, 11):
                ws.cell(row=r, column=c).fill = ALERT_FILL
        elif alerta == "RECORRENTE 5 MESES":
            c_alerta.font = Font(bold=True, color="0052FF")
        elif alerta == "PONTUAL ALTO VALOR":
            c_alerta.font = Font(bold=True, color="D97706")

        for c in range(1, 11):
            ws.cell(row=r, column=c).border = THIN_BORDER
            if r % 2 == 0 and not is_baixada:
                ws.cell(row=r, column=c).fill = ZEBRA_FILL
        r += 1

    for col, w in {1: 20, 2: 38, 3: 13, 4: 5, 5: 8, 6: 17, 7: 14, 8: 14, 9: 42, 10: 20}.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = f"A{start + 3}"
    ws.auto_filter.ref = f"A{start + 2}:J{r - 1}"

    # ────────────────────────────────────────────────────────────────
    # Aba 4: Alertas Pos-Baixa (todos os meses)
    # ────────────────────────────────────────────────────────────────
    ws = wb.create_sheet("Alertas Pos-Baixa")
    start = cabecalho(ws, 7)
    ws.cell(row=start, column=1, value="PAGAMENTOS POS-BAIXA - 5 MESES").font = TITLE_FONT
    ws.merge_cells(f"A{start}:G{start}")

    headers = ["Mes", "Data", "Valor (R$)", "Memo", "Razao Social", "Data Baixa", "Dias Apos Baixa"]
    r = start + 2
    for c, h in enumerate(headers, start=1):
        ws.cell(row=r, column=c, value=h)
    style_header(ws, r, 7)
    r += 1

    total_volume_pb = 0.0
    pos_baixa = []
    for mes, t, _ in todos:
        cnpj = _extrair_cnpj(t)
        if not cnpj or cnpj not in cache:
            continue
        info = cache[cnpj]
        sit = info.get("situacao", "")
        if "BAIXADA" not in sit and "INAPTA" not in sit:
            continue
        try:
            db_data = date.fromisoformat(info["data_situacao"][:10])
            dt_data = date.fromisoformat(t.data[:10])
        except (ValueError, KeyError, TypeError):
            continue
        if dt_data > db_data:
            dias = (dt_data - db_data).days
            pos_baixa.append((mes, t, cnpj, info, dias))

    pos_baixa.sort(key=lambda x: -x[4])  # mais dias apos primeiro
    for mes, t, cnpj, info, dias in pos_baixa:
        ws.cell(row=r, column=1, value=mes)
        ws.cell(row=r, column=2, value=t.data)
        cv = ws.cell(row=r, column=3, value=round(t.valor, 2))
        cv.number_format = "#,##0.00"
        cv.font = Font(bold=True, color="DC2626")
        ws.cell(row=r, column=4, value=(t.memo or "")[:40])
        ws.cell(row=r, column=5, value=info.get("razao_social", "")[:45])
        ws.cell(row=r, column=6, value=info.get("data_situacao", "")).font = Font(bold=True, color="DC2626")
        ws.cell(row=r, column=7, value=dias).number_format = "#,##0"
        for c in range(1, 8):
            ws.cell(row=r, column=c).border = THIN_BORDER
            ws.cell(row=r, column=c).fill = ALERT_FILL
        total_volume_pb += abs(t.valor)
        r += 1

    if pos_baixa:
        ws.cell(row=r, column=1, value=f"TOTAL ({len(pos_baixa)} alertas)").font = TOTAL_FONT
        ws.cell(row=r, column=3, value=round(total_volume_pb, 2)).number_format = "#,##0.00"
        ws.cell(row=r, column=3).font = TOTAL_FONT
        for c in range(1, 8):
            ws.cell(row=r, column=c).fill = TOTAL_FILL

    for col, w in {1: 12, 2: 12, 3: 14, 4: 38, 5: 42, 6: 12, 7: 16}.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = f"A{start + 3}"

    # ────────────────────────────────────────────────────────────────
    # Aba 5: Status Tributario Consolidado
    # ────────────────────────────────────────────────────────────────
    ws = wb.create_sheet("Status Tributario")
    start = cabecalho(ws, 7)
    ws.cell(row=start, column=1, value="STATUS TRIBUTARIO CONSOLIDADO - 5 MESES").font = TITLE_FONT
    ws.merge_cells(f"A{start}:G{start}")

    cat_count = Counter()
    cat_volume = defaultdict(float)
    cat_retencao = defaultdict(float)
    cat_por_mes = defaultdict(lambda: defaultdict(float))

    for mes, t, _ in todos:
        cnpj = _extrair_cnpj(t)
        info = cache.get(cnpj) if cnpj else None
        porte = (info.get("porte") if info else "") or ""
        trib = classificar_tributario(t.memo or "", t.nome or "", t.valor, cnpj or "", porte)
        cat_count[trib["categoria"]] += 1
        cat_volume[trib["categoria"]] += abs(t.valor)
        cat_retencao[trib["categoria"]] += trib["valor_retencao"]
        cat_por_mes[mes][trib["categoria"]] += trib["valor_retencao"]

    headers = ["Categoria Fiscal", "Qtd", "Volume (R$)", "Retencao Total (R$)",
               "JAN", "FEV", "MAR", "ABR", "MAI"]
    r = start + 2
    for c, h in enumerate(headers, start=1):
        ws.cell(row=r, column=c, value=h)
    style_header(ws, r, 9)
    r += 1

    CATS = ["RETENCAO_PJ", "RETENCAO_PF", "OPERACAO_CREDITO", "IOF", "JUROS",
            "PAGAMENTO_TRIBUTO", "TARIFA", "PIX_RECEBIDO", "BOLETO",
            "COMPRA_CARTAO", "NAO_TRIBUTAVEL", "OUTRO"]
    total_ret_5m = 0.0
    for cat in CATS:
        qtd = cat_count.get(cat, 0)
        if qtd == 0:
            continue
        c1 = ws.cell(row=r, column=1, value=cat)
        c1.font = Font(bold=True, color="DC2626" if cat.startswith("RETENCAO") else "0F172A")
        ws.cell(row=r, column=2, value=qtd).number_format = "#,##0"
        ws.cell(row=r, column=3, value=round(cat_volume[cat], 2)).number_format = "#,##0.00"
        cret = ws.cell(row=r, column=4, value=round(cat_retencao[cat], 2))
        cret.number_format = "#,##0.00"
        if cat_retencao[cat] > 0:
            cret.font = Font(bold=True, color="D97706")
            total_ret_5m += cat_retencao[cat]
        # Por mes
        for i, mes in enumerate(["JAN/2026", "FEV/2026", "MAR/2026", "ABR/2026", "MAI/2026"], start=5):
            val = cat_por_mes[mes].get(cat, 0)
            cm = ws.cell(row=r, column=i, value=round(val, 2) if val else "")
            if val:
                cm.number_format = "#,##0.00"
        for c in range(1, 10):
            ws.cell(row=r, column=c).border = THIN_BORDER
            if r % 2 == 0:
                ws.cell(row=r, column=c).fill = ZEBRA_FILL
        r += 1

    ws.cell(row=r, column=1, value="TOTAL").font = TOTAL_FONT
    ws.cell(row=r, column=2, value=sum(cat_count.values())).number_format = "#,##0"
    ws.cell(row=r, column=3, value=round(sum(cat_volume.values()), 2)).number_format = "#,##0.00"
    ws.cell(row=r, column=4, value=round(total_ret_5m, 2)).number_format = "#,##0.00"
    for i, mes in enumerate(["JAN/2026", "FEV/2026", "MAR/2026", "ABR/2026", "MAI/2026"], start=5):
        ws.cell(row=r, column=i, value=round(sum(cat_por_mes[mes].values()), 2)).number_format = "#,##0.00"
    for c in range(1, 10):
        ws.cell(row=r, column=c).fill = TOTAL_FILL
        ws.cell(row=r, column=c).font = TOTAL_FONT

    for col, w in {1: 22, 2: 8, 3: 17, 4: 18, 5: 13, 6: 13, 7: 13, 8: 13, 9: 13}.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = f"A{start + 3}"

    # ────────────────────────────────────────────────────────────────
    # Aba 6: Risk Score por Contraparte
    # ────────────────────────────────────────────────────────────────
    ws = wb.create_sheet("Risk Score Contraparte")
    start = cabecalho(ws, 7)
    ws.cell(row=start, column=1, value="RISK SCORE CONSOLIDADO POR CONTRAPARTE").font = TITLE_FONT
    ws.merge_cells(f"A{start}:G{start}")

    # Agrega risk score por CNPJ (media ponderada por volume)
    risk_por_cnpj = defaultdict(lambda: {"max_score": 0, "n": 0, "vol": 0.0, "razao": "", "sit": "", "alertas": set()})

    # Pre-calcula agg geral (para todas as transacoes)
    from api.matchers.cascata import Resultado, Transacao
    fake_disps = []
    for mes, t, r_class in todos:
        # Cria objeto similar a Disposicao com .transacao
        class FakeDisp:
            pass
        fd = FakeDisp()
        fd.transacao = t
        fd.disposicao = "NORMAL"
        fake_disps.append(fd)
    agg = calcular_agregados(fake_disps)

    for mes, t, _ in todos:
        cnpj = _extrair_cnpj(t)
        if not cnpj:
            continue
        info = cache.get(cnpj, {})
        sit = info.get("situacao", "")
        porte = info.get("porte", "")
        meio = detectar_meio(t.memo or "", t.nome or "")
        acumulado = agg.acumulado_mes.get((cnpj, t.data[:7]), 0.0)
        vr = detectar_valor_redondo(t.valor)
        sm = detectar_smurfing(cnpj, t.data, agg)
        car = detectar_carrossel(cnpj, agg)
        pv = detectar_primeira_vez(cnpj, t.data, agg)

        # Detecta pos-baixa
        disp_name = "NORMAL"
        if ("BAIXADA" in sit or "INAPTA" in sit) and info.get("data_situacao"):
            try:
                db_d = date.fromisoformat(info["data_situacao"][:10])
                dt_d = date.fromisoformat(t.data[:10])
                if dt_d > db_d:
                    disp_name = "ALERTA_POS_BAIXA"
            except (ValueError, TypeError):
                pass

        score, classe = calcular_risk_score(
            t.valor, disp_name, sit, porte, meio, vr, sm, car, pv, acumulado,
        )

        d = risk_por_cnpj[cnpj]
        d["max_score"] = max(d["max_score"], score)
        d["n"] += 1
        d["vol"] += abs(t.valor)
        d["razao"] = info.get("razao_social", "")
        d["sit"] = sit
        if classe in ("CRITICO", "ALTO"):
            d["alertas"].add(classe)
        if sm:
            d["alertas"].add("SMURFING")
        if car:
            d["alertas"].add("CARROSSEL")
        if disp_name == "ALERTA_POS_BAIXA":
            d["alertas"].add("POS_BAIXA")

    headers = ["CNPJ", "Razao Social", "Score Max", "Classe Max",
               "Transacoes", "Volume (R$)", "Alertas"]
    r = start + 2
    for c, h in enumerate(headers, start=1):
        ws.cell(row=r, column=c, value=h)
    style_header(ws, r, 7)
    r += 1

    # Top 100 por max_score, depois por volume
    top_risk = sorted(
        risk_por_cnpj.items(),
        key=lambda x: (-x[1]["max_score"], -x[1]["vol"])
    )[:100]

    for cnpj, d in top_risk:
        if d["max_score"] < 25:
            continue  # so mostra >= MEDIO
        score = d["max_score"]
        classe = "CRITICO" if score >= 70 else "ALTO" if score >= 50 else "MEDIO"

        fmt = f"{cnpj[:2]}.{cnpj[2:5]}.{cnpj[5:8]}/{cnpj[8:12]}-{cnpj[12:14]}"
        ws.cell(row=r, column=1, value=fmt).font = Font(name="Consolas", size=10)
        ws.cell(row=r, column=2, value=d["razao"][:45])
        c_score = ws.cell(row=r, column=3, value=score)
        c_score.number_format = "0"
        c_score.font = Font(bold=True, color="DC2626" if score >= 70 else "D97706" if score >= 50 else "0052FF")
        c_classe = ws.cell(row=r, column=4, value=classe)
        c_classe.font = Font(bold=True)
        fill_classe = {"CRITICO": "FEE2E2", "ALTO": "FEF3C7", "MEDIO": "DBEAFE"}.get(classe, "")
        if fill_classe:
            c_classe.fill = PatternFill("solid", fgColor=fill_classe)
        ws.cell(row=r, column=5, value=d["n"]).number_format = "#,##0"
        ws.cell(row=r, column=6, value=round(d["vol"], 2)).number_format = "#,##0.00"
        ws.cell(row=r, column=7, value=", ".join(sorted(d["alertas"])))
        for c in range(1, 8):
            ws.cell(row=r, column=c).border = THIN_BORDER
            if classe == "CRITICO":
                ws.cell(row=r, column=c).fill = ALERT_FILL
        r += 1

    for col, w in {1: 20, 2: 42, 3: 10, 4: 11, 5: 11, 6: 17, 7: 38}.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = f"A{start + 3}"
    ws.auto_filter.ref = f"A{start + 2}:G{r - 1}"

    # Salvar
    wb.save(str(OUT_XLSX))
    print(f"\nXLSX consolidado salvo: {OUT_XLSX}")
    print(f"  Total transacoes (5 meses): {n_total:,}")
    print(f"  Volume bruto: R$ {abs(cred_total) + abs(deb_total):,.2f}")
    print(f"  Alertas pos-baixa: {len(pos_baixa)}")
    print(f"  Retencao total estimada (5m): R$ {total_ret_5m:,.2f}")


def gerar_markdown(todos, saldos, cache) -> str:
    n_total = len(todos)
    cred_total = sum(s["cred"] for s in saldos.values())
    deb_total = sum(s["deb"] for s in saldos.values())
    saldo_inicial_jan = saldos["JAN/2026"]["saldo_final"] - (saldos["JAN/2026"]["cred"] + saldos["JAN/2026"]["deb"])
    saldo_final_mai = saldos["MAI/2026"]["saldo_final"]
    volume_bruto = abs(cred_total) + abs(deb_total)

    # Agregados
    por_cnpj = defaultdict(lambda: {"n": 0, "cred": 0.0, "deb": 0.0, "meses": set()})
    pos_baixa = []
    cat_count = Counter()
    cat_volume = defaultdict(float)
    cat_retencao = defaultdict(float)

    for mes, t, _ in todos:
        cnpj = _extrair_cnpj(t)
        if cnpj:
            d = por_cnpj[cnpj]
            d["n"] += 1
            d["meses"].add(mes)
            if t.valor > 0:
                d["cred"] += t.valor
            else:
                d["deb"] += t.valor
            # Pos-baixa
            info = cache.get(cnpj, {})
            sit = info.get("situacao", "")
            if ("BAIXADA" in sit or "INAPTA" in sit) and info.get("data_situacao"):
                try:
                    db_d = date.fromisoformat(info["data_situacao"][:10])
                    dt_d = date.fromisoformat(t.data[:10])
                    if dt_d > db_d:
                        pos_baixa.append((mes, t, cnpj, info, (dt_d - db_d).days))
                except (ValueError, TypeError):
                    pass
        # Tributario
        info = cache.get(cnpj) if cnpj else None
        porte = (info.get("porte") if info else "") or ""
        trib = classificar_tributario(t.memo or "", t.nome or "", t.valor, cnpj or "", porte)
        cat_count[trib["categoria"]] += 1
        cat_volume[trib["categoria"]] += abs(t.valor)
        cat_retencao[trib["categoria"]] += trib["valor_retencao"]

    total_ret = sum(cat_retencao.values())

    lines = [
        "# AUDITORIA CONSOLIDADA — 5 MESES",
        "",
        "**[ORGATEC] Conciliacao Bancaria · Auditoria**",
        "",
        "**Empresa:** [NAO CADASTRADO] | **CNPJ:** [PENDENTE] | **Socios:** [PENDENTE]",
        "",
        f"**Conta:** 158083-3 | **Agencia:** 3333-2 | **Banco:** {BANCO}",
        "",
        f"**Periodo:** 01/01/2026 a 14/05/2026 (4,5 meses) | **Gerado em:** {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        "",
        "---",
        "",
        "## 1. Resumo Executivo",
        "",
        "| Indicador | Valor |",
        "|---|---|",
        f"| Total de transacoes (5 meses) | {n_total:,} |",
        f"| Volume de creditos | R$ {cred_total:,.2f} |",
        f"| Volume de debitos | R$ {deb_total:,.2f} |",
        f"| Volume bruto movimentado | R$ {volume_bruto:,.2f} |",
        f"| Saldo inicial (01/01) | R$ {saldo_inicial_jan:,.2f} |",
        f"| Saldo final (14/05) | R$ {saldo_final_mai:,.2f} |",
        f"| Variacao do periodo | R$ {saldo_final_mai - saldo_inicial_jan:,.2f} |",
        f"| Volume anualizado projetado | R$ {volume_bruto * 12 / 4.5:,.2f} |",
        f"| Alertas pos-baixa | **{len(pos_baixa)}** |",
        f"| Retencao tributaria estimada | **R$ {total_ret:,.2f}** |",
        "",
        "## 2. Evolucao Mensal",
        "",
        "| Mes | Transacoes | Creditos (R$) | Debitos (R$) | Saldo Final |",
        "|---|---:|---:|---:|---:|",
    ]
    for mes in ["JAN/2026", "FEV/2026", "MAR/2026", "ABR/2026", "MAI/2026"]:
        s = saldos[mes]
        marker = " 🆕" if mes == "FEV/2026" else (" ⚠️" if s["saldo_final"] < 0 else "")
        lines.append(
            f"| {mes}{marker} | {s['n']:,} | {s['cred']:,.2f} | {s['deb']:,.2f} | {s['saldo_final']:,.2f} |"
        )

    # Top 20 contrapartes
    top_cnpjs = sorted(por_cnpj.items(), key=lambda x: -(x[1]["cred"] + abs(x[1]["deb"])))[:20]
    lines += [
        "",
        "## 3. Top 20 Contrapartes Consolidadas",
        "",
        "| # | CNPJ | Razao Social | UF | Volume (R$) | Meses | Alerta |",
        "|---|---|---|---|---:|:---:|---|",
    ]
    for i, (cnpj, d) in enumerate(top_cnpjs, start=1):
        info = cache.get(cnpj, {})
        razao = info.get("razao_social", "(nao enriquecido)")[:42]
        sit = info.get("situacao", "")
        uf = info.get("uf", "?") or "?"
        is_baixada = "BAIXADA" in sit or "INAPTA" in sit
        volume = d["cred"] + abs(d["deb"])
        n_meses = len(d["meses"])
        porte = info.get("porte", "")
        eh_mei = porte == "MICRO EMPRESA"

        alerta = ""
        if is_baixada:
            alerta = "🔴 BAIXADA"
        elif eh_mei and volume > 81_000:
            alerta = "🟠 MEI estourando teto"
        elif n_meses == 5:
            alerta = "🔵 Recorrente"

        cnpj_fmt = f"{cnpj[:2]}.{cnpj[2:5]}.{cnpj[5:8]}/{cnpj[8:12]}-{cnpj[12:14]}"
        lines.append(f"| {i} | {cnpj_fmt} | {razao} | {uf} | {volume:,.2f} | {n_meses}/5 | {alerta} |")

    # Alertas pos-baixa
    if pos_baixa:
        lines += [
            "",
            "## 4. 🚨 Alertas Pos-Baixa",
            "",
            f"**{len(pos_baixa)} transacoes** identificadas em CNPJs ja BAIXADOS no momento do pagamento.",
            "",
            "| Mes | Data | Valor (R$) | Razao Social | Data Baixa | Dias Apos |",
            "|---|---|---:|---|---|---:|",
        ]
        pos_baixa.sort(key=lambda x: -x[4])
        total_pb = 0.0
        for mes, t, cnpj, info, dias in pos_baixa:
            razao = info.get("razao_social", "")[:35]
            data_b = info.get("data_situacao", "")
            lines.append(f"| {mes} | {t.data} | **{t.valor:,.2f}** | {razao} | {data_b} | **{dias}** |")
            total_pb += abs(t.valor)
        lines.append(f"| | | **R$ {total_pb:,.2f}** | TOTAL | | |")

    # Status Tributario
    lines += [
        "",
        "## 5. Status Tributario Consolidado",
        "",
        "Classificacao fiscal automatica das 7.110 transacoes:",
        "",
        "| Categoria | Qtd | Volume (R$) | Retencao Estimada (R$) |",
        "|---|---:|---:|---:|",
    ]
    CATS_ORD = ["RETENCAO_PJ", "RETENCAO_PF", "OPERACAO_CREDITO", "IOF", "JUROS",
                "PAGAMENTO_TRIBUTO", "TARIFA", "PIX_RECEBIDO", "BOLETO",
                "COMPRA_CARTAO", "NAO_TRIBUTAVEL", "OUTRO"]
    for cat in CATS_ORD:
        qtd = cat_count.get(cat, 0)
        if qtd == 0:
            continue
        vol = cat_volume[cat]
        ret = cat_retencao[cat]
        marker = "**" if cat.startswith("RETENCAO") else ""
        ret_fmt = f"**{ret:,.2f}**" if ret > 0 else "—"
        lines.append(f"| {marker}{cat}{marker} | {qtd:,} | {vol:,.2f} | {ret_fmt} |")
    lines.append(f"| **TOTAL** | **{sum(cat_count.values()):,}** | **{sum(cat_volume.values()):,.2f}** | **R$ {total_ret:,.2f}** |")

    # Recomendacoes
    lines += [
        "",
        "## 6. Recomendacoes",
        "",
        "1. **Confirmar enquadramento tributario** da empresa auditada — volume anualizado projetado",
        f"   de R$ {volume_bruto * 12 / 4.5:,.2f} excede largamente o teto Simples (R$ 4,8M/ano).",
        "2. **Investigar relacao com LOCAR TRANSPORTE** (R$ 2,79M em transferencias) — confirmar se Renato",
        "   Costa Esperidiao Junior tem participacao na empresa auditada.",
        f"3. **Auditar MEIs estourando teto** — multiplos fornecedores PJ recebendo valores anualizados",
        "   superiores a R$ 81.000 (teto MEI).",
        f"4. **Apurar retencoes nao recolhidas** — R$ {total_ret:,.2f} sugeridos como obrigacao",
        "   tributaria pendente (PIS+COFINS+CSLL+IRRF+INSS).",
        f"5. **Investigar {len(pos_baixa)} alertas pos-baixa** — pagamentos a CNPJs ja extintos no",
        "   momento do pagamento (red flag forense).",
        "",
        "---",
        "",
        "*Sistema OrgConc/OrgNeural2 - cascata de 6 estagios + enriquecimento RFB via BrasilAPI.*",
    ]

    return "\n".join(lines)


def gerar_html(md_text: str) -> str:
    import markdown as mdlib
    body = mdlib.markdown(md_text, extensions=["tables", "fenced_code"])
    agora = datetime.now().strftime("%d/%m/%Y %H:%M")
    css = """
@page { size: A4 landscape; margin: 14mm 12mm 14mm 12mm;
  @bottom-right { content: "Pagina " counter(page) " de " counter(pages); font-size: 9px; color: #6B7280; } }
* { box-sizing: border-box; margin: 0; padding: 0; }
body { font-family: 'DejaVu Sans', Arial, sans-serif; font-size: 10pt; color: #1a202c; line-height: 1.55; }
.hd { background: linear-gradient(135deg, #0F172A, #0B1B3D 45%, #0052FF); color: #fff;
      padding: 22px 28px; border-radius: 12px; margin-bottom: 22px; }
.hd h1 { font-size: 20pt; font-family: 'DejaVu Serif', Georgia, serif; margin-bottom: 4px; }
.hd .tag { font-size: 9pt; opacity: 0.85; text-transform: uppercase; letter-spacing: 0.16em; }
h1 { font-size: 14pt; color: #0F172A; margin: 22px 0 8px; padding-bottom: 6px; border-bottom: 2px solid #BFDBFE; }
h2 { font-size: 12pt; color: #0052FF; margin: 18px 0 8px; padding-left: 10px; border-left: 3px solid #0EA5E9; }
table { width: 100%; border-collapse: collapse; margin: 10px 0 14px; font-size: 9pt; border-radius: 6px; overflow: hidden; }
th { background: linear-gradient(180deg, #0F172A, #1E3A8A); color: #fff; padding: 6px 9px; text-align: left; font-weight: 600; }
td { padding: 5px 9px; border-bottom: 1px solid #E2E8F0; vertical-align: top; }
tr:nth-child(even) td { background: #F8FAFC; }
strong { color: #0F172A; font-weight: 700; }
.ft { margin-top: 28px; padding-top: 12px; border-top: 1px solid #E2E8F0; font-size: 8.5pt; color: #94A3B8; }
"""
    logo_html = html_logo_inline()
    return f"""<!DOCTYPE html><html lang="pt-BR"><head><meta charset="UTF-8">
<title>Auditoria Consolidada 158083-3</title><style>{css}
.hd {{ display: flex; align-items: center; gap: 18px; }}
.hd-text {{ flex: 1; }}
</style></head>
<body>
<div class="hd">{logo_html}<div class="hd-text">
<h1>ORGATEC</h1>
<div class="tag">Auditoria Consolidada · 5 Meses · Conta 158083-3</div>
<div style="margin-top:10px;font-size:10pt;opacity:.92">Sicoob 756 · Agencia 3333-2 · Periodo 01/01 a 14/05/2026 · Gerado em {agora}</div>
</div></div>
{body}
<div class="ft">(c) ORGATEC Contabilidade e Auditoria - OrgConc v0.5.0</div>
</body></html>"""


async def gerar_pdf(html_text: str) -> bool:
    try:
        from playwright.async_api import async_playwright
        async with async_playwright() as p:
            browser = await p.chromium.launch()
            page = await browser.new_page()
            await page.set_content(html_text, wait_until="load")
            await page.pdf(
                path=str(OUT_PDF), format="A4", landscape=True,
                margin={"top": "14mm", "right": "12mm", "bottom": "14mm", "left": "12mm"},
                print_background=True,
            )
            await browser.close()
        return True
    except Exception as exc:  # noqa: BLE001
        print(f"PDF failed: {exc}")
        return False


async def main_async():
    print("Coletando dados de 5 OFXs...")
    todos, saldos, cache = coletar_dados()
    print(f"  {len(todos):,} transacoes lidas")

    gerar_xlsx(todos, saldos, cache)

    md_text = gerar_markdown(todos, saldos, cache)
    OUT_MD.write_text(md_text, encoding="utf-8")
    print(f"  MD:   {OUT_MD}")

    html_text = gerar_html(md_text)
    OUT_HTML.write_text(html_text, encoding="utf-8")
    print(f"  HTML: {OUT_HTML}")

    ok = await gerar_pdf(html_text)
    if ok:
        print(f"  PDF:  {OUT_PDF}")


if __name__ == "__main__":
    asyncio.run(main_async())
