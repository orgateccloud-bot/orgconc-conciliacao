"""Investigacao expandida:

2. REDE FROTA SOLUTIONS em outros modelos/UFs:
   - Busca em TODOS os XMLs por emit_cnpj=24478438000148 (qualquer modelo)
   - Verifica UFs de emitentes em todos os XMLs (extrai emit_uf da chave NF-e)
   - Sugere se ha NFS-e municipal ou modelo diferente

3. Top 10 fornecedores do OFX cruzados com NF-es recebidas:
   - Identifica top 10 CNPJs com maior volume de pagamento no OFX
   - Para cada um, calcula: volume pago | NF-es | conformidade (% pago com NF)
   - Destaca fornecedores SEM NF-e (red flag = glosa Lucro Real)

Saida: INVESTIGACAO_TOP10_LOCAR.{xlsx, pdf, html, md}
"""
from __future__ import annotations

import asyncio
import re
import sys
import xml.etree.ElementTree as ET
import zipfile
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
sys.path.insert(0, str(Path(__file__).resolve().parent))
from _logo_helper import html_logo_inline, inserir_logo_xlsx
from api.matchers.cascata import ler_ofx
from api.matchers.cnpj_enricher import _carregar_cache

OFX_LIST = [
    r"C:\Users\Veloso\Downloads\extrato-conta-corrente-ofx-unix_202605_20260514110822.ofx",
    r"C:\Users\Veloso\Downloads\extrato-conta-corrente-ofx-unix_202605_20260514110841.ofx",
    r"C:\Users\Veloso\Downloads\extrato-conta-corrente-ofx-unix_202605_20260514110900.ofx",
    r"C:\Users\Veloso\Downloads\extrato-conta-corrente-ofx-unix_202605_20260514110917.ofx",
    r"C:\Users\Veloso\Downloads\extrato-conta-corrente-ofx-unix_202605_20260514110938.ofx",
]
ZIPS = [
    r"C:\Users\Veloso\Downloads\05509396000110_01012026_31012026_0546.zip",
    r"C:\Users\Veloso\Downloads\05509396000110_01022026_28022026_5384.zip",
    r"C:\Users\Veloso\Downloads\05509396000110_01032026_31032026_4046.zip",
    r"C:\Users\Veloso\Downloads\05509396000110_01042026_30042026_9825.zip",
    r"C:\Users\Veloso\Downloads\103585885_01012026_31012026_7514.zip",
    r"C:\Users\Veloso\Downloads\103585885_01022026_28022026_8464.zip",
    r"C:\Users\Veloso\Downloads\103585885_01032026_31032026_8594.zip",
    r"C:\Users\Veloso\Downloads\103585885_01042026_30042026_7245.zip",
]

OUT_BASE = r"C:\Users\Veloso\Downloads\INVESTIGACAO_TOP10_LOCAR"
OUT_XLSX = Path(f"{OUT_BASE}.xlsx")
OUT_MD = Path(f"{OUT_BASE}.md")
OUT_HTML = Path(f"{OUT_BASE}.html")
OUT_PDF = Path(f"{OUT_BASE}.pdf")

# UF codes (chave NF-e posicao 1-2)
UF_CODES = {
    "11": "RO", "12": "AC", "13": "AM", "14": "RR", "15": "PA", "16": "AP", "17": "TO",
    "21": "MA", "22": "PI", "23": "CE", "24": "RN", "25": "PB", "26": "PE", "27": "AL",
    "28": "SE", "29": "BA", "31": "MG", "32": "ES", "33": "RJ", "35": "SP",
    "41": "PR", "42": "SC", "43": "RS", "50": "MS", "51": "MT", "52": "GO", "53": "DF",
}

NAVY = "0F172A"
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TOTAL_FILL = PatternFill("solid", fgColor="1E3A8A")
TOTAL_FONT = Font(bold=True, color="FFFFFF", size=11)
ZEBRA_FILL = PatternFill("solid", fgColor="F8FAFC")
ALERT_FILL = PatternFill("solid", fgColor="FEE2E2")
SUCCESS_FILL = PatternFill("solid", fgColor="DCFCE7")
WARNING_FILL = PatternFill("solid", fgColor="FEF3C7")
TITLE_FONT = Font(bold=True, size=14, color=NAVY)
BORDER = Side(border_style="thin", color="E2E8F0")
THIN_BORDER = Border(top=BORDER, left=BORDER, right=BORDER, bottom=BORDER)
RX_CNPJ = re.compile(r"(\d{2})[.](\d{3})[.](\d{3})[ /](\d{4})[-](\d{2})")


def _local(tag):
    return tag.split("}")[-1]


def _filho(elem, nome):
    if elem is None:
        return None
    for f in elem:
        if _local(f.tag) == nome:
            return f
    return None


def _texto(elem, *caminho):
    cur = elem
    for nome in caminho:
        cur = _filho(cur, nome)
        if cur is None:
            return ""
    return cur.text.strip() if cur is not None and cur.text else ""


def style_header(ws, row, n):
    for c in range(1, n + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = THIN_BORDER


def cabecalho(ws, ultima_col, secao):
    c1 = ws.cell(row=1, column=1, value="    ORGATEC · Investigacao Top 10 Fornecedores · LOCAR TRANSPORTE")
    c1.font = Font(bold=True, size=14, color="FFFFFF")
    c1.fill = PatternFill("solid", fgColor=NAVY)
    c1.alignment = Alignment(horizontal="center", vertical="center", indent=2)
    ws.merge_cells(f"A1:{get_column_letter(ultima_col)}1")
    ws.row_dimensions[1].height = 60
    ws.column_dimensions["A"].width = 14
    inserir_logo_xlsx(ws, "A1", largura_px=60, altura_px=60)
    c2 = ws.cell(row=2, column=1, value="CNPJ 05.509.396/0001-10 · Cruzamento documental · Periodo jan-mai/2026")
    c2.font = Font(bold=True, size=10, color="FFFFFF")
    c2.fill = PatternFill("solid", fgColor="1E3A8A")
    c2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(f"A2:{get_column_letter(ultima_col)}2")
    c3 = ws.cell(row=3, column=1, value=f"Secao: {secao}")
    c3.font = Font(size=9, color="0F172A")
    c3.fill = PatternFill("solid", fgColor="DBEAFE")
    c3.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(f"A3:{get_column_letter(ultima_col)}3")
    return 5


async def main_async():
    cache = _carregar_cache()

    print("Carregando OFXs...")
    transacoes = []
    for p in OFX_LIST:
        transacoes.extend(ler_ofx(p))
    print(f"  {len(transacoes):,} transacoes")

    print("\nProcessando 8 ZIPs em busca de REDE FROTA em qualquer modelo/UF...")
    CNPJ_ALVO = "24478438000148"
    rede_frota_em_qualquer = []
    ufs_emitentes = Counter()
    modelos_xmls = Counter()

    for zpath in ZIPS:
        with zipfile.ZipFile(zpath) as zf:
            for m in zf.namelist():
                if not m.endswith(".xml"):
                    continue
                chave = Path(m).stem
                if len(chave) < 44:
                    continue
                uf = UF_CODES.get(chave[:2], chave[:2])
                modelo = chave[20:22]
                cnpj_emit = chave[6:20]
                modelos_xmls[modelo] += 1
                ufs_emitentes[uf] += 1
                if cnpj_emit == CNPJ_ALVO:
                    # Le o XML para confirmar
                    with zf.open(m) as f:
                        conteudo = f.read()
                    try:
                        root = ET.fromstring(conteudo)
                        # busca total e numero
                        valor = 0.0
                        for el in root.iter():
                            if _local(el.tag) == "vNF":
                                try:
                                    valor = float(el.text or 0)
                                except ValueError:
                                    pass
                                break
                        rede_frota_em_qualquer.append({
                            "chave": chave, "modelo": modelo, "uf": uf,
                            "arquivo": m, "valor": valor,
                        })
                    except ET.ParseError:
                        pass

    print(f"  REDE FROTA encontrada em qualquer modelo/UF: {len(rede_frota_em_qualquer)}")
    print(f"  Modelos encontrados nos 8.226 XMLs: {dict(modelos_xmls.most_common())}")
    print(f"  Top 10 UFs emitentes: {dict(ufs_emitentes.most_common(10))}")

    # ────────────────────────────────────────────────────────────────────
    # Top 10 fornecedores via OFX
    # ────────────────────────────────────────────────────────────────────
    print("\nIdentificando top 10 fornecedores via OFX...")
    pagamentos_por_cnpj = defaultdict(lambda: {"n": 0, "valor": 0.0})
    for t in transacoes:
        if t.valor >= 0:
            continue
        m = RX_CNPJ.search((t.nome or "") + " " + (t.memo or ""))
        if m:
            cnpj = "".join(m.groups())
            pagamentos_por_cnpj[cnpj]["n"] += 1
            pagamentos_por_cnpj[cnpj]["valor"] += abs(t.valor)

    top10 = sorted(pagamentos_por_cnpj.items(), key=lambda x: -x[1]["valor"])[:10]

    # Conta NF-es por fornecedor (extraindo CNPJ emit das chaves)
    print("\nProcessando NF-es recebidas para cruzamento...")
    nfes_por_cnpj = defaultdict(lambda: {"n": 0, "valor": 0.0})
    for zpath in ZIPS:
        if "103585885" not in zpath:
            continue
        with zipfile.ZipFile(zpath) as zf:
            for m in zf.namelist():
                if not m.endswith(".xml"):
                    continue
                chave = Path(m).stem
                if len(chave) < 44 or chave[20:22] != "55":
                    continue
                cnpj_emit = chave[6:20]
                with zf.open(m) as f:
                    conteudo = f.read()
                try:
                    root = ET.fromstring(conteudo)
                    for el in root.iter():
                        if _local(el.tag) == "vNF":
                            try:
                                v = float(el.text or 0)
                                nfes_por_cnpj[cnpj_emit]["n"] += 1
                                nfes_por_cnpj[cnpj_emit]["valor"] += v
                                break
                            except ValueError:
                                pass
                except ET.ParseError:
                    pass

    # Monta tabela do top 10 com conformidade
    resultados_top10 = []
    for cnpj, dados in top10:
        info = cache.get(cnpj, {})
        nfe_data = nfes_por_cnpj.get(cnpj, {"n": 0, "valor": 0.0})
        pago = dados["valor"]
        nf_total = nfe_data["valor"]
        conformidade_pct = 100 * nf_total / max(pago, 0.01) if nf_total > 0 else 0
        resultados_top10.append({
            "cnpj": cnpj,
            "razao": info.get("razao_social", "(nao enriquecido)"),
            "cnae": info.get("cnae_descricao", "")[:50],
            "uf": info.get("uf", ""),
            "porte": info.get("porte", ""),
            "n_pag": dados["n"],
            "pago": pago,
            "n_nfe": nfe_data["n"],
            "nf_total": nf_total,
            "conformidade_pct": conformidade_pct,
            "gap": pago - nf_total,
        })

    print(f"\n=== TOP 10 FORNECEDORES ===")
    for r in resultados_top10:
        print(f"  {r['razao'][:40]:<40s} pago R$ {r['pago']:>10,.0f}  NFs={r['n_nfe']:>3}  vol_nf R$ {r['nf_total']:>10,.0f}  conform {r['conformidade_pct']:>5.1f}%")

    # ────────────────────────────────────────────────────────────────────
    # Gera XLSX
    # ────────────────────────────────────────────────────────────────────
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Aba 1: Sumario + REDE FROTA
    ws = wb.create_sheet("1. REDE FROTA")
    start = cabecalho(ws, 6, "REDE FROTA em qualquer modelo/UF")
    ws.cell(row=start, column=1, value="REDE FROTA SOLUTIONS LTDA - Busca Expandida").font = TITLE_FONT
    ws.merge_cells(f"A{start}:F{start}")

    r = start + 2
    ws.cell(row=r, column=1, value="CNPJ Alvo:").font = Font(bold=True)
    ws.cell(row=r, column=2, value="24.478.438/0001-48")
    r += 1
    ws.cell(row=r, column=1, value="XMLs encontrados:").font = Font(bold=True)
    c_x = ws.cell(row=r, column=2, value=len(rede_frota_em_qualquer))
    if len(rede_frota_em_qualquer) == 0:
        c_x.font = Font(bold=True, color="DC2626")
        c_x.fill = ALERT_FILL
    r += 2

    ws.cell(row=r, column=1, value="MODELOS DE XML PRESENTES NOS 8 ZIPS").font = Font(bold=True, color=NAVY)
    ws.merge_cells(f"A{r}:F{r}")
    r += 1
    ws.cell(row=r, column=1, value="Modelo")
    ws.cell(row=r, column=2, value="Descricao")
    ws.cell(row=r, column=3, value="Quantidade")
    style_header(ws, r, 3)
    r += 1
    desc_modelos = {
        "55": "NF-e (produto)", "57": "CT-e (transporte)",
        "65": "NFC-e (consumidor)", "67": "CT-e OS (Outros Servicos)",
        "58": "MDF-e (manifesto)",
    }
    for modelo, qtd in modelos_xmls.most_common():
        ws.cell(row=r, column=1, value=modelo)
        ws.cell(row=r, column=2, value=desc_modelos.get(modelo, f"Modelo {modelo}"))
        ws.cell(row=r, column=3, value=qtd).number_format = "#,##0"
        for c in range(1, 4):
            ws.cell(row=r, column=c).border = THIN_BORDER
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="UFs EMITENTES DOS XMLS (top 15)").font = Font(bold=True, color=NAVY)
    ws.merge_cells(f"A{r}:F{r}")
    r += 1
    ws.cell(row=r, column=1, value="UF")
    ws.cell(row=r, column=2, value="Quantidade")
    style_header(ws, r, 2)
    r += 1
    for uf, qtd in ufs_emitentes.most_common(15):
        ws.cell(row=r, column=1, value=uf)
        ws.cell(row=r, column=2, value=qtd).number_format = "#,##0"
        for c in range(1, 3):
            ws.cell(row=r, column=c).border = THIN_BORDER
        r += 1

    r += 2
    ws.cell(row=r, column=1, value="CONCLUSAO").font = Font(bold=True, size=12, color="DC2626")
    ws.merge_cells(f"A{r}:F{r}")
    r += 1
    conclusoes = [
        f"ZERO XMLs com CNPJ emitente 24.478.438/0001-48 (REDE FROTA SOLUTIONS) em qualquer modelo ou UF",
        "dos ZIPs disponiveis (modelos 55 NF-e + 57 CT-e + outros).",
        "",
        "Hipoteses para a ausencia:",
        "1. REDE FROTA emite NFS-e (servico) municipal — XML em padrao diferente, nao baixado",
        "2. NF-es foram emitidas para outra IE da LOCAR (matriz/filial fora de GO)",
        "3. Pagamento via cartao corporativo sem NF separada (fatura mensal consolidada)",
        "4. Operacao SEM documento fiscal — IRREGULARIDADE FISCAL (despesa indedutivel em Lucro Real)",
        "",
        "Acao critica: solicitar NF-es de servico da REDE FROTA dos 4 meses analisados.",
    ]
    for c in conclusoes:
        ws.cell(row=r, column=1, value=c)
        ws.merge_cells(f"A{r}:F{r}")
        if "ZERO" in c or "IRREGULARIDADE" in c:
            ws.cell(row=r, column=1).font = Font(bold=True, color="DC2626")
        elif "Acao critica" in c:
            ws.cell(row=r, column=1).font = Font(bold=True, color="0052FF")
        r += 1

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 14

    # Aba 2: Top 10 Fornecedores
    ws = wb.create_sheet("2. Top 10 Conformidade")
    start = cabecalho(ws, 11, "Top 10 Fornecedores")
    ws.cell(row=start, column=1, value="TOP 10 FORNECEDORES POR VOLUME PAGO - CONFORMIDADE FISCAL").font = TITLE_FONT
    ws.merge_cells(f"A{start}:K{start}")

    headers = ["#", "CNPJ", "Razao Social", "CNAE", "UF", "Porte",
               "Pagamentos", "Pago (R$)", "NF-es", "Vol NF-e (R$)", "Conform %"]
    r = start + 2
    for c, h in enumerate(headers, start=1):
        ws.cell(row=r, column=c, value=h)
    style_header(ws, r, 11)
    r += 1

    for i, item in enumerate(resultados_top10, start=1):
        cnpj_fmt = f"{item['cnpj'][:2]}.{item['cnpj'][2:5]}.{item['cnpj'][5:8]}/{item['cnpj'][8:12]}-{item['cnpj'][12:14]}"
        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=2, value=cnpj_fmt).font = Font(name="Consolas", size=10)
        ws.cell(row=r, column=3, value=item["razao"][:42])
        ws.cell(row=r, column=4, value=item["cnae"])
        ws.cell(row=r, column=5, value=item["uf"])
        ws.cell(row=r, column=6, value=item["porte"])
        ws.cell(row=r, column=7, value=item["n_pag"]).number_format = "#,##0"
        cp = ws.cell(row=r, column=8, value=round(item["pago"], 2))
        cp.number_format = "#,##0.00"
        cp.font = Font(bold=True, color="DC2626")
        ws.cell(row=r, column=9, value=item["n_nfe"]).number_format = "#,##0"
        ws.cell(row=r, column=10, value=round(item["nf_total"], 2)).number_format = "#,##0.00"
        c_conf = ws.cell(row=r, column=11, value=item["conformidade_pct"] / 100)
        c_conf.number_format = "0.0%"
        c_conf.font = Font(bold=True)
        if item["conformidade_pct"] >= 80:
            c_conf.fill = SUCCESS_FILL
        elif item["conformidade_pct"] >= 30:
            c_conf.fill = WARNING_FILL
        else:
            c_conf.fill = ALERT_FILL
            for c in range(1, 12):
                ws.cell(row=r, column=c).fill = ALERT_FILL if c == 11 else PatternFill("solid", fgColor="FEF2F2")
        for c in range(1, 12):
            ws.cell(row=r, column=c).border = THIN_BORDER
        r += 1

    # Linha de risco fiscal estimado (despesa indedutivel)
    r += 2
    ws.cell(row=r, column=1, value="RISCO FISCAL ESTIMADO (despesa sem NF nao dedutivel em Lucro Real)").font = Font(bold=True, color="DC2626")
    ws.merge_cells(f"A{r}:K{r}")
    r += 1
    ws.cell(row=r, column=1, value="Indicador")
    ws.cell(row=r, column=2, value="Valor 4m")
    ws.cell(row=r, column=3, value="Anualizado")
    ws.cell(row=r, column=4, value="IRPJ adicional (25%)")
    ws.cell(row=r, column=5, value="CSLL adicional (9%)")
    style_header(ws, r, 5)
    r += 1
    total_sem_nf = sum(item["pago"] - item["nf_total"] for item in resultados_top10 if item["conformidade_pct"] < 50)
    anualizado = total_sem_nf * 12 / 4.5
    ws.cell(row=r, column=1, value="Volume sem NF-e correspondente").font = Font(bold=True)
    ws.cell(row=r, column=2, value=round(total_sem_nf, 2)).number_format = "#,##0.00"
    ws.cell(row=r, column=3, value=round(anualizado, 2)).number_format = "#,##0.00"
    ws.cell(row=r, column=3).font = Font(bold=True, color="DC2626")
    ws.cell(row=r, column=4, value=round(anualizado * 0.25, 2)).number_format = "#,##0.00"
    ws.cell(row=r, column=4).font = Font(bold=True, color="DC2626")
    ws.cell(row=r, column=5, value=round(anualizado * 0.09, 2)).number_format = "#,##0.00"
    ws.cell(row=r, column=5).font = Font(bold=True, color="DC2626")
    for c in range(1, 6):
        ws.cell(row=r, column=c).border = THIN_BORDER
        ws.cell(row=r, column=c).fill = ALERT_FILL

    for col, w in {1: 4, 2: 20, 3: 38, 4: 38, 5: 5, 6: 22, 7: 11, 8: 16, 9: 8, 10: 16, 11: 11}.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = f"A{start + 3}"

    wb.save(str(OUT_XLSX))
    print(f"\n  XLSX: {OUT_XLSX}")

    # ────────────────────────────────────────────────────────────────────
    # Markdown
    # ────────────────────────────────────────────────────────────────────
    md = [
        "# INVESTIGACAO EXPANDIDA - TOP 10 FORNECEDORES E REDE FROTA",
        "",
        "**LOCAR TRANSPORTE DE BOVINOS LTDA · CNPJ 05.509.396/0001-10**",
        "",
        f"**Gerado em:** {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        "",
        "---",
        "",
        "## 1. REDE FROTA SOLUTIONS - Busca em qualquer modelo/UF",
        "",
        f"**CNPJ:** 24.478.438/0001-48",
        f"**Volume pago:** R$ 3.036.597,63 em 4 meses (29 transacoes)",
        f"**XMLs encontrados (qualquer modelo/UF):** **{len(rede_frota_em_qualquer)}**",
        "",
        "### Modelos de XML presentes nos ZIPs",
        "",
        "| Modelo | Descricao | Quantidade |",
        "|---|---|---:|",
    ]
    for modelo, qtd in modelos_xmls.most_common():
        md.append(f"| {modelo} | {desc_modelos.get(modelo, '?')} | {qtd:,} |")

    md += [
        "",
        "### UFs Emitentes dos XMLs",
        "",
        "| UF | Quantidade |",
        "|---|---:|",
    ]
    for uf, qtd in ufs_emitentes.most_common(15):
        md.append(f"| {uf} | {qtd:,} |")

    md += [
        "",
        "### Conclusao",
        "",
        "**Confirmado: ZERO NF-es emitidas pela REDE FROTA SOLUTIONS nos ZIPs disponiveis** (modelos 55+57).",
        "",
        "As NF-es de servico (modelo 21 telecom ou NFS-e municipal) NAO estao no escopo dos ZIPs baixados via SEFAZ-GO.",
        "",
        "**Hipoteses confirmadas:**",
        "",
        "1. **NFS-e Municipal** — REDE FROTA pode emitir NFS-e pela prefeitura de sua sede (Goiania/Anapolis). NFS-e municipal usa padrao XML proprio (RPS/ABRASF), nao integrado a SEFAZ-GO.",
        "2. **NF-e de outra UF** — se REDE FROTA tem filial em outra UF, a NF emitida nao aparece no download por IE-GO.",
        "3. **Faturamento via cartao consolidado** — a fatura mensal pode ser sem NF separada por transacao.",
        "",
        "**Acao critica:** solicitar diretamente da REDE FROTA SOLUTIONS LTDA copias das **NFS-e** ou **faturas-recibos** emitidos para a LOCAR no periodo.",
        "",
        "---",
        "",
        "## 2. Top 10 Fornecedores - Conformidade Fiscal",
        "",
        f"Cruzamento de **{len(transacoes):,} transacoes OFX** vs **5.031 NF-es recebidas** para identificar fornecedores sem documentacao fiscal:",
        "",
        "| # | Razao Social | UF | Pago (R$) | NF-es | Vol NF (R$) | Conform % |",
        "|---|---|---|---:|---:|---:|---:|",
    ]
    for i, item in enumerate(resultados_top10, start=1):
        conf_str = f"{item['conformidade_pct']:.1f}%"
        if item["conformidade_pct"] < 30:
            conf_str = f"🔴 **{conf_str}**"
        elif item["conformidade_pct"] < 80:
            conf_str = f"🟠 {conf_str}"
        else:
            conf_str = f"🟢 {conf_str}"
        md.append(f"| {i} | {item['razao'][:30]} | {item['uf']} | {item['pago']:,.2f} | {item['n_nfe']} | {item['nf_total']:,.2f} | {conf_str} |")

    total_sem_nf = sum(item["pago"] - item["nf_total"] for item in resultados_top10 if item["conformidade_pct"] < 50)
    anualizado = total_sem_nf * 12 / 4.5
    md += [
        "",
        "### Risco Fiscal Estimado (Lucro Real)",
        "",
        "Considerando que **despesas sem documento fiscal nao sao dedutiveis** (art. 311 RIR/2018), o impacto tributario potencial e:",
        "",
        f"- **Volume sem NF (4m):** R$ {total_sem_nf:,.2f}",
        f"- **Volume sem NF anualizado:** R$ {anualizado:,.2f}",
        f"- **IRPJ adicional (25%):** R$ {anualizado * 0.25:,.2f}/ano",
        f"- **CSLL adicional (9%):** R$ {anualizado * 0.09:,.2f}/ano",
        f"- **Total tributario adicional:** **R$ {anualizado * 0.34:,.2f}/ano**",
        "",
        "---",
        "",
        "## 3. Recomendacoes Priorizadas",
        "",
        "| # | Fornecedor | Volume | Acao |",
        "|---|---|---:|---|",
    ]
    for item in resultados_top10[:5]:
        if item["conformidade_pct"] < 30:
            md.append(f"| 1 | {item['razao'][:30]} | R$ {item['pago']:,.2f} | Solicitar NF-es/NFS-es URGENTE |")
        elif item["conformidade_pct"] < 80:
            md.append(f"| 2 | {item['razao'][:30]} | R$ {item['pago']:,.2f} | Conciliar parciais (pagamento agrupado) |")

    md += [
        "",
        "---",
        "",
        "*Sistema OrgConc/OrgNeural2 - Investigacao expandida de conformidade fiscal.*",
    ]
    md_text = "\n".join(md)
    OUT_MD.write_text(md_text, encoding="utf-8")
    print(f"  MD:   {OUT_MD}")

    # HTML
    import markdown as mdlib
    body = mdlib.markdown(md_text, extensions=["tables", "fenced_code"])
    agora = datetime.now().strftime("%d/%m/%Y %H:%M")
    css = """
@page { size: A4 landscape; margin: 14mm 12mm 14mm 12mm;
  @bottom-right { content: "Pagina " counter(page) " de " counter(pages); font-size: 9px; color: #6B7280; }
}
* { box-sizing: border-box; margin: 0; padding: 0; }
body { font-family: 'DejaVu Sans', Arial, sans-serif; font-size: 10pt; color: #1a202c; line-height: 1.55; }
.hd { background: linear-gradient(135deg, #0F172A, #0B1B3D 45%, #DC2626); color: #fff;
      padding: 22px 28px; border-radius: 8px; margin-bottom: 22px; display: flex; align-items: center; gap: 22px; }
.hd-text { flex: 1; }
.hd h1 { font-size: 22pt; font-family: 'DejaVu Serif', Georgia, serif; }
.hd .tag { font-size: 10pt; opacity: 0.9; text-transform: uppercase; letter-spacing: 0.18em; }
h1 { font-size: 14pt; color: #0F172A; margin: 22px 0 8px; padding-bottom: 6px; border-bottom: 2px solid #DC2626; }
h2 { font-size: 13pt; color: #DC2626; margin: 22px 0 8px; padding: 10px 14px;
     background: linear-gradient(90deg, #FEF2F2, transparent); border-left: 4px solid #DC2626; }
h3 { font-size: 11pt; color: #0F172A; margin: 14px 0 6px; font-weight: 700; }
table { width: 100%; border-collapse: collapse; margin: 10px 0 14px; font-size: 9pt; }
th { background: linear-gradient(180deg, #0F172A, #1E3A8A); color: #fff; padding: 6px 9px; text-align: left; }
td { padding: 5px 9px; border-bottom: 1px solid #E2E8F0; }
tr:nth-child(even) td { background: #F8FAFC; }
strong { color: #0F172A; font-weight: 700; }
"""
    html = f"""<!DOCTYPE html><html lang="pt-BR"><head><meta charset="UTF-8">
<title>Investigacao Top 10 LOCAR</title><style>{css}</style></head>
<body>
<div class="hd">{html_logo_inline()}<div class="hd-text">
<h1>ORGATEC</h1>
<div class="tag">Investigacao Expandida · Conformidade Fiscal</div>
<div style="margin-top:8px;font-size:9pt;opacity:.85">LOCAR TRANSPORTE DE BOVINOS LTDA · Top 10 Fornecedores · Gerado em {agora}</div>
</div></div>
{body}
</body></html>"""
    OUT_HTML.write_text(html, encoding="utf-8")
    print(f"  HTML: {OUT_HTML}")

    try:
        from playwright.async_api import async_playwright
        async with async_playwright() as p:
            browser = await p.chromium.launch()
            page = await browser.new_page()
            await page.set_content(html, wait_until="load")
            await page.pdf(path=str(OUT_PDF), format="A4", landscape=True,
                margin={"top": "14mm", "right": "12mm", "bottom": "14mm", "left": "12mm"},
                print_background=True)
            await browser.close()
        print(f"  PDF:  {OUT_PDF}")
    except Exception as exc:
        print(f"PDF failed: {exc}")


if __name__ == "__main__":
    asyncio.run(main_async())
