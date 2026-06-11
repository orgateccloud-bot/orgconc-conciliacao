"""Fase 0+1: substitui locale en-US por pt-BR e corrige acentos em laudo_forense.py."""
from __future__ import annotations
import re
from pathlib import Path

src = Path("api/services/laudo_forense.py")
code = src.read_text(encoding="utf-8")

# ── 0) Adicionar import report_utils ────────────────────────────────────────
old_imp = (
    "from openpyxl import Workbook\n"
    "from openpyxl.styles import Alignment, Border, Font, PatternFill, Side\n"
    "from openpyxl.utils import get_column_letter"
)
new_imp = (
    "from openpyxl import Workbook\n"
    "from openpyxl.styles import Alignment, Border, Font, PatternFill, Side\n"
    "from openpyxl.utils import get_column_letter\n\n"
    "from api.services.report_utils import format_brl, format_num, format_pct, font_faces_css"
)
if old_imp in code:
    code = code.replace(old_imp, new_imp, 1)
    print("OK: import report_utils adicionado")
else:
    print("AVISO: bloco import openpyxl nao encontrado")

# ── 1) Locale: R$ {v:,.2f} → R$ {format_brl(v)} (todos os lugares) ──────────
code = re.sub(
    r'R\$\s*\{([^}:]+):,\.2f\}',
    lambda m: f'R$ {{format_brl({m.group(1)})}}',
    code,
)
# {v:,.2f} sem R$ → {format_brl(v)}
code = re.sub(
    r'\{([^}:]+):,\.2f\}',
    lambda m: f'{{format_brl({m.group(1)})}}',
    code,
)
# {n:,.0f} → {format_brl(n)} (valor inteiro sem centavos mas ainda monetário)
code = re.sub(
    r'\{([^}:]+):,\.0f\}',
    lambda m: f'{{format_brl({m.group(1)})}}',
    code,
)
# {n:,} → {format_num(n)} (inteiro com separador de milhar)
code = re.sub(
    r'\{([^}:]+):,\}',
    lambda m: f'{{format_num({m.group(1)})}}',
    code,
)
# {v:.1f}% → {str(round(v,1)).replace('.', ',')}%
code = re.sub(
    r'\{([^}:]+):\d+\.\df\}%',
    lambda m: f'{{str(round({m.group(1)},1)).replace(".", ",")}}%',
    code,
)
print("OK: locale substituido")

# ── 2) Acentuacao: strings de secao e labels ─────────────────────────────────
accent_pairs = [
    # Títulos MD
    ('"## 1. Sumario Executivo"',     '"## 1. Sumário Executivo"'),
    ('"## 2. Identificacao Cadastral"', '"## 2. Identificação Cadastral"'),
    ('"### Dados da Pessoa Juridica (Contrato Social + Cartao CNPJ)"',
     '"### Dados da Pessoa Jurídica (Contrato Social + Cartão CNPJ)"'),
    ('"### Quadro Societario"', '"### Quadro Societário"'),
    ('"## 3. Evolucao Mensal"',  '"## 3. Evolução Mensal"'),
    ('"## 4. Risk Heatmap (Distribuicao por Classe)"', '"## 4. Risk Heatmap (Distribuição por Classe)"'),
    ('"## 7. Status Tributario Consolidado"', '"## 7. Status Tributário Consolidado"'),
    ('"## 9. Conclusao"', '"## 9. Conclusão"'),
    ('"## Conformidade Fiscal (OFX x NF-e x CT-e)"', '"## Conformidade Fiscal (OFX × NF-e × CT-e)"'),
    # Cabeçalhos de tabela
    ('"| Categoria | Qtd | Volume (R$) | Retencao (R$) |"',
     '"| Categoria | Qtd | Volume (R$) | Retenção (R$) |"'),
    ('"| Mes | Transacoes | Creditos (R$) | Debitos (R$) | Saldo Final |"',
     '"| Mês | Transações | Créditos (R$) | Débitos (R$) | Saldo Final |"'),
    ('"| Classe | Transacoes | % | Volume (R$) | Acao |"',
     '"| Classe | Transações | % | Volume (R$) | Ação |"'),
    ('"| Entidade | Trans | Creditos (R$) | Debitos (R$) | Volume (R$) |"',
     '"| Entidade | Trans | Créditos (R$) | Débitos (R$) | Volume (R$) |"'),
    ('"| Mes | Data | Valor (R$) | Razao Social | Data Baixa | Dias Apos |"',
     '"| Mês | Data | Valor (R$) | Razão Social | Data Baixa | Dias Após |"'),
    ('"| # | CNPJ | Razao Social | CNAE | Anualizado | Excesso |"',
     '"| # | CNPJ | Razão Social | CNAE | Anualizado | Excesso |"'),
    # Campos da empresa
    ('("Razao Social",',      '("Razão Social",'),
    ('("Razao Anterior",',    '("Razão Anterior",'),
    ('("Natureza Juridica",', '("Natureza Jurídica",'),
    ('("CNAE Secundario",',   '("CNAE Secundário",'),
    ('("Escritorio Admin",',  '("Escritório Admin",'),
    ('("Ultima Alteracao",',  '("Última Alteração",'),
    ('("Endereco Sede",',     '("Endereço Sede",'),
    ('("Situacao",',          '("Situação",'),
    ('("Porte Declarado",',   '("Porte Declarado",'),
    ('("Data Abertura",',     '("Data Abertura",'),
    # Função no quadro societário
    ('"- **Funcao:** Administrador unico por prazo indeterminado"',
     '"- **Função:** Administrador único por prazo indeterminado"'),
    # MEI labels
    ('"- **MEI Padrao** (outros CNAEs): teto **R$ 81.000/ano** (LC 123/2006)"',
     '"- **MEI Padrão** (outros CNAEs): teto **R$ 81.000/ano** (LC 123/2006)"'),
    # Risk heatmap ações
    ('"Revisao prioritaria"',   '"Revisão prioritária"'),
    # Conclusão
    ('"Os testes deterministicos aplicados apontam os seguintes **pontos de atencao**, que "',
     '"Os testes determinísticos aplicados apontam os seguintes **pontos de atenção**, que "'),
    ('"demandam verificacao documental e, se confirmados, regularizacao:"',
     '"demandam verificação documental e, se confirmados, regularização:"'),
    ('"Recomenda-se a verificacao dos itens acima e, quando aplicavel, o acionamento das "',
     '"Recomenda-se a verificação dos itens acima e, quando aplicável, o acionamento das "'),
    ('"medidas formais cabiveis (ex.: denuncia espontanea — CTN art. 138)."',
     '"medidas formais cabíveis (ex.: denúncia espontânea — CTN art. 138)."'),
    ('"Nos testes deterministicos aplicados a este recorte **nao foram identificados achados "',
     '"Nos testes determinísticos aplicados a este recorte **não foram identificados achados "'),
    ('"materiais** (regime compativel com o teto, sem pagamentos pos-baixa, sem retencoes "',
     '"materiais** (regime compatível com o teto, sem pagamentos pós-baixa, sem retenções "'),
    ('"estimadas e sem MEIs acima do teto). Ver ressalvas de escopo e enriquecimento cadastral.")',
     '"estimadas e sem MEIs acima do teto). Ver ressalvas de escopo e enriquecimento cadastral."'),
    ('"*Documento gerado pelo OrgConc — Sistema OrgAudi. Indicadores deterministicos; "',
     '"*Documento gerado pelo OrgConc — Sistema OrgAudi. Indicadores determinísticos; "'),
    ('"NAO constituem conclusao de auditoria sem verificacao documental.*"',
     '"NÃO constituem conclusão de auditoria sem verificação documental.*"'),
    # MEI - reclassificação
    ('"## 6. MEIs Fornecedores - Reclassificacao MEI-TAC vs Padrao"',
     '"## 6. MEIs Fornecedores — Reclassificação MEI-TAC vs Padrão"'),
    ('"**Aplicacao do limite correto** apos confirmacao do cliente que muitos sao caminhoneiros:"',
     '"**Aplicação do limite correto** após confirmação do cliente que muitos são caminhoneiros:"'),
    # Conformidade
    ('"Sem fornecedores em faixa critica de conformidade no recorte."',
     '"Sem fornecedores em faixa crítica de conformidade no recorte."'),
    # Pós-baixa título
    ('"## 8. Pagamentos Pos-Baixa ({len(stats[\'pos_baixa\'])} alertas)"',
     '"## 8. Pagamentos Pós-Baixa ({len(stats[\'pos_baixa\'])} alertas)"'),
    # Retencao no achado (sem R$ — só texto)
    ("f\"**Retencoes estimadas na fonte** — R$ {format_brl(stats['total_ret_5m'])} em \"",
     "f\"**Retenções estimadas na fonte** — R$ {format_brl(stats['total_ret_5m'])} em \""),
    # Risco de pejotizacao
    ('"risco de pejotizacao"', '"risco de pejotização"'),
    # Achado pos-baixa
    ("f\"**{len(stats['pos_baixa'])} pagamentos pos-baixa** — R$ {format_brl(total_pb)} a CNPJ ja baixado\"",
     "f\"**{len(stats['pos_baixa'])} pagamentos pós-baixa** — R$ {format_brl(total_pb)} a CNPJ já baixado\""),
]

misses = []
for old, new in accent_pairs:
    if old in code:
        code = code.replace(old, new)
    else:
        misses.append(old[:70])

if misses:
    print(f"AVISO: {len(misses)} strings nao encontradas:")
    for m in misses:
        print(f"  - {m!r}")
else:
    print("OK: todas as strings de acento substituidas")

# ── 3) Cabeçalho RELATORIO → RELATÓRIO ──────────────────────────────────────
code = code.replace(
    "f\"# RELATORIO INTEGRADO DE AUDITORIA — {_emp()['razao_social']}\"",
    "f\"# RELATÓRIO INTEGRADO DE AUDITORIA — {_emp()['razao_social']}\"",
)

# ── 4) Gravar ────────────────────────────────────────────────────────────────
src.write_text(code, encoding="utf-8")
print("OK: laudo_forense.py gravado.")
