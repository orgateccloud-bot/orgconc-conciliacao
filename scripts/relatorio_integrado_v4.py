"""Relatorio Integrado v4 - 14 abas (11 originais + 3 novas fiscais).

NOVO em v4:
- Aba 12: Conformidade Fiscal Top 10 (fornecedores com gap)
- Aba 13: Documentos Fiscais Processados (8.226 XMLs)
- Aba 14: Riscos Fiscais Consolidados (R$ 5,12M/ano)

Mantem as 11 abas originais do v2/v3 com hyperlinks navegaveis.
"""
from __future__ import annotations

import re
import sys
import xml.etree.ElementTree as ET
import zipfile
from collections import defaultdict
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
sys.path.insert(0, str(Path(__file__).resolve().parent))
from _logo_helper import inserir_logo_xlsx
from api.matchers.cascata import ler_ofx

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT_XLSX = r"C:\Users\Veloso\Downloads\RELATORIO_INTEGRADO_LOCAR_v4.xlsx"

OFX_LIST = [
    r"C:\Users\Veloso\Downloads\extrato-conta-corrente-ofx-unix_202605_20260514110822.ofx",
    r"C:\Users\Veloso\Downloads\extrato-conta-corrente-ofx-unix_202605_20260514110841.ofx",
    r"C:\Users\Veloso\Downloads\extrato-conta-corrente-ofx-unix_202605_20260514110900.ofx",
    r"C:\Users\Veloso\Downloads\extrato-conta-corrente-ofx-unix_202605_20260514110917.ofx",
    r"C:\Users\Veloso\Downloads\extrato-conta-corrente-ofx-unix_202605_20260514110938.ofx",
]
ZIPS_NFE = [
    r"C:\Users\Veloso\Downloads\103585885_01012026_31012026_7514.zip",
    r"C:\Users\Veloso\Downloads\103585885_01022026_28022026_8464.zip",
    r"C:\Users\Veloso\Downloads\103585885_01032026_31032026_8594.zip",
    r"C:\Users\Veloso\Downloads\103585885_01042026_30042026_7245.zip",
]
ZIPS_CTE = [
    r"C:\Users\Veloso\Downloads\05509396000110_01012026_31012026_0546.zip",
    r"C:\Users\Veloso\Downloads\05509396000110_01022026_28022026_5384.zip",
    r"C:\Users\Veloso\Downloads\05509396000110_01032026_31032026_4046.zip",
    r"C:\Users\Veloso\Downloads\05509396000110_01042026_30042026_9825.zip",
]


# Estilos
CINZA_ESCURO = "0F172A"
AZUL_PRIMARIO = "0052FF"
AZUL_CLARO = "EFF6FF"
VERDE = "10B981"
LARANJA = "F59E0B"
VERMELHO = "DC2626"
BRANCO = "FFFFFF"

FONT_TITULO = Font(name="Arial", size=18, bold=True, color=BRANCO)
FONT_SUBTITULO = Font(name="Arial", size=11, color=BRANCO, italic=True)
FONT_HEADER = Font(name="Arial", size=10, bold=True, color=BRANCO)
FONT_CORPO = Font(name="Arial", size=10)
FONT_NEGRITO = Font(name="Arial", size=10, bold=True)
FONT_LINK = Font(name="Arial", size=10, color=AZUL_PRIMARIO, underline="single")

FILL_CABECALHO = PatternFill("solid", fgColor=CINZA_ESCURO)
FILL_HEADER = PatternFill("solid", fgColor=AZUL_PRIMARIO)
FILL_ALT = PatternFill("solid", fgColor="F8FAFC")
FILL_DESTAQUE = PatternFill("solid", fgColor=AZUL_CLARO)
FILL_VERDE = PatternFill("solid", fgColor="DCFCE7")
FILL_LARANJA = PatternFill("solid", fgColor="FED7AA")
FILL_VERMELHO = PatternFill("solid", fgColor="FECACA")

BORDA_FINA = Border(
    left=Side(border_style="thin", color="E2E8F0"),
    right=Side(border_style="thin", color="E2E8F0"),
    top=Side(border_style="thin", color="E2E8F0"),
    bottom=Side(border_style="thin", color="E2E8F0"),
)


def _local(tag):
    return tag.split("}")[-1]


def _filho(elem, nome):
    if elem is None:
        return None
    for f in elem:
        if _local(f.tag) == nome:
            return f
    return None


def _texto(elem, *caminho):
    cur = elem
    for nome in caminho:
        cur = _filho(cur, nome)
        if cur is None:
            return ""
    return cur.text.strip() if cur is not None and cur.text else ""


def parse_nfe(conteudo):
    try:
        root = ET.fromstring(conteudo)
    except ET.ParseError:
        return None
    for el in root.iter():
        if _local(el.tag) == "infNFe":
            inf = el
            break
    else:
        return None
    ide = _filho(inf, "ide")
    emit = _filho(inf, "emit")
    total = _filho(inf, "total")
    icms_tot = _filho(total, "ICMSTot") if total is not None else None
    return {
        "chave": (inf.get("Id") or "").lstrip("NFe"),
        "data": (_texto(ide, "dhEmi") or _texto(ide, "dEmi"))[:10],
        "emit_cnpj": _texto(emit, "CNPJ"),
        "emit_nome": _texto(emit, "xNome"),
        "valor": float(_texto(icms_tot, "vNF") or 0) if icms_tot is not None else 0.0,
        "modelo": _texto(ide, "mod") or "55",
        "uf": _texto(emit, "enderEmit", "UF"),
    }


def parse_cte(conteudo):
    try:
        root = ET.fromstring(conteudo)
    except ET.ParseError:
        return None
    for el in root.iter():
        if _local(el.tag) in ("infCte", "infCTe"):
            inf = el
            break
    else:
        return None
    ide = _filho(inf, "ide")
    rem = _filho(inf, "rem")
    dest = _filho(inf, "dest")
    vprest = _filho(inf, "vPrest")
    return {
        "chave": (inf.get("Id") or "").lstrip("CTe"),
        "data": (_texto(ide, "dhEmi") or _texto(ide, "dEmi"))[:10],
        "rem_cnpj": _texto(rem, "CNPJ") if rem else "",
        "rem_nome": _texto(rem, "xNome") if rem else "",
        "dest_cnpj": _texto(dest, "CNPJ") if dest else "",
        "dest_nome": _texto(dest, "xNome") if dest else "",
        "valor": float(_texto(vprest, "vTPrest") or 0) if vprest else 0.0,
        "modelo": "57",
        "uf": _texto(ide, "UFIni"),
    }


def carregar_dados():
    print("Carregando OFX...")
    transacoes = []
    for p in OFX_LIST:
        transacoes.extend(ler_ofx(p))
    print(f"  {len(transacoes):,} transacoes OFX")

    print("Parseando NF-es...")
    nfes = []
    for path in ZIPS_NFE:
        with zipfile.ZipFile(path) as zf:
            for m in zf.namelist():
                if m.endswith(".xml"):
                    chave = Path(m).stem
                    if len(chave) >= 22 and chave[20:22] == "55":
                        with zf.open(m) as f:
                            doc = parse_nfe(f.read())
                        if doc:
                            nfes.append(doc)
    print(f"  {len(nfes):,} NF-es")

    print("Parseando CT-es...")
    ctes = []
    for path in ZIPS_CTE:
        with zipfile.ZipFile(path) as zf:
            for m in zf.namelist():
                if m.endswith(".xml"):
                    chave = Path(m).stem
                    if len(chave) >= 22 and chave[20:22] == "57":
                        with zf.open(m) as f:
                            doc = parse_cte(f.read())
                        if doc:
                            ctes.append(doc)
    print(f"  {len(ctes):,} CT-es")

    return transacoes, nfes, ctes


def _set_header(ws, titulo, subtitulo):
    ws.row_dimensions[1].height = 50
    ws.row_dimensions[2].height = 24
    ws["A1"] = ""
    ws["B1"] = titulo
    ws["B1"].font = FONT_TITULO
    ws["B1"].fill = FILL_CABECALHO
    ws["B1"].alignment = Alignment(vertical="center", indent=1)
    ws["B2"] = subtitulo
    ws["B2"].font = FONT_SUBTITULO
    ws["B2"].fill = FILL_CABECALHO
    ws["B2"].alignment = Alignment(vertical="center", indent=1)
    # Merge col B to end
    ws.merge_cells("B1:N1")
    ws.merge_cells("B2:N2")
    # Logo na A1
    inserir_logo_xlsx(ws, anchor="A1", largura_px=60, altura_px=60)
    ws["A1"].fill = FILL_CABECALHO
    ws["A2"].fill = FILL_CABECALHO


def _aplicar_header(ws, headers, row_start=4):
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row_start, column=col_idx, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDA_FINA


def _aplicar_zebra(ws, row_start, row_end, col_start, col_end):
    for r in range(row_start, row_end + 1):
        for c in range(col_start, col_end + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDA_FINA
            if r % 2 == 0:
                cell.fill = FILL_ALT


def aba_indice(wb, abas_meta):
    ws = wb.create_sheet("Indice", 0)
    _set_header(ws, "INDICE NAVEGAVEL", f"Relatorio Integrado v4 - LOCAR - Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    _aplicar_header(ws, ["#", "Aba", "Descricao", "Destaque"])
    for i, (nome, desc, destaque) in enumerate(abas_meta, 1):
        r = 4 + i
        ws.cell(row=r, column=1, value=i).alignment = Alignment(horizontal="center")
        link_cell = ws.cell(row=r, column=2, value=nome)
        link_cell.hyperlink = f"#'{nome}'!A1"
        link_cell.font = FONT_LINK
        ws.cell(row=r, column=3, value=desc).font = FONT_CORPO
        d = ws.cell(row=r, column=4, value=destaque)
        d.font = FONT_NEGRITO
        d.alignment = Alignment(horizontal="right")
        if "Critico" in destaque or "R$ 5" in destaque or "R$ 3" in destaque:
            d.fill = FILL_VERMELHO
        elif "R$" in destaque:
            d.fill = FILL_LARANJA
        elif "OK" in destaque or "CONFORME" in destaque:
            d.fill = FILL_VERDE
    _aplicar_zebra(ws, 5, 4 + len(abas_meta), 1, 4)
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 65
    ws.column_dimensions["D"].width = 24


def aba_sumario(wb):
    ws = wb.create_sheet("01_Sumario_Executivo")
    _set_header(ws, "SUMARIO EXECUTIVO", "LOCAR Transporte de Bovinos LTDA - 8 Constatacoes - v3")
    headers = ["#", "Achado", "Materialidade", "Risco Anualizado (R$)"]
    _aplicar_header(ws, headers)
    achados = [
        ("I", "Confirmacao regime Lucro Real", "Informativo", 0),
        ("II", "Historico exclusao administrativa Simples (2015-2018)", "Alto", 0),
        ("III", "Subcapitalizacao (capital R$ 400k vs R$ 187M giro)", "Critico", 0),
        ("IV", "Partes relacionadas sem lastro contratual", "Alto", 568_000),
        ("V", "5 MEI padrao com pequenos excessos", "Baixo", 12_000),
        ("VI", "Retencoes nao recolhidas (5 meses anualizado)", "Critico", 1_173_000),
        ("VII", "Pagamentos pos-baixa CNPJ (R$ 35,6k)", "Critico", 12_000),
        ("VIII", "Gap fiscal: REDE FROTA (R$ 3M sem NF) + 5 MEIs sem CT-e", "Critico", 3_359_000),
    ]
    for i, (num, descr, mat, risco) in enumerate(achados):
        r = 5 + i
        ws.cell(row=r, column=1, value=num).font = FONT_NEGRITO
        ws.cell(row=r, column=2, value=descr).font = FONT_CORPO
        cell_mat = ws.cell(row=r, column=3, value=mat)
        if mat == "Critico":
            cell_mat.fill = FILL_VERMELHO
            cell_mat.font = FONT_NEGRITO
        elif mat == "Alto":
            cell_mat.fill = FILL_LARANJA
        cell_risco = ws.cell(row=r, column=4, value=risco)
        cell_risco.number_format = '"R$ "#,##0.00;[Red]-"R$ "#,##0.00'
        cell_risco.alignment = Alignment(horizontal="right")
        if risco >= 1_000_000:
            cell_risco.fill = FILL_VERMELHO
            cell_risco.font = FONT_NEGRITO
    # Total
    total_r = 5 + len(achados)
    ws.cell(row=total_r, column=2, value="TOTAL ANUALIZADO").font = FONT_NEGRITO
    total_cell = ws.cell(row=total_r, column=4, value=sum(a[3] for a in achados))
    total_cell.number_format = '"R$ "#,##0.00'
    total_cell.font = FONT_NEGRITO
    total_cell.fill = FILL_CABECALHO
    total_cell.font = Font(name="Arial", size=10, bold=True, color=BRANCO)
    _aplicar_zebra(ws, 5, total_r, 1, 4)
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 24


def aba_conformidade_fiscal(wb, nfes, ctes, transacoes):
    """Aba 12: Conformidade Fiscal - Top fornecedores com gap"""
    ws = wb.create_sheet("12_Conformidade_Fiscal")
    _set_header(ws, "CONFORMIDADE FISCAL - TOP FORNECEDORES", "Cruzamento OFX x NF-e x CT-e - Score de Conformidade por Fornecedor")
    headers = ["#", "Fornecedor", "CNPJ", "Vol. Pago OFX", "Vol. NF-e Recebida", "Vol. CT-e", "Conformidade %", "Classe Risco"]
    _aplicar_header(ws, headers)

    # Agrupar pagamentos OFX por nome
    pag_por_nome = defaultdict(lambda: {"vol": 0.0, "n": 0})
    for t in transacoes:
        if t.valor < 0:
            nome = (t.nome or "").strip().upper()[:60]
            if nome:
                pag_por_nome[nome]["vol"] += abs(t.valor)
                pag_por_nome[nome]["n"] += 1

    # Agrupar NF-es por emit_nome
    nfe_por_nome = defaultdict(lambda: {"vol": 0.0, "n": 0, "cnpj": ""})
    for n in nfes:
        nome = (n["emit_nome"] or "").strip().upper()[:60]
        if nome:
            nfe_por_nome[nome]["vol"] += n["valor"]
            nfe_por_nome[nome]["n"] += 1
            nfe_por_nome[nome]["cnpj"] = n["emit_cnpj"]

    # CT-es destinatario
    cte_por_nome = defaultdict(lambda: {"vol": 0.0})
    for c in ctes:
        nome = (c["dest_nome"] or "").strip().upper()[:60]
        if nome:
            cte_por_nome[nome]["vol"] += c["valor"]

    # Top 30 fornecedores por volume pago
    top_pagamentos = sorted(pag_por_nome.items(), key=lambda x: -x[1]["vol"])[:30]

    rows_data = []
    for nome, info in top_pagamentos:
        vol_pago = info["vol"]
        # Busca fuzzy de NF-e pelo nome
        nfe_match = None
        for k_nfe in nfe_por_nome:
            if k_nfe[:30] == nome[:30] or nome[:20] in k_nfe or k_nfe[:20] in nome:
                nfe_match = nfe_por_nome[k_nfe]
                break
        vol_nfe = nfe_match["vol"] if nfe_match else 0.0
        cnpj = nfe_match["cnpj"] if nfe_match else "?"
        vol_cte = cte_por_nome.get(nome, {}).get("vol", 0.0)
        vol_doc = vol_nfe + vol_cte
        conformidade = (vol_doc / vol_pago * 100) if vol_pago else 0
        if conformidade >= 80:
            classe = "CONFORME"
        elif conformidade >= 50:
            classe = "MEDIO"
        elif conformidade >= 20:
            classe = "ALTO"
        else:
            classe = "CRITICO"
        rows_data.append((nome, cnpj, vol_pago, vol_nfe, vol_cte, conformidade, classe))

    for i, (nome, cnpj, vp, vn, vc, conf, classe) in enumerate(rows_data):
        r = 5 + i
        ws.cell(row=r, column=1, value=i + 1)
        ws.cell(row=r, column=2, value=nome[:50]).font = FONT_CORPO
        ws.cell(row=r, column=3, value=cnpj).font = FONT_CORPO
        c = ws.cell(row=r, column=4, value=vp)
        c.number_format = '"R$ "#,##0.00'
        c = ws.cell(row=r, column=5, value=vn)
        c.number_format = '"R$ "#,##0.00'
        c = ws.cell(row=r, column=6, value=vc)
        c.number_format = '"R$ "#,##0.00'
        c = ws.cell(row=r, column=7, value=conf / 100)
        c.number_format = "0.00%"
        c_cls = ws.cell(row=r, column=8, value=classe)
        if classe == "CRITICO":
            c_cls.fill = FILL_VERMELHO
            c_cls.font = FONT_NEGRITO
        elif classe == "ALTO":
            c_cls.fill = FILL_LARANJA
            c_cls.font = FONT_NEGRITO
        elif classe == "MEDIO":
            c_cls.fill = FILL_DESTAQUE
        else:
            c_cls.fill = FILL_VERDE

    _aplicar_zebra(ws, 5, 4 + len(rows_data), 1, 8)
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 48
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 14
    ws.freeze_panes = "A5"


def aba_documentos_fiscais(wb, nfes, ctes):
    """Aba 13: Documentos Fiscais Processados"""
    ws = wb.create_sheet("13_Documentos_Fiscais")
    _set_header(ws, "DOCUMENTOS FISCAIS PROCESSADOS", f"{len(nfes):,} NF-es + {len(ctes):,} CT-es = {len(nfes)+len(ctes):,} XMLs - jan-abr/2026")

    # Resumo por modelo
    headers = ["Modelo", "Tipo", "Qtde Documentos", "Valor Total (R$)", "Periodo", "Fontes ZIP"]
    _aplicar_header(ws, headers)

    nfe_vol = sum(n["valor"] for n in nfes)
    cte_vol = sum(c["valor"] for c in ctes)
    nfe_min = min((n["data"] for n in nfes if n["data"]), default="-")
    nfe_max = max((n["data"] for n in nfes if n["data"]), default="-")
    cte_min = min((c["data"] for c in ctes if c["data"]), default="-")
    cte_max = max((c["data"] for c in ctes if c["data"]), default="-")

    dados = [
        ("55", "NF-e (Nota Fiscal Eletronica - Compras)", len(nfes), nfe_vol, f"{nfe_min} a {nfe_max}", "4 ZIPs"),
        ("57", "CT-e (Conhecimento Transporte - Emitidos)", len(ctes), cte_vol, f"{cte_min} a {cte_max}", "4 ZIPs"),
    ]
    for i, (mod, tipo, qtd, val, per, fonte) in enumerate(dados):
        r = 5 + i
        ws.cell(row=r, column=1, value=mod).font = FONT_NEGRITO
        ws.cell(row=r, column=2, value=tipo)
        ws.cell(row=r, column=3, value=qtd).number_format = "#,##0"
        c = ws.cell(row=r, column=4, value=val)
        c.number_format = '"R$ "#,##0.00'
        ws.cell(row=r, column=5, value=per)
        ws.cell(row=r, column=6, value=fonte)
    r_total = 5 + len(dados)
    ws.cell(row=r_total, column=2, value="TOTAL").font = FONT_NEGRITO
    ws.cell(row=r_total, column=3, value=len(nfes) + len(ctes)).font = FONT_NEGRITO
    ws.cell(row=r_total, column=3).number_format = "#,##0"
    c = ws.cell(row=r_total, column=4, value=nfe_vol + cte_vol)
    c.number_format = '"R$ "#,##0.00'
    c.font = FONT_NEGRITO
    c.fill = FILL_DESTAQUE
    _aplicar_zebra(ws, 5, r_total, 1, 6)

    # Espacamento
    r_sec2 = r_total + 3

    # Top 10 emissores de NF-e (compras)
    ws.cell(row=r_sec2, column=1, value="TOP 10 EMISSORES DE NF-E (Fornecedores)").font = Font(name="Arial", size=12, bold=True, color=AZUL_PRIMARIO)
    headers2 = ["#", "Emitente", "CNPJ", "UF", "Qtd NF-es", "Valor Total"]
    for col_idx, h in enumerate(headers2, 1):
        c = ws.cell(row=r_sec2 + 1, column=col_idx, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = Alignment(horizontal="center")
        c.border = BORDA_FINA
    nfe_por_emit = defaultdict(lambda: {"qtd": 0, "vol": 0.0, "cnpj": "", "uf": ""})
    for n in nfes:
        k = n["emit_nome"][:50]
        nfe_por_emit[k]["qtd"] += 1
        nfe_por_emit[k]["vol"] += n["valor"]
        nfe_por_emit[k]["cnpj"] = n["emit_cnpj"]
        nfe_por_emit[k]["uf"] = n["uf"]
    top_emit = sorted(nfe_por_emit.items(), key=lambda x: -x[1]["vol"])[:10]
    for i, (nome, info) in enumerate(top_emit):
        r = r_sec2 + 2 + i
        ws.cell(row=r, column=1, value=i + 1)
        ws.cell(row=r, column=2, value=nome)
        ws.cell(row=r, column=3, value=info["cnpj"])
        ws.cell(row=r, column=4, value=info["uf"])
        ws.cell(row=r, column=5, value=info["qtd"]).number_format = "#,##0"
        c = ws.cell(row=r, column=6, value=info["vol"])
        c.number_format = '"R$ "#,##0.00'
    _aplicar_zebra(ws, r_sec2 + 2, r_sec2 + 1 + len(top_emit), 1, 6)

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 48
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 20


def aba_riscos_fiscais(wb):
    """Aba 14: Riscos Fiscais Consolidados"""
    ws = wb.create_sheet("14_Riscos_Fiscais")
    _set_header(ws, "RISCOS FISCAIS CONSOLIDADOS", "Estimativa anualizada de passivo tributario - Lucro Real - R$ 5,12M/ano")

    headers = ["Categoria", "Fundamento Legal", "Base Calculo (R$)", "Aliquota", "Risco Anual (R$)", "Prioridade"]
    _aplicar_header(ws, headers)
    riscos = [
        ("Retencoes nao recolhidas (5 meses x 12/5)", "IN RFB 1.234/2012", 11_167_000, "10,5% media", 1_173_000, "30 dias"),
        ("Distribuicao disfarcada lucros (DDL)", "RIR/2018 art. 464", 2_065_000, "27,5%", 568_000, "90 dias"),
        ("MEI padrao com excesso", "LC 123/2006", 35_000, "34%", 12_000, "60 dias"),
        ("Pagamentos pos-baixa CNPJ", "RIR/2018 art. 311", 35_627, "34%", 12_000, "30 dias"),
        ("REDE FROTA sem NF-e (anualizado)", "RIR/2018 art. 311", 8_840_000, "34% (IRPJ+CSLL)", 3_005_600, "30 dias"),
        ("MEIs caminhoneiros sem CT-e", "Decreto 8.324/2014", 1_053_000, "34% + 5% ICMS-ST", 358_020, "30 dias"),
    ]
    for i, (cat, fund, base, aliq, risco, prazo) in enumerate(riscos):
        r = 5 + i
        ws.cell(row=r, column=1, value=cat).font = FONT_CORPO
        ws.cell(row=r, column=2, value=fund)
        c = ws.cell(row=r, column=3, value=base)
        c.number_format = '"R$ "#,##0.00'
        ws.cell(row=r, column=4, value=aliq)
        c = ws.cell(row=r, column=5, value=risco)
        c.number_format = '"R$ "#,##0.00'
        c.font = FONT_NEGRITO
        if risco >= 1_000_000:
            c.fill = FILL_VERMELHO
        elif risco >= 300_000:
            c.fill = FILL_LARANJA
        prazo_c = ws.cell(row=r, column=6, value=prazo)
        if prazo == "30 dias":
            prazo_c.fill = FILL_VERMELHO
            prazo_c.font = FONT_NEGRITO

    # Total
    r_total = 5 + len(riscos)
    ws.cell(row=r_total, column=1, value="TOTAL ANUALIZADO").font = FONT_NEGRITO
    c = ws.cell(row=r_total, column=5, value=sum(x[4] for x in riscos))
    c.number_format = '"R$ "#,##0.00'
    c.fill = FILL_CABECALHO
    c.font = Font(name="Arial", size=11, bold=True, color=BRANCO)

    _aplicar_zebra(ws, 5, r_total, 1, 6)

    # Espacamento
    r_sec2 = r_total + 3
    ws.cell(row=r_sec2, column=1, value="MITIGACOES PROPOSTAS").font = Font(name="Arial", size=12, bold=True, color=AZUL_PRIMARIO)
    headers_mit = ["#", "Acao", "Categoria Reduzida", "Reducao Estimada (R$)"]
    for col_idx, h in enumerate(headers_mit, 1):
        c = ws.cell(row=r_sec2 + 1, column=col_idx, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = Alignment(horizontal="center")
        c.border = BORDA_FINA
    mitigacoes = [
        (1, "Obter NF-es REDE FROTA via SEFAZ Distribuicao DFe", "REDE FROTA", 3_005_600),
        (2, "Denuncia espontanea CTN 138 (afasta multa oficio)", "Retencoes", 879_750),
        (3, "Exigir CT-e dos MEIs caminhoneiros", "MEIs sem CT-e", 358_020),
        (4, "Documentar lastro contratual partes relacionadas", "DDL", 426_000),
        (5, "Substituir MEIs com excesso por PJ", "MEI excesso", 12_000),
    ]
    for i, (n, acao, cat, red) in enumerate(mitigacoes):
        r = r_sec2 + 2 + i
        ws.cell(row=r, column=1, value=n)
        ws.cell(row=r, column=2, value=acao)
        ws.cell(row=r, column=3, value=cat)
        c = ws.cell(row=r, column=4, value=red)
        c.number_format = '"R$ "#,##0.00'
        c.font = FONT_NEGRITO
        c.fill = FILL_VERDE
    r_tot_mit = r_sec2 + 2 + len(mitigacoes)
    ws.cell(row=r_tot_mit, column=2, value="POTENCIAL DE MITIGACAO").font = FONT_NEGRITO
    c = ws.cell(row=r_tot_mit, column=4, value=sum(m[3] for m in mitigacoes))
    c.number_format = '"R$ "#,##0.00'
    c.font = Font(name="Arial", size=11, bold=True, color=BRANCO)
    c.fill = FILL_CABECALHO
    _aplicar_zebra(ws, r_sec2 + 2, r_tot_mit, 1, 4)

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 14


def aba_resumo_geral(wb, transacoes, nfes, ctes):
    ws = wb.create_sheet("02_Resumo_Geral")
    _set_header(ws, "RESUMO GERAL DO PERIODO", "5 meses (jan-mai/2026) - SICOOB 158083-3")
    headers = ["Indicador", "Valor"]
    _aplicar_header(ws, headers)
    saidas = [t for t in transacoes if t.valor < 0]
    entradas = [t for t in transacoes if t.valor > 0]
    dados = [
        ("Periodo OFX", "01/01/2026 a 14/05/2026"),
        ("Total transacoes OFX", f"{len(transacoes):,}"),
        ("Transacoes de saida (pagamentos)", f"{len(saidas):,}"),
        ("Transacoes de entrada (recebimentos)", f"{len(entradas):,}"),
        ("Volume saidas (R$)", f"R$ {sum(abs(t.valor) for t in saidas):,.2f}"),
        ("Volume entradas (R$)", f"R$ {sum(t.valor for t in entradas):,.2f}"),
        ("Movimentacao bruta total (R$)", f"R$ {sum(abs(t.valor) for t in transacoes):,.2f}"),
        ("NF-es processadas (compras)", f"{len(nfes):,}"),
        ("Volume NF-es (R$)", f"R$ {sum(n['valor'] for n in nfes):,.2f}"),
        ("CT-es emitidos (transporte)", f"{len(ctes):,}"),
        ("Volume CT-es (R$)", f"R$ {sum(c['valor'] for c in ctes):,.2f}"),
        ("Documentos fiscais total", f"{len(nfes)+len(ctes):,}"),
    ]
    for i, (k, v) in enumerate(dados):
        r = 5 + i
        ws.cell(row=r, column=1, value=k).font = FONT_NEGRITO
        ws.cell(row=r, column=2, value=v)
    _aplicar_zebra(ws, 5, 4 + len(dados), 1, 2)
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 36


def aba_partes_relacionadas(wb):
    ws = wb.create_sheet("03_Partes_Relacionadas")
    _set_header(ws, "PARTES RELACIONADAS", "Movimentacoes intercompany identificadas")
    headers = ["Parte", "Transacoes", "Volume (R$)", "Natureza", "Risco"]
    _aplicar_header(ws, headers)
    dados = [
        ("LOCAR LOCADORA (CNPJ a confirmar)", 73, 6_733_631.85, "PIX MESMA TIT", "Alto"),
        ("LOCAR MAQUINAS E SERVICOS", 13, 249_947.18, "Pagamentos/Recebimentos", "Medio"),
        ("RENATO COSTA ESPERIDIAO JR (PF)", 201, 8_253_024.12, "Pro-labore/Dividendos/Mutuo", "Critico"),
    ]
    for i, (p, n, v, natureza, risco) in enumerate(dados):
        r = 5 + i
        ws.cell(row=r, column=1, value=p)
        ws.cell(row=r, column=2, value=n).alignment = Alignment(horizontal="center")
        c = ws.cell(row=r, column=3, value=v)
        c.number_format = '"R$ "#,##0.00'
        ws.cell(row=r, column=4, value=natureza)
        risk_c = ws.cell(row=r, column=5, value=risco)
        if risco == "Critico":
            risk_c.fill = FILL_VERMELHO
        elif risco == "Alto":
            risk_c.fill = FILL_LARANJA
    _aplicar_zebra(ws, 5, 4 + len(dados), 1, 5)
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 32
    ws.column_dimensions["E"].width = 12


def aba_meis(wb):
    ws = wb.create_sheet("04_MEIs_Reclassificados")
    _set_header(ws, "MEIs RECLASSIFICADOS", "MEI-TAC (R$ 251.600) vs MEI Padrao (R$ 81.000)")
    headers = ["Categoria", "Teto Anual", "Total", "Dentro", "Acima", "Excesso Acumulado"]
    _aplicar_header(ws, headers)
    dados = [
        ("MEI-TAC (Caminhoneiros)", 251_600, 65, 65, 0, 0),
        ("MEI Padrao (demais CNAEs)", 81_000, 260, 255, 5, 1_927_414),
    ]
    for i, row in enumerate(dados):
        r = 5 + i
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c_idx, value=val)
            if c_idx in (2, 6) and isinstance(val, (int, float)):
                cell.number_format = '"R$ "#,##0.00'
    _aplicar_zebra(ws, 5, 4 + len(dados), 1, 6)
    for col in "ABCDEF":
        ws.column_dimensions[col].width = 18


def aba_retencoes(wb):
    ws = wb.create_sheet("05_Retencoes")
    _set_header(ws, "RETENCOES NA FONTE NAO RECOLHIDAS", "5 meses anualizado")
    headers = ["Categoria", "Tributos", "Aliquota", "5 meses (R$)", "Anualizado (R$)"]
    _aplicar_header(ws, headers)
    dados = [
        ("Pagamentos a PJ (servicos)", "PIS+COFINS+CSLL+IRRF", "6,15%", 456_552.83, 1_095_727),
        ("Pagamentos a PF (autonomos)", "IRRF+INSS", "ate 27,5%", 32_164.40, 77_195),
    ]
    for i, (cat, trib, aliq, v5, va) in enumerate(dados):
        r = 5 + i
        ws.cell(row=r, column=1, value=cat)
        ws.cell(row=r, column=2, value=trib)
        ws.cell(row=r, column=3, value=aliq)
        c1 = ws.cell(row=r, column=4, value=v5)
        c1.number_format = '"R$ "#,##0.00'
        c2 = ws.cell(row=r, column=5, value=va)
        c2.number_format = '"R$ "#,##0.00'
    r_tot = 5 + len(dados)
    ws.cell(row=r_tot, column=1, value="TOTAL").font = FONT_NEGRITO
    c = ws.cell(row=r_tot, column=4, value=sum(x[3] for x in dados))
    c.number_format = '"R$ "#,##0.00'
    c.font = FONT_NEGRITO
    c2 = ws.cell(row=r_tot, column=5, value=sum(x[4] for x in dados))
    c2.number_format = '"R$ "#,##0.00'
    c2.font = FONT_NEGRITO
    c2.fill = FILL_VERMELHO
    _aplicar_zebra(ws, 5, r_tot, 1, 5)
    for col in "ABCDE":
        ws.column_dimensions[col].width = 22


def aba_baixados(wb):
    ws = wb.create_sheet("06_CNPJ_Baixados")
    _set_header(ws, "PAGAMENTOS A CNPJ BAIXADOS", "17 transacoes - R$ 35.626,89")
    headers = ["Fornecedor", "CNPJ", "Status RFB", "Pagamentos", "Volume", "Defasagem"]
    _aplicar_header(ws, headers)
    dados = [
        ("PERCIVAL DIAS DA SILVA", "63.567.345/0001-41", "BAIXADO 11/03/2026", 17, 35_626.89, "63 dias"),
    ]
    for i, row in enumerate(dados):
        r = 5 + i
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c_idx, value=val)
            if c_idx == 5:
                cell.number_format = '"R$ "#,##0.00'
                cell.fill = FILL_VERMELHO
    _aplicar_zebra(ws, 5, 4 + len(dados), 1, 6)
    for col in "ABCDEF":
        ws.column_dimensions[col].width = 22


def aba_top_pagamentos(wb, transacoes):
    ws = wb.create_sheet("07_Top_Pagamentos")
    _set_header(ws, "TOP 50 PAGAMENTOS", "Maiores saidas do periodo")
    headers = ["#", "Data", "Valor (R$)", "Beneficiario", "Memo"]
    _aplicar_header(ws, headers)
    saidas = sorted([t for t in transacoes if t.valor < 0], key=lambda x: x.valor)[:50]
    for i, t in enumerate(saidas):
        r = 5 + i
        ws.cell(row=r, column=1, value=i + 1)
        ws.cell(row=r, column=2, value=str(t.data))
        c = ws.cell(row=r, column=3, value=abs(t.valor))
        c.number_format = '"R$ "#,##0.00'
        ws.cell(row=r, column=4, value=(t.nome or "")[:50])
        ws.cell(row=r, column=5, value=(t.memo or "")[:80])
    _aplicar_zebra(ws, 5, 4 + len(saidas), 1, 5)
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 38
    ws.column_dimensions["E"].width = 60
    ws.freeze_panes = "A5"


def aba_top_recebimentos(wb, transacoes):
    ws = wb.create_sheet("08_Top_Recebimentos")
    _set_header(ws, "TOP 50 RECEBIMENTOS", "Maiores entradas do periodo")
    headers = ["#", "Data", "Valor (R$)", "Remetente", "Memo"]
    _aplicar_header(ws, headers)
    entradas = sorted([t for t in transacoes if t.valor > 0], key=lambda x: -x.valor)[:50]
    for i, t in enumerate(entradas):
        r = 5 + i
        ws.cell(row=r, column=1, value=i + 1)
        ws.cell(row=r, column=2, value=str(t.data))
        c = ws.cell(row=r, column=3, value=t.valor)
        c.number_format = '"R$ "#,##0.00'
        ws.cell(row=r, column=4, value=(t.nome or "")[:50])
        ws.cell(row=r, column=5, value=(t.memo or "")[:80])
    _aplicar_zebra(ws, 5, 4 + len(entradas), 1, 5)
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 38
    ws.column_dimensions["E"].width = 60
    ws.freeze_panes = "A5"


def aba_alvos_investigacao(wb):
    ws = wb.create_sheet("09_Alvos_Investigacao")
    _set_header(ws, "ALVOS DE INVESTIGACAO FORENSE", "3 alvos especificos - acoes recomendadas")
    headers = ["Alvo", "Tipo", "Volume", "Status", "Acao Recomendada"]
    _aplicar_header(ws, headers)
    dados = [
        ("THIAGO MARQUES DE AVILA", "PF", 1_273_894.13, "A INVESTIGAR", "Solicitar contratos/recibos"),
        ("GT PARTICIPACOES LTDA", "PJ", 1_625_236.75, "A IDENTIFICAR CNPJ", "Verificar quadro societario"),
        ("REDE FROTA SOLUTIONS LTDA", "PJ", 3_025_000.00, "NAO CONFORME", "Solicitar NF-es jan-mai/2026"),
    ]
    for i, (alvo, tp, vol, status, acao) in enumerate(dados):
        r = 5 + i
        ws.cell(row=r, column=1, value=alvo).font = FONT_NEGRITO
        ws.cell(row=r, column=2, value=tp)
        c = ws.cell(row=r, column=3, value=vol)
        c.number_format = '"R$ "#,##0.00'
        st = ws.cell(row=r, column=4, value=status)
        if "NAO CONFORME" in status:
            st.fill = FILL_VERMELHO
            st.font = FONT_NEGRITO
        else:
            st.fill = FILL_LARANJA
        ws.cell(row=r, column=5, value=acao)
    _aplicar_zebra(ws, 5, 4 + len(dados), 1, 5)
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 38


def aba_obrigacoes(wb):
    ws = wb.create_sheet("10_Obrigacoes_Acessorias")
    _set_header(ws, "OBRIGACOES ACESSORIAS - LUCRO REAL", "Lembretes mensais e anuais")
    headers = ["Obrigacao", "Periodicidade", "Fundamento", "Status"]
    _aplicar_header(ws, headers)
    dados = [
        ("LALUR Digital (e-LALUR)", "Anual", "IN RFB 1.422/2013", "A confirmar"),
        ("SPED-ECD (Escrituracao Contabil Digital)", "Anual", "IN RFB 1.420/2013", "A confirmar"),
        ("SPED-ECF (Escrituracao Contabil Fiscal)", "Anual", "IN RFB 1.422/2013", "A confirmar"),
        ("DCTF (Declaracao de Debitos)", "Mensal", "IN RFB 2.005/2021", "A confirmar"),
        ("EFD-Contribuicoes", "Mensal", "IN RFB 1.252/2012", "A confirmar"),
        ("DIRF (Imposto Retido na Fonte)", "Anual", "IN RFB 2.005/2021", "A confirmar"),
        ("IRPJ Trimestral", "Trimestral", "Lei 9.430/96", "A confirmar"),
    ]
    for i, (obr, per, fund, st) in enumerate(dados):
        r = 5 + i
        ws.cell(row=r, column=1, value=obr)
        ws.cell(row=r, column=2, value=per).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=3, value=fund)
        ws.cell(row=r, column=4, value=st)
    _aplicar_zebra(ws, 5, 4 + len(dados), 1, 4)
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 14


def aba_recomendacoes(wb):
    ws = wb.create_sheet("11_Recomendacoes")
    _set_header(ws, "RECOMENDACOES FORMAIS CONSOLIDADAS", "11 acoes ordenadas por prazo - v3")
    headers = ["#", "Acao", "Prazo", "Risco se nao executar"]
    _aplicar_header(ws, headers)
    dados = [
        (1, "Apurar e recolher retencoes (denuncia espontanea CTN 138)", "30 dias", "Multa 75-150% + juros"),
        (2, "Solicitar NF-es REDE FROTA + verificar SEFAZ Distribuicao DFe", "30 dias", "R$ 3M adicao LALUR"),
        (3, "Notificar 5 MEIs sobre CT-e obrigatorio", "30 dias", "R$ 358k LALUR + ICMS-ST"),
        (4, "Investigar pagamentos pos-baixa CNPJ", "30 dias", "Glosa LALUR + penal"),
        (5, "Implantar controle de retencoes na fonte", "30 dias", "Recorrencia"),
        (6, "Solicitar processo administrativo exclusao 2015-2018", "30 dias", "Antecedente fiscal"),
        (7, "Suspender pagamentos sem NF-e/CT-e", "45 dias", "Risco continuo"),
        (8, "Notificar MEIs padrao desenquadrados", "60 dias", "Responsabilidade solidaria"),
        (9, "Documentar lastro contratual partes relacionadas", "90 dias", "Adicoes obrigatorias LALUR"),
        (10, "Aumento de capital social", "120 dias", "Desconsideracao PJ"),
        (11, "Conferir obrigacoes acessorias (DCTF/EFD/SPED)", "30 dias", "Multas por atraso"),
    ]
    for n, acao, prazo, risco in dados:
        r = 4 + n
        ws.cell(row=r, column=1, value=n).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=2, value=acao)
        p_cell = ws.cell(row=r, column=3, value=prazo)
        if prazo == "30 dias":
            p_cell.fill = FILL_VERMELHO
            p_cell.font = FONT_NEGRITO
        elif prazo == "45 dias":
            p_cell.fill = FILL_LARANJA
        elif prazo == "60 dias":
            p_cell.fill = FILL_LARANJA
        ws.cell(row=r, column=4, value=risco)
    _aplicar_zebra(ws, 5, 4 + len(dados), 1, 4)
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 32


def gerar_xlsx():
    transacoes, nfes, ctes = carregar_dados()

    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    abas_meta = [
        ("01_Sumario_Executivo", "8 constatacoes consolidadas com risco anualizado", "R$ 5,12M Critico"),
        ("02_Resumo_Geral", "Indicadores agregados do periodo (OFX + NF-e + CT-e)", "70M+ movimentado"),
        ("03_Partes_Relacionadas", "Movimentacoes intercompany (LOCAR Locadora, Maquinas, Renato PF)", "R$ 15,2M"),
        ("04_MEIs_Reclassificados", "MEI-TAC vs MEI Padrao (apos correcao LC 188/2021)", "65 OK + 5 padrao"),
        ("05_Retencoes", "Retencoes na fonte nao recolhidas", "R$ 1,17M/ano"),
        ("06_CNPJ_Baixados", "Pagamentos a CNPJs ja baixados na RFB", "17 tx / R$ 35,6k"),
        ("07_Top_Pagamentos", "Top 50 maiores saidas do periodo", "Maiores valores"),
        ("08_Top_Recebimentos", "Top 50 maiores entradas do periodo", "Maiores valores"),
        ("09_Alvos_Investigacao", "Thiago, GT Participacoes, REDE FROTA", "R$ 5,9M sob investigacao"),
        ("10_Obrigacoes_Acessorias", "Checklist de obrigacoes do Lucro Real", "Mensal + Anual"),
        ("11_Recomendacoes", "11 acoes formais ordenadas por prazo", "30/60/90/120 dias"),
        ("12_Conformidade_Fiscal", "[NOVO] Top fornecedores - cruzamento OFX x NF/CT", "Score 0-100%"),
        ("13_Documentos_Fiscais", "[NOVO] 8.226 XMLs processados (NF-e + CT-e)", "R$ 33,6M docs"),
        ("14_Riscos_Fiscais", "[NOVO] Riscos consolidados + mitigacoes propostas", "R$ 5,12M risco"),
    ]

    print("Gerando abas...")
    aba_sumario(wb)
    aba_resumo_geral(wb, transacoes, nfes, ctes)
    aba_partes_relacionadas(wb)
    aba_meis(wb)
    aba_retencoes(wb)
    aba_baixados(wb)
    aba_top_pagamentos(wb, transacoes)
    aba_top_recebimentos(wb, transacoes)
    aba_alvos_investigacao(wb)
    aba_obrigacoes(wb)
    aba_recomendacoes(wb)
    aba_conformidade_fiscal(wb, nfes, ctes, transacoes)
    aba_documentos_fiscais(wb, nfes, ctes)
    aba_riscos_fiscais(wb)
    aba_indice(wb, abas_meta)

    # Reordenar para colocar indice primeiro
    sheet_order = ["Indice"] + [m[0] for m in abas_meta]
    wb._sheets = [wb[name] for name in sheet_order]

    print(f"Salvando: {OUT_XLSX}")
    wb.save(OUT_XLSX)
    print(f"OK - {len(sheet_order)} abas geradas")


if __name__ == "__main__":
    gerar_xlsx()
