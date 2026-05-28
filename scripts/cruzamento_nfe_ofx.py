"""Cruzamento individual de CT-es/NF-es com transacoes bancarias.

Para cada um dos 3.045 CT-es emitidos pela LOCAR:
- Busca recebimentos no OFX com mesmo CNPJ tomador + valor proximo + data proxima
- Classifica: CASADO / VALOR_DIVERGENTE / SEM_RECEBIMENTO

Para cada uma das 5.031 NF-es recebidas:
- Busca pagamentos no OFX com mesmo CNPJ emitente + valor proximo + data proxima
- Classifica: CASADO / VALOR_DIVERGENTE / SEM_PAGAMENTO

Gera XLSX com 4 abas:
1. Sumario do Cruzamento
2. Cruzamento Receita (CT-es vs Recebimentos)
3. Cruzamento Compras (NF-es vs Pagamentos)
4. Transacoes sem documento fiscal
"""
from __future__ import annotations

import asyncio
import re
import sys
import xml.etree.ElementTree as ET
import zipfile
from collections import Counter, defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
sys.path.insert(0, str(Path(__file__).resolve().parent))
from _logo_helper import html_logo_inline, inserir_logo_xlsx
from api.matchers.cascata import ler_ofx

ZIPS = [
    (r"C:\Users\Veloso\Downloads\05509396000110_01012026_31012026_0546.zip", "EMITIDOS"),
    (r"C:\Users\Veloso\Downloads\05509396000110_01022026_28022026_5384.zip", "EMITIDOS"),
    (r"C:\Users\Veloso\Downloads\05509396000110_01032026_31032026_4046.zip", "EMITIDOS"),
    (r"C:\Users\Veloso\Downloads\05509396000110_01042026_30042026_9825.zip", "EMITIDOS"),
    (r"C:\Users\Veloso\Downloads\103585885_01012026_31012026_7514.zip", "RECEBIDOS"),
    (r"C:\Users\Veloso\Downloads\103585885_01022026_28022026_8464.zip", "RECEBIDOS"),
    (r"C:\Users\Veloso\Downloads\103585885_01032026_31032026_8594.zip", "RECEBIDOS"),
    (r"C:\Users\Veloso\Downloads\103585885_01042026_30042026_7245.zip", "RECEBIDOS"),
]
OFX_LIST = [
    r"C:\Users\Veloso\Downloads\extrato-conta-corrente-ofx-unix_202605_20260514110822.ofx",
    r"C:\Users\Veloso\Downloads\extrato-conta-corrente-ofx-unix_202605_20260514110841.ofx",
    r"C:\Users\Veloso\Downloads\extrato-conta-corrente-ofx-unix_202605_20260514110900.ofx",
    r"C:\Users\Veloso\Downloads\extrato-conta-corrente-ofx-unix_202605_20260514110917.ofx",
]

OUT_BASE = r"C:\Users\Veloso\Downloads\CRUZAMENTO_NFE_OFX_LOCAR"
OUT_XLSX = Path(f"{OUT_BASE}.xlsx")
OUT_MD = Path(f"{OUT_BASE}.md")
OUT_HTML = Path(f"{OUT_BASE}.html")
OUT_PDF = Path(f"{OUT_BASE}.pdf")

# Tolerancias de matching
TOLERANCIA_VALOR = 0.50      # R$ 0,50 de diferenca aceita
TOLERANCIA_DIAS = 15         # janela de +/- 15 dias entre emissao e pagamento

RX_CNPJ = re.compile(r"(\d{2})[.](\d{3})[.](\d{3})[ /](\d{4})[-](\d{2})")

# Estilos
NAVY = "0F172A"
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TOTAL_FILL = PatternFill("solid", fgColor="1E3A8A")
TOTAL_FONT = Font(bold=True, color="FFFFFF", size=11)
ZEBRA_FILL = PatternFill("solid", fgColor="F8FAFC")
ALERT_FILL = PatternFill("solid", fgColor="FEE2E2")
SUCCESS_FILL = PatternFill("solid", fgColor="DCFCE7")
WARNING_FILL = PatternFill("solid", fgColor="FEF3C7")
TITLE_FONT = Font(bold=True, size=14, color=NAVY)
SUBTITLE_FONT = Font(bold=True, size=11, color=NAVY)
BORDER = Side(border_style="thin", color="E2E8F0")
THIN_BORDER = Border(top=BORDER, left=BORDER, right=BORDER, bottom=BORDER)


def _local(tag):
    return tag.split("}")[-1]


def _filho(elem, nome):
    if elem is None:
        return None
    for f in elem:
        if _local(f.tag) == nome:
            return f
    return None


def _texto(elem, *caminho):
    cur = elem
    for nome in caminho:
        cur = _filho(cur, nome)
        if cur is None:
            return ""
    return cur.text.strip() if cur is not None and cur.text else ""


def _achar(root, *nomes):
    for elem in root.iter():
        if _local(elem.tag) in nomes:
            return elem
    return None


def parse_nfe(conteudo):
    try:
        root = ET.fromstring(conteudo)
    except ET.ParseError:
        return None
    inf = _achar(root, "infNFe")
    if inf is None:
        return None
    ide = _filho(inf, "ide")
    emit = _filho(inf, "emit")
    total = _filho(inf, "total")
    icms_tot = _filho(total, "ICMSTot") if total is not None else None
    chave = (inf.get("Id") or "").lstrip("NFe")
    data = (_texto(ide, "dhEmi") or _texto(ide, "dEmi"))[:10]
    return {
        "tipo": "NF-e",
        "chave": chave,
        "numero": _texto(ide, "nNF"),
        "data": data,
        "emit_cnpj": _texto(emit, "CNPJ"),
        "emit_nome": _texto(emit, "xNome"),
        "valor": float(_texto(icms_tot, "vNF") or 0) if icms_tot is not None else 0.0,
    }


def parse_cte(conteudo):
    try:
        root = ET.fromstring(conteudo)
    except ET.ParseError:
        return None
    inf = _achar(root, "infCte", "infCTe")
    if inf is None:
        return None
    ide = _filho(inf, "ide")
    emit = _filho(inf, "emit")
    rem = _filho(inf, "rem")
    dest = _filho(inf, "dest")
    toma3 = _filho(inf, "toma3")
    toma4 = _filho(inf, "toma4")
    vprest = _filho(inf, "vPrest")

    # Tomador conforme indicacao
    # toma3: 0=Remetente, 1=Expedidor, 2=Recebedor, 3=Destinatario
    # toma4: campo customizado
    toma_cnpj = ""
    toma_nome = ""
    if toma4 is not None:
        toma_cnpj = _texto(toma4, "CNPJ")
        toma_nome = _texto(toma4, "xNome")
    elif toma3 is not None:
        cod_toma = _texto(toma3, "toma")
        # mapeia para rem/exped/receb/dest
        if cod_toma == "0" and rem is not None:
            toma_cnpj = _texto(rem, "CNPJ")
            toma_nome = _texto(rem, "xNome")
        elif cod_toma == "3" and dest is not None:
            toma_cnpj = _texto(dest, "CNPJ")
            toma_nome = _texto(dest, "xNome")

    # Default fallback
    if not toma_cnpj:
        if dest is not None:
            toma_cnpj = _texto(dest, "CNPJ")
            toma_nome = _texto(dest, "xNome")
        elif rem is not None:
            toma_cnpj = _texto(rem, "CNPJ")
            toma_nome = _texto(rem, "xNome")

    chave = (inf.get("Id") or "").lstrip("CTe")
    data = (_texto(ide, "dhEmi") or _texto(ide, "dEmi"))[:10]
    return {
        "tipo": "CT-e",
        "chave": chave,
        "numero": _texto(ide, "nCT"),
        "data": data,
        "emit_cnpj": _texto(emit, "CNPJ"),
        "emit_nome": _texto(emit, "xNome"),
        "toma_cnpj": toma_cnpj,
        "toma_nome": toma_nome,
        "valor": float(_texto(vprest, "vTPrest") or 0) if vprest is not None else 0.0,
    }


def processar_xmls():
    """Le todos os XMLs dos 8 ZIPs."""
    ctes, nfes = [], []
    for path, tipo in ZIPS:
        with zipfile.ZipFile(path) as zf:
            for member in zf.namelist():
                if not member.endswith(".xml"):
                    continue
                chave_base = Path(member).stem
                if len(chave_base) < 44:
                    continue
                modelo = chave_base[20:22]
                with zf.open(member) as f:
                    conteudo = f.read()
                if modelo == "55":
                    doc = parse_nfe(conteudo)
                    if doc:
                        nfes.append(doc)
                elif modelo == "57":
                    doc = parse_cte(conteudo)
                    if doc:
                        ctes.append(doc)
    return ctes, nfes


def _extrair_cnpj(texto):
    """Extrai CNPJ de qualquer texto."""
    if not texto:
        return None
    m = RX_CNPJ.search(texto)
    if m:
        return "".join(m.groups())
    return None


def cruzar_recebimentos(ctes, transacoes):
    """Para cada CT-e, busca recebimento correspondente no OFX."""
    # Indexa transacoes de credito por CNPJ
    creditos_por_cnpj = defaultdict(list)
    for t in transacoes:
        if t.valor <= 0:
            continue
        cnpj = _extrair_cnpj((t.nome or "") + " " + (t.memo or ""))
        if cnpj:
            creditos_por_cnpj[cnpj].append(t)

    resultados = []
    for cte in ctes:
        cnpj = cte["toma_cnpj"]
        valor_cte = cte["valor"]
        try:
            data_cte = date.fromisoformat(cte["data"]) if cte["data"] else None
        except ValueError:
            data_cte = None

        candidatos = creditos_por_cnpj.get(cnpj, [])
        match = None
        for t in candidatos:
            if abs(t.valor - valor_cte) > TOLERANCIA_VALOR:
                continue
            try:
                data_trn = date.fromisoformat(t.data[:10])
            except (ValueError, AttributeError):
                continue
            if data_cte and abs((data_trn - data_cte).days) > TOLERANCIA_DIAS:
                continue
            match = t
            break

        # Procura por valor sem restrir CNPJ se nao achou
        match_valor = None
        if not match and cnpj:
            for t in candidatos:
                if abs(t.valor - valor_cte) <= TOLERANCIA_VALOR * 100:  # tolerancia maior
                    match_valor = t
                    break

        if match:
            status = "CASADO"
        elif match_valor:
            status = "VALOR_DIVERGENTE"
        else:
            status = "SEM_RECEBIMENTO"

        resultados.append({
            "cte": cte,
            "match": match or match_valor,
            "status": status,
        })

    return resultados


def cruzar_pagamentos(nfes, transacoes):
    """Para cada NF-e, busca pagamento correspondente."""
    debitos_por_cnpj = defaultdict(list)
    for t in transacoes:
        if t.valor >= 0:
            continue
        cnpj = _extrair_cnpj((t.nome or "") + " " + (t.memo or ""))
        if cnpj:
            debitos_por_cnpj[cnpj].append(t)

    resultados = []
    for nfe in nfes:
        cnpj = nfe["emit_cnpj"]
        valor_nfe = nfe["valor"]
        try:
            data_nfe = date.fromisoformat(nfe["data"]) if nfe["data"] else None
        except ValueError:
            data_nfe = None

        candidatos = debitos_por_cnpj.get(cnpj, [])
        match = None
        for t in candidatos:
            if abs(abs(t.valor) - valor_nfe) > TOLERANCIA_VALOR:
                continue
            try:
                data_trn = date.fromisoformat(t.data[:10])
            except (ValueError, AttributeError):
                continue
            if data_nfe and abs((data_trn - data_nfe).days) > TOLERANCIA_DIAS:
                continue
            match = t
            break

        if match:
            status = "CASADO"
        elif candidatos:
            status = "CNPJ_SEM_VALOR_BATENDO"
        else:
            status = "SEM_PAGAMENTO"

        resultados.append({
            "nfe": nfe,
            "match": match,
            "status": status,
        })

    return resultados


# ════════════════════════════════════════════════════════════════════════
# Gerador XLSX
# ════════════════════════════════════════════════════════════════════════


def cabecalho(ws, ultima_col, secao):
    c1 = ws.cell(row=1, column=1, value="    ORGATEC · Cruzamento NF-e/OFX · LOCAR TRANSPORTE DE BOVINOS LTDA")
    c1.font = Font(bold=True, size=14, color="FFFFFF")
    c1.fill = PatternFill("solid", fgColor=NAVY)
    c1.alignment = Alignment(horizontal="center", vertical="center", indent=2)
    ws.merge_cells(f"A1:{get_column_letter(ultima_col)}1")
    ws.row_dimensions[1].height = 60
    ws.column_dimensions["A"].width = 14
    inserir_logo_xlsx(ws, "A1", largura_px=60, altura_px=60)
    c2 = ws.cell(row=2, column=1, value="CNPJ 05.509.396/0001-10 · Cruzamento documental: 8.226 XMLs x 7.110 transacoes bancarias")
    c2.font = Font(bold=True, size=10, color="FFFFFF")
    c2.fill = PatternFill("solid", fgColor="1E3A8A")
    c2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(f"A2:{get_column_letter(ultima_col)}2")
    c3 = ws.cell(row=3, column=1, value=f"Secao: {secao} · Tolerancia: R$ {TOLERANCIA_VALOR:.2f} valor / +-{TOLERANCIA_DIAS} dias data")
    c3.font = Font(size=9, color="0F172A")
    c3.fill = PatternFill("solid", fgColor="DBEAFE")
    c3.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(f"A3:{get_column_letter(ultima_col)}3")
    return 5


def style_header(ws, row, n):
    for c in range(1, n + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = THIN_BORDER


def gerar_xlsx(rec_results, pag_results, transacoes):
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # ── Aba 1: Sumario ─────────────────────────────────────────────────
    ws = wb.create_sheet("1. Sumario")
    start = cabecalho(ws, 7, "Sumario do Cruzamento")
    ws.cell(row=start, column=1, value="SUMARIO DO CRUZAMENTO FISCAL X BANCARIO").font = TITLE_FONT
    ws.merge_cells(f"A{start}:G{start}")

    # Recebimentos
    rec_casados = sum(1 for r in rec_results if r["status"] == "CASADO")
    rec_divergentes = sum(1 for r in rec_results if r["status"] == "VALOR_DIVERGENTE")
    rec_sem = sum(1 for r in rec_results if r["status"] == "SEM_RECEBIMENTO")
    rec_total_valor = sum(r["cte"]["valor"] for r in rec_results)
    rec_casado_valor = sum(r["cte"]["valor"] for r in rec_results if r["status"] == "CASADO")
    rec_sem_valor = sum(r["cte"]["valor"] for r in rec_results if r["status"] == "SEM_RECEBIMENTO")

    pag_casados = sum(1 for r in pag_results if r["status"] == "CASADO")
    pag_divergentes = sum(1 for r in pag_results if r["status"] == "CNPJ_SEM_VALOR_BATENDO")
    pag_sem = sum(1 for r in pag_results if r["status"] == "SEM_PAGAMENTO")
    pag_total_valor = sum(r["nfe"]["valor"] for r in pag_results)
    pag_casado_valor = sum(r["nfe"]["valor"] for r in pag_results if r["status"] == "CASADO")
    pag_sem_valor = sum(r["nfe"]["valor"] for r in pag_results if r["status"] == "SEM_PAGAMENTO")

    r = start + 2
    ws.cell(row=r, column=1, value="RECEITA — CT-es vs Recebimentos OFX").font = SUBTITLE_FONT
    ws.merge_cells(f"A{r}:G{r}")
    r += 1
    headers = ["Status", "Quantidade", "% Doc", "Valor (R$)", "% Valor", "Cor", "Acao"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=r, column=c, value=h)
    style_header(ws, r, 7)
    r += 1

    rec_total_n = len(rec_results)
    linhas_rec = [
        ("CASADO", rec_casados, rec_casado_valor, "verde", "OK - receita confirmada"),
        ("VALOR_DIVERGENTE", rec_divergentes, sum(r["cte"]["valor"] for r in rec_results if r["status"] == "VALOR_DIVERGENTE"), "amarelo", "Conciliar manual"),
        ("SEM_RECEBIMENTO", rec_sem, rec_sem_valor, "vermelho", "Investigar - receita nao recebida"),
    ]
    fills = {"verde": SUCCESS_FILL, "amarelo": WARNING_FILL, "vermelho": ALERT_FILL}
    for status, qtd, val, cor, acao in linhas_rec:
        ws.cell(row=r, column=1, value=status).font = Font(bold=True)
        ws.cell(row=r, column=2, value=qtd).number_format = "#,##0"
        ws.cell(row=r, column=3, value=qtd / max(rec_total_n, 1)).number_format = "0.0%"
        ws.cell(row=r, column=4, value=round(val, 2)).number_format = "#,##0.00"
        ws.cell(row=r, column=5, value=val / max(rec_total_valor, 1)).number_format = "0.0%"
        ws.cell(row=r, column=6, value=cor)
        ws.cell(row=r, column=7, value=acao)
        for c in range(1, 8):
            ws.cell(row=r, column=c).border = THIN_BORDER
            ws.cell(row=r, column=c).fill = fills[cor]
        r += 1
    ws.cell(row=r, column=1, value="TOTAL CT-es").font = TOTAL_FONT
    ws.cell(row=r, column=2, value=rec_total_n).number_format = "#,##0"
    ws.cell(row=r, column=4, value=round(rec_total_valor, 2)).number_format = "#,##0.00"
    for c in range(1, 8):
        ws.cell(row=r, column=c).fill = TOTAL_FILL
        ws.cell(row=r, column=c).font = TOTAL_FONT
    r += 3

    # Pagamentos
    ws.cell(row=r, column=1, value="COMPRAS — NF-es vs Pagamentos OFX").font = SUBTITLE_FONT
    ws.merge_cells(f"A{r}:G{r}")
    r += 1
    for c, h in enumerate(headers, start=1):
        ws.cell(row=r, column=c, value=h)
    style_header(ws, r, 7)
    r += 1
    pag_total_n = len(pag_results)
    linhas_pag = [
        ("CASADO", pag_casados, pag_casado_valor, "verde", "OK - pagamento confirmado"),
        ("CNPJ_SEM_VALOR_BATENDO", pag_divergentes, sum(r["nfe"]["valor"] for r in pag_results if r["status"] == "CNPJ_SEM_VALOR_BATENDO"), "amarelo", "Pagamento parcial ou agrupado"),
        ("SEM_PAGAMENTO", pag_sem, pag_sem_valor, "vermelho", "A pagar OU pago por outro canal"),
    ]
    for status, qtd, val, cor, acao in linhas_pag:
        ws.cell(row=r, column=1, value=status).font = Font(bold=True)
        ws.cell(row=r, column=2, value=qtd).number_format = "#,##0"
        ws.cell(row=r, column=3, value=qtd / max(pag_total_n, 1)).number_format = "0.0%"
        ws.cell(row=r, column=4, value=round(val, 2)).number_format = "#,##0.00"
        ws.cell(row=r, column=5, value=val / max(pag_total_valor, 1)).number_format = "0.0%"
        ws.cell(row=r, column=6, value=cor)
        ws.cell(row=r, column=7, value=acao)
        for c in range(1, 8):
            ws.cell(row=r, column=c).border = THIN_BORDER
            ws.cell(row=r, column=c).fill = fills[cor]
        r += 1
    ws.cell(row=r, column=1, value="TOTAL NF-es").font = TOTAL_FONT
    ws.cell(row=r, column=2, value=pag_total_n).number_format = "#,##0"
    ws.cell(row=r, column=4, value=round(pag_total_valor, 2)).number_format = "#,##0.00"
    for c in range(1, 8):
        ws.cell(row=r, column=c).fill = TOTAL_FILL
        ws.cell(row=r, column=c).font = TOTAL_FONT

    for col, w in {1: 24, 2: 14, 3: 10, 4: 18, 5: 10, 6: 12, 7: 38}.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = f"A{start + 2}"

    # ── Aba 2: Cruzamento Receita Detalhado ────────────────────────────
    ws = wb.create_sheet("2. Cruzamento Receita")
    start = cabecalho(ws, 9, "Cruzamento Receita")
    ws.cell(row=start, column=1, value="CT-ES EMITIDOS vs RECEBIMENTOS OFX").font = TITLE_FONT
    ws.merge_cells(f"A{start}:I{start}")

    headers = ["#", "Data CT-e", "Numero CT-e", "Tomador", "CNPJ", "Valor CT-e (R$)",
               "Status", "Data Pago", "Diferenca"]
    r = start + 2
    for c, h in enumerate(headers, start=1):
        ws.cell(row=r, column=c, value=h)
    style_header(ws, r, 9)
    r += 1

    fills_status = {
        "CASADO": SUCCESS_FILL,
        "VALOR_DIVERGENTE": WARNING_FILL,
        "SEM_RECEBIMENTO": ALERT_FILL,
    }
    for i, res in enumerate(rec_results[:500], start=1):
        cte = res["cte"]
        cnpj_fmt = ""
        if cte["toma_cnpj"] and len(cte["toma_cnpj"]) == 14:
            cnpj_fmt = f"{cte['toma_cnpj'][:2]}.{cte['toma_cnpj'][2:5]}.{cte['toma_cnpj'][5:8]}/{cte['toma_cnpj'][8:12]}-{cte['toma_cnpj'][12:14]}"
        match = res["match"]
        data_pago = match.data if match else ""
        try:
            diff = (date.fromisoformat(match.data[:10]) - date.fromisoformat(cte["data"])).days if match and cte["data"] else 0
        except (ValueError, AttributeError):
            diff = 0

        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=2, value=cte["data"])
        ws.cell(row=r, column=3, value=cte["numero"])
        ws.cell(row=r, column=4, value=(cte["toma_nome"] or "")[:38])
        ws.cell(row=r, column=5, value=cnpj_fmt).font = Font(name="Consolas", size=9)
        cv = ws.cell(row=r, column=6, value=round(cte["valor"], 2))
        cv.number_format = "#,##0.00"
        c_status = ws.cell(row=r, column=7, value=res["status"])
        c_status.font = Font(bold=True)
        ws.cell(row=r, column=8, value=data_pago)
        ws.cell(row=r, column=9, value=f"{diff:+d} dias" if match else "—")
        for c in range(1, 10):
            ws.cell(row=r, column=c).border = THIN_BORDER
            if res["status"] in fills_status:
                ws.cell(row=r, column=c).fill = fills_status[res["status"]]
        r += 1

    if len(rec_results) > 500:
        ws.cell(row=r, column=1, value=f"... + {len(rec_results) - 500} CT-es restantes (truncado para nao sobrecarregar XLSX)").font = Font(italic=True, color="64748B")

    for col, w in {1: 4, 2: 11, 3: 12, 4: 38, 5: 19, 6: 16, 7: 22, 8: 11, 9: 12}.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = f"A{start + 3}"
    ws.auto_filter.ref = f"A{start + 2}:I{r - 1}"

    # ── Aba 3: Cruzamento Compras Detalhado ────────────────────────────
    ws = wb.create_sheet("3. Cruzamento Compras")
    start = cabecalho(ws, 9, "Cruzamento Compras")
    ws.cell(row=start, column=1, value="NF-ES RECEBIDAS vs PAGAMENTOS OFX").font = TITLE_FONT
    ws.merge_cells(f"A{start}:I{start}")

    headers = ["#", "Data NF-e", "Numero NF-e", "Emitente", "CNPJ", "Valor NF-e (R$)",
               "Status", "Data Pago", "Diferenca"]
    r = start + 2
    for c, h in enumerate(headers, start=1):
        ws.cell(row=r, column=c, value=h)
    style_header(ws, r, 9)
    r += 1

    fills_pag = {
        "CASADO": SUCCESS_FILL,
        "CNPJ_SEM_VALOR_BATENDO": WARNING_FILL,
        "SEM_PAGAMENTO": ALERT_FILL,
    }
    for i, res in enumerate(pag_results[:500], start=1):
        nfe = res["nfe"]
        cnpj_fmt = ""
        if nfe["emit_cnpj"] and len(nfe["emit_cnpj"]) == 14:
            cnpj_fmt = f"{nfe['emit_cnpj'][:2]}.{nfe['emit_cnpj'][2:5]}.{nfe['emit_cnpj'][5:8]}/{nfe['emit_cnpj'][8:12]}-{nfe['emit_cnpj'][12:14]}"
        match = res["match"]
        data_pago = match.data if match else ""
        try:
            diff = (date.fromisoformat(match.data[:10]) - date.fromisoformat(nfe["data"])).days if match and nfe["data"] else 0
        except (ValueError, AttributeError):
            diff = 0

        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=2, value=nfe["data"])
        ws.cell(row=r, column=3, value=nfe["numero"])
        ws.cell(row=r, column=4, value=(nfe["emit_nome"] or "")[:38])
        ws.cell(row=r, column=5, value=cnpj_fmt).font = Font(name="Consolas", size=9)
        cv = ws.cell(row=r, column=6, value=round(nfe["valor"], 2))
        cv.number_format = "#,##0.00"
        c_status = ws.cell(row=r, column=7, value=res["status"])
        c_status.font = Font(bold=True)
        ws.cell(row=r, column=8, value=data_pago)
        ws.cell(row=r, column=9, value=f"{diff:+d} dias" if match else "—")
        for c in range(1, 10):
            ws.cell(row=r, column=c).border = THIN_BORDER
            if res["status"] in fills_pag:
                ws.cell(row=r, column=c).fill = fills_pag[res["status"]]
        r += 1

    if len(pag_results) > 500:
        ws.cell(row=r, column=1, value=f"... + {len(pag_results) - 500} NF-es restantes (truncado)").font = Font(italic=True, color="64748B")

    for col, w in {1: 4, 2: 11, 3: 12, 4: 38, 5: 19, 6: 16, 7: 25, 8: 11, 9: 12}.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = f"A{start + 3}"
    ws.auto_filter.ref = f"A{start + 2}:I{r - 1}"

    # ── Aba 4: Transacoes Sem NF ────────────────────────────────────────
    ws = wb.create_sheet("4. Transacoes Sem NF")
    start = cabecalho(ws, 6, "Transacoes sem Documento Fiscal")
    ws.cell(row=start, column=1, value="TRANSACOES OFX SEM CT-E/NF-E CORRESPONDENTE").font = TITLE_FONT
    ws.merge_cells(f"A{start}:F{start}")

    # Identifica transacoes que NAO foram casadas
    matches_ids = set()
    for r in rec_results:
        if r["match"]:
            matches_ids.add(id(r["match"]))
    for r in pag_results:
        if r["match"]:
            matches_ids.add(id(r["match"]))

    sem_nf = [t for t in transacoes if id(t) not in matches_ids]

    headers = ["#", "Data", "Tipo", "Valor (R$)", "Memo", "Nome"]
    r = start + 2
    for c, h in enumerate(headers, start=1):
        ws.cell(row=r, column=c, value=h)
    style_header(ws, r, 6)
    r += 1

    sem_nf_sorted = sorted(sem_nf, key=lambda t: -abs(t.valor))[:200]
    for i, t in enumerate(sem_nf_sorted, start=1):
        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=2, value=t.data)
        ws.cell(row=r, column=3, value=t.tipo)
        cv = ws.cell(row=r, column=4, value=round(t.valor, 2))
        cv.number_format = "#,##0.00"
        cv.font = Font(color=("DC2626" if t.valor < 0 else "16A34A"))
        ws.cell(row=r, column=5, value=(t.memo or "")[:35])
        ws.cell(row=r, column=6, value=(t.nome or "")[:35])
        for c in range(1, 7):
            ws.cell(row=r, column=c).border = THIN_BORDER
            if r % 2 == 0:
                ws.cell(row=r, column=c).fill = ZEBRA_FILL
        r += 1

    for col, w in {1: 4, 2: 11, 3: 8, 4: 14, 5: 35, 6: 35}.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = f"A{start + 3}"

    wb.save(str(OUT_XLSX))
    print(f"  XLSX: {OUT_XLSX}")
    return {
        "rec_casados": rec_casados, "rec_divergentes": rec_divergentes, "rec_sem": rec_sem,
        "rec_total_valor": rec_total_valor, "rec_casado_valor": rec_casado_valor, "rec_sem_valor": rec_sem_valor,
        "pag_casados": pag_casados, "pag_divergentes": pag_divergentes, "pag_sem": pag_sem,
        "pag_total_valor": pag_total_valor, "pag_casado_valor": pag_casado_valor, "pag_sem_valor": pag_sem_valor,
        "rec_total_n": len(rec_results), "pag_total_n": len(pag_results),
        "sem_nf_total": len(sem_nf),
    }


def gerar_md(stats):
    pct_rec_casado = 100 * stats["rec_casados"] / max(stats["rec_total_n"], 1)
    pct_pag_casado = 100 * stats["pag_casados"] / max(stats["pag_total_n"], 1)
    return f"""# CRUZAMENTO NF-E/CT-E x OFX — LOCAR TRANSPORTE DE BOVINOS LTDA

**Auditoria Documental · Cruzamento Fiscal-Bancario**

---

**Gerado em:** {datetime.now().strftime('%d/%m/%Y %H:%M')}
**Empresa:** LOCAR TRANSPORTE DE BOVINOS LTDA · CNPJ 05.509.396/0001-10
**Documentos:** 8.226 XMLs · 7.110 transacoes bancarias · Periodo jan-abr/2026
**Tolerancia:** R$ {TOLERANCIA_VALOR:.2f} de valor / +-{TOLERANCIA_DIAS} dias de data

---

## 1. Sumario do Cruzamento

### Receita — CT-es vs Recebimentos no OFX

| Status | Quantidade | % | Valor (R$) | Acao |
|---|---:|---:|---:|---|
| CASADO | {stats['rec_casados']:,} | {pct_rec_casado:.1f}% | {stats['rec_casado_valor']:,.2f} | OK |
| VALOR DIVERGENTE | {stats['rec_divergentes']:,} | — | — | Conciliar manual |
| **SEM RECEBIMENTO** | **{stats['rec_sem']:,}** | — | **{stats['rec_sem_valor']:,.2f}** | **Investigar** |
| **TOTAL** | **{stats['rec_total_n']:,}** | 100% | **{stats['rec_total_valor']:,.2f}** | — |

### Compras — NF-es vs Pagamentos no OFX

| Status | Quantidade | % | Valor (R$) | Acao |
|---|---:|---:|---:|---|
| CASADO | {stats['pag_casados']:,} | {pct_pag_casado:.1f}% | {stats['pag_casado_valor']:,.2f} | OK |
| CNPJ S/ VALOR | {stats['pag_divergentes']:,} | — | — | Pagamento parcial/agrupado |
| **SEM PAGAMENTO** | **{stats['pag_sem']:,}** | — | **{stats['pag_sem_valor']:,.2f}** | **A pagar OU outro canal** |
| **TOTAL** | **{stats['pag_total_n']:,}** | 100% | **{stats['pag_total_valor']:,.2f}** | — |

## 2. Interpretacao dos Achados

### Taxa de Casamento Direto

- **CT-es (Receita)**: {pct_rec_casado:.1f}% casados diretamente com recebimentos no OFX
- **NF-es (Compras)**: {pct_pag_casado:.1f}% casadas diretamente com pagamentos no OFX

### Hipoteses para os SEM_RECEBIMENTO ({stats['rec_sem']:,} CT-es totalizando R$ {stats['rec_sem_valor']:,.2f}):

1. **Recebimento em outra conta** (a LOCAR pode ter outras contas alem da 158083-3 analisada)
2. **Recebimento via boleto** intermediado por banco diferente do Sicoob
3. **Pagamento agrupado** (varios CT-es pagos em 1 unica transacao)
4. **Inadimplencia** (CT-es emitidos mas nao recebidos)
5. **Receita fictica** (red flag - emissao sem contraprestacao real) - **investigar**

### Hipoteses para os SEM_PAGAMENTO ({stats['pag_sem']:,} NF-es totalizando R$ {stats['pag_sem_valor']:,.2f}):

1. **Pagamento em outra conta** da LOCAR
2. **Compras a prazo** (NF-e emitida mas pagamento futuro)
3. **Compras pagas em dinheiro** (sem rastreio bancario)
4. **Pagamento via cartao** (de outra conta)
5. **NF-es nao reconhecidas** (emitidas erroneamente contra a LOCAR)

## 3. Transacoes Bancarias sem Documento Fiscal

{stats['sem_nf_total']:,} transacoes bancarias NAO foram casadas a nenhum CT-e ou NF-e.

**Possiveis explicacoes:**

- Tarifas bancarias, IOF, tributos, pro-labore, dividendos
- Transferencias internas (MESMA TIT)
- Movimentacoes com partes relacionadas (LOCAR LOCADORA, MAQUINAS, Renato)
- Despesas sem documento fiscal eletronico (sujeitas a glosa em fiscalizacao Lucro Real)

**Recomendacao:** Verificar Aba 4 do XLSX para identificar transacoes de **alto valor** sem NF correspondente.

## 4. Conclusoes Tributarias

### Receita Fiscal Confirmada vs Receita Bancaria

- Receita CT-e (formal): R$ {stats['rec_total_valor']:,.2f}
- Receita CT-e CASADA no OFX: R$ {stats['rec_casado_valor']:,.2f} ({100 * stats['rec_casado_valor'] / max(stats['rec_total_valor'], 1):.1f}%)
- Receita CT-e SEM recebimento bancario: R$ {stats['rec_sem_valor']:,.2f} ({100 * stats['rec_sem_valor'] / max(stats['rec_total_valor'], 1):.1f}%)

### Risco Fiscal Adicional

Se a receita CT-e nao confirmada (R$ {stats['rec_sem_valor']:,.2f}) for de fato ficticia ou nao recebida, ha:

- Possivel **emissao de CT-e sem prestacao real** (Lei 8.137/90)
- **PIS/COFINS pagos sem fundamento** (Lucro Real - regime competencia)
- Necessidade de **cancelamento das CT-es** OU **emissao de carta de correcao** dentro do prazo legal

---

*Sistema OrgConc/OrgNeural2 v0.5.0 - Cruzamento documental fiscal-bancario.*
"""


def gerar_html(md_text):
    import markdown as mdlib
    body = mdlib.markdown(md_text, extensions=["tables", "fenced_code"])
    agora = datetime.now().strftime("%d/%m/%Y %H:%M")
    css = """
@page { size: A4 landscape; margin: 14mm 12mm 14mm 12mm;
  @bottom-right { content: "Pagina " counter(page) " de " counter(pages); font-size: 9px; color: #6B7280; }
  @bottom-left { content: "Cruzamento NF-e/OFX · LOCAR TRANSPORTE"; font-size: 9px; color: #6B7280; }
}
* { box-sizing: border-box; margin: 0; padding: 0; }
body { font-family: 'DejaVu Sans', Arial, sans-serif; font-size: 10pt; color: #1a202c; line-height: 1.55; }
.hd { background: linear-gradient(135deg, #0F172A, #0B1B3D 45%, #0052FF); color: #fff;
      padding: 22px 28px; border-radius: 8px; margin-bottom: 22px; display: flex; align-items: center; gap: 22px; }
.hd-text { flex: 1; }
.hd h1 { font-size: 22pt; font-family: 'DejaVu Serif', Georgia, serif; }
.hd .tag { font-size: 10pt; opacity: 0.9; text-transform: uppercase; letter-spacing: 0.18em; }
h1 { font-size: 14pt; color: #0F172A; margin: 22px 0 8px; padding-bottom: 6px; border-bottom: 2px solid #0052FF; }
h2 { font-size: 12pt; color: #0052FF; margin: 18px 0 8px; padding-left: 10px; border-left: 3px solid #0EA5E9; }
h3 { font-size: 11pt; color: #0F172A; margin: 14px 0 6px; }
table { width: 100%; border-collapse: collapse; margin: 10px 0 14px; font-size: 9pt; }
th { background: linear-gradient(180deg, #0F172A, #1E3A8A); color: #fff; padding: 6px 9px; text-align: left; }
td { padding: 5px 9px; border-bottom: 1px solid #E2E8F0; }
tr:nth-child(even) td { background: #F8FAFC; }
strong { color: #0F172A; font-weight: 700; }
"""
    return f"""<!DOCTYPE html><html lang="pt-BR"><head><meta charset="UTF-8">
<title>Cruzamento NF-e/OFX LOCAR</title><style>{css}</style></head>
<body>
<div class="hd">{html_logo_inline()}<div class="hd-text">
<h1>ORGATEC</h1>
<div class="tag">Cruzamento Documental · NF-e/CT-e vs OFX</div>
<div style="margin-top:8px;font-size:9pt;opacity:.85">LOCAR TRANSPORTE DE BOVINOS LTDA · CNPJ 05.509.396/0001-10 · Gerado em {agora}</div>
</div></div>
{body}
</body></html>"""


async def gerar_pdf(html_text):
    try:
        from playwright.async_api import async_playwright
        async with async_playwright() as p:
            browser = await p.chromium.launch()
            page = await browser.new_page()
            await page.set_content(html_text, wait_until="load")
            await page.pdf(
                path=str(OUT_PDF), format="A4", landscape=True,
                margin={"top": "14mm", "right": "12mm", "bottom": "14mm", "left": "12mm"},
                print_background=True,
            )
            await browser.close()
        return True
    except Exception as exc:
        print(f"PDF failed: {exc}")
        return False


async def main_async():
    print("Carregando 7.110 transacoes OFX...")
    transacoes = []
    for path in OFX_LIST:
        transacoes.extend(ler_ofx(path))
    print(f"  {len(transacoes):,} transacoes")

    print("\nProcessando 8.226 XMLs...")
    ctes, nfes = processar_xmls()
    print(f"  {len(ctes):,} CT-es + {len(nfes):,} NF-es")

    print("\nCruzando CT-es com recebimentos...")
    rec_results = cruzar_recebimentos(ctes, transacoes)
    casados = sum(1 for r in rec_results if r["status"] == "CASADO")
    print(f"  {casados:,}/{len(rec_results):,} CT-es casados ({100*casados/max(len(rec_results),1):.1f}%)")

    print("\nCruzando NF-es com pagamentos...")
    pag_results = cruzar_pagamentos(nfes, transacoes)
    casados_p = sum(1 for r in pag_results if r["status"] == "CASADO")
    print(f"  {casados_p:,}/{len(pag_results):,} NF-es casadas ({100*casados_p/max(len(pag_results),1):.1f}%)")

    print("\nGerando XLSX...")
    stats = gerar_xlsx(rec_results, pag_results, transacoes)

    print("Gerando MD...")
    md = gerar_md(stats)
    OUT_MD.write_text(md, encoding="utf-8")
    print(f"  MD:   {OUT_MD}")

    print("Gerando HTML...")
    html = gerar_html(md)
    OUT_HTML.write_text(html, encoding="utf-8")
    print(f"  HTML: {OUT_HTML}")

    print("Gerando PDF...")
    if await gerar_pdf(html):
        print(f"  PDF:  {OUT_PDF}")


if __name__ == "__main__":
    asyncio.run(main_async())
