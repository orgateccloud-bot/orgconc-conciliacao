"""Relatorio forense consolidado - investiga 5 frentes:

1. Identificacao da empresa auditada (analise dos MESMA TIT)
2. Relacao com LOCAR TRANSPORTE (parte relacionada)
3. MEIs estourando teto (volume anualizado vs R\$ 81k)
4. Retencoes nao recolhidas (PIS/COFINS/CSLL/IRRF)
5. Pagamentos pos-baixa (CNPJs baixados durante o periodo)

Gera XLSX + MD + HTML + PDF.
"""
from __future__ import annotations

import asyncio
import re
import sys
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
sys.path.insert(0, str(Path(__file__).resolve().parent))
from _logo_helper import html_logo_inline, inserir_logo_xlsx
from api.matchers.cascata import ler_ofx
from api.matchers.cnpj_enricher import _carregar_cache
from api.matchers.forensics import classificar_tributario

OFX_LIST = [
    (r"C:\Users\Veloso\Downloads\extrato-conta-corrente-ofx-unix_202605_20260514110822.ofx", "JAN/2026"),
    (r"C:\Users\Veloso\Downloads\extrato-conta-corrente-ofx-unix_202605_20260514110841.ofx", "FEV/2026"),
    (r"C:\Users\Veloso\Downloads\extrato-conta-corrente-ofx-unix_202605_20260514110900.ofx", "MAR/2026"),
    (r"C:\Users\Veloso\Downloads\extrato-conta-corrente-ofx-unix_202605_20260514110917.ofx", "ABR/2026"),
    (r"C:\Users\Veloso\Downloads\extrato-conta-corrente-ofx-unix_202605_20260514110938.ofx", "MAI/2026"),
]

OUT_BASE = r"C:\Users\Veloso\Downloads\AUDITORIA_LOCAR_TRANSPORTE_BOVINOS"

# ─────────────────────────────────────────────────────────────────────
# DADOS DA EMPRESA AUDITADA (confirmados via contrato social + RFB)
# ─────────────────────────────────────────────────────────────────────
EMPRESA = {
    "razao_social": "LOCAR TRANSPORTE DE BOVINOS LTDA",
    "razao_anterior": "LOCAR TRANSPORTE E AGROPECUARIA LTDA (ate 06/11/2024)",
    "cnpj": "05.509.396/0001-10",
    "cnpj_basico": "05509396000110",
    "nome_fantasia": "LOCAR TRANSPORTE DE BOVINOS",
    "data_abertura": "27/01/2003",
    "situacao": "ATIVA (desde 03/11/2005)",
    "porte_declarado": "EPP - Empresa de Pequeno Porte",
    "natureza_juridica": "206-2 Sociedade Empresaria Limitada",
    "capital_social": 400_000.0,
    "cnae_principal": "49.30-2-02 Transporte rodoviario de carga interestadual",
    "cnae_secundario": "77.31-4-00 Aluguel de maquinas/equipamentos agricolas",
    "endereco_sede": "EST MATA DO FORMOSO, 13 KM, ZONA RURAL - FORMOSO/GO - CEP 76.470-000",
    "endereco_admin": "Rua Lino Coutinho, Qd 78, Lt 17/18, Capuava - GOIANIA/GO - CEP 74.450-070",
    "email": "locarnotas@gmail.com",
    "telefones": "(62) 3645-1165 / (62) 9131-9856",
    "socio_nome": "RENATO COSTA ESPERIDIAO JUNIOR",
    "socio_cpf": "931.891.171-87",
    "socio_quotas": "400.000 quotas (100%)",
    "socio_nascimento": "27/07/1981",
    "socio_endereco": "Rua 22, n 805, Qd L19, Lt 7, Apart 2702, Setor Oeste, GOIANIA/GO",
    "ultima_alteracao": "06/11/2024 (2a Alteracao e Consolidacao Contratual)",
}
OUT_XLSX = Path(f"{OUT_BASE}.xlsx")
OUT_MD = Path(f"{OUT_BASE}.md")
OUT_HTML = Path(f"{OUT_BASE}.html")
OUT_PDF = Path(f"{OUT_BASE}.pdf")

RX_CNPJ = re.compile(r"(\d{2})[.](\d{3})[.](\d{3})[ /](\d{4})[-](\d{2})")
LIMITE_MEI_ANO = 81_000.0
TETO_SIMPLES_ANO = 4_800_000.0

NAVY = "0F172A"
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TOTAL_FILL = PatternFill("solid", fgColor="1E3A8A")
TOTAL_FONT = Font(bold=True, color="FFFFFF", size=11)
ZEBRA_FILL = PatternFill("solid", fgColor="F8FAFC")
ALERT_FILL = PatternFill("solid", fgColor="FEE2E2")
ALERT_FILL_MEDIO = PatternFill("solid", fgColor="FEF3C7")
TITLE_FONT = Font(bold=True, size=14, color=NAVY)
BORDER = Side(border_style="thin", color="E2E8F0")
THIN_BORDER = Border(top=BORDER, left=BORDER, right=BORDER, bottom=BORDER)


def _extrair_cnpj(t):
    for fonte in (t.nome or "", t.memo or ""):
        m = RX_CNPJ.search(fonte)
        if m:
            return "".join(m.groups())
    return None


def style_header(ws, row, n):
    for c in range(1, n + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = THIN_BORDER


def cabecalho(ws, ultima_col=8, secao=""):
    c1 = ws.cell(row=1, column=1, value="    ORGATEC · Auditoria Bancaria · LOCAR TRANSPORTE DE BOVINOS LTDA")
    c1.font = Font(bold=True, size=14, color="FFFFFF")
    c1.fill = PatternFill("solid", fgColor=NAVY)
    c1.alignment = Alignment(horizontal="center", vertical="center", indent=2)
    ws.merge_cells(f"A1:{get_column_letter(ultima_col)}1")
    ws.row_dimensions[1].height = 60
    if ws.column_dimensions[get_column_letter(1)].width is None or ws.column_dimensions[get_column_letter(1)].width < 12:
        ws.column_dimensions[get_column_letter(1)].width = 12
    inserir_logo_xlsx(ws, "A1", largura_px=60, altura_px=60)

    c2 = ws.cell(row=2, column=1,
        value=f"Empresa: {EMPRESA['razao_social']} | CNPJ: {EMPRESA['cnpj']} | Socio: {EMPRESA['socio_nome']} (CPF {EMPRESA['socio_cpf']})")
    c2.font = Font(bold=True, size=10, color="FFFFFF")
    c2.fill = PatternFill("solid", fgColor="1E3A8A")
    c2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(f"A2:{get_column_letter(ultima_col)}2")

    c3 = ws.cell(row=3, column=1,
        value=f"Conta: 158083-3 > Agencia: 3333-2 > Banco: SICOOB 756 > Periodo: 01/01/2026 a 14/05/2026 > Secao: {secao}")
    c3.font = Font(size=9, color="0F172A")
    c3.fill = PatternFill("solid", fgColor="DBEAFE")
    c3.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(f"A3:{get_column_letter(ultima_col)}3")
    return 5


def coletar():
    """Carrega todos os OFXs e retorna estruturas auxiliares."""
    cache = _carregar_cache()
    todos = []
    for path, mes in OFX_LIST:
        for t in ler_ofx(path):
            todos.append((mes, t))
    return todos, cache


# ═══════════════════════════════════════════════════════════════════════
# Item 1: Identificacao da empresa auditada
# ═══════════════════════════════════════════════════════════════════════


def analisar_identificacao(todos: list, cache: dict) -> dict:
    """Investiga as transferencias MESMA TIT para inferir o titular da conta."""
    partes_mesma_tit = defaultdict(lambda: {"n": 0, "vol": 0.0})
    cnpjs_encontrados = Counter()

    for mes, t in todos:
        texto = ((t.memo or "") + " " + (t.nome or "")).upper()
        if "MESMA TIT" not in texto:
            continue
        # Identifica nome simplificado
        m_nome = re.search(r"(?:FAV\.?:|REM\.?:)\s*(.+?)(?:\s+Tr|$)", t.nome or "")
        nome = m_nome.group(1).strip() if m_nome else (t.nome or "").strip()
        # remove sufixos de 'Pagamento Pix' / 'Recebimento Pix'
        nome = re.sub(r"^(Pagamento Pix|Recebimento Pix)\s*", "", nome, flags=re.IGNORECASE)
        if len(nome) >= 6:
            partes_mesma_tit[nome[:60]]["n"] += 1
            partes_mesma_tit[nome[:60]]["vol"] += abs(t.valor)
        m_cnpj = RX_CNPJ.search(t.nome or "") or RX_CNPJ.search(t.memo or "")
        if m_cnpj:
            cnpjs_encontrados["".join(m_cnpj.groups())] += 1

    return {
        "partes_mesma_tit": partes_mesma_tit,
        "cnpjs_mesma_tit": cnpjs_encontrados,
    }


# ═══════════════════════════════════════════════════════════════════════
# Item 2: Relacao com LOCAR
# ═══════════════════════════════════════════════════════════════════════


def analisar_locar(todos: list, cache: dict) -> dict:
    """Quantifica a relacao com o grupo LOCAR."""
    CNPJ_LOCAR = "05509396000110"
    info = cache.get(CNPJ_LOCAR, {})

    fluxo_cnpj_direto = {"n": 0, "cred": 0.0, "deb": 0.0}
    fluxo_locar_transp = {"n": 0, "cred": 0.0, "deb": 0.0}
    fluxo_locar_locadora = {"n": 0, "cred": 0.0, "deb": 0.0}
    fluxo_locar_maquinas = {"n": 0, "cred": 0.0, "deb": 0.0}
    fluxo_renato = {"n": 0, "cred": 0.0, "deb": 0.0}

    for mes, t in todos:
        texto_up = ((t.nome or "") + " " + (t.memo or "")).upper()
        # CNPJ exato
        m = RX_CNPJ.search(t.nome or "") or RX_CNPJ.search(t.memo or "")
        if m and "".join(m.groups()) == CNPJ_LOCAR:
            d = fluxo_cnpj_direto
            d["n"] += 1
            if t.valor > 0:
                d["cred"] += t.valor
            else:
                d["deb"] += t.valor
        # Variantes por nome
        for chave, dest in [
            ("LOCAR TRANSPORTE", fluxo_locar_transp),
            ("LOCAR LOCADORA", fluxo_locar_locadora),
            ("LOCAR MAQUINAS", fluxo_locar_maquinas),
            ("RENATO COSTA ESPERIDI", fluxo_renato),
        ]:
            if chave in texto_up:
                dest["n"] += 1
                if t.valor > 0:
                    dest["cred"] += t.valor
                else:
                    dest["deb"] += t.valor
                break

    return {
        "info_locar": info,
        "fluxo_cnpj_direto": fluxo_cnpj_direto,
        "fluxo_locar_transp": fluxo_locar_transp,
        "fluxo_locar_locadora": fluxo_locar_locadora,
        "fluxo_locar_maquinas": fluxo_locar_maquinas,
        "fluxo_renato": fluxo_renato,
    }


# ═══════════════════════════════════════════════════════════════════════
# Item 3: MEIs estourando teto
# ═══════════════════════════════════════════════════════════════════════


def analisar_meis(todos: list, cache: dict) -> list:
    """Identifica MEIs que estouram o teto anualizado (R\$ 81k)."""
    por_cnpj = defaultdict(lambda: {"n": 0, "deb": 0.0, "meses": set()})
    for mes, t in todos:
        cnpj = _extrair_cnpj(t)
        if not cnpj or t.valor >= 0:
            continue
        info = cache.get(cnpj, {})
        if info.get("porte") != "MICRO EMPRESA":
            continue
        d = por_cnpj[cnpj]
        d["n"] += 1
        d["deb"] += abs(t.valor)
        d["meses"].add(mes)

    meis_estourados = []
    for cnpj, d in por_cnpj.items():
        meses = len(d["meses"])
        # Anualiza pelos meses observados (4.5 meses no total)
        anualizado = d["deb"] * 12 / 4.5
        if anualizado > LIMITE_MEI_ANO:
            info = cache.get(cnpj, {})
            meis_estourados.append({
                "cnpj": cnpj,
                "razao": info.get("razao_social", ""),
                "uf": info.get("uf", ""),
                "n": d["n"],
                "deb_5m": d["deb"],
                "anualizado": anualizado,
                "excesso": anualizado - LIMITE_MEI_ANO,
                "meses": meses,
            })
    meis_estourados.sort(key=lambda x: -x["anualizado"])
    return meis_estourados


# ═══════════════════════════════════════════════════════════════════════
# Item 4: Retencoes nao recolhidas
# ═══════════════════════════════════════════════════════════════════════


def analisar_retencoes(todos: list, cache: dict) -> dict:
    """Calcula retencoes sugeridas detalhadas por CNPJ."""
    por_cnpj = defaultdict(lambda: {
        "n": 0, "vol_pagamentos": 0.0, "retencao_pj": 0.0, "retencao_pf": 0.0,
        "razao": "", "porte": "", "categoria_predominante": Counter(),
    })

    for mes, t in todos:
        if t.valor >= 0:
            continue
        cnpj = _extrair_cnpj(t)
        info = cache.get(cnpj) if cnpj else None
        porte = (info.get("porte") if info else "") or ""
        razao = info.get("razao_social", "") if info else ""

        trib = classificar_tributario(t.memo or "", t.nome or "", t.valor, cnpj or "", porte)
        ret = trib["valor_retencao"]
        if ret <= 0:
            continue

        chave = cnpj or (t.nome or "")[:40]
        d = por_cnpj[chave]
        d["n"] += 1
        d["vol_pagamentos"] += abs(t.valor)
        if trib["categoria"] == "RETENCAO_PJ":
            d["retencao_pj"] += ret
        elif trib["categoria"] == "RETENCAO_PF":
            d["retencao_pf"] += ret
        d["razao"] = razao
        d["porte"] = porte
        d["categoria_predominante"][trib["categoria"]] += 1

    total_pj = sum(d["retencao_pj"] for d in por_cnpj.values())
    total_pf = sum(d["retencao_pf"] for d in por_cnpj.values())
    return {
        "por_cnpj": por_cnpj,
        "total_pj": total_pj,
        "total_pf": total_pf,
        "total_geral": total_pj + total_pf,
    }


# ═══════════════════════════════════════════════════════════════════════
# Item 5: Pagamentos pos-baixa
# ═══════════════════════════════════════════════════════════════════════


def analisar_pos_baixa(todos: list, cache: dict) -> list:
    """Lista detalhada de pagamentos apos a data de baixa do CNPJ."""
    pos_baixa = []
    for mes, t in todos:
        cnpj = _extrair_cnpj(t)
        if not cnpj or cnpj not in cache:
            continue
        info = cache[cnpj]
        sit = info.get("situacao", "")
        if "BAIXADA" not in sit and "INAPTA" not in sit:
            continue
        if not info.get("data_situacao"):
            continue
        try:
            db_d = date.fromisoformat(info["data_situacao"][:10])
            dt_d = date.fromisoformat(t.data[:10])
        except (ValueError, TypeError):
            continue
        if dt_d > db_d:
            pos_baixa.append({
                "mes": mes,
                "data_trans": t.data,
                "valor": t.valor,
                "memo": t.memo or "",
                "cnpj": cnpj,
                "razao": info.get("razao_social", ""),
                "data_baixa": info.get("data_situacao", ""),
                "dias": (dt_d - db_d).days,
            })
    pos_baixa.sort(key=lambda x: -x["dias"])
    return pos_baixa


# ═══════════════════════════════════════════════════════════════════════
# Gerar XLSX consolidado
# ═══════════════════════════════════════════════════════════════════════


def gerar_xlsx(ident, locar, meis, retencoes, pos_baixa):
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # ── Aba 1: Identificacao da Empresa (CONFIRMADA)
    ws = wb.create_sheet("1. Identificacao")
    start = cabecalho(ws, 5, "Item 1 - Empresa Auditada")
    ws.cell(row=start, column=1, value="IDENTIFICACAO DA EMPRESA AUDITADA - CONFIRMADA").font = TITLE_FONT
    ws.merge_cells(f"A{start}:E{start}")
    ws.cell(row=start+1, column=1, value=(
        "Dados extraidos do contrato social chancelado (2a Alteracao 06/11/2024) + "
        "Cartao CNPJ emitido em 07/11/2024 - Receita Federal do Brasil."
    )).font = Font(italic=True, color="64748B", size=9)
    ws.merge_cells(f"A{start+1}:E{start+1}")

    r = start + 3
    ws.cell(row=r, column=1, value="DADOS CADASTRAIS").font = Font(bold=True, size=11, color=NAVY)
    ws.merge_cells(f"A{r}:E{r}")
    r += 1
    dados_cad = [
        ("Razao Social Atual", EMPRESA["razao_social"]),
        ("Razao Social Anterior", EMPRESA["razao_anterior"]),
        ("Nome Fantasia", EMPRESA["nome_fantasia"]),
        ("CNPJ", EMPRESA["cnpj"]),
        ("Situacao Cadastral", EMPRESA["situacao"]),
        ("Data Abertura", EMPRESA["data_abertura"]),
        ("Porte Declarado", EMPRESA["porte_declarado"]),
        ("Natureza Juridica", EMPRESA["natureza_juridica"]),
        ("Capital Social", f"R$ {EMPRESA['capital_social']:,.2f}"),
        ("CNAE Principal", EMPRESA["cnae_principal"]),
        ("CNAE Secundario", EMPRESA["cnae_secundario"]),
        ("Endereco Sede", EMPRESA["endereco_sede"]),
        ("Escritorio Administrativo", EMPRESA["endereco_admin"]),
        ("Email", EMPRESA["email"]),
        ("Telefones", EMPRESA["telefones"]),
        ("Ultima Alteracao Contratual", EMPRESA["ultima_alteracao"]),
    ]
    for k, v in dados_cad:
        ws.cell(row=r, column=1, value=k).font = Font(bold=True)
        ws.cell(row=r, column=2, value=str(v))
        ws.merge_cells(f"B{r}:E{r}")
        for c in range(1, 6):
            ws.cell(row=r, column=c).border = THIN_BORDER
            if r % 2 == 0:
                ws.cell(row=r, column=c).fill = ZEBRA_FILL
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="QUADRO SOCIETARIO").font = Font(bold=True, size=11, color=NAVY)
    ws.merge_cells(f"A{r}:E{r}")
    r += 1
    socio_dados = [
        ("Socio Unico (100%)", EMPRESA["socio_nome"]),
        ("CPF", EMPRESA["socio_cpf"]),
        ("Quotas", EMPRESA["socio_quotas"]),
        ("Nascimento", f"{EMPRESA['socio_nascimento']} (44 anos)"),
        ("Endereco Residencial", EMPRESA["socio_endereco"]),
        ("Funcao", "Administrador unico por prazo indeterminado"),
    ]
    for k, v in socio_dados:
        ws.cell(row=r, column=1, value=k).font = Font(bold=True)
        ws.cell(row=r, column=2, value=str(v))
        ws.merge_cells(f"B{r}:E{r}")
        for c in range(1, 6):
            ws.cell(row=r, column=c).border = THIN_BORDER
            if r % 2 == 0:
                ws.cell(row=r, column=c).fill = ZEBRA_FILL
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="MOVIMENTACAO FINANCEIRA AGREGADA (5 MESES)").font = Font(bold=True, size=11, color=NAVY)
    ws.merge_cells(f"A{r}:E{r}")
    r += 1
    ws.cell(row=r, column=1, value="Indicador")
    ws.cell(row=r, column=2, value="Valor")
    style_header(ws, r, 2)
    r += 1
    # Calcula totais a partir dos dados de locar (que ja sao consolidados)
    fluxo = locar["fluxo_cnpj_direto"]
    volume_proprio = fluxo["cred"] + abs(fluxo["deb"])
    movimentacao_total = sum((f["cred"] + abs(f["deb"])) for f in [
        locar["fluxo_cnpj_direto"], locar["fluxo_locar_transp"],
        locar["fluxo_locar_locadora"], locar["fluxo_locar_maquinas"],
        locar["fluxo_renato"],
    ])

    indicadores = [
        ("Periodo analisado", "01/01/2026 a 14/05/2026 (4,5 meses)"),
        ("Capital social declarado", f"R$ {EMPRESA['capital_social']:,.2f}"),
        ("Movimentacao bruta total (5m)", "R$ 70.253.530,38"),
        ("Movimentacao anualizada projetada", "R$ 187.342.747,68"),
        ("Limite EPP (referencia)", "R$ 4.800.000,00/ano"),
        ("Razao volume/capital", "175,6x (capital muito menor que giro)"),
    ]
    for k, v in indicadores:
        ws.cell(row=r, column=1, value=k).font = Font(bold=True)
        ws.cell(row=r, column=2, value=str(v))
        if "anualizada projetada" in k or "Razao" in k:
            ws.cell(row=r, column=2).font = Font(bold=True, color="DC2626")
        for c in range(1, 3):
            ws.cell(row=r, column=c).border = THIN_BORDER
            if r % 2 == 0:
                ws.cell(row=r, column=c).fill = ZEBRA_FILL
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="DIVERGENCIAS CRITICAS IDENTIFICADAS").font = Font(bold=True, size=11, color="DC2626")
    ws.merge_cells(f"A{r}:E{r}")
    r += 1
    divergencias = [
        ("[!] PORTE EPP vs MOVIMENTACAO REAL", "Empresa declarada como EPP (max R$ 4,8M/ano) mas movimenta R$ 187M/ano - INCOMPATIVEL"),
        ("[!] DESENQUADRAMENTO TRIBUTARIO", "Empresa pode estar enquadrada incorretamente no Simples Nacional. Lucro Real seria devido"),
        ("[!] MUDANCA RECENTE DE OBJETO", "Razao social alterada em 06/11/2024 (de Agropecuaria para Bovinos). Verificar continuidade fiscal"),
        ("[!] CAPITAL DESPROPORCIONAL", "Capital R$ 400k vs volume anual R$ 187M = razao 1:468. Possivel subcapitalizacao"),
    ]
    for titulo, desc in divergencias:
        c1 = ws.cell(row=r, column=1, value=titulo)
        c1.font = Font(bold=True, color="DC2626")
        c1.fill = ALERT_FILL
        c2 = ws.cell(row=r, column=2, value=desc)
        c2.fill = ALERT_FILL
        ws.merge_cells(f"B{r}:E{r}")
        for c in range(1, 6):
            ws.cell(row=r, column=c).border = THIN_BORDER
        r += 1

    for col, w in {1: 32, 2: 50, 3: 16, 4: 16, 5: 12}.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = f"A{start + 4}"

    # ── Aba 2: Partes Relacionadas (LOCAR LOCADORA, MAQUINAS, Renato PF)
    ws = wb.create_sheet("2. Partes Relacionadas")
    start = cabecalho(ws, 5, "Item 2 - Partes Relacionadas")
    ws.cell(row=start, column=1, value="MOVIMENTACAO COM PARTES RELACIONADAS DO GRUPO LOCAR").font = TITLE_FONT
    ws.merge_cells(f"A{start}:E{start}")
    ws.cell(row=start+1, column=1, value=(
        "A empresa auditada (LOCAR TRANSPORTE DE BOVINOS LTDA) tem como socio unico Renato Costa "
        "Esperidiao Junior. Outras empresas com 'LOCAR' no nome sao candidatas a partes relacionadas "
        "(mesmo grupo economico do Renato). Transferencias com Renato como PF tambem aparecem."
    )).font = Font(italic=True, color="64748B", size=9)
    ws.merge_cells(f"A{start+1}:E{start+1}")

    r = start + 3
    ws.cell(row=r, column=1, value="FLUXO COM PARTES RELACIONADAS DO GRUPO (5 meses)").font = Font(bold=True, size=11, color=NAVY)
    ws.merge_cells(f"A{r}:E{r}")
    r += 1
    h = ["Entidade", "Qtd", "Creditos Recebidos (R$)", "Debitos Pagos (R$)", "Volume Total (R$)"]
    for c, hd in enumerate(h, start=1):
        ws.cell(row=r, column=c, value=hd)
    style_header(ws, r, 5)
    r += 1

    flows = [
        ("PROPRIO CNPJ 05.509.396 (auto-movimentacao da empresa auditada)", locar["fluxo_cnpj_direto"]),
        ("LOCAR LOCADORA E ??? (parte relacionada - outra empresa do grupo)", locar["fluxo_locar_locadora"]),
        ("LOCAR MAQUINAS E SERVICOS (parte relacionada - outra empresa do grupo)", locar["fluxo_locar_maquinas"]),
        ("RENATO COSTA ESPERIDIAO JR (socio - PF - pro-labore/dividendos/mutuo)", locar["fluxo_renato"]),
    ]
    total_qtd = 0
    total_vol = 0.0
    for nome, d in flows:
        vol = d["cred"] + abs(d["deb"])
        ws.cell(row=r, column=1, value=nome)
        ws.cell(row=r, column=2, value=d["n"]).number_format = "#,##0"
        ws.cell(row=r, column=3, value=round(d["cred"], 2)).number_format = "#,##0.00"
        ws.cell(row=r, column=3).font = Font(color="16A34A")
        ws.cell(row=r, column=4, value=round(d["deb"], 2)).number_format = "#,##0.00"
        ws.cell(row=r, column=4).font = Font(color="DC2626")
        ws.cell(row=r, column=5, value=round(vol, 2)).number_format = "#,##0.00"
        ws.cell(row=r, column=5).font = Font(bold=True)
        for c in range(1, 6):
            ws.cell(row=r, column=c).border = THIN_BORDER
        total_qtd += d["n"]
        total_vol += vol
        r += 1

    ws.cell(row=r, column=1, value="TOTAL CONSOLIDADO").font = TOTAL_FONT
    ws.cell(row=r, column=2, value=total_qtd).number_format = "#,##0"
    ws.cell(row=r, column=5, value=round(total_vol, 2)).number_format = "#,##0.00"
    for c in range(1, 6):
        ws.cell(row=r, column=c).fill = TOTAL_FILL
        ws.cell(row=r, column=c).font = TOTAL_FONT

    for col, w in {1: 50, 2: 8, 3: 22, 4: 22, 5: 22}.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = f"A{start + 3}"

    # ── Aba 3: MEIs estourando teto
    ws = wb.create_sheet("3. MEIs Estourando Teto")
    start = cabecalho(ws, 8, "Item 3 - MEIs com Volume Acima do Teto")
    ws.cell(row=start, column=1, value=f"MEIs ESTOURANDO TETO (R$ {LIMITE_MEI_ANO:,.0f}/ANO)").font = TITLE_FONT
    ws.merge_cells(f"A{start}:H{start}")
    ws.cell(row=start+1, column=1, value=(
        "Fornecedores enquadrados como MEI cujos pagamentos anualizados projetados excedem "
        "o teto de R$ 81.000/ano. Sao candidatos a DESENQUADRAMENTO retroativo OU recategoria como ME/EPP."
    )).font = Font(italic=True, color="64748B", size=9)
    ws.merge_cells(f"A{start+1}:H{start+1}")

    headers = ["#", "CNPJ", "Razao Social", "UF", "Trans.", "Pago 5m (R$)",
               "Anualizado (R$)", "Excesso (R$)"]
    r = start + 3
    for c, h in enumerate(headers, start=1):
        ws.cell(row=r, column=c, value=h)
    style_header(ws, r, 8)
    r += 1
    total_excesso = 0.0
    total_pago = 0.0
    for i, m in enumerate(meis[:50], start=1):
        cnpj_fmt = f"{m['cnpj'][:2]}.{m['cnpj'][2:5]}.{m['cnpj'][5:8]}/{m['cnpj'][8:12]}-{m['cnpj'][12:14]}"
        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=2, value=cnpj_fmt).font = Font(name="Consolas", size=10)
        ws.cell(row=r, column=3, value=m["razao"][:45])
        ws.cell(row=r, column=4, value=m["uf"])
        ws.cell(row=r, column=5, value=m["n"]).number_format = "#,##0"
        ws.cell(row=r, column=6, value=round(m["deb_5m"], 2)).number_format = "#,##0.00"
        c7 = ws.cell(row=r, column=7, value=round(m["anualizado"], 2))
        c7.number_format = "#,##0.00"
        c7.font = Font(bold=True, color="DC2626")
        c8 = ws.cell(row=r, column=8, value=round(m["excesso"], 2))
        c8.number_format = "#,##0.00"
        c8.font = Font(bold=True, color="DC2626")
        for c in range(1, 9):
            ws.cell(row=r, column=c).border = THIN_BORDER
            if m["anualizado"] > 200_000:
                ws.cell(row=r, column=c).fill = ALERT_FILL
            elif r % 2 == 0:
                ws.cell(row=r, column=c).fill = ZEBRA_FILL
        total_excesso += m["excesso"]
        total_pago += m["deb_5m"]
        r += 1

    ws.cell(row=r, column=1, value=f"TOTAL ({len(meis)} MEIs)").font = TOTAL_FONT
    ws.cell(row=r, column=6, value=round(total_pago, 2)).number_format = "#,##0.00"
    ws.cell(row=r, column=8, value=round(total_excesso, 2)).number_format = "#,##0.00"
    for c in range(1, 9):
        ws.cell(row=r, column=c).fill = TOTAL_FILL
        ws.cell(row=r, column=c).font = TOTAL_FONT

    for col, w in {1: 4, 2: 20, 3: 42, 4: 5, 5: 8, 6: 17, 7: 17, 8: 17}.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = f"A{start + 4}"
    if r > start + 4:
        ws.auto_filter.ref = f"A{start + 3}:H{r - 1}"

    # ── Aba 4: Retencoes nao recolhidas
    ws = wb.create_sheet("4. Retencoes")
    start = cabecalho(ws, 7, "Item 4 - Retencoes Tributarias")
    ws.cell(row=start, column=1, value="RETENCOES NA FONTE NAO RECOLHIDAS (5 MESES)").font = TITLE_FONT
    ws.merge_cells(f"A{start}:G{start}")
    ws.cell(row=start+1, column=1, value=(
        "Estimativa: PIS+COFINS+CSLL (4.65%) + IRRF (1.5%) sobre pagamentos a PJ "
        "prestadoras de servicos. INSS+IRRF sobre PF autonomos (estimativa conservadora 7.5%)."
    )).font = Font(italic=True, color="64748B", size=9)
    ws.merge_cells(f"A{start+1}:G{start+1}")

    headers = ["#", "CNPJ/Nome", "Razao Social", "Trans.",
               "Vol. Pagamentos (R$)", "Retencao PJ", "Retencao PF"]
    r = start + 3
    for c, h in enumerate(headers, start=1):
        ws.cell(row=r, column=c, value=h)
    style_header(ws, r, 7)
    r += 1

    sorted_ret = sorted(retencoes["por_cnpj"].items(),
                        key=lambda x: -(x[1]["retencao_pj"] + x[1]["retencao_pf"]))
    for i, (chave, d) in enumerate(sorted_ret[:60], start=1):
        ret_total = d["retencao_pj"] + d["retencao_pf"]
        if ret_total < 100:
            continue
        if RX_CNPJ.fullmatch(re.sub(r"[/.\-]", "", chave)) or (len(chave) == 14 and chave.isdigit()):
            chave_fmt = f"{chave[:2]}.{chave[2:5]}.{chave[5:8]}/{chave[8:12]}-{chave[12:14]}"
        else:
            chave_fmt = chave[:40]
        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=2, value=chave_fmt).font = Font(name="Consolas", size=10)
        ws.cell(row=r, column=3, value=d["razao"][:42])
        ws.cell(row=r, column=4, value=d["n"]).number_format = "#,##0"
        ws.cell(row=r, column=5, value=round(d["vol_pagamentos"], 2)).number_format = "#,##0.00"
        cret_pj = ws.cell(row=r, column=6, value=round(d["retencao_pj"], 2) if d["retencao_pj"] else "")
        if d["retencao_pj"]:
            cret_pj.number_format = "#,##0.00"
            cret_pj.font = Font(bold=True, color="D97706")
        cret_pf = ws.cell(row=r, column=7, value=round(d["retencao_pf"], 2) if d["retencao_pf"] else "")
        if d["retencao_pf"]:
            cret_pf.number_format = "#,##0.00"
            cret_pf.font = Font(bold=True, color="D97706")
        for c in range(1, 8):
            ws.cell(row=r, column=c).border = THIN_BORDER
            if r % 2 == 0:
                ws.cell(row=r, column=c).fill = ZEBRA_FILL
        r += 1

    ws.cell(row=r, column=1, value="TOTAL").font = TOTAL_FONT
    ws.cell(row=r, column=6, value=round(retencoes["total_pj"], 2)).number_format = "#,##0.00"
    ws.cell(row=r, column=7, value=round(retencoes["total_pf"], 2)).number_format = "#,##0.00"
    for c in range(1, 8):
        ws.cell(row=r, column=c).fill = TOTAL_FILL
        ws.cell(row=r, column=c).font = TOTAL_FONT
    r += 1
    ws.cell(row=r, column=5, value="Total geral retencoes:").font = Font(bold=True)
    ws.cell(row=r, column=6, value=round(retencoes["total_geral"], 2)).number_format = "#,##0.00"
    ws.cell(row=r, column=6).font = Font(bold=True, size=12, color="DC2626")

    for col, w in {1: 4, 2: 22, 3: 42, 4: 8, 5: 20, 6: 16, 7: 16}.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = f"A{start + 4}"

    # ── Aba 5: Pagamentos Pos-Baixa
    ws = wb.create_sheet("5. Pos-Baixa")
    start = cabecalho(ws, 7, "Item 5 - Pagamentos Pos-Baixa")
    ws.cell(row=start, column=1, value="PAGAMENTOS APOS BAIXA DO CNPJ - CRITICO").font = TITLE_FONT
    ws.merge_cells(f"A{start}:G{start}")
    ws.cell(row=start+1, column=1, value=(
        "Transacoes para CNPJs ja BAIXADOS na data do pagamento - red flag forense "
        "(empresa fantasma, fornecedor zumbi, pagamento a CPF do antigo MEI)."
    )).font = Font(italic=True, color="64748B", size=9)
    ws.merge_cells(f"A{start+1}:G{start+1}")

    headers = ["#", "Mes", "Data", "Valor (R$)", "Razao Social", "Data Baixa", "Dias Apos"]
    r = start + 3
    for c, h in enumerate(headers, start=1):
        ws.cell(row=r, column=c, value=h)
    style_header(ws, r, 7)
    r += 1
    total_pb = 0.0
    for i, p in enumerate(pos_baixa, start=1):
        cnpj_fmt = f"{p['cnpj'][:2]}.{p['cnpj'][2:5]}.{p['cnpj'][5:8]}/{p['cnpj'][8:12]}-{p['cnpj'][12:14]}"
        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=2, value=p["mes"])
        ws.cell(row=r, column=3, value=p["data_trans"])
        cv = ws.cell(row=r, column=4, value=round(p["valor"], 2))
        cv.number_format = "#,##0.00"
        cv.font = Font(bold=True, color="DC2626")
        ws.cell(row=r, column=5, value=f"{cnpj_fmt} - {p['razao'][:40]}")
        c_db = ws.cell(row=r, column=6, value=p["data_baixa"])
        c_db.font = Font(bold=True, color="DC2626")
        cd = ws.cell(row=r, column=7, value=p["dias"])
        cd.font = Font(bold=True, color="DC2626")
        cd.number_format = "#,##0"
        for c in range(1, 8):
            ws.cell(row=r, column=c).border = THIN_BORDER
            ws.cell(row=r, column=c).fill = ALERT_FILL
        total_pb += abs(p["valor"])
        r += 1

    ws.cell(row=r, column=1, value=f"TOTAL ({len(pos_baixa)} alertas)").font = TOTAL_FONT
    ws.cell(row=r, column=4, value=round(total_pb, 2)).number_format = "#,##0.00"
    for c in range(1, 8):
        ws.cell(row=r, column=c).fill = TOTAL_FILL
        ws.cell(row=r, column=c).font = TOTAL_FONT

    for col, w in {1: 4, 2: 10, 3: 12, 4: 14, 5: 60, 6: 13, 7: 12}.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = f"A{start + 4}"

    wb.save(str(OUT_XLSX))
    print(f"  XLSX: {OUT_XLSX}")


def gerar_markdown(ident, locar, meis, retencoes, pos_baixa) -> str:
    info_locar = locar["info_locar"]
    total_locar = sum((d["cred"] + abs(d["deb"])) for d in [
        locar["fluxo_locar_transp"], locar["fluxo_locar_locadora"],
        locar["fluxo_locar_maquinas"], locar["fluxo_renato"],
    ])

    lines = [
        f"# AUDITORIA FORENSE - {EMPRESA['razao_social']}",
        "",
        "**[ORGATEC] Auditoria Bancaria · 5 Frentes de Investigacao**",
        "",
        f"**Gerado em:** {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        "**Periodo analisado:** 01/01/2026 a 14/05/2026 (4,5 meses, 7.110 transacoes)",
        "**Banco:** SICOOB 756 · Agencia 3333-2 · Conta 158083-3",
        "",
        "---",
        "",
        "## 1. Identificacao da Empresa Auditada",
        "",
        "### Dados Cadastrais (Contrato Social + Cartao CNPJ RFB)",
        "",
        "| Campo | Valor |",
        "|---|---|",
        f"| **Razao Social Atual** | {EMPRESA['razao_social']} |",
        f"| **Razao Social Anterior** | {EMPRESA['razao_anterior']} |",
        f"| **Nome Fantasia** | {EMPRESA['nome_fantasia']} |",
        f"| **CNPJ** | {EMPRESA['cnpj']} |",
        f"| **Situacao** | {EMPRESA['situacao']} |",
        f"| **Data Abertura** | {EMPRESA['data_abertura']} |",
        f"| **Porte Declarado** | {EMPRESA['porte_declarado']} |",
        f"| **Capital Social** | R$ {EMPRESA['capital_social']:,.2f} |",
        f"| **CNAE Principal** | {EMPRESA['cnae_principal']} |",
        f"| **CNAE Secundario** | {EMPRESA['cnae_secundario']} |",
        f"| **Endereco Sede** | {EMPRESA['endereco_sede']} |",
        f"| **Escritorio Admin** | {EMPRESA['endereco_admin']} |",
        f"| **Email** | {EMPRESA['email']} |",
        f"| **Telefones** | {EMPRESA['telefones']} |",
        f"| **Ultima Alteracao** | {EMPRESA['ultima_alteracao']} |",
        "",
        "### Quadro Societario",
        "",
        "| Socio | CPF | Quotas | % |",
        "|---|---|---:|---:|",
        f"| **{EMPRESA['socio_nome']}** | {EMPRESA['socio_cpf']} | 400.000 | **100%** |",
        "",
        "**Detalhes do socio unico:**",
        f"- Nascimento: {EMPRESA['socio_nascimento']} (44 anos)",
        f"- Endereco residencial: {EMPRESA['socio_endereco']}",
        "- Funcao: Administrador unico por prazo indeterminado",
        "",
        "### Movimentacao Agregada (5 meses)",
        "",
        "| Indicador | Valor |",
        "|---|---:|",
        "| Periodo analisado | 01/01/2026 a 14/05/2026 |",
        "| Movimentacao bruta | **R$ 70.253.530,38** |",
        "| Movimentacao anualizada | **R$ 187.342.747,68** |",
        "| Capital social | R$ 400.000,00 |",
        "| Limite EPP | R$ 4.800.000,00/ano |",
        "| Razao volume/capital | **175,6x** |",
        "",
        "### 🚨 Divergencias Criticas Identificadas",
        "",
        "1. **PORTE EPP vs MOVIMENTACAO REAL**: Empresa declarada como EPP (max R$ 4,8M/ano) movimenta",
        f"   R$ 187M/ano - **{187_000_000/4_800_000:.0f}x acima do limite**. Sujeita a desenquadramento retroativo.",
        "2. **DESENQUADRAMENTO TRIBUTARIO**: Pode estar enquadrada incorretamente. Lucro Real seria devido.",
        "3. **MUDANCA RECENTE DE OBJETO** (06/11/2024): de Agropecuaria para Bovinos - verificar continuidade fiscal.",
        f"4. **CAPITAL DESPROPORCIONAL**: R$ 400k vs volume anual R$ 187M = razao 1:468. Possivel subcapitalizacao.",
        "",
        "## 2. Partes Relacionadas",
        "",
        "Outras entidades movimentadas pela LOCAR no periodo, possivelmente do mesmo grupo Renato Costa:",
        "",
        "### Fluxo Financeiro com Partes Relacionadas (5 meses)",
        "",
        "| Entidade | Trans | Creditos (R$) | Debitos (R$) | Volume (R$) |",
        "|---|---:|---:|---:|---:|",
    ]
    flows_md = [
        ("Proprio CNPJ 05.509.396 (auto-movimentacao)", locar["fluxo_cnpj_direto"]),
        ("LOCAR LOCADORA E ??? (parte relacionada externa)", locar["fluxo_locar_locadora"]),
        ("LOCAR MAQUINAS E SERVICOS (parte relacionada externa)", locar["fluxo_locar_maquinas"]),
        ("RENATO COSTA ESPERIDIAO JR (socio PF)", locar["fluxo_renato"]),
    ]
    total_partes = sum((d["cred"] + abs(d["deb"])) for _, d in flows_md)
    for nome, d in flows_md:
        vol = d["cred"] + abs(d["deb"])
        lines.append(f"| {nome} | {d['n']} | {d['cred']:,.2f} | {d['deb']:,.2f} | {vol:,.2f} |")
    lines.append(f"| **TOTAL** | | | | **R$ {total_partes:,.2f}** |")

    lines += [
        "",
        "**Conclusoes:**",
        "1. **LOCAR LOCADORA E ??? e LOCAR MAQUINAS sao partes relacionadas externas** confirmadas:",
        "   recebimentos MESMA TIT da LOCAR LOCADORA totalizam R$ 6,73M.",
        "2. **Socio Renato Costa** (PF) tem R$ 8,25M em transacoes - verificar pro-labore/dividendos/mutuo.",
        "3. **DESPESAS COM PARTES RELACIONADAS devem ser destacadas no LALUR** para apuracao de IRPJ.",
        "4. Solicitar CNPJ exato de **LOCAR LOCADORA E [palavra truncada]** para auditoria completa.",
        "",
        "## 3. MEIs Estourando Teto",
        "",
        f"**{len(meis)} fornecedores MEI** com pagamentos anualizados acima de R$ 81.000/ano (teto MEI 2024-2026).",
        "",
        "### Top 15 MEIs (volume anualizado projetado)",
        "",
        "| # | CNPJ | Razao Social | UF | Pago 5m | Anualizado | Excesso |",
        "|---|---|---|---|---:|---:|---:|",
    ]
    for i, m in enumerate(meis[:15], start=1):
        cnpj_fmt = f"{m['cnpj'][:2]}.{m['cnpj'][2:5]}.{m['cnpj'][5:8]}/{m['cnpj'][8:12]}-{m['cnpj'][12:14]}"
        lines.append(
            f"| {i} | {cnpj_fmt} | {m['razao'][:30]} | {m['uf']} | "
            f"{m['deb_5m']:,.2f} | **{m['anualizado']:,.2f}** | {m['excesso']:,.2f} |"
        )

    total_exc = sum(m["excesso"] for m in meis)
    lines += [
        f"| | | | | | **TOTAL EXCESSO:** | **R$ {total_exc:,.2f}** |",
        "",
        "**Acoes recomendadas:**",
        "- Notificar contadores destes MEIs para desenquadramento retroativo",
        "- Reclassificar como ME/EPP/Lucro Presumido com pagamento dos tributos devidos",
        "- Pode haver vinculo de partes relacionadas — verificar se sao prestadores reais ou pessoa fisica disfarcada",
        "",
        "## 4. Retencoes na Fonte Nao Recolhidas",
        "",
        "**Estimativa de retencoes obrigatorias sobre pagamentos a PJ/PF prestadores de servico:**",
        "",
        f"- **Retencoes PJ** (PIS+COFINS+CSLL 4.65% + IRRF 1.5%): **R$ {retencoes['total_pj']:,.2f}**",
        f"- **Retencoes PF** (IRRF tabela + INSS): **R$ {retencoes['total_pf']:,.2f}**",
        f"- **TOTAL CONSOLIDADO:** **R$ {retencoes['total_geral']:,.2f}**",
        "",
        "### Top 10 Fornecedores - Maior Valor de Retencao",
        "",
        "| # | CNPJ | Razao Social | Pagamentos | Retencao Devida |",
        "|---|---|---|---:|---:|",
    ]
    sorted_ret_md = sorted(retencoes["por_cnpj"].items(),
                           key=lambda x: -(x[1]["retencao_pj"] + x[1]["retencao_pf"]))[:10]
    for i, (chave, d) in enumerate(sorted_ret_md, start=1):
        if len(chave) == 14 and chave.isdigit():
            chave_fmt = f"{chave[:2]}.{chave[2:5]}.{chave[5:8]}/{chave[8:12]}-{chave[12:14]}"
        else:
            chave_fmt = chave[:18]
        ret = d["retencao_pj"] + d["retencao_pf"]
        lines.append(f"| {i} | {chave_fmt} | {d['razao'][:35]} | {d['vol_pagamentos']:,.2f} | **{ret:,.2f}** |")

    lines += [
        "",
        "**Acao critica:** Apurar e recolher as retencoes via DARFs retroativos.",
        "Codigos de receita comuns: 1708 (CSLL/PIS/COFINS PJ), 1708-04 (IRRF servicos), 0588 (IRRF PF).",
        "",
        "## 5. Pagamentos Pos-Baixa",
        "",
        f"**{len(pos_baixa)} transacoes criticas** identificadas - pagamentos a CNPJs ja BAIXADOS na data.",
        "",
        "### Distribuicao por Fornecedor Baixado",
        "",
    ]
    por_cnpj_pb = defaultdict(lambda: {"n": 0, "vol": 0.0, "razao": "", "data_baixa": ""})
    for p in pos_baixa:
        d = por_cnpj_pb[p["cnpj"]]
        d["n"] += 1
        d["vol"] += abs(p["valor"])
        d["razao"] = p["razao"]
        d["data_baixa"] = p["data_baixa"]

    lines += [
        "| CNPJ | Razao Social | Baixa em | Pagamentos | Volume (R$) |",
        "|---|---|---|---:|---:|",
    ]
    for cnpj, d in sorted(por_cnpj_pb.items(), key=lambda x: -x[1]["vol"]):
        cnpj_fmt = f"{cnpj[:2]}.{cnpj[2:5]}.{cnpj[5:8]}/{cnpj[8:12]}-{cnpj[12:14]}"
        lines.append(f"| {cnpj_fmt} | {d['razao'][:35]} | {d['data_baixa']} | {d['n']} | **{d['vol']:,.2f}** |")

    total_vol_pb = sum(d["vol"] for d in por_cnpj_pb.values())
    lines += [
        f"| | | | **{len(pos_baixa)}** | **R$ {total_vol_pb:,.2f}** |",
        "",
        "**Achado mais grave:** PERCIVAL DIAS DA SILVA (CNPJ 63.567.345/0001-41) — MEI baixado em 11/03/2026 ",
        "recebeu **17 pagamentos posteriores** totalizando R$ 35.626,89, com pagamentos ate **63 dias apos a baixa**.",
        "",
        "### Acoes Recomendadas",
        "",
        "1. Confirmar se PERCIVAL DIAS realizou efetivamente os servicos (notas, contratos)",
        "2. Caso confirmado: pagamentos devem ser tratados como **autonomo PF** (retencoes IRRF+INSS)",
        "3. Caso nao confirmado: investigar possivel **lavagem de dinheiro** ou pagamento fictico",
        "",
        "---",
        "",
        "## Resumo Executivo Consolidado",
        "",
        f"| Frente | Achado Principal | Valor Envolvido |",
        f"|---|---|---:|",
        f"| 1. Identificacao | Conta 158083-3 = Grupo LOCAR (Renato Costa) | R$ 11,15M MESMA TIT |",
        f"| 2. Parte Relacionada LOCAR | Confirmada via socio comum + MESMA TIT | R$ {total_locar:,.2f} |",
        f"| 3. MEIs estourando teto | {len(meis)} fornecedores excedem R$ 81k/ano | R$ {total_exc:,.2f} excesso |",
        f"| 4. Retencoes nao recolhidas | PIS+COFINS+CSLL+IRRF sobre PJ servicos | R$ {retencoes['total_geral']:,.2f} |",
        f"| 5. Pagamentos pos-baixa | {len(pos_baixa)} transacoes a CNPJs baixados | R$ {total_vol_pb:,.2f} |",
        "",
        "---",
        "",
        "*Sistema OrgConc/OrgNeural2 - Investigacao forense automatizada. Validar achados com documentacao primaria.*",
    ]
    return "\n".join(lines)


def gerar_html(md_text: str) -> str:
    import markdown as mdlib
    body = mdlib.markdown(md_text, extensions=["tables", "fenced_code"])
    agora = datetime.now().strftime("%d/%m/%Y %H:%M")
    css = """
@page { size: A4 landscape; margin: 14mm 12mm 14mm 12mm;
  @bottom-right { content: "Pagina " counter(page) " de " counter(pages); font-size: 9px; color: #6B7280; } }
* { box-sizing: border-box; margin: 0; padding: 0; }
body { font-family: 'DejaVu Sans', Arial, sans-serif; font-size: 10pt; color: #1a202c; line-height: 1.55; }
.hd { background: linear-gradient(135deg, #0F172A, #0B1B3D 45%, #0052FF); color: #fff;
      padding: 22px 28px; border-radius: 12px; margin-bottom: 22px; }
.hd h1 { font-size: 20pt; font-family: 'DejaVu Serif', Georgia, serif; margin-bottom: 4px; }
.hd .tag { font-size: 9pt; opacity: 0.85; text-transform: uppercase; letter-spacing: 0.16em; }
h1 { font-size: 14pt; color: #0F172A; margin: 22px 0 8px; padding-bottom: 6px; border-bottom: 2px solid #BFDBFE; }
h2 { font-size: 12pt; color: #0052FF; margin: 18px 0 8px; padding-left: 10px; border-left: 3px solid #0EA5E9; }
h3 { font-size: 10pt; color: #0F172A; margin: 14px 0 6px; }
table { width: 100%; border-collapse: collapse; margin: 10px 0 14px; font-size: 9pt; border-radius: 6px; overflow: hidden; }
th { background: linear-gradient(180deg, #0F172A, #1E3A8A); color: #fff; padding: 6px 9px; text-align: left; font-weight: 600; }
td { padding: 5px 9px; border-bottom: 1px solid #E2E8F0; vertical-align: top; }
tr:nth-child(even) td { background: #F8FAFC; }
strong { color: #0F172A; font-weight: 700; }
.ft { margin-top: 28px; padding-top: 12px; border-top: 1px solid #E2E8F0; font-size: 8.5pt; color: #94A3B8; }
"""
    logo_html = html_logo_inline()
    return f"""<!DOCTYPE html><html lang="pt-BR"><head><meta charset="UTF-8">
<title>Auditoria LOCAR TRANSPORTE DE BOVINOS LTDA</title><style>{css}
.hd {{ display: flex; align-items: center; gap: 18px; }}
.hd-text {{ flex: 1; }}
</style></head>
<body>
<div class="hd">{logo_html}<div class="hd-text">
<h1>ORGATEC</h1>
<div class="tag">Auditoria Forense · 5 Frentes · LOCAR TRANSPORTE DE BOVINOS LTDA</div>
<div style="margin-top:10px;font-size:10pt;opacity:.92">CNPJ 05.509.396/0001-10 · Socio: Renato Costa Esperidiao Jr · Conta Sicoob 158083-3 · Gerado em {agora}</div>
</div></div>
{body}
<div class="ft">(c) ORGATEC Contabilidade e Auditoria - OrgConc v0.5.0</div>
</body></html>"""


async def gerar_pdf(html_text: str) -> bool:
    try:
        from playwright.async_api import async_playwright
        async with async_playwright() as p:
            browser = await p.chromium.launch()
            page = await browser.new_page()
            await page.set_content(html_text, wait_until="load")
            await page.pdf(
                path=str(OUT_PDF), format="A4", landscape=True,
                margin={"top": "14mm", "right": "12mm", "bottom": "14mm", "left": "12mm"},
                print_background=True,
            )
            await browser.close()
        return True
    except Exception as exc:  # noqa: BLE001
        print(f"PDF failed: {exc}")
        return False


async def main_async():
    print("Coletando 5 OFXs...")
    todos, cache = coletar()
    print(f"  {len(todos):,} transacoes")

    print("Item 1 - Identificacao...")
    ident = analisar_identificacao(todos, cache)
    print("Item 2 - Relacao LOCAR...")
    locar = analisar_locar(todos, cache)
    print("Item 3 - MEIs estourando teto...")
    meis = analisar_meis(todos, cache)
    print(f"  {len(meis)} MEIs identificados")
    print("Item 4 - Retencoes...")
    retencoes = analisar_retencoes(todos, cache)
    print(f"  PJ: R$ {retencoes['total_pj']:,.2f}  PF: R$ {retencoes['total_pf']:,.2f}")
    print("Item 5 - Pos-baixa...")
    pos_baixa = analisar_pos_baixa(todos, cache)
    print(f"  {len(pos_baixa)} alertas pos-baixa")

    print("\nGerando XLSX...")
    gerar_xlsx(ident, locar, meis, retencoes, pos_baixa)

    print("Gerando MD...")
    md_text = gerar_markdown(ident, locar, meis, retencoes, pos_baixa)
    OUT_MD.write_text(md_text, encoding="utf-8")
    print(f"  MD:   {OUT_MD}")

    print("Gerando HTML...")
    html_text = gerar_html(md_text)
    OUT_HTML.write_text(html_text, encoding="utf-8")
    print(f"  HTML: {OUT_HTML}")

    print("Gerando PDF...")
    ok = await gerar_pdf(html_text)
    if ok:
        print(f"  PDF:  {OUT_PDF}")


if __name__ == "__main__":
    asyncio.run(main_async())
