"""Relatorio integrado LOCAR TRANSPORTE: combina o modelo MENSAL com o FORENSE.

Funde:
- AUDIT_LOCAR_158083-3_*.xlsx (modelo mensal: Resumo, Transacoes, Disposicoes,
  Risk, CNPJs, Partes Relacionadas, Status Tributario - 7 abas por mes)
- AUDITORIA_LOCAR_TRANSPORTE_BOVINOS.xlsx (modelo forense: Identificacao,
  Partes Relacionadas, MEIs, Retencoes, Pos-Baixa - 5 abas)

Saida unica em 11 abas + PDF + HTML + MD - relatorio completo cobrindo
identificacao cadastral, movimentacao detalhada, classificacao forense,
analise tributaria e achados criticos.
"""
from __future__ import annotations

import asyncio
import re
import sys
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
sys.path.insert(0, str(Path(__file__).resolve().parent))

from _logo_helper import html_logo_inline, inserir_logo_xlsx
from api.matchers.cascata import Disposicao, classificar, ler_ofx
from api.matchers.cnpj_enricher import _carregar_cache, _salvar_cache, enriquecer_um
from api.matchers.forensics import (
    calcular_agregados,
    calcular_risk_score,
    classificar_tributario,
    detectar_carrossel,
    detectar_meio,
    detectar_primeira_vez,
    detectar_smurfing,
    detectar_valor_redondo,
    hash_linha,
    periodo_fiscal,
)

import httpx

# ════════════════════════════════════════════════════════════════════════
# Dados da empresa auditada (confirmados)
# ════════════════════════════════════════════════════════════════════════

EMPRESA = {
    "razao_social": "LOCAR TRANSPORTE DE BOVINOS LTDA",
    "razao_anterior": "LOCAR TRANSPORTE E AGROPECUARIA LTDA (ate 06/11/2024)",
    "cnpj": "05.509.396/0001-10",
    "cnpj_basico": "05509396000110",
    "nome_fantasia": "LOCAR TRANSPORTE DE BOVINOS",
    "data_abertura": "27/01/2003",
    "situacao": "ATIVA (desde 03/11/2005)",
    "porte_declarado": "EPP - Empresa de Pequeno Porte",
    "natureza_juridica": "206-2 Sociedade Empresaria Limitada",
    "capital_social": 400_000.0,
    "cnae_principal": "49.30-2-02 Transporte rodoviario de carga interestadual",
    "cnae_secundario": "77.31-4-00 Aluguel de maquinas/equipamentos agricolas",
    "endereco_sede": "EST MATA DO FORMOSO, 13 KM, ZONA RURAL - FORMOSO/GO - CEP 76.470-000",
    "endereco_admin": "Rua Lino Coutinho, Qd 78, Lt 17/18, Capuava - GOIANIA/GO - CEP 74.450-070",
    "email": "locarnotas@gmail.com",
    "telefones": "(62) 3645-1165 / (62) 9131-9856",
    "socio_nome": "RENATO COSTA ESPERIDIAO JUNIOR",
    "socio_cpf": "931.891.171-87",
    "socio_quotas": "400.000 quotas (100%)",
    "socio_nascimento": "27/07/1981",
    "socio_endereco": "Rua 22, n 805, Qd L19, Lt 7, Apart 2702, Setor Oeste, GOIANIA/GO",
    "ultima_alteracao": "06/11/2024 (2a Alteracao e Consolidacao Contratual)",
}

OFX_LIST = [
    (r"C:\Users\Veloso\Downloads\extrato-conta-corrente-ofx-unix_202605_20260514110822.ofx", "JAN/2026"),
    (r"C:\Users\Veloso\Downloads\extrato-conta-corrente-ofx-unix_202605_20260514110841.ofx", "FEV/2026"),
    (r"C:\Users\Veloso\Downloads\extrato-conta-corrente-ofx-unix_202605_20260514110900.ofx", "MAR/2026"),
    (r"C:\Users\Veloso\Downloads\extrato-conta-corrente-ofx-unix_202605_20260514110917.ofx", "ABR/2026"),
    (r"C:\Users\Veloso\Downloads\extrato-conta-corrente-ofx-unix_202605_20260514110938.ofx", "MAI/2026"),
]

OUT_BASE = r"C:\Users\Veloso\Downloads\RELATORIO_INTEGRADO_LOCAR_v3"
OUT_XLSX = Path(f"{OUT_BASE}.xlsx")
OUT_MD = Path(f"{OUT_BASE}.md")
OUT_HTML = Path(f"{OUT_BASE}.html")
OUT_PDF = Path(f"{OUT_BASE}.pdf")

RX_CNPJ = re.compile(r"(\d{2})[.](\d{3})[.](\d{3})[ /](\d{4})[-](\d{2})")
LIMITE_MEI_PADRAO = 81_000.0
LIMITE_MEI_TAC = 251_600.0   # MEI Transportador Autonomo de Cargas (LC 188/2021)
CNAES_TRANSPORTE = ("4930", "5320", "4911")  # Transporte rodoviario carga, courier, ferroviario


def _limite_mei_por_cnae(cnae: str) -> float:
    """Retorna teto MEI conforme CNAE: TAC para transporte, padrao p/ outros."""
    cnae_clean = (cnae or "").replace(".", "").replace("-", "").replace("/", "")
    if any(cnae_clean.startswith(c) for c in CNAES_TRANSPORTE):
        return LIMITE_MEI_TAC
    return LIMITE_MEI_PADRAO

NAVY = "0F172A"
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TOTAL_FILL = PatternFill("solid", fgColor="1E3A8A")
TOTAL_FONT = Font(bold=True, color="FFFFFF", size=11)
ZEBRA_FILL = PatternFill("solid", fgColor="F8FAFC")
ALERT_FILL = PatternFill("solid", fgColor="FEE2E2")
ALERT_FILL_MEDIO = PatternFill("solid", fgColor="FEF3C7")
INFO_FILL = PatternFill("solid", fgColor="DBEAFE")
SUCCESS_FILL = PatternFill("solid", fgColor="DCFCE7")
TITLE_FONT = Font(bold=True, size=14, color=NAVY)
SUBTITLE_FONT = Font(bold=True, size=11, color=NAVY)
BORDER = Side(border_style="thin", color="E2E8F0")
THIN_BORDER = Border(top=BORDER, left=BORDER, right=BORDER, bottom=BORDER)


def _extrair_cnpj(t):
    for fonte in (t.nome or "", t.memo or ""):
        m = RX_CNPJ.search(fonte)
        if m:
            return "".join(m.groups())
    return None


def style_header(ws, row, n_cols):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = THIN_BORDER


def cabecalho(ws, ultima_col, secao):
    c1 = ws.cell(row=1, column=1, value="    ORGATEC · Relatorio Integrado de Auditoria · LOCAR TRANSPORTE DE BOVINOS LTDA")
    c1.font = Font(bold=True, size=14, color="FFFFFF")
    c1.fill = PatternFill("solid", fgColor=NAVY)
    c1.alignment = Alignment(horizontal="center", vertical="center", indent=2)
    ws.merge_cells(f"A1:{get_column_letter(ultima_col)}1")
    ws.row_dimensions[1].height = 60
    if ws.column_dimensions["A"].width is None or ws.column_dimensions["A"].width < 12:
        ws.column_dimensions["A"].width = 12
    inserir_logo_xlsx(ws, "A1", largura_px=60, altura_px=60)

    c2 = ws.cell(row=2, column=1,
        value=f"Empresa: {EMPRESA['razao_social']} | CNPJ: {EMPRESA['cnpj']} | Socio: {EMPRESA['socio_nome']} (CPF {EMPRESA['socio_cpf']})")
    c2.font = Font(bold=True, size=10, color="FFFFFF")
    c2.fill = PatternFill("solid", fgColor="1E3A8A")
    c2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(f"A2:{get_column_letter(ultima_col)}2")

    c3 = ws.cell(row=3, column=1,
        value=f"Conta: 158083-3 > Agencia: 3333-2 > Banco: SICOOB 756 > Periodo: 01/01/2026 a 14/05/2026 > Secao: {secao}")
    c3.font = Font(size=9, color="0F172A")
    c3.fill = INFO_FILL
    c3.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(f"A3:{get_column_letter(ultima_col)}3")
    return 5


# ════════════════════════════════════════════════════════════════════════
# Coleta de dados (carrega todos os 5 OFXs)
# ════════════════════════════════════════════════════════════════════════


async def coletar_dados():
    print("Coletando 5 OFXs e enriquecendo CNPJs...")
    cache = _carregar_cache()
    todos = []
    saldos = {}

    for path, mes in OFX_LIST:
        txs = ler_ofx(path)
        cred = sum(t.valor for t in txs if t.valor > 0)
        deb = sum(t.valor for t in txs if t.valor < 0)
        raw = Path(path).read_text(encoding="latin-1", errors="ignore")
        bal = re.search(r"<BALAMT>([\d.\-]+)", raw)
        saldo_final = float(bal.group(1)) if bal else 0.0
        saldos[mes] = {
            "saldo_final": saldo_final, "n": len(txs),
            "cred": cred, "deb": deb,
        }
        for t in txs:
            r = classificar(t)
            todos.append((mes, t, r))

    # Enriquecer CNPJs faltantes
    cnpjs_unicos = {_extrair_cnpj(t) for _, t, _ in todos if _extrair_cnpj(t)}
    cnpjs_unicos.discard(None)
    cnpjs_a_enriquecer = [c for c in cnpjs_unicos if c not in cache]
    if cnpjs_a_enriquecer:
        print(f"  Enriquecendo {len(cnpjs_a_enriquecer)} CNPJs novos via BrasilAPI...")
        semaforo = asyncio.Semaphore(2)
        async with httpx.AsyncClient(timeout=httpx.Timeout(10.0, connect=5.0)) as client:
            async def _job(c):
                await enriquecer_um(c, cache, client, None, semaforo)
            await asyncio.gather(*[_job(c) for c in cnpjs_a_enriquecer])
        _salvar_cache(cache)

    print(f"  {len(todos):,} transacoes carregadas | {len(cnpjs_unicos)} CNPJs no cache")
    return todos, saldos, cache


# ════════════════════════════════════════════════════════════════════════
# Gerar XLSX integrado (11 abas)
# ════════════════════════════════════════════════════════════════════════


def gerar_xlsx(todos, saldos, cache):
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # ── Pre-calculos ────────────────────────────────────────────────────
    n_total = len(todos)
    cred_total = sum(s["cred"] for s in saldos.values())
    deb_total = sum(s["deb"] for s in saldos.values())
    saldo_ini_jan = saldos["JAN/2026"]["saldo_final"] - (saldos["JAN/2026"]["cred"] + saldos["JAN/2026"]["deb"])
    saldo_fim_mai = saldos["MAI/2026"]["saldo_final"]
    volume_bruto = abs(cred_total) + abs(deb_total)

    # Disposicoes com classificacao forense
    todas_disps = []
    fake_disps_para_agg = []
    for mes, t, r in todos:
        cnpj = _extrair_cnpj(t)
        info = cache.get(cnpj, {}) if cnpj else {}
        sit = info.get("situacao", "")
        porte = info.get("porte", "")
        meio = detectar_meio(t.memo or "", t.nome or "")
        razao = info.get("razao_social", "")

        # Detecta pos-baixa
        disp_name = "NORMAL"
        flag = ""
        if ("BAIXADA" in sit or "INAPTA" in sit) and info.get("data_situacao"):
            try:
                db_d = date.fromisoformat(info["data_situacao"][:10])
                dt_d = date.fromisoformat(t.data[:10])
                if dt_d > db_d:
                    disp_name = "ALERTA_POS_BAIXA"
                    flag = f"Pos-baixa: {(dt_d - db_d).days}d apos {info['data_situacao']}"
            except (ValueError, TypeError):
                pass

        class _D:
            pass
        d = _D()
        d.transacao = t
        d.estagio = r.estagio
        d.disposicao = disp_name
        d.contraparte = razao
        d.flag = flag
        d.meio = meio
        d.cnpj = cnpj
        d.info_cnpj = info
        d.mes = mes
        todas_disps.append(d)
        fake_disps_para_agg.append(d)

    agg = calcular_agregados(fake_disps_para_agg)

    # ════════════════════════════════════════════════════════════════════
    # Aba 1: Capa / Indice
    # ════════════════════════════════════════════════════════════════════
    ws = wb.create_sheet("1. Capa")
    start = cabecalho(ws, 6, "Capa e Indice")
    ws.cell(row=start, column=1, value="RELATORIO INTEGRADO DE AUDITORIA BANCARIA").font = TITLE_FONT
    ws.merge_cells(f"A{start}:F{start}")
    ws.cell(row=start+1, column=1, value=f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}").font = Font(italic=True, color="64748B", size=9)
    ws.merge_cells(f"A{start+1}:F{start+1}")

    r = start + 3
    ws.cell(row=r, column=1, value="INDICE DE SECOES").font = SUBTITLE_FONT
    ws.merge_cells(f"A{r}:F{r}")
    r += 1
    ws.cell(row=r, column=1, value="#")
    ws.cell(row=r, column=2, value="Secao")
    ws.cell(row=r, column=3, value="Conteudo")
    style_header(ws, r, 3)
    r += 1

    indice = [
        ("1", "Capa", "Indice, sumario executivo, totais", "1. Capa"),
        ("2", "Identificacao Cadastral", "Dados RFB + contrato social + quadro societario", "2. Identificacao"),
        ("3", "Resumo Executivo", "KPIs principais e evolucao mensal", "3. Resumo Executivo"),
        ("4", "Transacoes", "7.110 lancamentos com saldo acumulado e contraparte", "4. Transacoes"),
        ("5", "Disposicoes Forenses", "Classificacao em 27 colunas + Risk Score", "5. Disposicoes"),
        ("6", "Risk Heatmap", "Distribuicao por classe (CRITICO/ALTO/MEDIO/BAIXO)", "6. Risk Heatmap"),
        ("7", "CNPJs Enriquecidos", "Contrapartes identificadas via RFB / BrasilAPI", "7. CNPJs"),
        ("8", "Partes Relacionadas", "LOCAR LOCADORA + MAQUINAS + Renato PF + MESMA TIT", "8. Partes Relacionadas"),
        ("9", "MEIs Estourando Teto", "32 fornecedores PJ acima de R$ 81k/ano", "9. MEIs Teto"),
        ("10", "Status Tributario", "Categorias fiscais + retencoes estimadas", "10. Status Tributario"),
        ("11", "Pagamentos Pos-Baixa", "Transacoes a CNPJs ja baixados", "11. Pos-Baixa"),
    ]
    for num, sec, desc, sheet_name in indice:
        ws.cell(row=r, column=1, value=num)
        c_sec = ws.cell(row=r, column=2, value=sec)
        c_sec.font = Font(bold=True, color="0052FF", underline="single")
        # Hyperlink para a aba correspondente
        c_sec.hyperlink = f"#'{sheet_name}'!A1"
        c_sec.style = "Hyperlink"
        ws.cell(row=r, column=3, value=desc)
        ws.merge_cells(f"C{r}:F{r}")
        for c in range(1, 7):
            ws.cell(row=r, column=c).border = THIN_BORDER
            if r % 2 == 0:
                ws.cell(row=r, column=c).fill = ZEBRA_FILL
        r += 1

    # Sumario rapido
    r += 2
    ws.cell(row=r, column=1, value="SUMARIO EXECUTIVO").font = SUBTITLE_FONT
    ws.merge_cells(f"A{r}:F{r}")
    r += 1
    resumo = [
        ("Periodo analisado", "01/01/2026 a 14/05/2026 (4,5 meses)"),
        ("Total de transacoes", f"{n_total:,}"),
        ("Volume bruto movimentado", f"R$ {volume_bruto:,.2f}"),
        ("Volume anualizado projetado", f"R$ {volume_bruto * 12 / 4.5:,.2f}"),
        ("Saldo inicial (01/01)", f"R$ {saldo_ini_jan:,.2f}"),
        ("Saldo final (14/05)", f"R$ {saldo_fim_mai:,.2f}"),
        ("Variacao do periodo", f"R$ {saldo_fim_mai - saldo_ini_jan:,.2f}"),
        ("CNPJs identificados", f"{sum(1 for d in todas_disps if d.cnpj)}"),
        ("Alertas pos-baixa", f"{sum(1 for d in todas_disps if d.disposicao == 'ALERTA_POS_BAIXA')}"),
    ]
    for k, v in resumo:
        ws.cell(row=r, column=1, value=k).font = Font(bold=True)
        c = ws.cell(row=r, column=2, value=v)
        if "anualizado" in k or "Volume bruto" in k:
            c.font = Font(bold=True, color="DC2626")
        ws.merge_cells(f"B{r}:F{r}")
        for col in range(1, 7):
            ws.cell(row=r, column=col).border = THIN_BORDER
            if r % 2 == 0:
                ws.cell(row=r, column=col).fill = ZEBRA_FILL
        r += 1

    for col, w in {1: 6, 2: 26, 3: 50, 4: 12, 5: 12, 6: 12}.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = f"A{start + 2}"

    # ════════════════════════════════════════════════════════════════════
    # Aba 2: Identificacao Cadastral
    # ════════════════════════════════════════════════════════════════════
    ws = wb.create_sheet("2. Identificacao")
    start = cabecalho(ws, 5, "Identificacao Cadastral")
    ws.cell(row=start, column=1, value="DADOS CADASTRAIS DA EMPRESA AUDITADA").font = TITLE_FONT
    ws.merge_cells(f"A{start}:E{start}")
    ws.cell(row=start+1, column=1, value="Fontes: Contrato Social (2a Alteracao 06/11/2024) + Cartao CNPJ RFB (07/11/2024)").font = Font(italic=True, color="64748B", size=9)
    ws.merge_cells(f"A{start+1}:E{start+1}")

    r = start + 3
    ws.cell(row=r, column=1, value="DADOS DA PESSOA JURIDICA").font = SUBTITLE_FONT
    ws.merge_cells(f"A{r}:E{r}")
    r += 1
    dados_pj = [
        ("Razao Social Atual", EMPRESA["razao_social"]),
        ("Razao Social Anterior", EMPRESA["razao_anterior"]),
        ("Nome Fantasia", EMPRESA["nome_fantasia"]),
        ("CNPJ", EMPRESA["cnpj"]),
        ("Situacao Cadastral", EMPRESA["situacao"]),
        ("Data de Abertura", EMPRESA["data_abertura"]),
        ("Porte Declarado", EMPRESA["porte_declarado"]),
        ("Natureza Juridica", EMPRESA["natureza_juridica"]),
        ("Capital Social", f"R$ {EMPRESA['capital_social']:,.2f}"),
        ("CNAE Principal", EMPRESA["cnae_principal"]),
        ("CNAE Secundario", EMPRESA["cnae_secundario"]),
        ("Endereco Sede", EMPRESA["endereco_sede"]),
        ("Escritorio Administrativo", EMPRESA["endereco_admin"]),
        ("Email", EMPRESA["email"]),
        ("Telefones", EMPRESA["telefones"]),
        ("Ultima Alteracao Contratual", EMPRESA["ultima_alteracao"]),
    ]
    for k, v in dados_pj:
        ws.cell(row=r, column=1, value=k).font = Font(bold=True)
        ws.cell(row=r, column=2, value=str(v))
        ws.merge_cells(f"B{r}:E{r}")
        for c in range(1, 6):
            ws.cell(row=r, column=c).border = THIN_BORDER
            if r % 2 == 0:
                ws.cell(row=r, column=c).fill = ZEBRA_FILL
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="QUADRO SOCIETARIO").font = SUBTITLE_FONT
    ws.merge_cells(f"A{r}:E{r}")
    r += 1
    socio = [
        ("Socio Unico (100%)", EMPRESA["socio_nome"]),
        ("CPF", EMPRESA["socio_cpf"]),
        ("Participacao", EMPRESA["socio_quotas"]),
        ("Data de Nascimento", f"{EMPRESA['socio_nascimento']} (44 anos)"),
        ("Endereco Residencial", EMPRESA["socio_endereco"]),
        ("Funcao", "Administrador unico por prazo indeterminado"),
    ]
    for k, v in socio:
        ws.cell(row=r, column=1, value=k).font = Font(bold=True)
        ws.cell(row=r, column=2, value=str(v))
        ws.merge_cells(f"B{r}:E{r}")
        for c in range(1, 6):
            ws.cell(row=r, column=c).border = THIN_BORDER
            if r % 2 == 0:
                ws.cell(row=r, column=c).fill = ZEBRA_FILL
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="DIVERGENCIAS IDENTIFICADAS").font = Font(bold=True, size=11, color="DC2626")
    ws.merge_cells(f"A{r}:E{r}")
    r += 1
    divergencias = [
        ("[!] Porte EPP vs Movimentacao", "Limite EPP: R$ 4,8M/ano | Real: R$ 187M/ano (39x acima)"),
        ("[!] Capital vs Giro", "Capital R$ 400k | Giro anual R$ 187M (razao 1:468)"),
        ("[!] Desenquadramento Tributario", "Provavelmente Lucro Real seria devido"),
        ("[!] Mudanca de Razao Social", "06/11/2024 (1 mes antes do periodo) - estreitamento de objeto"),
    ]
    for k, v in divergencias:
        ws.cell(row=r, column=1, value=k).font = Font(bold=True, color="DC2626")
        ws.cell(row=r, column=1).fill = ALERT_FILL
        ws.cell(row=r, column=2, value=v).fill = ALERT_FILL
        ws.merge_cells(f"B{r}:E{r}")
        for c in range(1, 6):
            ws.cell(row=r, column=c).border = THIN_BORDER
        r += 1

    for col, w in {1: 32, 2: 60, 3: 10, 4: 10, 5: 10}.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = f"A{start + 2}"

    # ════════════════════════════════════════════════════════════════════
    # Aba 3: Resumo Executivo (KPIs + Evolucao Mensal)
    # ════════════════════════════════════════════════════════════════════
    ws = wb.create_sheet("3. Resumo Executivo")
    start = cabecalho(ws, 7, "Resumo Executivo")
    ws.cell(row=start, column=1, value="INDICADORES PRINCIPAIS").font = TITLE_FONT
    ws.merge_cells(f"A{start}:G{start}")

    r = start + 2
    ws.cell(row=r, column=1, value="Indicador")
    ws.cell(row=r, column=2, value="Valor")
    style_header(ws, r, 2)
    r += 1
    kpis = [
        ("Total de transacoes (5 meses)", n_total),
        ("Volume de creditos", cred_total),
        ("Volume de debitos", deb_total),
        ("Volume bruto movimentado", volume_bruto),
        ("Saldo inicial (01/01)", saldo_ini_jan),
        ("Saldo final (14/05)", saldo_fim_mai),
        ("Variacao do periodo", saldo_fim_mai - saldo_ini_jan),
        ("Volume anualizado projetado", volume_bruto * 12 / 4.5),
        ("Limite EPP (referencia)", 4_800_000),
        ("Multiplo do teto EPP", volume_bruto * 12 / 4.5 / 4_800_000),
    ]
    for k, v in kpis:
        ws.cell(row=r, column=1, value=k).font = Font(bold=True)
        c = ws.cell(row=r, column=2, value=v)
        if isinstance(v, (int, float)) and "Volume" in k or "Saldo" in k or "Variacao" in k or "Limite" in k:
            c.number_format = "#,##0.00"
        elif "Multiplo" in k:
            c.number_format = "0.0\\x"
            c.font = Font(bold=True, color="DC2626")
        for col in range(1, 3):
            ws.cell(row=r, column=col).border = THIN_BORDER
            if r % 2 == 0:
                ws.cell(row=r, column=col).fill = ZEBRA_FILL
        r += 1

    # Evolucao mensal
    r += 2
    ws.cell(row=r, column=1, value="EVOLUCAO MENSAL").font = SUBTITLE_FONT
    ws.merge_cells(f"A{r}:G{r}")
    r += 1
    headers_m = ["Mes", "Transacoes", "Creditos (R$)", "Debitos (R$)", "Fluxo Liquido", "Saldo Final", "Var. Mes Ant."]
    for c, h in enumerate(headers_m, start=1):
        ws.cell(row=r, column=c, value=h)
    style_header(ws, r, 7)
    r += 1

    saldo_anterior = saldo_ini_jan
    for mes in ["JAN/2026", "FEV/2026", "MAR/2026", "ABR/2026", "MAI/2026"]:
        s = saldos[mes]
        var = s["saldo_final"] - saldo_anterior
        ws.cell(row=r, column=1, value=mes).font = Font(bold=True)
        ws.cell(row=r, column=2, value=s["n"]).number_format = "#,##0"
        ws.cell(row=r, column=3, value=round(s["cred"], 2)).number_format = "#,##0.00"
        ws.cell(row=r, column=3).font = Font(color="16A34A")
        ws.cell(row=r, column=4, value=round(s["deb"], 2)).number_format = "#,##0.00"
        ws.cell(row=r, column=4).font = Font(color="DC2626")
        fl = s["cred"] + s["deb"]
        ws.cell(row=r, column=5, value=round(fl, 2)).number_format = "#,##0.00"
        ws.cell(row=r, column=5).font = Font(bold=True, color="DC2626" if fl < 0 else "16A34A")
        ws.cell(row=r, column=6, value=round(s["saldo_final"], 2)).number_format = "#,##0.00"
        ws.cell(row=r, column=6).font = Font(bold=True, color="DC2626" if s["saldo_final"] < 0 else "0F172A")
        ws.cell(row=r, column=7, value=round(var, 2)).number_format = "#,##0.00"
        ws.cell(row=r, column=7).font = Font(color="DC2626" if var < 0 else "16A34A")
        for c in range(1, 8):
            ws.cell(row=r, column=c).border = THIN_BORDER
            if r % 2 == 0:
                ws.cell(row=r, column=c).fill = ZEBRA_FILL
        saldo_anterior = s["saldo_final"]
        r += 1

    # Linha total
    ws.cell(row=r, column=1, value="TOTAL").font = TOTAL_FONT
    ws.cell(row=r, column=2, value=n_total).number_format = "#,##0"
    ws.cell(row=r, column=3, value=round(cred_total, 2)).number_format = "#,##0.00"
    ws.cell(row=r, column=4, value=round(deb_total, 2)).number_format = "#,##0.00"
    ws.cell(row=r, column=5, value=round(cred_total + deb_total, 2)).number_format = "#,##0.00"
    ws.cell(row=r, column=6, value=round(saldo_fim_mai, 2)).number_format = "#,##0.00"
    for c in range(1, 8):
        ws.cell(row=r, column=c).fill = TOTAL_FILL
        ws.cell(row=r, column=c).font = TOTAL_FONT

    for col, w in {1: 28, 2: 14, 3: 18, 4: 18, 5: 18, 6: 18, 7: 18}.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = f"A{start + 2}"

    # ════════════════════════════════════════════════════════════════════
    # Aba 4: Transacoes (extrato com saldo acumulado)
    # ════════════════════════════════════════════════════════════════════
    ws = wb.create_sheet("4. Transacoes")
    start = cabecalho(ws, 9, "Transacoes")
    ws.cell(row=start, column=1, value=f"EXTRATO DETALHADO - {n_total:,} TRANSACOES").font = TITLE_FONT
    ws.merge_cells(f"A{start}:I{start}")
    ws.cell(row=start+1, column=1, value=f"Saldo inicial: R$ {saldo_ini_jan:,.2f} | Saldo final: R$ {saldo_fim_mai:,.2f}").font = Font(italic=True, color="64748B", size=9)
    ws.merge_cells(f"A{start+1}:I{start+1}")

    headers_t = ["#", "Mes", "Data", "Tipo", "Valor (R$)", "Memo", "Nome", "Contraparte (RFB)", "Saldo Acumulado (R$)"]
    r = start + 3
    for c, h in enumerate(headers_t, start=1):
        ws.cell(row=r, column=c, value=h)
    style_header(ws, r, 9)
    r += 1

    txs_ord = sorted(todas_disps, key=lambda x: x.transacao.data)
    saldo_corrente = saldo_ini_jan
    for i, d in enumerate(txs_ord, start=1):
        t = d.transacao
        saldo_corrente += t.valor
        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=2, value=d.mes)
        ws.cell(row=r, column=3, value=t.data)
        ws.cell(row=r, column=4, value=t.tipo)
        cv = ws.cell(row=r, column=5, value=round(t.valor, 2))
        cv.number_format = "#,##0.00"
        cv.font = Font(color=("DC2626" if t.valor < 0 else "16A34A"))
        ws.cell(row=r, column=6, value=t.memo or "")
        ws.cell(row=r, column=7, value=t.nome or "")
        cc = ws.cell(row=r, column=8, value=d.contraparte or "")
        if d.disposicao == "ALERTA_POS_BAIXA":
            cc.font = Font(bold=True, color="DC2626")
        cs = ws.cell(row=r, column=9, value=round(saldo_corrente, 2))
        cs.number_format = "#,##0.00"
        cs.font = Font(bold=True, color=("DC2626" if saldo_corrente < 0 else "0F172A"))
        for c in range(1, 10):
            ws.cell(row=r, column=c).border = THIN_BORDER
            if d.disposicao == "ALERTA_POS_BAIXA":
                ws.cell(row=r, column=c).fill = ALERT_FILL
        r += 1

    for col, w in {1: 5, 2: 10, 3: 12, 4: 8, 5: 14, 6: 32, 7: 30, 8: 35, 9: 20}.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = f"A{start + 4}"
    ws.auto_filter.ref = f"A{start + 3}:I{r - 1}"

    # ════════════════════════════════════════════════════════════════════
    # Aba 5: Disposicoes Forenses (27 colunas)
    # ════════════════════════════════════════════════════════════════════
    ws = wb.create_sheet("5. Disposicoes")
    start = cabecalho(ws, 27, "Disposicoes Forenses")
    ws.cell(row=start, column=1, value="DISPOSICOES POR TRANSACAO - Auditoria Forense").font = TITLE_FONT
    ws.merge_cells(f"A{start}:AA{start}")
    ws.cell(row=start+1, column=1, value="Eixos: A=Compliance | B=Identificacao | C=Padroes | D=Risk Score | E=Rastreabilidade").font = Font(italic=True, color="64748B", size=9)
    ws.merge_cells(f"A{start+1}:AA{start+1}")

    headers = [
        "Data", "Tipo", "Valor (R$)", "Memo", "Nome (banco)",
        "FITID", "CheckNum", "Meio",
        "CNPJ", "Contraparte (RFB)",
        "CNAE", "UF", "Municipio", "Porte",
        "Acumulado Mes (R$)", "1a Vez?", "Valor Redondo", "Smurfing", "Carrossel",
        "Disposicao", "Flag",
        "Risk Score", "Risk Class",
        "Periodo Fiscal", "Hash Linha", "Status Revisao", "Comentario Revisor",
    ]
    r = start + 3
    for c, h in enumerate(headers, start=1):
        ws.cell(row=r, column=c, value=h)
    style_header(ws, r, len(headers))
    r += 1

    for d in sorted(todas_disps, key=lambda x: x.transacao.data):
        t = d.transacao
        cnpj = d.cnpj
        cnpj_fmt = ""
        if cnpj:
            cnpj_fmt = f"{cnpj[:2]}.{cnpj[2:5]}.{cnpj[5:8]}/{cnpj[8:12]}-{cnpj[12:14]}"
        info = d.info_cnpj
        cnae = info.get("cnae_descricao", "")
        uf = info.get("uf", "")
        municipio = info.get("municipio", "")
        porte = info.get("porte", "")
        sit = info.get("situacao", "")
        meio = d.meio
        mes = t.data[:7]
        acum = agg.acumulado_mes.get((cnpj, mes), 0.0) if cnpj else 0.0
        pv = detectar_primeira_vez(cnpj, t.data, agg)
        vr = detectar_valor_redondo(t.valor)
        sm = detectar_smurfing(cnpj, t.data, agg)
        car = detectar_carrossel(cnpj, agg)
        score, classe = calcular_risk_score(t.valor, d.disposicao, sit, porte, meio, vr, sm, car, pv, acum)
        pf = periodo_fiscal(t.data)
        h_linha = hash_linha(t.data, t.valor, t.memo or "", t.fitid or "")
        is_alerta = d.disposicao == "ALERTA_POS_BAIXA"
        is_critico = classe == "CRITICO"

        # Preenche linha
        ws.cell(row=r, column=1, value=t.data)
        ws.cell(row=r, column=2, value=t.tipo)
        cv = ws.cell(row=r, column=3, value=round(t.valor, 2))
        cv.number_format = "#,##0.00"
        cv.font = Font(color=("DC2626" if t.valor < 0 else "16A34A"), bold=is_alerta or is_critico)
        ws.cell(row=r, column=4, value=t.memo or "")
        ws.cell(row=r, column=5, value=t.nome or "")
        ws.cell(row=r, column=6, value=t.fitid or "").font = Font(name="Consolas", size=9, color="64748B")
        ws.cell(row=r, column=7, value=t.checknum or "")
        ws.cell(row=r, column=8, value=meio)
        ws.cell(row=r, column=9, value=cnpj_fmt).font = Font(name="Consolas", size=10)
        ws.cell(row=r, column=10, value=d.contraparte or "")
        ws.cell(row=r, column=11, value=cnae[:60])
        ws.cell(row=r, column=12, value=uf)
        ws.cell(row=r, column=13, value=municipio)
        ws.cell(row=r, column=14, value=porte)
        ws.cell(row=r, column=15, value=round(acum, 2)).number_format = "#,##0.00"
        ws.cell(row=r, column=16, value=pv)
        ws.cell(row=r, column=17, value=vr)
        ws.cell(row=r, column=18, value=sm)
        ws.cell(row=r, column=19, value=car)
        ws.cell(row=r, column=20, value=d.disposicao)
        ws.cell(row=r, column=21, value=d.flag or "")
        c_score = ws.cell(row=r, column=22, value=score)
        c_score.number_format = "0"
        c_classe = ws.cell(row=r, column=23, value=classe)
        c_classe.font = Font(bold=True)
        fill_classe = {"CRITICO": "FEE2E2", "ALTO": "FEF3C7", "MEDIO": "DBEAFE", "BAIXO": "DCFCE7"}.get(classe)
        if fill_classe:
            c_classe.fill = PatternFill("solid", fgColor=fill_classe)
        ws.cell(row=r, column=24, value=pf)
        ws.cell(row=r, column=25, value=h_linha).font = Font(name="Consolas", size=8, color="94A3B8")
        ws.cell(row=r, column=26, value="PENDENTE").font = Font(italic=True, color="64748B", size=9)
        ws.cell(row=r, column=27, value="")

        for c in range(1, 28):
            ws.cell(row=r, column=c).border = THIN_BORDER
            if is_alerta and c != 23:
                ws.cell(row=r, column=c).fill = ALERT_FILL
        r += 1

    larguras = {1: 11, 2: 7, 3: 13, 4: 28, 5: 28, 6: 18, 7: 10, 8: 11, 9: 19, 10: 32,
                11: 38, 12: 5, 13: 22, 14: 22, 15: 16, 16: 8, 17: 13, 18: 22, 19: 11,
                20: 22, 21: 32, 22: 10, 23: 10, 24: 12, 25: 17, 26: 13, 27: 30}
    for col, w in larguras.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = f"F{start + 4}"
    ws.auto_filter.ref = f"A{start + 3}:AA{r - 1}"

    # ════════════════════════════════════════════════════════════════════
    # Aba 6: Risk Heatmap
    # ════════════════════════════════════════════════════════════════════
    ws = wb.create_sheet("6. Risk Heatmap")
    start = cabecalho(ws, 6, "Risk Heatmap")
    ws.cell(row=start, column=1, value="DISTRIBUICAO POR CLASSE DE RISCO").font = TITLE_FONT
    ws.merge_cells(f"A{start}:F{start}")

    classe_counts = {"CRITICO": [0, 0.0], "ALTO": [0, 0.0], "MEDIO": [0, 0.0], "BAIXO": [0, 0.0]}
    for d in todas_disps:
        t = d.transacao
        cnpj = d.cnpj
        info = d.info_cnpj
        sit = info.get("situacao", "")
        porte = info.get("porte", "")
        meio = d.meio
        mes = t.data[:7]
        acum = agg.acumulado_mes.get((cnpj, mes), 0.0) if cnpj else 0.0
        pv = detectar_primeira_vez(cnpj, t.data, agg)
        vr = detectar_valor_redondo(t.valor)
        sm = detectar_smurfing(cnpj, t.data, agg)
        car = detectar_carrossel(cnpj, agg)
        _, classe = calcular_risk_score(t.valor, d.disposicao, sit, porte, meio, vr, sm, car, pv, acum)
        classe_counts[classe][0] += 1
        classe_counts[classe][1] += abs(t.valor)

    r = start + 2
    headers_h = ["Classe", "Qtd Transacoes", "% do Total", "Volume (R$)", "% Volume", "Acao Sugerida"]
    for c, h in enumerate(headers_h, start=1):
        ws.cell(row=r, column=c, value=h)
    style_header(ws, r, 6)
    r += 1

    cores = {
        "CRITICO": ("DC2626", "FEE2E2", "Auditoria imediata - investigar"),
        "ALTO":    ("D97706", "FEF3C7", "Revisao prioritaria"),
        "MEDIO":   ("0052FF", "DBEAFE", "Conferir em lote"),
        "BAIXO":   ("16A34A", "DCFCE7", "Auto-aprovar apos confirmacao"),
    }
    total_qtd = sum(v[0] for v in classe_counts.values())
    total_vol = sum(v[1] for v in classe_counts.values())
    for classe in ("CRITICO", "ALTO", "MEDIO", "BAIXO"):
        qtd, vol = classe_counts[classe]
        cor, fill_cor, acao = cores[classe]
        c1 = ws.cell(row=r, column=1, value=classe)
        c1.font = Font(bold=True, color=cor)
        c1.fill = PatternFill("solid", fgColor=fill_cor)
        ws.cell(row=r, column=2, value=qtd).number_format = "#,##0"
        ws.cell(row=r, column=3, value=qtd / max(total_qtd, 1)).number_format = "0.0%"
        ws.cell(row=r, column=4, value=round(vol, 2)).number_format = "#,##0.00"
        ws.cell(row=r, column=5, value=vol / max(total_vol, 1)).number_format = "0.0%"
        ws.cell(row=r, column=6, value=acao)
        for c in range(1, 7):
            ws.cell(row=r, column=c).border = THIN_BORDER
        r += 1

    ws.cell(row=r, column=1, value="TOTAL").font = TOTAL_FONT
    ws.cell(row=r, column=2, value=total_qtd).number_format = "#,##0"
    ws.cell(row=r, column=4, value=round(total_vol, 2)).number_format = "#,##0.00"
    for c in range(1, 7):
        ws.cell(row=r, column=c).fill = TOTAL_FILL
        ws.cell(row=r, column=c).font = TOTAL_FONT

    for col, w in {1: 12, 2: 16, 3: 12, 4: 18, 5: 11, 6: 36}.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = f"A{start + 2}"

    # ════════════════════════════════════════════════════════════════════
    # Aba 7: CNPJs Enriquecidos
    # ════════════════════════════════════════════════════════════════════
    ws = wb.create_sheet("7. CNPJs")
    start = cabecalho(ws, 8, "CNPJs Enriquecidos")
    ws.cell(row=start, column=1, value="CONTRAPARTES IDENTIFICADAS VIA BASE RFB").font = TITLE_FONT
    ws.merge_cells(f"A{start}:H{start}")

    cnpjs_unicos_usados = {d.cnpj for d in todas_disps if d.cnpj}
    cnpjs_unicos_usados.discard(None)

    headers_c = ["CNPJ", "Razao Social", "Situacao", "Data Situacao", "UF",
                 "Municipio", "CNAE", "Porte"]
    r = start + 2
    for c, h in enumerate(headers_c, start=1):
        ws.cell(row=r, column=c, value=h)
    style_header(ws, r, 8)
    r += 1

    # Agrega por aparicoes
    aparicoes = Counter(d.cnpj for d in todas_disps if d.cnpj)
    for cnpj in sorted(cnpjs_unicos_usados, key=lambda c: -aparicoes[c]):
        info = cache.get(cnpj, {})
        fmt = f"{cnpj[:2]}.{cnpj[2:5]}.{cnpj[5:8]}/{cnpj[8:12]}-{cnpj[12:14]}"
        is_baixada = "BAIXADA" in info.get("situacao", "") or "INAPTA" in info.get("situacao", "")
        ws.cell(row=r, column=1, value=fmt).font = Font(name="Consolas", size=10)
        ws.cell(row=r, column=2, value=info.get("razao_social", "(nao enriquecido)"))
        c_sit = ws.cell(row=r, column=3, value=info.get("situacao", ""))
        if is_baixada:
            c_sit.font = Font(bold=True, color="DC2626")
        ws.cell(row=r, column=4, value=info.get("data_situacao", ""))
        ws.cell(row=r, column=5, value=info.get("uf", ""))
        ws.cell(row=r, column=6, value=info.get("municipio", ""))
        ws.cell(row=r, column=7, value=info.get("cnae_descricao", "")[:55])
        ws.cell(row=r, column=8, value=info.get("porte", ""))
        for c in range(1, 9):
            ws.cell(row=r, column=c).border = THIN_BORDER
            if is_baixada:
                ws.cell(row=r, column=c).fill = ALERT_FILL
            elif r % 2 == 0:
                ws.cell(row=r, column=c).fill = ZEBRA_FILL
        r += 1

    for col, w in {1: 20, 2: 42, 3: 14, 4: 13, 5: 5, 6: 22, 7: 38, 8: 24}.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = f"A{start + 3}"
    ws.auto_filter.ref = f"A{start + 2}:H{r - 1}"

    # ════════════════════════════════════════════════════════════════════
    # Aba 8: Partes Relacionadas
    # ════════════════════════════════════════════════════════════════════
    ws = wb.create_sheet("8. Partes Relacionadas")
    start = cabecalho(ws, 5, "Partes Relacionadas")
    ws.cell(row=start, column=1, value="MOVIMENTACAO COM PARTES RELACIONADAS DO GRUPO LOCAR").font = TITLE_FONT
    ws.merge_cells(f"A{start}:E{start}")

    # Calcula fluxos
    fluxos = {
        "Proprio CNPJ (auto-movimentacao 05.509.396)": {"n": 0, "cred": 0.0, "deb": 0.0},
        "LOCAR LOCADORA E ??? (parte relacionada)": {"n": 0, "cred": 0.0, "deb": 0.0},
        "LOCAR MAQUINAS E SERVICOS (parte relacionada)": {"n": 0, "cred": 0.0, "deb": 0.0},
        "RENATO COSTA ESPERIDIAO JR (socio PF)": {"n": 0, "cred": 0.0, "deb": 0.0},
    }
    for d in todas_disps:
        texto_up = ((d.transacao.nome or "") + " " + (d.transacao.memo or "")).upper()
        if d.cnpj == EMPRESA["cnpj_basico"]:
            target = fluxos["Proprio CNPJ (auto-movimentacao 05.509.396)"]
        elif "LOCAR LOCADORA" in texto_up:
            target = fluxos["LOCAR LOCADORA E ??? (parte relacionada)"]
        elif "LOCAR MAQUINAS" in texto_up:
            target = fluxos["LOCAR MAQUINAS E SERVICOS (parte relacionada)"]
        elif "RENATO COSTA ESPERIDI" in texto_up:
            target = fluxos["RENATO COSTA ESPERIDIAO JR (socio PF)"]
        else:
            continue
        target["n"] += 1
        if d.transacao.valor > 0:
            target["cred"] += d.transacao.valor
        else:
            target["deb"] += d.transacao.valor

    r = start + 2
    h = ["Entidade", "Qtd", "Creditos (R$)", "Debitos (R$)", "Volume Total (R$)"]
    for c, hd in enumerate(h, start=1):
        ws.cell(row=r, column=c, value=hd)
    style_header(ws, r, 5)
    r += 1
    total_pr = 0.0
    for nome, dados in fluxos.items():
        vol = dados["cred"] + abs(dados["deb"])
        ws.cell(row=r, column=1, value=nome)
        ws.cell(row=r, column=2, value=dados["n"]).number_format = "#,##0"
        ws.cell(row=r, column=3, value=round(dados["cred"], 2)).number_format = "#,##0.00"
        ws.cell(row=r, column=3).font = Font(color="16A34A")
        ws.cell(row=r, column=4, value=round(dados["deb"], 2)).number_format = "#,##0.00"
        ws.cell(row=r, column=4).font = Font(color="DC2626")
        ws.cell(row=r, column=5, value=round(vol, 2)).number_format = "#,##0.00"
        ws.cell(row=r, column=5).font = Font(bold=True)
        for c in range(1, 6):
            ws.cell(row=r, column=c).border = THIN_BORDER
            if r % 2 == 0:
                ws.cell(row=r, column=c).fill = ZEBRA_FILL
        total_pr += vol
        r += 1

    ws.cell(row=r, column=1, value="TOTAL").font = TOTAL_FONT
    ws.cell(row=r, column=5, value=round(total_pr, 2)).number_format = "#,##0.00"
    for c in range(1, 6):
        ws.cell(row=r, column=c).fill = TOTAL_FILL
        ws.cell(row=r, column=c).font = TOTAL_FONT

    for col, w in {1: 50, 2: 8, 3: 22, 4: 22, 5: 22}.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = f"A{start + 2}"

    # ════════════════════════════════════════════════════════════════════
    # Aba 9: MEIs Teto (com diferenciacao MEI-TAC vs MEI Padrao)
    # ════════════════════════════════════════════════════════════════════
    ws = wb.create_sheet("9. MEIs Teto")
    start = cabecalho(ws, 10, "MEIs Estourando Teto (MEI-TAC vs Padrao)")
    ws.cell(row=start, column=1, value="ANALISE DE MEIs - MEI-TAC (R$ 251.600/ano) vs MEI Padrao (R$ 81.000/ano)").font = TITLE_FONT
    ws.merge_cells(f"A{start}:J{start}")
    ws.cell(row=start+1, column=1, value="Lei Complementar 188/2021: MEI-TAC (caminhoneiros, CNAEs 4930-*, 5320-*, 4911-*) tem teto de R$ 251.600/ano. MEI padrao mantem R$ 81.000/ano.").font = Font(italic=True, color="64748B", size=9)
    ws.merge_cells(f"A{start+1}:J{start+1}")

    # Agrega todos os MEIs (TAC e padrao)
    por_cnpj_mei = defaultdict(lambda: {"n": 0, "deb": 0.0})
    for d in todas_disps:
        if d.transacao.valor >= 0 or not d.cnpj:
            continue
        if d.info_cnpj.get("porte") != "MICRO EMPRESA":
            continue
        por_cnpj_mei[d.cnpj]["n"] += 1
        por_cnpj_mei[d.cnpj]["deb"] += abs(d.transacao.valor)

    # Classifica cada MEI conforme CNAE
    meis_tac_ok, meis_tac_estourados = [], []
    meis_padrao_ok, meis_padrao_estourados = [], []
    for cnpj, dd in por_cnpj_mei.items():
        info = cache.get(cnpj, {})
        cnae = info.get("cnae_principal", "")
        anualizado = dd["deb"] * 12 / 4.5
        teto = _limite_mei_por_cnae(cnae)
        eh_tac = teto == LIMITE_MEI_TAC
        excesso = anualizado - teto if anualizado > teto else 0.0
        item = {
            "cnpj": cnpj, "razao": info.get("razao_social", ""),
            "cnae": cnae, "cnae_desc": info.get("cnae_descricao", ""),
            "uf": info.get("uf", ""), "n": dd["n"],
            "deb_5m": dd["deb"], "anualizado": anualizado,
            "teto": teto, "excesso": excesso, "eh_tac": eh_tac,
        }
        if eh_tac and excesso > 0:
            meis_tac_estourados.append(item)
        elif eh_tac:
            meis_tac_ok.append(item)
        elif excesso > 0:
            meis_padrao_estourados.append(item)
        else:
            meis_padrao_ok.append(item)

    # Sumario
    r = start + 3
    ws.cell(row=r, column=1, value="SUMARIO DA RECLASSIFICACAO").font = SUBTITLE_FONT
    ws.merge_cells(f"A{r}:J{r}")
    r += 1
    hdr_sum = ["Categoria", "Teto Legal", "Total MEIs", "Dentro do Teto", "Acima do Teto", "Status"]
    for c, h in enumerate(hdr_sum, start=1):
        ws.cell(row=r, column=c, value=h)
    style_header(ws, r, 6)
    r += 1
    linhas_sum = [
        ("MEI-TAC (caminhoneiros)", f"R$ {LIMITE_MEI_TAC:,.0f}/ano",
         len(meis_tac_ok) + len(meis_tac_estourados),
         len(meis_tac_ok), len(meis_tac_estourados),
         "OK" if not meis_tac_estourados else "ATENCAO"),
        ("MEI Padrao (outros CNAEs)", f"R$ {LIMITE_MEI_PADRAO:,.0f}/ano",
         len(meis_padrao_ok) + len(meis_padrao_estourados),
         len(meis_padrao_ok), len(meis_padrao_estourados),
         "OK" if not meis_padrao_estourados else "ATENCAO"),
        ("TOTAL", "—",
         len(por_cnpj_mei), len(meis_tac_ok) + len(meis_padrao_ok),
         len(meis_tac_estourados) + len(meis_padrao_estourados), "—"),
    ]
    for cat, teto, total, dentro, acima, status in linhas_sum:
        c1 = ws.cell(row=r, column=1, value=cat)
        c1.font = Font(bold=True)
        ws.cell(row=r, column=2, value=teto)
        ws.cell(row=r, column=3, value=total).number_format = "#,##0"
        ws.cell(row=r, column=4, value=dentro).number_format = "#,##0"
        ws.cell(row=r, column=4).font = Font(color="16A34A", bold=True)
        ws.cell(row=r, column=5, value=acima).number_format = "#,##0"
        if acima > 0:
            ws.cell(row=r, column=5).font = Font(color="DC2626", bold=True)
        c_status = ws.cell(row=r, column=6, value=status)
        if status == "OK":
            c_status.fill = SUCCESS_FILL
            c_status.font = Font(bold=True, color="16A34A")
        elif status == "ATENCAO":
            c_status.fill = ALERT_FILL_MEDIO
            c_status.font = Font(bold=True, color="D97706")
        for c in range(1, 7):
            ws.cell(row=r, column=c).border = THIN_BORDER
        if cat == "TOTAL":
            for c in range(1, 7):
                ws.cell(row=r, column=c).fill = TOTAL_FILL
                ws.cell(row=r, column=c).font = TOTAL_FONT
        r += 1

    # Detalhamento dos casos acima do teto
    r += 1
    ws.cell(row=r, column=1, value="DETALHAMENTO - CASOS ACIMA DO TETO").font = SUBTITLE_FONT
    ws.merge_cells(f"A{r}:J{r}")
    r += 1

    headers = ["#", "Tipo", "CNPJ", "Razao Social", "CNAE", "UF",
               "Trans.", "Pago 5m (R$)", "Anualizado (R$)", "Excesso (R$)"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=r, column=c, value=h)
    style_header(ws, r, 10)
    r += 1

    todos_estourados = (
        [(m, "MEI-TAC") for m in meis_tac_estourados] +
        [(m, "MEI Padrao") for m in meis_padrao_estourados]
    )
    todos_estourados.sort(key=lambda x: -x[0]["anualizado"])
    total_excesso = 0.0

    if not todos_estourados:
        ws.cell(row=r, column=1, value="(nenhum MEI excede o teto legal correspondente)").font = Font(italic=True, color="16A34A")
    else:
        for i, (m, tipo) in enumerate(todos_estourados, start=1):
            cnpj_fmt = f"{m['cnpj'][:2]}.{m['cnpj'][2:5]}.{m['cnpj'][5:8]}/{m['cnpj'][8:12]}-{m['cnpj'][12:14]}"
            ws.cell(row=r, column=1, value=i)
            c_tipo = ws.cell(row=r, column=2, value=tipo)
            c_tipo.font = Font(bold=True)
            ws.cell(row=r, column=3, value=cnpj_fmt).font = Font(name="Consolas", size=10)
            ws.cell(row=r, column=4, value=m["razao"][:45])
            ws.cell(row=r, column=5, value=m["cnae_desc"][:35] if m["cnae_desc"] else m["cnae"])
            ws.cell(row=r, column=6, value=m["uf"])
            ws.cell(row=r, column=7, value=m["n"]).number_format = "#,##0"
            ws.cell(row=r, column=8, value=round(m["deb_5m"], 2)).number_format = "#,##0.00"
            c9 = ws.cell(row=r, column=9, value=round(m["anualizado"], 2))
            c9.number_format = "#,##0.00"
            c9.font = Font(bold=True, color="DC2626")
            c10 = ws.cell(row=r, column=10, value=round(m["excesso"], 2))
            c10.number_format = "#,##0.00"
            c10.font = Font(bold=True, color="DC2626")
            for c in range(1, 11):
                ws.cell(row=r, column=c).border = THIN_BORDER
                if m["excesso"] > 100_000:
                    ws.cell(row=r, column=c).fill = ALERT_FILL
                elif r % 2 == 0:
                    ws.cell(row=r, column=c).fill = ZEBRA_FILL
            total_excesso += m["excesso"]
            r += 1

        ws.cell(row=r, column=1, value=f"TOTAL ({len(todos_estourados)} MEIs)").font = TOTAL_FONT
        ws.cell(row=r, column=10, value=round(total_excesso, 2)).number_format = "#,##0.00"
        for c in range(1, 11):
            ws.cell(row=r, column=c).fill = TOTAL_FILL
            ws.cell(row=r, column=c).font = TOTAL_FONT

    for col, w in {1: 4, 2: 12, 3: 20, 4: 38, 5: 32, 6: 5, 7: 8, 8: 16, 9: 17, 10: 17}.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = f"A{start + 4}"
    # Mantem para compatibilidade com codigo posterior
    meis = [m for m, _ in todos_estourados]

    # ════════════════════════════════════════════════════════════════════
    # Aba 10: Status Tributario
    # ════════════════════════════════════════════════════════════════════
    ws = wb.create_sheet("10. Status Tributario")
    start = cabecalho(ws, 9, "Status Tributario")
    ws.cell(row=start, column=1, value="STATUS TRIBUTARIO CONSOLIDADO - 5 MESES").font = TITLE_FONT
    ws.merge_cells(f"A{start}:I{start}")

    cat_count = Counter()
    cat_volume = defaultdict(float)
    cat_retencao = defaultdict(float)
    cat_por_mes = defaultdict(lambda: defaultdict(float))

    for d in todas_disps:
        t = d.transacao
        porte = d.info_cnpj.get("porte", "")
        trib = classificar_tributario(t.memo or "", t.nome or "", t.valor, d.cnpj or "", porte)
        cat_count[trib["categoria"]] += 1
        cat_volume[trib["categoria"]] += abs(t.valor)
        cat_retencao[trib["categoria"]] += trib["valor_retencao"]
        cat_por_mes[d.mes][trib["categoria"]] += trib["valor_retencao"]

    r = start + 2
    headers = ["Categoria", "Qtd", "Volume (R$)", "Retencao (R$)", "JAN", "FEV", "MAR", "ABR", "MAI"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=r, column=c, value=h)
    style_header(ws, r, 9)
    r += 1

    CATS = ["RETENCAO_PJ", "RETENCAO_PF", "OPERACAO_CREDITO", "IOF", "JUROS",
            "PAGAMENTO_TRIBUTO", "TARIFA", "PIX_RECEBIDO", "BOLETO",
            "COMPRA_CARTAO", "NAO_TRIBUTAVEL", "OUTRO"]
    total_ret_5m = 0.0
    for cat in CATS:
        qtd = cat_count.get(cat, 0)
        if qtd == 0:
            continue
        c1 = ws.cell(row=r, column=1, value=cat)
        c1.font = Font(bold=True, color="DC2626" if cat.startswith("RETENCAO") else "0F172A")
        ws.cell(row=r, column=2, value=qtd).number_format = "#,##0"
        ws.cell(row=r, column=3, value=round(cat_volume[cat], 2)).number_format = "#,##0.00"
        cret = ws.cell(row=r, column=4, value=round(cat_retencao[cat], 2))
        cret.number_format = "#,##0.00"
        if cat_retencao[cat] > 0:
            cret.font = Font(bold=True, color="D97706")
            total_ret_5m += cat_retencao[cat]
        for i, mes in enumerate(["JAN/2026", "FEV/2026", "MAR/2026", "ABR/2026", "MAI/2026"], start=5):
            val = cat_por_mes[mes].get(cat, 0)
            cm = ws.cell(row=r, column=i, value=round(val, 2) if val else "")
            if val:
                cm.number_format = "#,##0.00"
        for c in range(1, 10):
            ws.cell(row=r, column=c).border = THIN_BORDER
            if r % 2 == 0:
                ws.cell(row=r, column=c).fill = ZEBRA_FILL
        r += 1

    ws.cell(row=r, column=1, value="TOTAL").font = TOTAL_FONT
    ws.cell(row=r, column=2, value=sum(cat_count.values())).number_format = "#,##0"
    ws.cell(row=r, column=3, value=round(sum(cat_volume.values()), 2)).number_format = "#,##0.00"
    ws.cell(row=r, column=4, value=round(total_ret_5m, 2)).number_format = "#,##0.00"
    for i, mes in enumerate(["JAN/2026", "FEV/2026", "MAR/2026", "ABR/2026", "MAI/2026"], start=5):
        ws.cell(row=r, column=i, value=round(sum(cat_por_mes[mes].values()), 2)).number_format = "#,##0.00"
    for c in range(1, 10):
        ws.cell(row=r, column=c).fill = TOTAL_FILL
        ws.cell(row=r, column=c).font = TOTAL_FONT

    for col, w in {1: 22, 2: 8, 3: 17, 4: 16, 5: 13, 6: 13, 7: 13, 8: 13, 9: 13}.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = f"A{start + 2}"

    # ════════════════════════════════════════════════════════════════════
    # Aba 11: Pagamentos Pos-Baixa
    # ════════════════════════════════════════════════════════════════════
    ws = wb.create_sheet("11. Pos-Baixa")
    start = cabecalho(ws, 7, "Pagamentos Pos-Baixa")
    ws.cell(row=start, column=1, value="PAGAMENTOS APOS BAIXA DO CNPJ - CRITICO").font = TITLE_FONT
    ws.merge_cells(f"A{start}:G{start}")
    ws.cell(row=start+1, column=1, value="Transacoes a CNPJs ja BAIXADOS na data do pagamento - red flag forense").font = Font(italic=True, color="64748B", size=9)
    ws.merge_cells(f"A{start+1}:G{start+1}")

    pos_baixa = []
    for d in todas_disps:
        if d.disposicao == "ALERTA_POS_BAIXA":
            try:
                db_d = date.fromisoformat(d.info_cnpj["data_situacao"][:10])
                dt_d = date.fromisoformat(d.transacao.data[:10])
                pos_baixa.append({
                    "mes": d.mes, "t": d.transacao, "cnpj": d.cnpj,
                    "info": d.info_cnpj, "dias": (dt_d - db_d).days,
                })
            except (ValueError, TypeError, KeyError):
                pass

    pos_baixa.sort(key=lambda x: -x["dias"])

    headers = ["#", "Mes", "Data", "Valor (R$)", "Razao Social", "Data Baixa", "Dias Apos"]
    r = start + 3
    for c, h in enumerate(headers, start=1):
        ws.cell(row=r, column=c, value=h)
    style_header(ws, r, 7)
    r += 1
    total_pb = 0.0
    for i, p in enumerate(pos_baixa, start=1):
        cnpj_fmt = f"{p['cnpj'][:2]}.{p['cnpj'][2:5]}.{p['cnpj'][5:8]}/{p['cnpj'][8:12]}-{p['cnpj'][12:14]}"
        razao = p["info"].get("razao_social", "")[:35]
        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=2, value=p["mes"])
        ws.cell(row=r, column=3, value=p["t"].data)
        cv = ws.cell(row=r, column=4, value=round(p["t"].valor, 2))
        cv.number_format = "#,##0.00"
        cv.font = Font(bold=True, color="DC2626")
        ws.cell(row=r, column=5, value=f"{cnpj_fmt} - {razao}")
        ws.cell(row=r, column=6, value=p["info"].get("data_situacao", "")).font = Font(bold=True, color="DC2626")
        ws.cell(row=r, column=7, value=p["dias"]).font = Font(bold=True, color="DC2626")
        ws.cell(row=r, column=7).number_format = "#,##0"
        for c in range(1, 8):
            ws.cell(row=r, column=c).border = THIN_BORDER
            ws.cell(row=r, column=c).fill = ALERT_FILL
        total_pb += abs(p["t"].valor)
        r += 1

    ws.cell(row=r, column=1, value=f"TOTAL ({len(pos_baixa)} alertas)").font = TOTAL_FONT
    ws.cell(row=r, column=4, value=round(total_pb, 2)).number_format = "#,##0.00"
    for c in range(1, 8):
        ws.cell(row=r, column=c).fill = TOTAL_FILL
        ws.cell(row=r, column=c).font = TOTAL_FONT

    for col, w in {1: 4, 2: 10, 3: 12, 4: 14, 5: 60, 6: 13, 7: 12}.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = f"A{start + 4}"

    wb.save(str(OUT_XLSX))
    print(f"  XLSX: {OUT_XLSX}")

    return {
        "n_total": n_total, "volume_bruto": volume_bruto,
        "cred_total": cred_total, "deb_total": deb_total,
        "saldo_ini": saldo_ini_jan, "saldo_fim": saldo_fim_mai,
        "saldos": saldos, "todas_disps": todas_disps,
        "classe_counts": classe_counts, "cat_count": cat_count,
        "cat_volume": cat_volume, "cat_retencao": cat_retencao,
        "fluxos": fluxos, "meis": meis, "pos_baixa": pos_baixa,
        "total_ret_5m": total_ret_5m,
    }


# ════════════════════════════════════════════════════════════════════════
# Geracao MD / HTML / PDF
# ════════════════════════════════════════════════════════════════════════


def gerar_md(stats):
    n = stats["n_total"]
    cred = stats["cred_total"]
    deb = stats["deb_total"]
    vol = stats["volume_bruto"]
    saldo_ini = stats["saldo_ini"]
    saldo_fim = stats["saldo_fim"]

    lines = [
        f"# RELATORIO INTEGRADO DE AUDITORIA — {EMPRESA['razao_social']}",
        "",
        "**ORGATEC · Contabilidade · Auditoria · Compliance**",
        "",
        f"**Gerado em:** {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        f"**Periodo:** 01/01/2026 a 14/05/2026 (4,5 meses, {n:,} transacoes)",
        f"**Banco:** SICOOB 756 · Agencia 3333-2 · Conta 158083-3",
        "",
        "---",
        "",
        "## 1. Capa e Sumario Executivo",
        "",
        "| Indicador | Valor |",
        "|---|---:|",
        f"| Total de transacoes | {n:,} |",
        f"| Volume de creditos | R$ {cred:,.2f} |",
        f"| Volume de debitos | R$ {deb:,.2f} |",
        f"| Volume bruto movimentado | **R$ {vol:,.2f}** |",
        f"| Saldo inicial (01/01) | R$ {saldo_ini:,.2f} |",
        f"| Saldo final (14/05) | R$ {saldo_fim:,.2f} |",
        f"| Variacao do periodo | R$ {saldo_fim - saldo_ini:,.2f} |",
        f"| **Volume anualizado projetado** | **R$ {vol * 12 / 4.5:,.2f}** |",
        f"| Limite EPP (referencia) | R$ 4.800.000,00 |",
        f"| **Multiplo do teto EPP** | **{vol * 12 / 4.5 / 4_800_000:.1f}x** |",
        f"| Alertas pos-baixa | {len(stats['pos_baixa'])} |",
        f"| Retencao estimada (5m) | R$ {stats['total_ret_5m']:,.2f} |",
        "",
        "## 2. Identificacao Cadastral",
        "",
        "### Dados da Pessoa Juridica (Contrato Social + Cartao CNPJ)",
        "",
        "| Campo | Valor |",
        "|---|---|",
    ]
    dados_pj = [
        ("Razao Social", EMPRESA["razao_social"]),
        ("Razao Anterior", EMPRESA["razao_anterior"]),
        ("Nome Fantasia", EMPRESA["nome_fantasia"]),
        ("CNPJ", EMPRESA["cnpj"]),
        ("Situacao", EMPRESA["situacao"]),
        ("Data Abertura", EMPRESA["data_abertura"]),
        ("Porte Declarado", EMPRESA["porte_declarado"]),
        ("Natureza Juridica", EMPRESA["natureza_juridica"]),
        ("Capital Social", f"R$ {EMPRESA['capital_social']:,.2f}"),
        ("CNAE Principal", EMPRESA["cnae_principal"]),
        ("CNAE Secundario", EMPRESA["cnae_secundario"]),
        ("Endereco Sede", EMPRESA["endereco_sede"]),
        ("Escritorio Admin", EMPRESA["endereco_admin"]),
        ("Email", EMPRESA["email"]),
        ("Telefones", EMPRESA["telefones"]),
        ("Ultima Alteracao", EMPRESA["ultima_alteracao"]),
    ]
    for k, v in dados_pj:
        lines.append(f"| **{k}** | {v} |")

    lines += [
        "",
        "### Quadro Societario",
        "",
        f"| Socio | CPF | Quotas | % |",
        f"|---|---|---:|---:|",
        f"| **{EMPRESA['socio_nome']}** | {EMPRESA['socio_cpf']} | 400.000 | **100%** |",
        "",
        f"- **Nascimento:** {EMPRESA['socio_nascimento']} (44 anos)",
        f"- **Endereco:** {EMPRESA['socio_endereco']}",
        "- **Funcao:** Administrador unico por prazo indeterminado",
        "",
        "## 3. Evolucao Mensal",
        "",
        "| Mes | Transacoes | Creditos (R$) | Debitos (R$) | Saldo Final |",
        "|---|---:|---:|---:|---:|",
    ]
    for mes in ["JAN/2026", "FEV/2026", "MAR/2026", "ABR/2026", "MAI/2026"]:
        s = stats["saldos"][mes]
        lines.append(f"| {mes} | {s['n']:,} | {s['cred']:,.2f} | {s['deb']:,.2f} | {s['saldo_final']:,.2f} |")
    lines.append(f"| **TOTAL** | **{n:,}** | **{cred:,.2f}** | **{deb:,.2f}** | **{saldo_fim:,.2f}** |")

    # Risk Heatmap
    lines += [
        "",
        "## 4. Risk Heatmap (Distribuicao por Classe)",
        "",
        "| Classe | Transacoes | % | Volume (R$) | Acao |",
        "|---|---:|---:|---:|---|",
    ]
    cores_md = {"CRITICO": "🔴", "ALTO": "🟠", "MEDIO": "🔵", "BAIXO": "🟢"}
    acoes_md = {"CRITICO": "Auditoria imediata", "ALTO": "Revisao prioritaria",
                "MEDIO": "Conferir em lote", "BAIXO": "Auto-aprovar"}
    total_qtd = sum(v[0] for v in stats["classe_counts"].values())
    for classe in ("CRITICO", "ALTO", "MEDIO", "BAIXO"):
        qtd, vol_c = stats["classe_counts"][classe]
        pct = 100 * qtd / max(total_qtd, 1)
        lines.append(f"| {cores_md[classe]} {classe} | {qtd:,} | {pct:.1f}% | {vol_c:,.2f} | {acoes_md[classe]} |")

    # Partes Relacionadas
    lines += [
        "",
        "## 5. Partes Relacionadas",
        "",
        "| Entidade | Trans | Creditos (R$) | Debitos (R$) | Volume (R$) |",
        "|---|---:|---:|---:|---:|",
    ]
    total_pr = 0.0
    for nome, dados in stats["fluxos"].items():
        v = dados["cred"] + abs(dados["deb"])
        lines.append(f"| {nome} | {dados['n']} | {dados['cred']:,.2f} | {dados['deb']:,.2f} | {v:,.2f} |")
        total_pr += v
    lines.append(f"| **TOTAL** | | | | **R$ {total_pr:,.2f}** |")

    # MEIs (com reclassificacao MEI-TAC vs Padrao)
    lines += [
        "",
        f"## 6. MEIs Fornecedores - Reclassificacao MEI-TAC vs Padrao",
        "",
        "**Aplicacao do limite correto** apos confirmacao do cliente que muitos sao caminhoneiros:",
        "",
        f"- **MEI-TAC** (caminhoneiros, CNAEs 4930-*, 5320-*, 4911-*): teto **R$ 251.600/ano** (LC 188/2021)",
        f"- **MEI Padrao** (outros CNAEs): teto **R$ 81.000/ano** (LC 123/2006)",
        "",
        f"### Casos acima do teto correspondente ({len(stats['meis'])} fornecedores)",
        "",
    ]
    if stats["meis"]:
        lines += [
            "| # | CNPJ | Razao Social | CNAE | Anualizado | Excesso |",
            "|---|---|---|---|---:|---:|",
        ]
        total_exc = 0.0
        for i, m in enumerate(stats["meis"][:15], start=1):
            fmt = f"{m['cnpj'][:2]}.{m['cnpj'][2:5]}.{m['cnpj'][5:8]}/{m['cnpj'][8:12]}-{m['cnpj'][12:14]}"
            cnae_desc = m.get("cnae_desc", "") or m.get("cnae", "")
            lines.append(f"| {i} | {fmt} | {m['razao'][:30]} | {cnae_desc[:25]} | **R$ {m['anualizado']:,.2f}** | R$ {m['excesso']:,.2f} |")
            total_exc += m["excesso"]
        lines.append(f"| | | | **TOTAL EXCESSO:** | | **R$ {total_exc:,.2f}** |")
    else:
        lines.append("**Nenhum MEI excede o teto legal correspondente.**")

    # Status Tributario
    lines += [
        "",
        "## 7. Status Tributario Consolidado",
        "",
        "| Categoria | Qtd | Volume (R$) | Retencao (R$) |",
        "|---|---:|---:|---:|",
    ]
    CATS_ORD = ["RETENCAO_PJ", "RETENCAO_PF", "OPERACAO_CREDITO", "IOF", "JUROS",
                "PAGAMENTO_TRIBUTO", "TARIFA", "PIX_RECEBIDO", "BOLETO",
                "COMPRA_CARTAO", "NAO_TRIBUTAVEL", "OUTRO"]
    for cat in CATS_ORD:
        qtd = stats["cat_count"].get(cat, 0)
        if qtd == 0:
            continue
        vol_c = stats["cat_volume"][cat]
        ret = stats["cat_retencao"][cat]
        ret_fmt = f"**{ret:,.2f}**" if ret > 0 else "—"
        lines.append(f"| {cat} | {qtd:,} | {vol_c:,.2f} | {ret_fmt} |")
    lines.append(f"| **TOTAL** | **{sum(stats['cat_count'].values()):,}** | | **R$ {stats['total_ret_5m']:,.2f}** |")

    # Pos-Baixa
    lines += [
        "",
        f"## 8. Pagamentos Pos-Baixa ({len(stats['pos_baixa'])} alertas)",
        "",
        "| Mes | Data | Valor (R$) | Razao Social | Data Baixa | Dias Apos |",
        "|---|---|---:|---|---|---:|",
    ]
    total_pb = 0.0
    for p in stats["pos_baixa"]:
        cnpj_fmt = f"{p['cnpj'][:2]}.{p['cnpj'][2:5]}.{p['cnpj'][5:8]}/{p['cnpj'][8:12]}-{p['cnpj'][12:14]}"
        razao = p["info"].get("razao_social", "")[:30]
        lines.append(f"| {p['mes']} | {p['t'].data} | **{p['t'].valor:,.2f}** | {cnpj_fmt} - {razao} | {p['info'].get('data_situacao','')} | **{p['dias']}** |")
        total_pb += abs(p["t"].valor)
    lines.append(f"| | | **R$ {total_pb:,.2f}** | TOTAL | | |")

    # Conclusao
    lines += [
        "",
        "## 9. Conclusao",
        "",
        "Os achados consolidados desta auditoria evidenciam **riscos tributarios e contabeis significativos** que demandam regularizacao imediata:",
        "",
        f"1. **Desenquadramento EPP retroativo** — empresa movimenta {vol * 12 / 4.5 / 4_800_000:.0f}x o limite anual permitido",
        f"2. **Retencoes na fonte nao recolhidas** — R$ {stats['total_ret_5m']:,.2f} estimados em 5 meses",
        f"3. **{len(stats['meis'])} MEIs com volume acima do teto** — risco de pejotizacao",
        f"4. **{len(stats['pos_baixa'])} pagamentos pos-baixa** — R$ {total_pb:,.2f} a fornecedor baixado",
        f"5. **R$ {total_pr:,.2f} com partes relacionadas** — necessita lastro contratual documentado",
        "",
        "Recomenda-se acionamento das medidas formais descritas na **Carta de Constatacao** (documento anexo).",
        "",
        "---",
        "",
        f"*Documento gerado pelo OrgConc/OrgNeural2 v0.5.0 - Sistema integrado de auditoria bancaria.*",
    ]

    return "\n".join(lines), {"total_pr": total_pr, "total_exc": total_exc, "total_pb": total_pb}


def gerar_html(md_text):
    import markdown as mdlib
    body = mdlib.markdown(md_text, extensions=["tables", "fenced_code"])
    agora = datetime.now().strftime("%d/%m/%Y %H:%M")
    css = """
@page { size: A4 landscape; margin: 14mm 12mm 14mm 12mm;
  @bottom-right { content: "Pagina " counter(page) " de " counter(pages); font-size: 9px; color: #6B7280; }
  @bottom-left { content: "Relatorio Integrado · LOCAR TRANSPORTE DE BOVINOS LTDA"; font-size: 9px; color: #6B7280; }
}
* { box-sizing: border-box; margin: 0; padding: 0; }
body { font-family: 'DejaVu Sans', Arial, sans-serif; font-size: 10pt; color: #1a202c; line-height: 1.55; }
.hd { background: linear-gradient(135deg, #0F172A, #0B1B3D 45%, #0052FF); color: #fff;
      padding: 24px 32px; border-radius: 4px; margin-bottom: 24px; display: flex; align-items: center; gap: 22px; }
.hd-text { flex: 1; }
.hd h1 { font-size: 22pt; font-family: 'DejaVu Serif', Georgia, serif; margin-bottom: 4px; letter-spacing: 1px; }
.hd .tag { font-size: 10pt; opacity: 0.9; text-transform: uppercase; letter-spacing: 0.18em; }
.hd .meta { font-size: 9pt; opacity: 0.85; margin-top: 8px; }
h1 { font-size: 14pt; color: #0F172A; margin: 28px 0 10px; padding-bottom: 8px; border-bottom: 2px solid #0052FF; font-family: 'DejaVu Serif', Georgia, serif; }
h2 { font-size: 12pt; color: #0F172A; margin: 22px 0 8px; padding: 8px 14px; background: linear-gradient(90deg, #F0F7FF, transparent); border-left: 4px solid #0052FF; }
h3 { font-size: 10.5pt; color: #0052FF; margin: 14px 0 6px; font-weight: 700; }
p { margin-bottom: 6px; text-align: justify; }
table { width: 100%; border-collapse: collapse; margin: 8px 0 14px; font-size: 9pt; border-radius: 6px; overflow: hidden; box-shadow: 0 1px 3px rgba(0,0,0,0.06); }
th { background: linear-gradient(180deg, #0F172A, #1E3A8A); color: #fff; padding: 6px 10px; text-align: left; font-weight: 600; }
td { padding: 5px 10px; border-bottom: 1px solid #E2E8F0; }
tr:nth-child(even) td { background: #F8FAFC; }
strong { color: #0F172A; font-weight: 700; }
ul, ol { padding-left: 22px; margin-bottom: 8px; }
li { margin-bottom: 3px; }
em { color: #64748B; font-size: 8.5pt; }
hr { border: none; border-top: 1px solid #CBD5E1; margin: 16px 0; }
"""
    return f"""<!DOCTYPE html><html lang="pt-BR"><head><meta charset="UTF-8">
<title>Relatorio Integrado · LOCAR TRANSPORTE DE BOVINOS LTDA</title><style>{css}</style></head>
<body>
<div class="hd">{html_logo_inline()}<div class="hd-text">
<h1>ORGATEC</h1>
<div class="tag">Relatorio Integrado · Auditoria Bancaria</div>
<div class="meta">LOCAR TRANSPORTE DE BOVINOS LTDA · CNPJ 05.509.396/0001-10 · Conta Sicoob 158083-3 · Gerado em {agora}</div>
</div></div>
{body}
</body></html>"""


async def gerar_pdf(html_text):
    try:
        from playwright.async_api import async_playwright
        async with async_playwright() as p:
            browser = await p.chromium.launch()
            page = await browser.new_page()
            await page.set_content(html_text, wait_until="load")
            await page.pdf(
                path=str(OUT_PDF), format="A4", landscape=True,
                margin={"top": "14mm", "right": "12mm", "bottom": "14mm", "left": "12mm"},
                print_background=True,
            )
            await browser.close()
        return True
    except Exception as exc:  # noqa: BLE001
        print(f"PDF failed: {exc}")
        return False


async def main_async():
    todos, saldos, cache = await coletar_dados()
    print("Gerando XLSX integrado (11 abas)...")
    stats = gerar_xlsx(todos, saldos, cache)

    print("Gerando MD...")
    md, totais = gerar_md(stats)
    OUT_MD.write_text(md, encoding="utf-8")
    print(f"  MD:   {OUT_MD}")

    print("Gerando HTML...")
    html = gerar_html(md)
    OUT_HTML.write_text(html, encoding="utf-8")
    print(f"  HTML: {OUT_HTML}")

    print("Gerando PDF...")
    if await gerar_pdf(html):
        print(f"  PDF:  {OUT_PDF}")


if __name__ == "__main__":
    asyncio.run(main_async())
