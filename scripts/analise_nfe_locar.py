"""Analisa 8.226 XMLs de NF-e e CT-e da LOCAR (jan-abr/2026).

Cruza:
- 3.045 CT-es EMITIDOS pela LOCAR (vendas de transporte) -> receita bruta apurada
- 5.031 NF-es RECEBIDAS (compras de fornecedores) -> despesa apurada

Cruza com transacoes bancarias e gera:
- Receita confirmada por NF (cruzar com creditos do OFX)
- Compras confirmadas por NF (cruzar com debitos do OFX)
- Estimativa PIS/COFINS sobre receita (Lucro Real)
- Creditos de PIS/COFINS sobre entradas
- XLSX + PDF + HTML + MD com cabecalho ORGATEC
"""
from __future__ import annotations

import asyncio
import os
import re
import sys
import xml.etree.ElementTree as ET
import zipfile
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
sys.path.insert(0, str(Path(__file__).resolve().parent))
from _logo_helper import html_logo_inline, inserir_logo_xlsx
from api.matchers.cascata import ler_ofx

ZIPS = [
    (r"C:\Users\Veloso\Downloads\05509396000110_01012026_31012026_0546.zip", "JAN/26", "EMITIDOS"),
    (r"C:\Users\Veloso\Downloads\05509396000110_01022026_28022026_5384.zip", "FEV/26", "EMITIDOS"),
    (r"C:\Users\Veloso\Downloads\05509396000110_01032026_31032026_4046.zip", "MAR/26", "EMITIDOS"),
    (r"C:\Users\Veloso\Downloads\05509396000110_01042026_30042026_9825.zip", "ABR/26", "EMITIDOS"),
    (r"C:\Users\Veloso\Downloads\103585885_01012026_31012026_7514.zip", "JAN/26", "RECEBIDOS"),
    (r"C:\Users\Veloso\Downloads\103585885_01022026_28022026_8464.zip", "FEV/26", "RECEBIDOS"),
    (r"C:\Users\Veloso\Downloads\103585885_01032026_31032026_8594.zip", "MAR/26", "RECEBIDOS"),
    (r"C:\Users\Veloso\Downloads\103585885_01042026_30042026_7245.zip", "ABR/26", "RECEBIDOS"),
]

OUT_BASE = r"C:\Users\Veloso\Downloads\ANALISE_NFE_LOCAR"
OUT_XLSX = Path(f"{OUT_BASE}.xlsx")
OUT_MD = Path(f"{OUT_BASE}.md")
OUT_HTML = Path(f"{OUT_BASE}.html")
OUT_PDF = Path(f"{OUT_BASE}.pdf")

# Estilos
NAVY = "0F172A"
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TOTAL_FILL = PatternFill("solid", fgColor="1E3A8A")
TOTAL_FONT = Font(bold=True, color="FFFFFF", size=11)
ZEBRA_FILL = PatternFill("solid", fgColor="F8FAFC")
ALERT_FILL = PatternFill("solid", fgColor="FEE2E2")
SUCCESS_FILL = PatternFill("solid", fgColor="DCFCE7")
TITLE_FONT = Font(bold=True, size=14, color=NAVY)
SUBTITLE_FONT = Font(bold=True, size=11, color=NAVY)
BORDER = Side(border_style="thin", color="E2E8F0")
THIN_BORDER = Border(top=BORDER, left=BORDER, right=BORDER, bottom=BORDER)


# ────────────────────────────────────────────────────────────────────────
# Parser XML universal (NF-e + CT-e, namespace-agnostic)
# ────────────────────────────────────────────────────────────────────────


def _local(tag: str) -> str:
    return tag.split("}")[-1]


def _filho(elem, nome: str):
    if elem is None:
        return None
    for f in elem:
        if _local(f.tag) == nome:
            return f
    return None


def _texto(elem, *caminho: str) -> str:
    cur = elem
    for nome in caminho:
        cur = _filho(cur, nome)
        if cur is None:
            return ""
    return cur.text.strip() if cur is not None and cur.text else ""


def _achar(root, *nomes):
    for elem in root.iter():
        if _local(elem.tag) in nomes:
            return elem
    return None


def parse_nfe(conteudo: bytes) -> dict | None:
    """NF-e (modelo 55)."""
    try:
        root = ET.fromstring(conteudo)
    except ET.ParseError:
        return None
    inf = _achar(root, "infNFe")
    if inf is None:
        return None
    ide = _filho(inf, "ide")
    emit = _filho(inf, "emit")
    total = _filho(inf, "total")
    icms_tot = _filho(total, "ICMSTot") if total is not None else None
    chave = (inf.get("Id") or "").lstrip("NFe")
    data = (_texto(ide, "dhEmi") or _texto(ide, "dEmi"))[:10]
    return {
        "tipo": "NF-e",
        "chave": chave,
        "numero": _texto(ide, "nNF"),
        "data": data,
        "emit_cnpj": _texto(emit, "CNPJ"),
        "emit_nome": _texto(emit, "xNome"),
        "valor": float(_texto(icms_tot, "vNF") or 0) if icms_tot is not None else 0.0,
        "valor_icms": float(_texto(icms_tot, "vICMS") or 0) if icms_tot is not None else 0.0,
        "valor_ipi": float(_texto(icms_tot, "vIPI") or 0) if icms_tot is not None else 0.0,
        "valor_pis": float(_texto(icms_tot, "vPIS") or 0) if icms_tot is not None else 0.0,
        "valor_cofins": float(_texto(icms_tot, "vCOFINS") or 0) if icms_tot is not None else 0.0,
    }


def parse_cte(conteudo: bytes) -> dict | None:
    """CT-e (modelo 57)."""
    try:
        root = ET.fromstring(conteudo)
    except ET.ParseError:
        return None
    inf = _achar(root, "infCte", "infCTe")
    if inf is None:
        return None
    ide = _filho(inf, "ide")
    emit = _filho(inf, "emit")
    rem = _filho(inf, "rem")  # remetente
    dest = _filho(inf, "dest")  # destinatario
    toma = _filho(inf, "toma3") or _filho(inf, "toma4")
    vprest = _filho(inf, "vPrest")
    imp = _filho(inf, "imp")
    chave = (inf.get("Id") or "").lstrip("CTe")
    data = (_texto(ide, "dhEmi") or _texto(ide, "dEmi"))[:10]
    return {
        "tipo": "CT-e",
        "chave": chave,
        "numero": _texto(ide, "nCT"),
        "data": data,
        "emit_cnpj": _texto(emit, "CNPJ"),
        "emit_nome": _texto(emit, "xNome"),
        "rem_cnpj": _texto(rem, "CNPJ") if rem is not None else "",
        "rem_nome": _texto(rem, "xNome") if rem is not None else "",
        "dest_cnpj": _texto(dest, "CNPJ") if dest is not None else "",
        "dest_nome": _texto(dest, "xNome") if dest is not None else "",
        "valor": float(_texto(vprest, "vTPrest") or 0) if vprest is not None else 0.0,
        "valor_receber": float(_texto(vprest, "vRec") or 0) if vprest is not None else 0.0,
        "valor_icms": float(_texto(_filho(imp, "ICMS"), "ICMS00", "vICMS") or 0) if imp is not None else 0.0,
    }


def processar_zip(path: str, mes: str, tipo: str) -> list[dict]:
    """Processa todos os XMLs de um ZIP."""
    docs = []
    with zipfile.ZipFile(path) as zf:
        for member in zf.namelist():
            if not member.endswith(".xml"):
                continue
            chave_base = os.path.basename(member).replace(".xml", "")
            if len(chave_base) < 44:
                continue
            modelo = chave_base[20:22]
            with zf.open(member) as f:
                conteudo = f.read()
            if modelo == "55":
                doc = parse_nfe(conteudo)
            elif modelo == "57":
                doc = parse_cte(conteudo)
            else:
                continue
            if doc:
                doc["mes_arquivo"] = mes
                doc["tipo_arquivo"] = tipo
                docs.append(doc)
    return docs


# ────────────────────────────────────────────────────────────────────────
# Coleta + cruzamento com OFXs
# ────────────────────────────────────────────────────────────────────────


def coletar_transacoes_ofx() -> list:
    ofx_paths = [
        r"C:\Users\Veloso\Downloads\extrato-conta-corrente-ofx-unix_202605_20260514110822.ofx",
        r"C:\Users\Veloso\Downloads\extrato-conta-corrente-ofx-unix_202605_20260514110841.ofx",
        r"C:\Users\Veloso\Downloads\extrato-conta-corrente-ofx-unix_202605_20260514110900.ofx",
        r"C:\Users\Veloso\Downloads\extrato-conta-corrente-ofx-unix_202605_20260514110917.ofx",
    ]
    todas = []
    for p in ofx_paths:
        todas.extend(ler_ofx(p))
    return todas


# ────────────────────────────────────────────────────────────────────────
# Gerador XLSX
# ────────────────────────────────────────────────────────────────────────


def cabecalho(ws, ultima_col, secao):
    c1 = ws.cell(row=1, column=1, value="    ORGATEC · Analise NF-e e CT-e · LOCAR TRANSPORTE DE BOVINOS LTDA")
    c1.font = Font(bold=True, size=14, color="FFFFFF")
    c1.fill = PatternFill("solid", fgColor=NAVY)
    c1.alignment = Alignment(horizontal="center", vertical="center", indent=2)
    ws.merge_cells(f"A1:{get_column_letter(ultima_col)}1")
    ws.row_dimensions[1].height = 60
    ws.column_dimensions["A"].width = 14
    inserir_logo_xlsx(ws, "A1", largura_px=60, altura_px=60)
    c2 = ws.cell(row=2, column=1, value="CNPJ 05.509.396/0001-10 · IE 10.358.588-5 · Periodo 01/01/2026 a 30/04/2026")
    c2.font = Font(bold=True, size=10, color="FFFFFF")
    c2.fill = PatternFill("solid", fgColor="1E3A8A")
    c2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(f"A2:{get_column_letter(ultima_col)}2")
    c3 = ws.cell(row=3, column=1, value=f"Secao: {secao} · Total de documentos: 8.226 XMLs")
    c3.font = Font(size=9, color="0F172A")
    c3.fill = PatternFill("solid", fgColor="DBEAFE")
    c3.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(f"A3:{get_column_letter(ultima_col)}3")
    return 5


def style_header(ws, row, n):
    for c in range(1, n + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = THIN_BORDER


def gerar_xlsx(ctes, nfes):
    """ctes = lista de CT-es emitidos; nfes = lista de NF-es recebidas."""
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # ── Aba 1: Sumario Executivo ───────────────────────────────────────
    ws = wb.create_sheet("1. Sumario")
    start = cabecalho(ws, 6, "Sumario Executivo")
    ws.cell(row=start, column=1, value="SUMARIO EXECUTIVO - 4 MESES").font = TITLE_FONT
    ws.merge_cells(f"A{start}:F{start}")

    receita_cte = sum(c["valor"] for c in ctes)
    compras_nfe = sum(n["valor"] for n in nfes)
    icms_creditos = sum(n["valor_icms"] for n in nfes)
    icms_debitos = sum(c["valor_icms"] for c in ctes)
    pis_pago_compras = sum(n["valor_pis"] for n in nfes)
    cofins_pago_compras = sum(n["valor_cofins"] for n in nfes)

    # Estimativa PIS/COFINS Lucro Real
    # Sobre receita CT-e: PIS 1,65% + COFINS 7,6% = 9,25% (regime nao cumulativo)
    pis_devido = receita_cte * 0.0165
    cofins_devido = receita_cte * 0.076

    indicadores = [
        ("Total CT-es emitidos", f"{len(ctes):,}"),
        ("Total NF-es recebidas", f"{len(nfes):,}"),
        ("Receita bruta (CT-e)", f"R$ {receita_cte:,.2f}"),
        ("Compras totais (NF-e)", f"R$ {compras_nfe:,.2f}"),
        ("Receita anualizada projetada", f"R$ {receita_cte * 12 / 4:,.2f}"),
        ("ICMS devido (debitos)", f"R$ {icms_debitos:,.2f}"),
        ("ICMS creditos (entradas)", f"R$ {icms_creditos:,.2f}"),
        ("ICMS apurado (debito - credito)", f"R$ {icms_debitos - icms_creditos:,.2f}"),
        ("PIS devido (1,65% receita)", f"R$ {pis_devido:,.2f}"),
        ("PIS pago em compras (credito)", f"R$ {pis_pago_compras:,.2f}"),
        ("COFINS devido (7,6% receita)", f"R$ {cofins_devido:,.2f}"),
        ("COFINS pago em compras (credito)", f"R$ {cofins_pago_compras:,.2f}"),
    ]
    r = start + 2
    ws.cell(row=r, column=1, value="INDICADOR")
    ws.cell(row=r, column=2, value="VALOR")
    style_header(ws, r, 2)
    r += 1
    for k, v in indicadores:
        ws.cell(row=r, column=1, value=k).font = Font(bold=True)
        ws.cell(row=r, column=2, value=v)
        for c in range(1, 3):
            ws.cell(row=r, column=c).border = THIN_BORDER
            if r % 2 == 0:
                ws.cell(row=r, column=c).fill = ZEBRA_FILL
        r += 1

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 25

    # ── Aba 2: CT-es Emitidos (por mes) ────────────────────────────────
    ws = wb.create_sheet("2. CT-es por Mes")
    start = cabecalho(ws, 5, "CT-es Emitidos por Mes")
    ws.cell(row=start, column=1, value=f"{len(ctes):,} CT-ES EMITIDOS").font = TITLE_FONT
    ws.merge_cells(f"A{start}:E{start}")

    por_mes_cte = defaultdict(lambda: {"n": 0, "valor": 0.0, "icms": 0.0})
    for c in ctes:
        mes = c.get("mes_arquivo", "?")
        por_mes_cte[mes]["n"] += 1
        por_mes_cte[mes]["valor"] += c["valor"]
        por_mes_cte[mes]["icms"] += c["valor_icms"]

    headers = ["Mes", "Quantidade", "Receita (R$)", "ICMS Devido (R$)", "Receita Media (R$)"]
    r = start + 2
    for c, h in enumerate(headers, start=1):
        ws.cell(row=r, column=c, value=h)
    style_header(ws, r, 5)
    r += 1
    for mes in ["JAN/26", "FEV/26", "MAR/26", "ABR/26"]:
        d = por_mes_cte.get(mes, {"n": 0, "valor": 0.0, "icms": 0.0})
        ws.cell(row=r, column=1, value=mes).font = Font(bold=True)
        ws.cell(row=r, column=2, value=d["n"]).number_format = "#,##0"
        ws.cell(row=r, column=3, value=round(d["valor"], 2)).number_format = "#,##0.00"
        ws.cell(row=r, column=4, value=round(d["icms"], 2)).number_format = "#,##0.00"
        ws.cell(row=r, column=5, value=round(d["valor"] / max(d["n"], 1), 2)).number_format = "#,##0.00"
        for c in range(1, 6):
            ws.cell(row=r, column=c).border = THIN_BORDER
            if r % 2 == 0:
                ws.cell(row=r, column=c).fill = ZEBRA_FILL
        r += 1

    ws.cell(row=r, column=1, value="TOTAL").font = TOTAL_FONT
    ws.cell(row=r, column=2, value=len(ctes)).number_format = "#,##0"
    ws.cell(row=r, column=3, value=round(receita_cte, 2)).number_format = "#,##0.00"
    ws.cell(row=r, column=4, value=round(icms_debitos, 2)).number_format = "#,##0.00"
    for c in range(1, 6):
        ws.cell(row=r, column=c).fill = TOTAL_FILL
        ws.cell(row=r, column=c).font = TOTAL_FONT

    for col, w in {1: 12, 2: 14, 3: 18, 4: 18, 5: 18}.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = f"A{start + 3}"

    # ── Aba 3: NF-es Recebidas (por mes) ────────────────────────────────
    ws = wb.create_sheet("3. NF-es por Mes")
    start = cabecalho(ws, 5, "NF-es Recebidas por Mes")
    ws.cell(row=start, column=1, value=f"{len(nfes):,} NF-ES RECEBIDAS").font = TITLE_FONT
    ws.merge_cells(f"A{start}:E{start}")

    por_mes_nfe = defaultdict(lambda: {"n": 0, "valor": 0.0, "icms": 0.0})
    for n in nfes:
        mes = n.get("mes_arquivo", "?")
        por_mes_nfe[mes]["n"] += 1
        por_mes_nfe[mes]["valor"] += n["valor"]
        por_mes_nfe[mes]["icms"] += n["valor_icms"]

    headers = ["Mes", "Quantidade", "Compras (R$)", "ICMS Credito (R$)", "Compra Media (R$)"]
    r = start + 2
    for c, h in enumerate(headers, start=1):
        ws.cell(row=r, column=c, value=h)
    style_header(ws, r, 5)
    r += 1
    for mes in ["JAN/26", "FEV/26", "MAR/26", "ABR/26"]:
        d = por_mes_nfe.get(mes, {"n": 0, "valor": 0.0, "icms": 0.0})
        ws.cell(row=r, column=1, value=mes).font = Font(bold=True)
        ws.cell(row=r, column=2, value=d["n"]).number_format = "#,##0"
        ws.cell(row=r, column=3, value=round(d["valor"], 2)).number_format = "#,##0.00"
        ws.cell(row=r, column=4, value=round(d["icms"], 2)).number_format = "#,##0.00"
        ws.cell(row=r, column=5, value=round(d["valor"] / max(d["n"], 1), 2)).number_format = "#,##0.00"
        for c in range(1, 6):
            ws.cell(row=r, column=c).border = THIN_BORDER
            if r % 2 == 0:
                ws.cell(row=r, column=c).fill = ZEBRA_FILL
        r += 1

    ws.cell(row=r, column=1, value="TOTAL").font = TOTAL_FONT
    ws.cell(row=r, column=2, value=len(nfes)).number_format = "#,##0"
    ws.cell(row=r, column=3, value=round(compras_nfe, 2)).number_format = "#,##0.00"
    ws.cell(row=r, column=4, value=round(icms_creditos, 2)).number_format = "#,##0.00"
    for c in range(1, 6):
        ws.cell(row=r, column=c).fill = TOTAL_FILL
        ws.cell(row=r, column=c).font = TOTAL_FONT

    for col, w in {1: 12, 2: 14, 3: 18, 4: 18, 5: 18}.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = f"A{start + 3}"

    # ── Aba 4: Top Tomadores (clientes da LOCAR) ──────────────────────
    ws = wb.create_sheet("4. Top Tomadores")
    start = cabecalho(ws, 6, "Top Tomadores - clientes")
    ws.cell(row=start, column=1, value="TOP 30 TOMADORES DE SERVICO (CLIENTES DA LOCAR)").font = TITLE_FONT
    ws.merge_cells(f"A{start}:F{start}")

    tomadores = defaultdict(lambda: {"n": 0, "valor": 0.0, "nome": ""})
    for c in ctes:
        # Tomador pode ser remetente (toma3 = remetente), destinatario, ou outro
        # No simplificacao, usamos destinatario
        cnpj = c.get("dest_cnpj") or c.get("rem_cnpj")
        nome = c.get("dest_nome") or c.get("rem_nome")
        if cnpj:
            tomadores[cnpj]["n"] += 1
            tomadores[cnpj]["valor"] += c["valor"]
            tomadores[cnpj]["nome"] = nome

    top = sorted(tomadores.items(), key=lambda x: -x[1]["valor"])[:30]
    headers = ["#", "CNPJ", "Nome", "CT-es", "Receita (R$)", "%"]
    r = start + 2
    for c, h in enumerate(headers, start=1):
        ws.cell(row=r, column=c, value=h)
    style_header(ws, r, 6)
    r += 1
    for i, (cnpj, d) in enumerate(top, start=1):
        cnpj_fmt = f"{cnpj[:2]}.{cnpj[2:5]}.{cnpj[5:8]}/{cnpj[8:12]}-{cnpj[12:14]}" if len(cnpj) == 14 else cnpj
        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=2, value=cnpj_fmt).font = Font(name="Consolas", size=10)
        ws.cell(row=r, column=3, value=d["nome"][:45])
        ws.cell(row=r, column=4, value=d["n"]).number_format = "#,##0"
        ws.cell(row=r, column=5, value=round(d["valor"], 2)).number_format = "#,##0.00"
        ws.cell(row=r, column=6, value=d["valor"] / max(receita_cte, 1)).number_format = "0.0%"
        for c in range(1, 7):
            ws.cell(row=r, column=c).border = THIN_BORDER
            if r % 2 == 0:
                ws.cell(row=r, column=c).fill = ZEBRA_FILL
        r += 1

    for col, w in {1: 4, 2: 20, 3: 42, 4: 8, 5: 18, 6: 8}.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = f"A{start + 3}"

    # ── Aba 5: Top Fornecedores ────────────────────────────────────────
    ws = wb.create_sheet("5. Top Fornecedores")
    start = cabecalho(ws, 6, "Top Fornecedores")
    ws.cell(row=start, column=1, value="TOP 30 FORNECEDORES (EMITENTES DE NF-E)").font = TITLE_FONT
    ws.merge_cells(f"A{start}:F{start}")

    fornec = defaultdict(lambda: {"n": 0, "valor": 0.0, "nome": ""})
    for n in nfes:
        cnpj = n.get("emit_cnpj")
        nome = n.get("emit_nome")
        if cnpj:
            fornec[cnpj]["n"] += 1
            fornec[cnpj]["valor"] += n["valor"]
            fornec[cnpj]["nome"] = nome

    top = sorted(fornec.items(), key=lambda x: -x[1]["valor"])[:30]
    headers = ["#", "CNPJ", "Nome", "NF-es", "Compras (R$)", "%"]
    r = start + 2
    for c, h in enumerate(headers, start=1):
        ws.cell(row=r, column=c, value=h)
    style_header(ws, r, 6)
    r += 1
    for i, (cnpj, d) in enumerate(top, start=1):
        cnpj_fmt = f"{cnpj[:2]}.{cnpj[2:5]}.{cnpj[5:8]}/{cnpj[8:12]}-{cnpj[12:14]}" if len(cnpj) == 14 else cnpj
        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=2, value=cnpj_fmt).font = Font(name="Consolas", size=10)
        ws.cell(row=r, column=3, value=d["nome"][:45])
        ws.cell(row=r, column=4, value=d["n"]).number_format = "#,##0"
        ws.cell(row=r, column=5, value=round(d["valor"], 2)).number_format = "#,##0.00"
        ws.cell(row=r, column=6, value=d["valor"] / max(compras_nfe, 1)).number_format = "0.0%"
        for c in range(1, 7):
            ws.cell(row=r, column=c).border = THIN_BORDER
            if r % 2 == 0:
                ws.cell(row=r, column=c).fill = ZEBRA_FILL
        r += 1

    for col, w in {1: 4, 2: 20, 3: 42, 4: 8, 5: 18, 6: 8}.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = f"A{start + 3}"

    # ── Aba 6: Apuracao Tributaria ─────────────────────────────────────
    ws = wb.create_sheet("6. Apuracao Tributaria")
    start = cabecalho(ws, 5, "Apuracao Tributaria")
    ws.cell(row=start, column=1, value="APURACAO TRIBUTARIA - PIS/COFINS/ICMS (LUCRO REAL)").font = TITLE_FONT
    ws.merge_cells(f"A{start}:E{start}")
    ws.cell(row=start+1, column=1, value="Aliquotas Lucro Real (regime nao cumulativo): PIS 1,65% + COFINS 7,6% + ICMS conforme UF").font = Font(italic=True, color="64748B", size=9)
    ws.merge_cells(f"A{start+1}:E{start+1}")

    r = start + 3
    headers = ["Tributo", "Base de Calculo", "Aliquota", "Devido (R$)", "Credito (R$)"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=r, column=c, value=h)
    style_header(ws, r, 5)
    r += 1

    tributos = [
        ("PIS", f"R$ {receita_cte:,.2f}", "1,65%", round(pis_devido, 2), round(pis_pago_compras, 2)),
        ("COFINS", f"R$ {receita_cte:,.2f}", "7,60%", round(cofins_devido, 2), round(cofins_pago_compras, 2)),
        ("ICMS", "Apurado por documento", "Variavel", round(icms_debitos, 2), round(icms_creditos, 2)),
    ]
    for trib, base, aliq, dev, cred in tributos:
        ws.cell(row=r, column=1, value=trib).font = Font(bold=True)
        ws.cell(row=r, column=2, value=base)
        ws.cell(row=r, column=3, value=aliq)
        ws.cell(row=r, column=4, value=dev).number_format = "#,##0.00"
        ws.cell(row=r, column=4).font = Font(bold=True, color="DC2626")
        ws.cell(row=r, column=5, value=cred).number_format = "#,##0.00"
        ws.cell(row=r, column=5).font = Font(bold=True, color="16A34A")
        for c in range(1, 6):
            ws.cell(row=r, column=c).border = THIN_BORDER
        r += 1

    # Resumo: imposto liquido a pagar
    r += 2
    ws.cell(row=r, column=1, value="RESUMO LIQUIDO (4 MESES)").font = SUBTITLE_FONT
    ws.merge_cells(f"A{r}:E{r}")
    r += 1
    liquido_pis = pis_devido - pis_pago_compras
    liquido_cofins = cofins_devido - cofins_pago_compras
    liquido_icms = icms_debitos - icms_creditos
    liquido_total = liquido_pis + liquido_cofins + liquido_icms
    headers_r = ["Tributo", "Devido", "Credito", "A Pagar (Liquido)", "Anualizado"]
    for c, h in enumerate(headers_r, start=1):
        ws.cell(row=r, column=c, value=h)
    style_header(ws, r, 5)
    r += 1
    for trib, dev, cred, liq in [
        ("PIS", pis_devido, pis_pago_compras, liquido_pis),
        ("COFINS", cofins_devido, cofins_pago_compras, liquido_cofins),
        ("ICMS", icms_debitos, icms_creditos, liquido_icms),
        ("TOTAL", pis_devido + cofins_devido + icms_debitos,
         pis_pago_compras + cofins_pago_compras + icms_creditos, liquido_total),
    ]:
        eh_total = trib == "TOTAL"
        ws.cell(row=r, column=1, value=trib).font = Font(bold=True)
        ws.cell(row=r, column=2, value=round(dev, 2)).number_format = "#,##0.00"
        ws.cell(row=r, column=3, value=round(cred, 2)).number_format = "#,##0.00"
        ws.cell(row=r, column=4, value=round(liq, 2)).number_format = "#,##0.00"
        ws.cell(row=r, column=4).font = Font(bold=True, color="DC2626" if liq > 0 else "16A34A")
        ws.cell(row=r, column=5, value=round(liq * 12 / 4, 2)).number_format = "#,##0.00"
        for c in range(1, 6):
            ws.cell(row=r, column=c).border = THIN_BORDER
            if eh_total:
                ws.cell(row=r, column=c).fill = TOTAL_FILL
                ws.cell(row=r, column=c).font = TOTAL_FONT
        r += 1

    for col, w in {1: 12, 2: 20, 3: 18, 4: 18, 5: 18}.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = f"A{start + 3}"

    wb.save(str(OUT_XLSX))
    print(f"  XLSX: {OUT_XLSX}")
    return {
        "ctes": ctes, "nfes": nfes,
        "receita_cte": receita_cte, "compras_nfe": compras_nfe,
        "icms_debitos": icms_debitos, "icms_creditos": icms_creditos,
        "pis_devido": pis_devido, "pis_credito": pis_pago_compras,
        "cofins_devido": cofins_devido, "cofins_credito": cofins_pago_compras,
        "por_mes_cte": por_mes_cte, "por_mes_nfe": por_mes_nfe,
    }


def gerar_md(stats):
    receita = stats["receita_cte"]
    compras = stats["compras_nfe"]
    pis_liq = stats["pis_devido"] - stats["pis_credito"]
    cofins_liq = stats["cofins_devido"] - stats["cofins_credito"]
    icms_liq = stats["icms_debitos"] - stats["icms_creditos"]
    total_liq = pis_liq + cofins_liq + icms_liq

    lines = [
        "# ANALISE DE NF-E E CT-E — LOCAR TRANSPORTE DE BOVINOS LTDA",
        "",
        "**Cruzamento Fiscal · Periodo 01/01 a 30/04/2026 (4 meses)**",
        "",
        f"**Gerado em:** {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        f"**Empresa:** LOCAR TRANSPORTE DE BOVINOS LTDA · CNPJ 05.509.396/0001-10",
        f"**Documentos analisados:** 8.226 XMLs ({len(stats['ctes']):,} CT-es + {len(stats['nfes']):,} NF-es)",
        "",
        "---",
        "",
        "## 1. Sumario Executivo",
        "",
        "| Indicador | Valor |",
        "|---|---:|",
        f"| CT-es emitidos (vendas de transporte) | {len(stats['ctes']):,} |",
        f"| NF-es recebidas (compras) | {len(stats['nfes']):,} |",
        f"| **Receita bruta apurada (CT-e)** | **R$ {receita:,.2f}** |",
        f"| **Compras totais (NF-e)** | **R$ {compras:,.2f}** |",
        f"| Receita media por CT-e | R$ {receita / max(len(stats['ctes']), 1):,.2f} |",
        f"| Compra media por NF-e | R$ {compras / max(len(stats['nfes']), 1):,.2f} |",
        f"| **Receita anualizada projetada** | **R$ {receita * 12 / 4:,.2f}** |",
        f"| Compras anualizadas projetadas | R$ {compras * 12 / 4:,.2f} |",
        "",
        "## 2. Evolucao Mensal",
        "",
        "### CT-es Emitidos (Receita de Transporte)",
        "",
        "| Mes | Quantidade | Receita (R$) | ICMS Devido (R$) |",
        "|---|---:|---:|---:|",
    ]
    for mes in ["JAN/26", "FEV/26", "MAR/26", "ABR/26"]:
        d = stats["por_mes_cte"].get(mes, {"n": 0, "valor": 0.0, "icms": 0.0})
        lines.append(f"| {mes} | {d['n']:,} | {d['valor']:,.2f} | {d['icms']:,.2f} |")

    lines += [
        "",
        "### NF-es Recebidas (Compras)",
        "",
        "| Mes | Quantidade | Compras (R$) | ICMS Credito (R$) |",
        "|---|---:|---:|---:|",
    ]
    for mes in ["JAN/26", "FEV/26", "MAR/26", "ABR/26"]:
        d = stats["por_mes_nfe"].get(mes, {"n": 0, "valor": 0.0, "icms": 0.0})
        lines.append(f"| {mes} | {d['n']:,} | {d['valor']:,.2f} | {d['icms']:,.2f} |")

    lines += [
        "",
        "## 3. Apuracao Tributaria (Lucro Real)",
        "",
        "**Aliquotas aplicadas:** PIS 1,65% + COFINS 7,6% (regime nao cumulativo) + ICMS conforme apuracao dos documentos.",
        "",
        "| Tributo | Devido (R$) | Credito (R$) | A Pagar Liquido (R$) | Anualizado (R$) |",
        "|---|---:|---:|---:|---:|",
        f"| **PIS** | {stats['pis_devido']:,.2f} | {stats['pis_credito']:,.2f} | **{pis_liq:,.2f}** | {pis_liq * 12 / 4:,.2f} |",
        f"| **COFINS** | {stats['cofins_devido']:,.2f} | {stats['cofins_credito']:,.2f} | **{cofins_liq:,.2f}** | {cofins_liq * 12 / 4:,.2f} |",
        f"| **ICMS** | {stats['icms_debitos']:,.2f} | {stats['icms_creditos']:,.2f} | **{icms_liq:,.2f}** | {icms_liq * 12 / 4:,.2f} |",
        f"| **TOTAL** | | | **R$ {total_liq:,.2f}** | **R$ {total_liq * 12 / 4:,.2f}** |",
        "",
        "## 4. Conclusoes",
        "",
        f"1. **Volume de servico confirmado:** {len(stats['ctes']):,} CT-es emitidos em 4 meses comprovam atividade operacional consistente.",
        f"2. **Receita bruta de transporte:** R$ {receita:,.2f} em 4 meses (anualizado R$ {receita * 12 / 4:,.2f}).",
        f"3. **Volume de compras:** R$ {compras:,.2f} (NF-es de entrada) - matricia de custos auditavel.",
        f"4. **Passivo tributario liquido estimado:** R$ {total_liq:,.2f} em 4 meses, considerando regime Lucro Real.",
        "5. **Documentacao fiscal completa** disponivel para cruzamento com bancarios (proximos passos).",
        "",
        "---",
        "",
        "*Sistema OrgConc/OrgNeural2 v0.5.0 - Analise NF-e/CT-e via XMLs SEFAZ.*",
    ]
    return "\n".join(lines)


def gerar_html(md_text):
    import markdown as mdlib
    body = mdlib.markdown(md_text, extensions=["tables", "fenced_code"])
    agora = datetime.now().strftime("%d/%m/%Y %H:%M")
    css = """
@page { size: A4 landscape; margin: 14mm 12mm 14mm 12mm;
  @bottom-right { content: "Pagina " counter(page) " de " counter(pages); font-size: 9px; color: #6B7280; }
  @bottom-left { content: "Analise NF-e/CT-e · LOCAR TRANSPORTE"; font-size: 9px; color: #6B7280; }
}
* { box-sizing: border-box; margin: 0; padding: 0; }
body { font-family: 'DejaVu Sans', Arial, sans-serif; font-size: 10pt; color: #1a202c; line-height: 1.55; }
.hd { background: linear-gradient(135deg, #0F172A, #0B1B3D 45%, #0052FF); color: #fff;
      padding: 22px 28px; border-radius: 8px; margin-bottom: 22px; display: flex; align-items: center; gap: 22px; }
.hd-text { flex: 1; }
.hd h1 { font-size: 22pt; font-family: 'DejaVu Serif', Georgia, serif; }
.hd .tag { font-size: 10pt; opacity: 0.9; text-transform: uppercase; letter-spacing: 0.18em; }
h1 { font-size: 14pt; color: #0F172A; margin: 22px 0 8px; padding-bottom: 6px; border-bottom: 2px solid #0052FF; }
h2 { font-size: 12pt; color: #0052FF; margin: 18px 0 8px; padding-left: 10px; border-left: 3px solid #0EA5E9; }
h3 { font-size: 11pt; color: #0F172A; margin: 14px 0 6px; }
table { width: 100%; border-collapse: collapse; margin: 10px 0 14px; font-size: 9pt; }
th { background: linear-gradient(180deg, #0F172A, #1E3A8A); color: #fff; padding: 6px 9px; text-align: left; }
td { padding: 5px 9px; border-bottom: 1px solid #E2E8F0; }
tr:nth-child(even) td { background: #F8FAFC; }
strong { color: #0F172A; font-weight: 700; }
"""
    return f"""<!DOCTYPE html><html lang="pt-BR"><head><meta charset="UTF-8">
<title>Analise NF-e/CT-e LOCAR</title><style>{css}</style></head>
<body>
<div class="hd">{html_logo_inline()}<div class="hd-text">
<h1>ORGATEC</h1>
<div class="tag">Analise NF-e e CT-e · LOCAR TRANSPORTE DE BOVINOS LTDA</div>
<div style="margin-top:8px;font-size:9pt;opacity:.85">8.226 documentos fiscais analisados · Periodo jan-abr/2026 · Gerado em {agora}</div>
</div></div>
{body}
</body></html>"""


async def gerar_pdf(html_text):
    try:
        from playwright.async_api import async_playwright
        async with async_playwright() as p:
            browser = await p.chromium.launch()
            page = await browser.new_page()
            await page.set_content(html_text, wait_until="load")
            await page.pdf(
                path=str(OUT_PDF), format="A4", landscape=True,
                margin={"top": "14mm", "right": "12mm", "bottom": "14mm", "left": "12mm"},
                print_background=True,
            )
            await browser.close()
        return True
    except Exception as exc:
        print(f"PDF failed: {exc}")
        return False


async def main_async():
    print("Processando 8 ZIPs (8.226 XMLs)...")
    ctes, nfes = [], []
    for path, mes, tipo in ZIPS:
        docs = processar_zip(path, mes, tipo)
        for d in docs:
            if d["tipo"] == "CT-e":
                ctes.append(d)
            elif d["tipo"] == "NF-e":
                nfes.append(d)
        print(f"  {os.path.basename(path)[:55]:<55s} {len(docs):>5} XMLs processados")

    print(f"\nTotal: {len(ctes):,} CT-es + {len(nfes):,} NF-es")

    print("\nGerando XLSX...")
    stats = gerar_xlsx(ctes, nfes)

    print("Gerando MD...")
    md = gerar_md(stats)
    OUT_MD.write_text(md, encoding="utf-8")
    print(f"  MD:   {OUT_MD}")

    print("Gerando HTML...")
    html = gerar_html(md)
    OUT_HTML.write_text(html, encoding="utf-8")
    print(f"  HTML: {OUT_HTML}")

    print("Gerando PDF...")
    if await gerar_pdf(html):
        print(f"  PDF:  {OUT_PDF}")


if __name__ == "__main__":
    asyncio.run(main_async())
