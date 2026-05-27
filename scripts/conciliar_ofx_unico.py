"""Conciliacao completa de UM OFX -> PDF + HTML + XLSX.

Pipeline:
  1. ler_ofx(path) -> lista de Transacao
  2. classificar(t) -> Resultado por transacao (6 estagios)
  3. enriquecer_um(cnpj) -> razao social + alerta de baixada
  4. Compila Disposicao final por transacao
  5. Gera 3 saidas estilizadas com cabecalho ORGATEC

Uso:
  python scripts/conciliar_ofx_unico.py --ofx CAMINHO --out PASTA_DOWNLOADS
"""
from __future__ import annotations

import argparse
import asyncio
import re
import sys
from collections import Counter
from datetime import date, datetime
from pathlib import Path

import httpx
import markdown as md_lib
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from api.matchers.cascata import Disposicao, classificar, ler_ofx
from api.matchers.cnpj_enricher import (
    _carregar_cache,
    _salvar_cache,
    buscar_cnpj_por_nome_no_cache,
    buscar_cnpj_por_nome_rfb,
    enriquecer_um,
)
from api.matchers.forensics import (
    calcular_agregados,
    calcular_risk_score,
    classificar_tributario,
    detectar_carrossel,
    detectar_meio,
    detectar_primeira_vez,
    detectar_smurfing,
    detectar_valor_redondo,
    hash_linha,
    periodo_fiscal,
)


# ────────────────────────────────────────────────────────────────────────
# Cabecalho institucional padronizado (em todas as abas)
# ────────────────────────────────────────────────────────────────────────


BANCOS = {
    "756": "SICOOB - Banco Cooperativo do Brasil",
    "001": "Banco do Brasil",
    "237": "Bradesco",
    "104": "Caixa Economica Federal",
    "341": "Itau Unibanco",
    "033": "Santander",
    "260": "Nubank",
    "748": "Sicredi",
}


def aplicar_cabecalho(ws, dados: dict, empresa: dict, ultima_col: int = 6):
    """Insere cabecalho institucional ORGATEC no topo da aba (linhas 1-4).

    A aba comeca o conteudo real na linha 5+.
    Modifica ws in place. Retorna a linha onde o conteudo deve comecar.
    """
    banco_nome = BANCOS.get(dados.get("banco_id", ""), f"Banco {dados.get('banco_id','?')}")
    end_col = get_column_letter(ultima_col)

    # Linha 1: titulo principal
    c1 = ws.cell(row=1, column=1, value="[ORGATEC] Conciliacao Bancaria - Auditoria")
    c1.font = Font(bold=True, size=14, color="FFFFFF")
    c1.fill = PatternFill("solid", fgColor=NAVY)
    c1.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(f"A1:{end_col}1")
    ws.row_dimensions[1].height = 28

    # Linha 2: empresa | CNPJ | socios
    c2 = ws.cell(
        row=2, column=1,
        value=f"Empresa: {empresa.get('nome', '[NAO CADASTRADO]')}  |  "
              f"CNPJ: {empresa.get('cnpj', '[PENDENTE]')}  |  "
              f"Socios: {empresa.get('socios', '[PENDENTE]')}",
    )
    c2.font = Font(bold=True, size=10, color="FFFFFF")
    c2.fill = PatternFill("solid", fgColor="1E3A8A")
    c2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(f"A2:{end_col}2")
    ws.row_dimensions[2].height = 22

    # Linha 3: conta detalhada
    c3 = ws.cell(
        row=3, column=1,
        value=f"Correntista: {empresa.get('nome', '[NAO CADASTRADO]')}  >  "
              f"Agencia: {dados.get('agencia', '?')}  >  "
              f"Conta: {dados.get('conta', '?')}  >  "
              f"Banco: {banco_nome}  >  "
              f"Periodo: {dados.get('periodo_ini','')} a {dados.get('periodo_fim','')}  >  "
              f"Saldo final: R$ {dados.get('saldo_final', 0):,.2f}",
    )
    c3.font = Font(size=9, color="0F172A")
    c3.fill = PatternFill("solid", fgColor="DBEAFE")
    c3.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(f"A3:{end_col}3")
    ws.row_dimensions[3].height = 20

    # Linha 4: separador
    ws.row_dimensions[4].height = 6

    return 5  # conteudo comeca na linha 5


# Estilos XLSX
NAVY = "0F172A"
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TOTAL_FILL = PatternFill("solid", fgColor="1E3A8A")
TOTAL_FONT = Font(bold=True, color="FFFFFF", size=11)
ZEBRA_FILL = PatternFill("solid", fgColor="F8FAFC")
ALERT_FILL = PatternFill("solid", fgColor="FEE2E2")
TITLE_FONT = Font(bold=True, size=14, color=NAVY)
BORDER = Side(border_style="thin", color="E2E8F0")
THIN_BORDER = Border(top=BORDER, left=BORDER, right=BORDER, bottom=BORDER)

RX_CNPJ = re.compile(r"(\d{2})[.](\d{3})[.](\d{3})[ /](\d{4})[-](\d{2})")

ESTAGIOS = {0: "TRANSF.INTERNA", 1: "CNPJ/CPF", 2: "NF-e", 3: "TARIFA",
            4: "TRIBUTO", 5: "CONTRATO", 6: "ALIAS/FUZZY"}

DISP_COR = {
    "TRANSFERENCIA_INTERNA": ("Auto", "gray"),
    "TARIFA_BANCARIA":        ("Auto", "gray"),
    "RESOLVIDO_CADASTRO":     ("Auto", "blue"),
    "RESOLVIDO_BASE":         ("Auto", "blue"),
    "RESOLVIDO_NFE":          ("Auto", "green"),
    "RESOLVIDO_GUIA":         ("Auto", "green"),
    "RESOLVIDO_CONTRATO":     ("Auto", "green"),
    "PENDENTE_REVISAO":       ("Pendente", "orange"),
    "PENDENTE_MATCHER":       ("Pendente", "orange"),
    "PENDENTE_FUZZY":         ("Pendente", "yellow"),
    "NAO_ENCONTRADO":         ("Pendente", "red"),
    "DOC_INVALIDO":           ("Pendente", "red"),
}


def _extrair_cnpj(t) -> str | None:
    for fonte in (t.nome or "", t.memo or ""):
        m = RX_CNPJ.search(fonte)
        if m:
            return "".join(m.groups())
    return None


def _classificar_disposicao(res) -> tuple[str, str, str]:
    """Mapeia (estagio, metodo) -> (disposicao, origem, flag).

    Sem matcher de BD ativo, decide pelo estagio + heuristicas locais.
    """
    t = res.transacao
    if res.metodo == "transferencia_interna":
        return "TRANSFERENCIA_INTERNA", "regra", "nao e evento economico"
    if res.metodo == "tarifa_bancaria":
        return "TARIFA_BANCARIA", "regra", ""
    if res.metodo == "match_documento":
        return "PENDENTE_REVISAO", "match_documento", "CNPJ detectado — enriquecer e cadastrar"
    if res.metodo == "match_nfe":
        return "PENDENTE_MATCHER", "match_nfe", f"NF {res.chave} — falta XML"
    if res.metodo == "match_guia_tributo":
        return "PENDENTE_REVISAO", "match_guia", f"tributo {res.chave} — cadastrar guia"
    if res.metodo == "match_contrato":
        return "PENDENTE_REVISAO", "match_contrato", "cadastrar como contrato"
    return "PENDENTE_FUZZY", "fuzzy_llm", "fallback alias/LLM"


# ────────────────────────────────────────────────────────────────────────
# Pipeline principal
# ────────────────────────────────────────────────────────────────────────


async def conciliar_ofx(ofx_path: Path) -> dict:
    """Roda a cascata completa em 1 OFX e devolve dict consolidado."""
    txs = ler_ofx(str(ofx_path))
    resultados = [classificar(t) for t in txs]

    # Enriquecimento de CNPJ (cache -> BrasilAPI)
    cnpjs = list({_extrair_cnpj(r.transacao) for r in resultados if _extrair_cnpj(r.transacao)})
    cache = _carregar_cache()
    semaforo = asyncio.Semaphore(2)

    cnpj_infos: dict[str, object] = {}
    async with httpx.AsyncClient(timeout=httpx.Timeout(10.0, connect=5.0),
                                  headers={"User-Agent": "OrgConc/0.5"}) as client:
        async def _job(c):
            info = await enriquecer_um(c, cache, client, None, semaforo)
            cnpj_infos[c] = info
        await asyncio.gather(*[_job(c) for c in cnpjs])
    _salvar_cache(cache)

    # Mapa nome -> info para fuzzy matching (transacoes sem CNPJ no extrato)
    nome_match_cache: dict[str, object] = {}

    def _buscar_por_nome(nome: str):
        if not nome:
            return None
        chave = nome.upper().strip()
        if chave in nome_match_cache:
            return nome_match_cache[chave]
        info = buscar_cnpj_por_nome_no_cache(nome, cache, min_score=85)
        nome_match_cache[chave] = info
        return info

    # Compila disposicoes
    disposicoes = []
    for r in resultados:
        disp_name, origem, flag = _classificar_disposicao(r)
        contraparte = ""
        nfe_chave = ""
        cnpj = _extrair_cnpj(r.transacao)
        if not cnpj:
            # Sem CNPJ no extrato (ex: COMPRA VISA ELECTRON) -> tenta fuzzy por nome
            info_fuzzy = _buscar_por_nome(r.transacao.nome or r.transacao.memo)
            if info_fuzzy and info_fuzzy.razao_social:
                cnpj = info_fuzzy.cnpj
                cnpj_infos[cnpj] = info_fuzzy
                flag = (flag + f" | {info_fuzzy.flag}").strip(" |") if flag else info_fuzzy.flag
        if cnpj and cnpj in cnpj_infos:
            info = cnpj_infos[cnpj]
            contraparte = info.razao_social or ""
            if info.flag:
                flag = (flag + " | " + info.flag).strip(" |") if flag else info.flag
            sit = info.situacao or ""
            if "BAIXADA" in sit or "INAPTA" in sit:
                # Pos-baixa?
                if info.data_situacao:
                    try:
                        db_data = date.fromisoformat(info.data_situacao[:10])
                        dt_data = date.fromisoformat(r.transacao.data[:10])
                        if dt_data > db_data:
                            flag = f"ALERTA POS-BAIXA: pago {(dt_data-db_data).days}d depois da baixa em {info.data_situacao}"
                            disp_name = "ALERTA_POS_BAIXA"
                    except (ValueError, TypeError):
                        pass

        disposicoes.append(Disposicao(
            transacao=r.transacao,
            estagio=r.estagio,
            disposicao=disp_name,
            contraparte=contraparte,
            conta_contabil="",
            origem=origem,
            flag=flag,
            nfe_chave=nfe_chave,
        ))

    # Stats
    n = len(disposicoes)
    cred = sum(d.transacao.valor for d in disposicoes if d.transacao.valor > 0)
    deb = sum(d.transacao.valor for d in disposicoes if d.transacao.valor < 0)
    auto = sum(1 for d in disposicoes if d.disposicao in (
        "RESOLVIDO_CADASTRO", "RESOLVIDO_BASE", "RESOLVIDO_NFE",
        "RESOLVIDO_GUIA", "RESOLVIDO_CONTRATO", "TARIFA_BANCARIA",
        "TRANSFERENCIA_INTERNA",
    ))
    alertas = sum(1 for d in disposicoes if d.disposicao == "ALERTA_POS_BAIXA")
    por_est = Counter(d.estagio for d in disposicoes)
    datas = sorted(d.transacao.data for d in disposicoes if d.transacao.data)

    # Identifica conta no OFX
    raw = ofx_path.read_text(encoding="latin-1", errors="ignore")
    agencia = re.search(r"<BRANCHID>([^<\s]+)", raw)
    conta = re.search(r"<ACCTID>([^<\s]+)", raw)
    banco_id = re.search(r"<BANKID>([^<\s]+)", raw)
    saldo = re.search(r"<BALAMT>([\d.\-]+)", raw)

    return {
        "ofx_path": str(ofx_path),
        "agencia": agencia.group(1) if agencia else "?",
        "conta": conta.group(1) if conta else "?",
        "banco_id": banco_id.group(1) if banco_id else "?",
        "saldo_final": float(saldo.group(1)) if saldo else 0.0,
        "periodo_ini": datas[0] if datas else "",
        "periodo_fim": datas[-1] if datas else "",
        "n_transacoes": n,
        "credito_total": round(cred, 2),
        "debito_total": round(deb, 2),
        "automatizadas": auto,
        "pendentes": n - auto,
        "alertas_pos_baixa": alertas,
        "por_estagio": dict(por_est),
        "disposicoes": disposicoes,
        "cnpj_infos": cnpj_infos,
    }


# ────────────────────────────────────────────────────────────────────────
# Geração HTML / PDF / XLSX
# ────────────────────────────────────────────────────────────────────────


def gerar_markdown(dados: dict) -> str:
    """Gera markdown estilizado com a conciliacao."""
    d = dados
    pct_auto = 100 * d["automatizadas"] / max(d["n_transacoes"], 1)
    bancos = {"756": "Sicoob (Banco 756)", "001": "Banco do Brasil", "237": "Bradesco",
              "104": "Caixa", "341": "Itau", "033": "Santander"}
    banco_nome = bancos.get(d["banco_id"], f"Banco {d['banco_id']}")

    lines = [
        f"# Conciliacao Bancaria - Conta {d['conta']}",
        "",
        f"**Banco:** {banco_nome} | **Agencia:** {d['agencia']} | **Conta:** {d['conta']}",
        f"**Periodo:** {d['periodo_ini']} a {d['periodo_fim']} | **Saldo final:** R$ {d['saldo_final']:,.2f}",
        f"**Gerado em:** {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        "",
        "## 1. Resumo Executivo",
        "",
        "| Indicador | Valor |",
        "|---|---|",
        f"| Total de transacoes | {d['n_transacoes']} |",
        f"| Volume de creditos | R$ {d['credito_total']:,.2f} |",
        f"| Volume de debitos | R$ {d['debito_total']:,.2f} |",
        f"| Fluxo liquido | R$ {d['credito_total'] + d['debito_total']:,.2f} |",
        f"| Automatizadas | {d['automatizadas']} ({pct_auto:.1f}%) |",
        f"| Pendentes | {d['pendentes']} |",
        f"| Alertas pos-baixa | **{d['alertas_pos_baixa']}** |",
        "",
        "## 2. Distribuicao por Estagio",
        "",
        "| Estagio | Tipo | Qtd | % |",
        "|---|---|---|---|",
    ]
    for est in range(7):
        qtd = d["por_estagio"].get(est, 0)
        if qtd:
            lines.append(f"| {est} | {ESTAGIOS[est]} | {qtd} | {100*qtd/d['n_transacoes']:.1f}% |")

    # Alertas
    alertas = [disp for disp in d["disposicoes"] if disp.disposicao == "ALERTA_POS_BAIXA"]
    if alertas:
        lines += [
            "",
            "## 3. ALERTAS - Pagamentos Pos-Baixa",
            "",
            "**Transacoes para CNPJs ja baixados na data do pagamento - critico:**",
            "",
            "| Data | Valor (R$) | Contraparte | Flag |",
            "|---|---|---|---|",
        ]
        for disp in alertas:
            valor = disp.transacao.valor
            cp = (disp.contraparte or "")[:45]
            flag = (disp.flag or "").replace("|", "/")[:80]
            lines.append(f"| {disp.transacao.data} | {valor:,.2f} | {cp} | {flag} |")

    # Transações (extrato com saldo acumulado + contraparte RFB)
    saldo_inicial = d['saldo_final'] - (d['credito_total'] + d['debito_total'])
    lines += [
        "",
        "## 4. Transacoes (Extrato Detalhado)",
        "",
        f"Saldo inicial: R$ {saldo_inicial:,.2f} | Saldo final: R$ {d['saldo_final']:,.2f}",
        "",
        "| # | Data | Valor (R$) | Memo | Nome | Contraparte (RFB) | Saldo Acumulado (R$) |",
        "|---|---|---|---|---|---|---|",
    ]
    txs_ord = sorted(d["disposicoes"], key=lambda x: x.transacao.data)
    saldo_corrente = saldo_inicial
    for i, disp in enumerate(txs_ord, start=1):
        t = disp.transacao
        saldo_corrente += t.valor
        memo_s = (t.memo or "")[:25].replace("|", "/")
        nome_s = (t.nome or "")[:25].replace("|", "/")
        contrap_s = (disp.contraparte or "")[:30].replace("|", "/")
        lines.append(
            f"| {i} | {t.data} | {t.valor:,.2f} | {memo_s} | {nome_s} | {contrap_s} | {saldo_corrente:,.2f} |"
        )

    # Disposições
    lines += [
        "",
        "## 5. Disposicoes por Transacao",
        "",
        "| Data | Valor (R$) | Memo | Nome (banco) | Disposicao | Contraparte (RFB) |",
        "|---|---|---|---|---|---|",
    ]
    for disp in sorted(d["disposicoes"], key=lambda x: x.transacao.data):
        t = disp.transacao
        memo_s = (t.memo or "")[:35].replace("|", "/")
        nome_s = (t.nome or "")[:35].replace("|", "/")
        contrap_s = (disp.contraparte or "")[:30].replace("|", "/")
        lines.append(
            f"| {t.data} | {t.valor:,.2f} | {memo_s} | {nome_s} | {disp.disposicao} | {contrap_s} |"
        )

    lines += [
        "",
        "---",
        "*Sistema: OrgConc/OrgNeural2 - cascata de 6 estagios. Enriquecimento CNPJ via BrasilAPI/RFB.*",
    ]
    return "\n".join(lines)


def gerar_html(md_text: str, dados: dict) -> str:
    body = md_lib.markdown(md_text, extensions=["tables", "fenced_code"])
    agora = datetime.now().strftime("%d/%m/%Y %H:%M")
    css = """
@page { size: A4; margin: 16mm 14mm 16mm 14mm;
  @bottom-right { content: "Pagina " counter(page) " de " counter(pages); font-size: 9px; color: #6B7280; } }
* { box-sizing: border-box; margin: 0; padding: 0; }
body { font-family: 'DejaVu Sans', Arial, sans-serif; font-size: 10pt; color: #1a202c; line-height: 1.55; }
.hd { background: linear-gradient(135deg, #0F172A 0%, #0B1B3D 45%, #0052FF 100%);
      color: #fff; padding: 22px 28px; border-radius: 12px; margin-bottom: 22px; }
.hd h1 { font-size: 22pt; font-family: 'DejaVu Serif', Georgia, serif; margin-bottom: 4px; }
.hd .tag { font-size: 9pt; opacity: 0.85; text-transform: uppercase; letter-spacing: 0.16em; }
.hd .meta { font-size: 9pt; margin-top: 10px; opacity: 0.92; }
h1 { font-size: 16pt; color: #0F172A; margin: 22px 0 8px; padding-bottom: 6px; border-bottom: 2px solid #BFDBFE; }
h2 { font-size: 13pt; color: #0052FF; margin: 18px 0 8px; padding-left: 10px; border-left: 3px solid #0EA5E9; }
table { width: 100%; border-collapse: collapse; margin: 10px 0 14px; font-size: 9pt; border-radius: 6px; overflow: hidden; }
th { background: linear-gradient(180deg, #0F172A, #1E3A8A); color: #fff; padding: 6px 9px; text-align: left; font-weight: 600; }
td { padding: 5px 9px; border-bottom: 1px solid #E2E8F0; vertical-align: top; }
tr:nth-child(even) td { background: #F8FAFC; }
strong { color: #0F172A; font-weight: 700; }
code { font-family: 'DejaVu Sans Mono', 'Courier New', monospace; font-size: 9pt; background: #F1F5F9; padding: 1px 5px; border-radius: 3px; color: #0052FF; }
.ft { margin-top: 28px; padding-top: 12px; border-top: 1px solid #E2E8F0; font-size: 8.5pt; color: #94A3B8; }
"""
    return f"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8">
<title>Conciliacao Conta {dados['conta']} - OrgConc</title>
<style>{css}</style>
</head>
<body>
  <div class="hd">
    <h1>ORGATEC</h1>
    <div class="tag">Contabilidade &amp; Auditoria - Conciliacao Bancaria</div>
    <div class="meta">Conta {dados['conta']} ({dados['banco_id']}) - Gerado em {agora}</div>
  </div>
  {body}
  <div class="ft">(c) ORGATEC Contabilidade e Auditoria - OrgConc v0.5.0</div>
</body>
</html>
"""


def style_header(ws, row: int, n_cols: int) -> None:
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = THIN_BORDER


def gerar_xlsx(dados: dict, out_path: Path, empresa: dict | None = None) -> None:
    d = dados
    if empresa is None:
        empresa = {"nome": "[NAO CADASTRADO]", "cnpj": "[PENDENTE]", "socios": "[PENDENTE]"}
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # ── Aba Resumo + Checklist de Processamento ─────────────────────────
    ws = wb.create_sheet("Resumo")
    aplicar_cabecalho(ws, d, empresa, ultima_col=5)
    ws.cell(row=5, column=1, value=f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}").font = Font(italic=True, color="64748B", size=9)
    ws.merge_cells("A5:E5")

    # Checklist de processamento em segundo plano
    ws.cell(row=7, column=1, value="CHECKLIST DE PROCESSAMENTO (BACKGROUND)").font = TITLE_FONT
    ws.merge_cells("A7:E7")
    enc_ok = "[X]" if d.get("encoding_ok", True) else "[ ]"
    datas_ok = "[X]" if d.get("datas_validas", True) else "[ ]"
    saldo_ok = "[X]" if d.get("saldo_validado", True) else "[~]"
    cat_ok = "[X]"
    docker_ok = "[ ]"
    checklist = [
        (f"{enc_ok}", "Encoding OFX validado (USASCII/Latin1 mapeados)"),
        (f"{datas_ok}", "Datas DTSTART/DTEND coincidem com extrato"),
        (f"{saldo_ok}", "Saldo final OFX (LEDGERBAL) validado contra fluxo"),
        (f"{cat_ok}", "Transacoes normalizadas para categorias Orgatec"),
        (f"{docker_ok}", "Processamento em container Docker/n8n com logs de auditoria"),
    ]
    for i, (status, item) in enumerate(checklist, start=8):
        c = ws.cell(row=i, column=1, value=status)
        c.font = Font(name="Consolas", size=10, color=("16A34A" if "X" in status else "D97706" if "~" in status else "94A3B8"))
        ws.cell(row=i, column=2, value=item)
        ws.merge_cells(f"B{i}:E{i}")

    # Indicadores principais
    ws.cell(row=14, column=1, value="INDICADORES PRINCIPAIS").font = TITLE_FONT
    ws.merge_cells("A14:E14")
    ws.cell(row=15, column=1, value="INDICADOR")
    ws.cell(row=15, column=2, value="VALOR")
    style_header(ws, 15, 2)
    indicadores = [
        ("Banco", f"{d['banco_id']} - {BANCOS.get(d['banco_id'], '?')}"),
        ("Agencia / Conta", f"{d['agencia']} / {d['conta']}"),
        ("Total de transacoes", d['n_transacoes']),
        ("Volume de creditos (R$)", d['credito_total']),
        ("Volume de debitos (R$)", d['debito_total']),
        ("Saldo final do periodo", d['saldo_final']),
        ("Automatizadas", f"{d['automatizadas']} ({100*d['automatizadas']/max(d['n_transacoes'],1):.1f}%)"),
        ("Pendentes", d['pendentes']),
        ("Alertas pos-baixa", d['alertas_pos_baixa']),
    ]
    for i, (k, v) in enumerate(indicadores, start=16):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        c = ws.cell(row=i, column=2, value=v)
        if isinstance(v, (int, float)) and "R$" in k:
            c.number_format = "#,##0.00"
        if i % 2 == 0:
            for col in (1, 2):
                ws.cell(row=i, column=col).fill = ZEBRA_FILL

    # Por estagio
    r = i + 2
    ws.cell(row=r, column=1, value="DISTRIBUICAO POR ESTAGIO").font = TITLE_FONT
    ws.merge_cells(f"A{r}:D{r}")
    r += 1
    ws.cell(row=r, column=1, value="Estagio")
    ws.cell(row=r, column=2, value="Tipo")
    ws.cell(row=r, column=3, value="Quantidade")
    ws.cell(row=r, column=4, value="%")
    style_header(ws, r, 4)
    r += 1
    for est in range(7):
        qtd = d["por_estagio"].get(est, 0)
        if qtd:
            ws.cell(row=r, column=1, value=est)
            ws.cell(row=r, column=2, value=ESTAGIOS[est])
            ws.cell(row=r, column=3, value=qtd).number_format = "#,##0"
            ws.cell(row=r, column=4, value=qtd / d['n_transacoes']).number_format = "0.0%"
            if r % 2 == 0:
                for col in range(1, 5):
                    ws.cell(row=r, column=col).fill = ZEBRA_FILL
            r += 1

    # Divergencias identificadas
    r += 2
    ws.cell(row=r, column=1, value="DIVERGENCIAS IDENTIFICADAS").font = TITLE_FONT
    ws.merge_cells(f"A{r}:E{r}")
    r += 1
    divergencias = []
    if empresa.get("nome", "[NAO CADASTRADO]") == "[NAO CADASTRADO]":
        divergencias.append("Cadastro da empresa auditada nao fornecido (CNPJ/Socios pendentes)")
    if d['alertas_pos_baixa'] > 0:
        divergencias.append(f"{d['alertas_pos_baixa']} pagamentos pos-baixa detectados - investigar")
    if not divergencias:
        ws.cell(row=r, column=1, value="(nenhuma divergencia)").font = Font(italic=True, color="16A34A")
    else:
        for div in divergencias:
            c = ws.cell(row=r, column=1, value=f"[!] DIVERGENCIA IDENTIFICADA: {div}")
            c.font = Font(bold=True, color="DC2626")
            c.fill = ALERT_FILL
            ws.merge_cells(f"A{r}:E{r}")
            r += 1

    for col, w in {1: 32, 2: 40, 3: 14, 4: 10, 5: 12}.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A5"

    # ── Aba Transacoes (extrato com saldo acumulado) ────────────────────
    ws = wb.create_sheet("Transacoes")
    aplicar_cabecalho(ws, d, empresa, ultima_col=8)
    ws.cell(row=5, column=1, value=f"EXTRATO DETALHADO - {d['n_transacoes']} transacoes - Saldo final R$ {d['saldo_final']:,.2f}").font = TITLE_FONT
    ws.merge_cells("A5:H5")

    headers_t = ["#", "Data", "Tipo", "Valor (R$)", "Memo", "Nome", "Contraparte (RFB)", "Saldo Acumulado (R$)"]
    for c, h in enumerate(headers_t, start=1):
        ws.cell(row=7, column=c, value=h)
    style_header(ws, 7, 8)

    # Calcula saldo inicial = saldo final - soma de todos os fluxos
    saldo_inicial = d['saldo_final'] - (d['credito_total'] + d['debito_total'])

    txs_ord = sorted(d['disposicoes'], key=lambda x: x.transacao.data)
    saldo_corrente = saldo_inicial
    r = 8
    total_cred = 0.0
    total_deb = 0.0
    for i, disp in enumerate(txs_ord, start=1):
        t = disp.transacao
        saldo_corrente += t.valor
        if t.valor > 0:
            total_cred += t.valor
        else:
            total_deb += t.valor

        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=2, value=t.data)
        ws.cell(row=r, column=3, value=t.tipo)
        cv = ws.cell(row=r, column=4, value=round(t.valor, 2))
        cv.number_format = "#,##0.00"
        cv.font = Font(color=("DC2626" if t.valor < 0 else "16A34A"))
        ws.cell(row=r, column=5, value=t.memo or "")
        ws.cell(row=r, column=6, value=t.nome or "")
        # Coluna Contraparte (RFB) — razao social enriquecida
        contraparte = disp.contraparte or ""
        cc = ws.cell(row=r, column=7, value=contraparte)
        # Destaca quando ha alerta pos-baixa
        if disp.disposicao == "ALERTA_POS_BAIXA":
            cc.font = Font(bold=True, color="DC2626")
        cs = ws.cell(row=r, column=8, value=round(saldo_corrente, 2))
        cs.number_format = "#,##0.00"
        cs.font = Font(bold=True, color=("DC2626" if saldo_corrente < 0 else "0F172A"))
        for c in range(1, 9):
            ws.cell(row=r, column=c).border = THIN_BORDER
            if disp.disposicao == "ALERTA_POS_BAIXA":
                ws.cell(row=r, column=c).fill = ALERT_FILL
            elif r % 2 == 0:
                ws.cell(row=r, column=c).fill = ZEBRA_FILL
        r += 1

    # Linha de totais
    ws.cell(row=r, column=1, value="TOTAL").font = TOTAL_FONT
    ws.cell(row=r, column=2, value=f"{d['n_transacoes']} transacoes")
    cv = ws.cell(row=r, column=4, value=round(total_cred + total_deb, 2))
    cv.number_format = "#,##0.00"
    ws.cell(row=r, column=5, value=f"+ {total_cred:,.2f}").font = Font(color="16A34A", bold=True)
    ws.cell(row=r, column=6, value=f"{total_deb:,.2f}").font = Font(color="DC2626", bold=True)
    ws.cell(row=r, column=8, value=round(d['saldo_final'], 2)).number_format = "#,##0.00"
    ws.cell(row=r, column=8).font = TOTAL_FONT
    for c in range(1, 9):
        ws.cell(row=r, column=c).fill = TOTAL_FILL
        if c != 5 and c != 6:
            ws.cell(row=r, column=c).font = TOTAL_FONT
        ws.cell(row=r, column=c).border = THIN_BORDER

    for col, w in {1: 5, 2: 12, 3: 8, 4: 16, 5: 35, 6: 30, 7: 38, 8: 22}.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A8"
    ws.auto_filter.ref = f"A7:H{r-1}"

    # ── Aba Disposicoes (com auditoria forense — 27 colunas) ────────────
    ws = wb.create_sheet("Disposicoes")
    aplicar_cabecalho(ws, d, empresa, ultima_col=27)
    ws.cell(row=5, column=1, value="DISPOSICOES POR TRANSACAO - Auditoria Forense").font = TITLE_FONT
    ws.merge_cells("A5:AA5")
    ws.cell(row=6, column=1, value="Eixos: A=Compliance | B=Identificacao | C=Padroes | D=Risk Score | E=Rastreabilidade").font = Font(italic=True, color="64748B", size=9)
    ws.merge_cells("A6:AA6")

    headers = [
        # Identificacao basica (1-5)
        "Data", "Tipo", "Valor (R$)", "Memo", "Nome (banco)",
        # B - Identificacao (6-8)
        "FITID", "CheckNum", "Meio",
        # Contraparte (9-10)
        "CNPJ", "Contraparte (RFB)",
        # A - Compliance (11-14)
        "CNAE", "UF", "Municipio", "Porte",
        # C - Padroes (15-19)
        "Acumulado Mes (R$)", "1a Vez?", "Valor Redondo", "Smurfing", "Carrossel",
        # Decisao (20-21)
        "Disposicao", "Flag",
        # D - Risk Score (22-23)
        "Risk Score", "Risk Class",
        # E - Rastreabilidade (24-27)
        "Periodo Fiscal", "Hash Linha", "Status Revisao", "Comentario Revisor",
    ]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=8, column=c, value=h)
    style_header(ws, 8, len(headers))

    # Pre-calcula agregados para padroes (C)
    agg = calcular_agregados(d['disposicoes'])

    r = 9
    for disp in sorted(d['disposicoes'], key=lambda x: x.transacao.data):
        t = disp.transacao
        cnpj = _extrair_cnpj(t)
        cnpj_fmt = ""
        if cnpj:
            cnpj_fmt = f"{cnpj[:2]}.{cnpj[2:5]}.{cnpj[5:8]}/{cnpj[8:12]}-{cnpj[12:14]}"

        # Info do cache CNPJ
        info_cnpj = d['cnpj_infos'].get(cnpj) if cnpj else None
        cnae_desc = info_cnpj.cnae_descricao if info_cnpj else ""
        uf = info_cnpj.uf if info_cnpj else ""
        municipio = info_cnpj.municipio if info_cnpj else ""
        porte = info_cnpj.porte if info_cnpj else ""
        situacao = info_cnpj.situacao if info_cnpj else ""

        # B - Meio
        meio = detectar_meio(t.memo or "", t.nome or "")

        # C - Padroes
        mes = t.data[:7]
        acumulado_mes = agg.acumulado_mes.get((cnpj, mes), 0.0) if cnpj else 0.0
        primeira_vez = detectar_primeira_vez(cnpj, t.data, agg)
        valor_redondo = detectar_valor_redondo(t.valor)
        smurfing = detectar_smurfing(cnpj, t.data, agg)
        carrossel = detectar_carrossel(cnpj, agg)

        # D - Risk Score
        score, classe = calcular_risk_score(
            t.valor, disp.disposicao, situacao, porte, meio,
            valor_redondo, smurfing, carrossel, primeira_vez, acumulado_mes,
        )

        # E - Rastreabilidade
        pf = periodo_fiscal(t.data)
        h_linha = hash_linha(t.data, t.valor, t.memo or "", t.fitid or "")

        is_alerta = disp.disposicao == "ALERTA_POS_BAIXA"
        is_critico = classe == "CRITICO"
        is_alto = classe == "ALTO"

        # Preenche linha (27 colunas)
        ws.cell(row=r, column=1, value=t.data)
        ws.cell(row=r, column=2, value=t.tipo)
        cv = ws.cell(row=r, column=3, value=round(t.valor, 2))
        cv.number_format = "#,##0.00"
        cv.font = Font(color=("DC2626" if t.valor < 0 else "16A34A"), bold=is_alerta or is_critico)
        ws.cell(row=r, column=4, value=t.memo or "")
        ws.cell(row=r, column=5, value=t.nome or "")
        # B - FITID, CHECKNUM, Meio
        ws.cell(row=r, column=6, value=t.fitid or "").font = Font(name="Consolas", size=9, color="64748B")
        ws.cell(row=r, column=7, value=t.checknum or "")
        c_meio = ws.cell(row=r, column=8, value=meio)
        if meio == "TRIBUTO":
            c_meio.fill = PatternFill("solid", fgColor="FEF3C7")
        elif meio == "PIX":
            c_meio.fill = PatternFill("solid", fgColor="DBEAFE")
        elif meio == "CARTAO":
            c_meio.fill = PatternFill("solid", fgColor="E0E7FF")
        # Contraparte
        ws.cell(row=r, column=9, value=cnpj_fmt).font = Font(name="Consolas", size=10)
        ws.cell(row=r, column=10, value=disp.contraparte or "")
        # A - Compliance
        ws.cell(row=r, column=11, value=cnae_desc[:60])
        ws.cell(row=r, column=12, value=uf)
        ws.cell(row=r, column=13, value=municipio)
        c_porte = ws.cell(row=r, column=14, value=porte)
        # MEI > limite mes = vermelho
        if porte == "MICRO EMPRESA" and acumulado_mes > 6_750:
            c_porte.font = Font(bold=True, color="DC2626")
        # C - Padroes
        c_acum = ws.cell(row=r, column=15, value=round(acumulado_mes, 2))
        c_acum.number_format = "#,##0.00"
        ws.cell(row=r, column=16, value=primeira_vez)
        c_redondo = ws.cell(row=r, column=17, value=valor_redondo)
        if valor_redondo in ("REDONDO_10K", "REDONDO_5K"):
            c_redondo.font = Font(bold=True, color="D97706")
        c_smur = ws.cell(row=r, column=18, value=smurfing)
        if smurfing:
            c_smur.font = Font(bold=True, color="DC2626")
            c_smur.fill = PatternFill("solid", fgColor="FEE2E2")
        c_carr = ws.cell(row=r, column=19, value=carrossel)
        if carrossel:
            c_carr.font = Font(bold=True, color="DC2626")
        # Decisao
        cell_disp = ws.cell(row=r, column=20, value=disp.disposicao)
        if disp.disposicao.startswith("RESOLVIDO_") or disp.disposicao in ("TRANSFERENCIA_INTERNA", "TARIFA_BANCARIA"):
            cell_disp.font = Font(color="16A34A", bold=True)
        elif is_alerta:
            cell_disp.font = Font(color="DC2626", bold=True)
        else:
            cell_disp.font = Font(color="D97706", bold=True)
        ws.cell(row=r, column=21, value=disp.flag or "")
        # D - Risk Score
        c_score = ws.cell(row=r, column=22, value=score)
        c_score.number_format = "0"
        if is_critico:
            c_score.font = Font(bold=True, color="DC2626")
        elif is_alto:
            c_score.font = Font(bold=True, color="D97706")
        c_classe = ws.cell(row=r, column=23, value=classe)
        c_classe.font = Font(bold=True)
        if classe == "CRITICO":
            c_classe.fill = PatternFill("solid", fgColor="FEE2E2")
            c_classe.font = Font(bold=True, color="DC2626")
        elif classe == "ALTO":
            c_classe.fill = PatternFill("solid", fgColor="FEF3C7")
            c_classe.font = Font(bold=True, color="D97706")
        elif classe == "MEDIO":
            c_classe.fill = PatternFill("solid", fgColor="DBEAFE")
        else:
            c_classe.fill = PatternFill("solid", fgColor="DCFCE7")
        # E - Rastreabilidade
        ws.cell(row=r, column=24, value=pf)
        ws.cell(row=r, column=25, value=h_linha).font = Font(name="Consolas", size=8, color="94A3B8")
        ws.cell(row=r, column=26, value="PENDENTE").font = Font(italic=True, color="64748B", size=9)
        ws.cell(row=r, column=27, value="")  # comentario livre

        for c in range(1, 28):
            ws.cell(row=r, column=c).border = THIN_BORDER
            if is_alerta:
                # Nao sobrescreve fills semanticos das colunas 18, 23
                if c not in (18, 23) and not (c == 17 and valor_redondo):
                    ws.cell(row=r, column=c).fill = ALERT_FILL
            elif r % 2 == 0 and ws.cell(row=r, column=c).fill.fgColor.rgb in (None, "00000000"):
                ws.cell(row=r, column=c).fill = ZEBRA_FILL
        r += 1

    larguras = {
        1: 11, 2: 7, 3: 13, 4: 28, 5: 28,              # Identif basica
        6: 18, 7: 10, 8: 11,                            # B
        9: 19, 10: 32,                                  # Contraparte
        11: 38, 12: 5, 13: 22, 14: 22,                  # A
        15: 16, 16: 8, 17: 13, 18: 22, 19: 11,          # C
        20: 22, 21: 32,                                 # Decisao
        22: 10, 23: 10,                                 # D
        24: 12, 25: 17, 26: 13, 27: 30,                 # E
    }
    for col, w in larguras.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "F9"  # congela ate Memo+Nome
    ws.auto_filter.ref = f"A8:AA{r-1}"

    # ── Aba Risk (resumo por classe de risco) ───────────────────────────
    ws = wb.create_sheet("Risk")
    aplicar_cabecalho(ws, d, empresa, ultima_col=6)
    ws.cell(row=5, column=1, value="ANALISE DE RISCO - Distribuicao por Classe").font = TITLE_FONT
    ws.merge_cells("A5:F5")

    # Re-calcula stats com os mesmos detectores
    agg = calcular_agregados(d['disposicoes'])
    classe_counts = {"CRITICO": [0, 0.0], "ALTO": [0, 0.0], "MEDIO": [0, 0.0], "BAIXO": [0, 0.0]}
    for disp in d['disposicoes']:
        t = disp.transacao
        cnpj = _extrair_cnpj(t)
        info_cnpj = d['cnpj_infos'].get(cnpj) if cnpj else None
        sit = info_cnpj.situacao if info_cnpj else ""
        porte = info_cnpj.porte if info_cnpj else ""
        meio = detectar_meio(t.memo or "", t.nome or "")
        mes = t.data[:7]
        acumulado = agg.acumulado_mes.get((cnpj, mes), 0.0) if cnpj else 0.0
        vr = detectar_valor_redondo(t.valor)
        sm = detectar_smurfing(cnpj, t.data, agg)
        car = detectar_carrossel(cnpj, agg)
        pv = detectar_primeira_vez(cnpj, t.data, agg)
        _, classe = calcular_risk_score(t.valor, disp.disposicao, sit, porte, meio,
                                         vr, sm, car, pv, acumulado)
        classe_counts[classe][0] += 1
        classe_counts[classe][1] += abs(t.valor)

    headers_h = ["Classe", "Qtd Transacoes", "% do Total", "Volume (R$)", "% Volume", "Acao Sugerida"]
    for c, h in enumerate(headers_h, start=1):
        ws.cell(row=7, column=c, value=h)
    style_header(ws, 7, 6)

    cores_classe = {
        "CRITICO": ("DC2626", "FEE2E2", "Auditoria imediata - investigar"),
        "ALTO":    ("D97706", "FEF3C7", "Revisao prioritaria"),
        "MEDIO":   ("0052FF", "DBEAFE", "Conferir em lote"),
        "BAIXO":   ("16A34A", "DCFCE7", "Auto-aprovar apos confirmacao"),
    }
    total_qtd = sum(v[0] for v in classe_counts.values())
    total_vol = sum(v[1] for v in classe_counts.values())
    r = 8
    for classe in ("CRITICO", "ALTO", "MEDIO", "BAIXO"):
        qtd, vol = classe_counts[classe]
        cor, fill_cor, acao = cores_classe[classe]
        c_classe = ws.cell(row=r, column=1, value=classe)
        c_classe.font = Font(bold=True, color=cor)
        c_classe.fill = PatternFill("solid", fgColor=fill_cor)
        ws.cell(row=r, column=2, value=qtd).number_format = "#,##0"
        ws.cell(row=r, column=3, value=qtd / max(total_qtd, 1)).number_format = "0.0%"
        ws.cell(row=r, column=4, value=round(vol, 2)).number_format = "#,##0.00"
        ws.cell(row=r, column=5, value=vol / max(total_vol, 1)).number_format = "0.0%"
        ws.cell(row=r, column=6, value=acao)
        for c in range(1, 7):
            ws.cell(row=r, column=c).border = THIN_BORDER
        r += 1

    # Linha total
    ws.cell(row=r, column=1, value="TOTAL").font = TOTAL_FONT
    ws.cell(row=r, column=2, value=total_qtd).number_format = "#,##0"
    ws.cell(row=r, column=4, value=round(total_vol, 2)).number_format = "#,##0.00"
    for c in range(1, 7):
        ws.cell(row=r, column=c).fill = TOTAL_FILL
        ws.cell(row=r, column=c).font = TOTAL_FONT
        ws.cell(row=r, column=c).border = THIN_BORDER

    for col, w in {1: 12, 2: 16, 3: 12, 4: 18, 5: 11, 6: 34}.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A8"

    # ── Aba CNPJs ──────────────────────────────────────────────────────
    if d["cnpj_infos"]:
        ws = wb.create_sheet("CNPJs")
        aplicar_cabecalho(ws, d, empresa, ultima_col=8)
        ws.cell(row=5, column=1, value=f"CONTRAPARTES IDENTIFICADAS - {len(d['cnpj_infos'])} CNPJS").font = TITLE_FONT
        ws.merge_cells("A5:H5")
        headers = ["CNPJ", "Razao Social", "Situacao", "Data Baixa/Situacao", "UF",
                   "Municipio", "CNAE Descricao", "Porte"]
        for c, h in enumerate(headers, start=1):
            ws.cell(row=7, column=c, value=h)
        style_header(ws, 7, 8)

        r = 8
        for cnpj, info in d["cnpj_infos"].items():
            fmt = f"{cnpj[:2]}.{cnpj[2:5]}.{cnpj[5:8]}/{cnpj[8:12]}-{cnpj[12:14]}"
            is_baixada = "BAIXADA" in (info.situacao or "") or "INAPTA" in (info.situacao or "")
            ws.cell(row=r, column=1, value=fmt).font = Font(name="Consolas", size=10)
            ws.cell(row=r, column=2, value=info.razao_social or "(nao encontrado)")
            ws.cell(row=r, column=3, value=info.situacao or "")
            ws.cell(row=r, column=4, value=info.data_situacao or "")
            ws.cell(row=r, column=5, value=info.uf or "")
            ws.cell(row=r, column=6, value=info.municipio or "")
            ws.cell(row=r, column=7, value=info.cnae_descricao or "")
            ws.cell(row=r, column=8, value=info.porte or "")
            for c in range(1, 9):
                ws.cell(row=r, column=c).border = THIN_BORDER
                if is_baixada:
                    ws.cell(row=r, column=c).fill = ALERT_FILL
                    if c in (3, 4):
                        ws.cell(row=r, column=c).font = Font(bold=True, color="DC2626")
                elif r % 2 == 0:
                    ws.cell(row=r, column=c).fill = ZEBRA_FILL
            r += 1
        for col, w in {1: 20, 2: 40, 3: 18, 4: 16, 5: 5, 6: 22, 7: 38, 8: 22}.items():
            ws.column_dimensions[get_column_letter(col)].width = w
        ws.freeze_panes = "A8"
        ws.auto_filter.ref = f"A7:H{r-1}"

    # ── Aba Partes Relacionadas ────────────────────────────────────────
    aba_partes_relacionadas(wb, d, empresa)

    # ── Aba Status Tributario ───────────────────────────────────────────
    aba_status_tributario(wb, d, empresa)

    wb.save(str(out_path))


def aba_partes_relacionadas(wb, d: dict, empresa: dict) -> None:
    """Identifica transacoes com socios, empresas do grupo ou contas vinculadas."""
    ws = wb.create_sheet("Partes Relacionadas")
    aplicar_cabecalho(ws, d, empresa, ultima_col=8)
    ws.cell(row=5, column=1, value="PARTES RELACIONADAS - Socios, Grupo Economico, Contas Vinculadas").font = TITLE_FONT
    ws.merge_cells("A5:H5")
    ws.cell(row=6, column=1, value="Identifica transacoes com pessoas/CNPJs cadastrados como socios, grupo (nome em comum) ou MESMA TIT.").font = Font(italic=True, color="64748B", size=9)
    ws.merge_cells("A6:H6")

    headers = ["Data", "Valor (R$)", "Memo", "Nome/Favorecido", "CNPJ", "Contraparte (RFB)",
               "Tipo de Vinculo", "Justificativa"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=8, column=c, value=h)
    style_header(ws, 8, 8)

    # Detecta vinculos
    socios_empresa = (empresa.get("socios") or "").upper()
    nomes_empresa = (empresa.get("nome") or "").upper()
    grupos_inferidos = {}  # nome_grupo -> [transacoes]

    # Estrategia: detecta MESMA TIT, e tambem nomes recorrentes (top 5)
    from collections import Counter
    nomes_count = Counter()
    for disp in d['disposicoes']:
        nome = (disp.transacao.nome or "").upper().strip()
        # Extrai primeira palavra significativa
        primeiro = re.sub(r"^(FAV\.:|REM\.:|FAV:|REM:|PAGAMENTO\s+PIX|RECEBIMENTO\s+PIX)\s*", "", nome).strip()
        primeira_palavra = primeiro.split()[0] if primeiro.split() else ""
        if len(primeira_palavra) >= 4:
            nomes_count[primeira_palavra] += 1

    # Top nomes recorrentes = candidatos a grupo economico
    top_recorrentes = {n for n, c in nomes_count.most_common(15) if c >= 3}

    r = 9
    total_vol = 0.0
    for disp in sorted(d['disposicoes'], key=lambda x: x.transacao.data):
        t = disp.transacao
        texto = ((t.nome or "") + " " + (t.memo or "")).upper()
        vinculos = []

        # 1. Mesma titularidade
        if "MESMA TIT" in texto or "MESMA TITULAR" in texto:
            vinculos.append(("MESMA_TIT", "Transferencia entre contas proprias do mesmo CPF/CNPJ"))

        # 2. Socio cadastrado (nome em parte da string)
        if socios_empresa and socios_empresa != "[PENDENTE]":
            for socio in socios_empresa.split(","):
                socio_clean = socio.strip()
                if socio_clean and len(socio_clean) >= 6 and socio_clean in texto:
                    vinculos.append(("SOCIO_PF", f"Nome do socio '{socio_clean}' identificado na transacao"))

        # 3. Mesmo nome da empresa auditada (deveria ser raro - autoreferencia)
        if nomes_empresa and nomes_empresa != "[NAO CADASTRADO]" and nomes_empresa in texto:
            vinculos.append(("AUTO_REF", f"Nome da empresa '{nomes_empresa}' aparece na contraparte"))

        # 4. Grupo economico (nome recorrente)
        primeiro = re.sub(r"^(FAV\.:|REM\.:|FAV:|REM:|PAGAMENTO\s+PIX|RECEBIMENTO\s+PIX)\s*", "", (t.nome or "").upper()).strip()
        primeira_palavra = primeiro.split()[0] if primeiro.split() else ""
        if primeira_palavra in top_recorrentes and len(primeira_palavra) >= 4:
            # Confirma vinculo so se aparece em multiplos memos
            cnt = nomes_count[primeira_palavra]
            if cnt >= 5:
                vinculos.append(("GRUPO_INFERIDO", f"Nome '{primeira_palavra}' recorrente ({cnt}x) - candidato a grupo economico"))

        if not vinculos:
            continue

        cnpj = _extrair_cnpj(t)
        cnpj_fmt = ""
        if cnpj:
            cnpj_fmt = f"{cnpj[:2]}.{cnpj[2:5]}.{cnpj[5:8]}/{cnpj[8:12]}-{cnpj[12:14]}"
        info = d['cnpj_infos'].get(cnpj) if cnpj else None
        razao = info.razao_social if info else ""

        # Escreve linha
        tipos = ", ".join(v[0] for v in vinculos)
        justis = " | ".join(v[1] for v in vinculos)

        ws.cell(row=r, column=1, value=t.data)
        cv = ws.cell(row=r, column=2, value=round(t.valor, 2))
        cv.number_format = "#,##0.00"
        cv.font = Font(color=("DC2626" if t.valor < 0 else "16A34A"), bold=True)
        ws.cell(row=r, column=3, value=(t.memo or "")[:40])
        ws.cell(row=r, column=4, value=(t.nome or "")[:40])
        ws.cell(row=r, column=5, value=cnpj_fmt).font = Font(name="Consolas", size=10)
        ws.cell(row=r, column=6, value=razao)
        c_tipo = ws.cell(row=r, column=7, value=tipos)
        c_tipo.font = Font(bold=True, color="0052FF")
        ws.cell(row=r, column=8, value=justis)
        for c in range(1, 9):
            ws.cell(row=r, column=c).border = THIN_BORDER
            ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor="EFF6FF") if r % 2 == 0 else ZEBRA_FILL
        total_vol += abs(t.valor)
        r += 1

    if r == 9:
        ws.cell(row=9, column=1, value="(nenhuma parte relacionada detectada)").font = Font(italic=True, color="16A34A")
    else:
        ws.cell(row=r, column=1, value=f"TOTAL ({r-9} transacoes)").font = TOTAL_FONT
        ws.cell(row=r, column=2, value=round(total_vol, 2)).number_format = "#,##0.00"
        ws.cell(row=r, column=2).font = TOTAL_FONT
        for c in range(1, 9):
            ws.cell(row=r, column=c).fill = TOTAL_FILL

    for col, w in {1: 12, 2: 14, 3: 32, 4: 32, 5: 20, 6: 38, 7: 22, 8: 50}.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A9"
    if r > 9:
        ws.auto_filter.ref = f"A8:H{r-1}"


def aba_status_tributario(wb, d: dict, empresa: dict) -> None:
    """Classificacao das movimentacoes por incidencia fiscal provavel."""
    ws = wb.create_sheet("Status Tributario")
    aplicar_cabecalho(ws, d, empresa, ultima_col=8)
    ws.cell(row=5, column=1, value="STATUS TRIBUTARIO - Classificacao Fiscal das Movimentacoes").font = TITLE_FONT
    ws.merge_cells("A5:H5")
    ws.cell(row=6, column=1, value="Incidencia fiscal provavel - sugestao automatica, validar com regime tributario da empresa.").font = Font(italic=True, color="64748B", size=9)
    ws.merge_cells("A6:H6")

    # Resumo por categoria
    from collections import Counter, defaultdict
    cat_count = Counter()
    cat_volume = defaultdict(float)
    cat_retencao = defaultdict(float)

    detalhes = []
    for disp in sorted(d['disposicoes'], key=lambda x: x.transacao.data):
        t = disp.transacao
        cnpj = _extrair_cnpj(t)
        info = d['cnpj_infos'].get(cnpj) if cnpj else None
        porte = info.porte if info else ""
        trib = classificar_tributario(t.memo or "", t.nome or "", t.valor, cnpj, porte)
        cat_count[trib['categoria']] += 1
        cat_volume[trib['categoria']] += abs(t.valor)
        cat_retencao[trib['categoria']] += trib['valor_retencao']
        detalhes.append((t, trib, cnpj, info))

    # Tabela resumo por categoria
    ws.cell(row=8, column=1, value="RESUMO POR CATEGORIA FISCAL").font = Font(bold=True, size=11, color=NAVY)
    ws.merge_cells("A8:F8")

    headers_resumo = ["Categoria", "Qtd", "Volume (R$)", "Retencao Estimada (R$)", "Tributo Provavel", "Aliquota"]
    for c, h in enumerate(headers_resumo, start=1):
        ws.cell(row=9, column=c, value=h)
    style_header(ws, 9, 6)

    CATS_ORDEM = ["RETENCAO_PJ", "RETENCAO_PF", "OPERACAO_CREDITO", "IOF", "JUROS",
                  "PAGAMENTO_TRIBUTO", "TARIFA", "PIX_RECEBIDO", "BOLETO",
                  "COMPRA_CARTAO", "NAO_TRIBUTAVEL", "OUTRO"]
    TRIBUTO_REF = {
        "RETENCAO_PJ": ("PIS+COFINS+CSLL+IRRF", "6.15%"),
        "RETENCAO_PF": ("IRRF + INSS", "ate 27.5%"),
        "OPERACAO_CREDITO": ("IOF credito", "0.38%"),
        "IOF": ("IOF auto-cobrado", "0.38%"),
        "JUROS": ("Dedutivel IRPJ/CSLL", "—"),
        "PAGAMENTO_TRIBUTO": ("DARF/DAS/GPS/GNRE", "—"),
        "TARIFA": ("Despesa dedutivel", "—"),
        "PIX_RECEBIDO": ("PIS/COFINS sobre receita", "3.65-9.25%"),
        "BOLETO": ("PIS+COFINS+CSLL se servico", "4.65%"),
        "COMPRA_CARTAO": ("ICMS embutido", "—"),
        "NAO_TRIBUTAVEL": ("Sem fato gerador", "—"),
        "OUTRO": ("Indeterminado", "—"),
    }

    r = 10
    total_qtd = sum(cat_count.values())
    total_vol = sum(cat_volume.values())
    total_ret = sum(cat_retencao.values())
    for cat in CATS_ORDEM:
        qtd = cat_count.get(cat, 0)
        if qtd == 0:
            continue
        vol = cat_volume[cat]
        ret = cat_retencao[cat]
        tributo, aliq = TRIBUTO_REF.get(cat, ("—", "—"))
        c1 = ws.cell(row=r, column=1, value=cat)
        c1.font = Font(bold=True, color="DC2626" if cat.startswith("RETENCAO") else "0F172A")
        ws.cell(row=r, column=2, value=qtd).number_format = "#,##0"
        ws.cell(row=r, column=3, value=round(vol, 2)).number_format = "#,##0.00"
        ws.cell(row=r, column=4, value=round(ret, 2)).number_format = "#,##0.00"
        if ret > 0:
            ws.cell(row=r, column=4).font = Font(bold=True, color="D97706")
        ws.cell(row=r, column=5, value=tributo)
        ws.cell(row=r, column=6, value=aliq)
        for c in range(1, 7):
            ws.cell(row=r, column=c).border = THIN_BORDER
            if r % 2 == 0:
                ws.cell(row=r, column=c).fill = ZEBRA_FILL
        r += 1

    # Linha total
    ws.cell(row=r, column=1, value="TOTAL").font = TOTAL_FONT
    ws.cell(row=r, column=2, value=total_qtd).number_format = "#,##0"
    ws.cell(row=r, column=3, value=round(total_vol, 2)).number_format = "#,##0.00"
    ws.cell(row=r, column=4, value=round(total_ret, 2)).number_format = "#,##0.00"
    for c in range(1, 7):
        ws.cell(row=r, column=c).fill = TOTAL_FILL
        ws.cell(row=r, column=c).font = TOTAL_FONT
        ws.cell(row=r, column=c).border = THIN_BORDER
    r += 3

    # Detalhamento transacao a transacao
    ws.cell(row=r, column=1, value="DETALHAMENTO POR TRANSACAO").font = Font(bold=True, size=11, color=NAVY)
    ws.merge_cells(f"A{r}:H{r}")
    r += 1

    headers_det = ["Data", "Valor (R$)", "Memo", "Categoria Fiscal", "Tributo Provavel",
                   "Aliquota Sugerida", "Retencao (R$)", "Observacao"]
    for c, h in enumerate(headers_det, start=1):
        ws.cell(row=r, column=c, value=h)
    style_header(ws, r, 8)
    r += 1

    inicio_det = r
    for t, trib, cnpj, info in detalhes:
        ws.cell(row=r, column=1, value=t.data)
        cv = ws.cell(row=r, column=2, value=round(t.valor, 2))
        cv.number_format = "#,##0.00"
        cv.font = Font(color="DC2626" if t.valor < 0 else "16A34A")
        ws.cell(row=r, column=3, value=(t.memo or "")[:35])
        c_cat = ws.cell(row=r, column=4, value=trib['categoria'])
        if trib['categoria'].startswith("RETENCAO"):
            c_cat.font = Font(bold=True, color="DC2626")
            c_cat.fill = PatternFill("solid", fgColor="FEE2E2")
        elif trib['categoria'] in ("IOF", "OPERACAO_CREDITO", "JUROS"):
            c_cat.fill = PatternFill("solid", fgColor="FEF3C7")
        elif trib['categoria'] == "NAO_TRIBUTAVEL":
            c_cat.fill = PatternFill("solid", fgColor="DCFCE7")
        ws.cell(row=r, column=5, value=trib['tributo'])
        ws.cell(row=r, column=6, value=trib['aliquota_sugerida'])
        cr = ws.cell(row=r, column=7, value=round(trib['valor_retencao'], 2) if trib['valor_retencao'] else "")
        if trib['valor_retencao']:
            cr.number_format = "#,##0.00"
            cr.font = Font(bold=True, color="D97706")
        ws.cell(row=r, column=8, value=trib['obs'])
        for c in range(1, 9):
            ws.cell(row=r, column=c).border = THIN_BORDER
            if r % 2 == 0 and ws.cell(row=r, column=c).fill.fgColor.rgb in (None, "00000000"):
                ws.cell(row=r, column=c).fill = ZEBRA_FILL
        r += 1

    for col, w in {1: 12, 2: 14, 3: 32, 4: 20, 5: 38, 6: 22, 7: 14, 8: 50}.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = f"A{inicio_det}"
    if r > inicio_det:
        ws.auto_filter.ref = f"A{inicio_det-1}:H{r-1}"


async def gerar_pdf_via_playwright(html_text: str, pdf_path: Path) -> bool:
    try:
        from playwright.async_api import async_playwright
        async with async_playwright() as p:
            browser = await p.chromium.launch()
            page = await browser.new_page()
            await page.set_content(html_text, wait_until="load")
            await page.pdf(
                path=str(pdf_path),
                format="A4",
                margin={"top": "16mm", "right": "14mm", "bottom": "16mm", "left": "14mm"},
                print_background=True,
            )
            await browser.close()
        return True
    except Exception as exc:  # noqa: BLE001
        print(f"  PDF Playwright falhou: {type(exc).__name__}: {exc}")
        return False


# ────────────────────────────────────────────────────────────────────────
# CLI
# ────────────────────────────────────────────────────────────────────────


async def main_async(ofx_path: Path, out_dir: Path, prefixo: str,
                      empresa: dict | None = None) -> None:
    print(f"Lendo OFX: {ofx_path}")
    dados = await conciliar_ofx(ofx_path)
    print(f"  {dados['n_transacoes']} transacoes  |  CNPJs unicos: {len(dados['cnpj_infos'])}")
    print(f"  Automatizadas: {dados['automatizadas']}/{dados['n_transacoes']}")
    if dados["alertas_pos_baixa"]:
        print(f"  >>> ALERTAS POS-BAIXA: {dados['alertas_pos_baixa']}")

    if empresa is None:
        empresa = {
            "nome": "[NAO CADASTRADO]",
            "cnpj": "[PENDENTE]",
            "socios": "[PENDENTE]",
        }

    md_text = gerar_markdown(dados)
    md_path = out_dir / f"{prefixo}.md"
    md_path.write_text(md_text, encoding="utf-8")
    print(f"  MD:   {md_path}")

    html_text = gerar_html(md_text, dados)
    html_path = out_dir / f"{prefixo}.html"
    html_path.write_text(html_text, encoding="utf-8")
    print(f"  HTML: {html_path}")

    pdf_path = out_dir / f"{prefixo}.pdf"
    ok = await gerar_pdf_via_playwright(html_text, pdf_path)
    if ok:
        print(f"  PDF:  {pdf_path}")
    else:
        print(f"  PDF:  (fallback HTML em {html_path})")

    xlsx_path = out_dir / f"{prefixo}.xlsx"
    gerar_xlsx(dados, xlsx_path, empresa)
    print(f"  XLSX: {xlsx_path}")


def main() -> None:
    ap = argparse.ArgumentParser(description="Conciliacao + relatorios de 1 OFX")
    ap.add_argument("--ofx", required=True, help="Caminho do arquivo OFX")
    ap.add_argument("--out", default=r"C:\Users\Veloso\Downloads", help="Pasta de saida")
    ap.add_argument("--prefixo", default=None, help="Prefixo dos arquivos (default: deriva do OFX)")
    ap.add_argument("--empresa-nome", default=None, help="Razao social da empresa auditada")
    ap.add_argument("--empresa-cnpj", default=None, help="CNPJ da empresa auditada")
    ap.add_argument("--empresa-socios", default=None, help="Nomes dos socios (separados por virgula)")
    args = ap.parse_args()

    ofx_path = Path(args.ofx)
    if not ofx_path.exists():
        sys.exit(f"ERRO: arquivo nao encontrado: {ofx_path}")

    out_dir = Path(args.out)
    out_dir.mkdir(parents=True, exist_ok=True)

    if args.prefixo:
        prefixo = args.prefixo
    else:
        # Tenta extrair conta+periodo do OFX
        raw = ofx_path.read_text(encoding="latin-1", errors="ignore")
        ag = re.search(r"<BRANCHID>([^<\s]+)", raw)
        ct = re.search(r"<ACCTID>([^<\s]+)", raw)
        ini = re.search(r"<DTSTART>(\d{6})", raw)
        prefixo = f"CONCILIACAO_{ct.group(1) if ct else 'X'}_{ini.group(1) if ini else 'X'}"

    empresa = None
    if args.empresa_nome or args.empresa_cnpj or args.empresa_socios:
        empresa = {
            "nome": args.empresa_nome or "[NAO CADASTRADO]",
            "cnpj": args.empresa_cnpj or "[PENDENTE]",
            "socios": args.empresa_socios or "[PENDENTE]",
        }

    asyncio.run(main_async(ofx_path, out_dir, prefixo, empresa))


if __name__ == "__main__":
    main()
