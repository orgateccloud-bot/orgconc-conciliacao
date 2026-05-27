"""Gera RELATORIO_CONSOLIDADO.xlsx com 7 abas a partir dos OFXs reais."""
from __future__ import annotations

import re
import sys
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

sys.path.insert(0, r"C:\OrgConc")
from api.matchers.cascata import classificar, ler_ofx

ARQUIVOS = [
    (r"C:\Users\Veloso\Downloads\extrato-conta-corrente-ofx-unix_202605_20260514110822 (1).ofx", "JAN/2026", "158083-3"),
    (r"C:\Users\Veloso\Downloads\extrato-conta-corrente-ofx-unix_202605_20260514110900 (1).ofx", "MAR/2026", "158083-3"),
    (r"C:\Users\Veloso\Downloads\extrato-conta-corrente-ofx-unix_202605_20260514110917 (1).ofx", "ABR/2026", "158083-3"),
    (r"C:\Users\Veloso\Downloads\extrato-conta-corrente-ofx-unix_202605_20260514110938 (1).ofx", "MAI/2026", "158083-3"),
    (r"C:\Users\Veloso\Downloads\extrato-conta-corrente-ofx-unix_202605_20260504172522.ofx", "ABR/2026", "9695-4"),
    (r"C:\Users\Veloso\Downloads\extrato-conta-corrente-ofx-unix_202605_20260504172614.ofx", "ABR/2026", "51785-2"),
]

ESTAGIOS = {0: "TRANSF.INTERNA", 1: "CNPJ/CPF", 2: "NF-e", 3: "TARIFA",
            4: "TRIBUTO", 5: "CONTRATO", 6: "ALIAS/FUZZY"}
STATUS_MAP = {
    0: "Auto (regra)", 1: "Auto se CNPJ cadastrado", 2: "Auto se XML disponivel",
    3: "Auto (regra)", 4: "Auto se guia cadastrada",
    5: "Auto se contrato cadastrado", 6: "Auto se alias / fuzzy LLM",
}
RX_CNPJ = re.compile(r"\b(\d{2}\.\d{3}\.\d{3}[ /]?\d{4}-?\d{2})\b")

OUT_PATH = Path(r"C:\Users\Veloso\Downloads\RELATORIO_CONSOLIDADO.xlsx")

# Estilos reutilizáveis
NAVY = "0F172A"
BLUE = "0052FF"
SKY = "0EA5E9"
BORDER_LIGHT = Side(border_style="thin", color="E2E8F0")
THIN_BORDER = Border(top=BORDER_LIGHT, left=BORDER_LIGHT, right=BORDER_LIGHT, bottom=BORDER_LIGHT)
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TOTAL_FILL = PatternFill("solid", fgColor="1E3A8A")
TOTAL_FONT = Font(bold=True, color="FFFFFF", size=11)
ZEBRA_FILL = PatternFill("solid", fgColor="F8FAFC")
TITLE_FONT = Font(bold=True, size=14, color=NAVY)


def style_header_row(ws, row: int, n_cols: int) -> None:
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = THIN_BORDER


def style_data_row(ws, row: int, n_cols: int, zebra: bool = False) -> None:
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.border = THIN_BORDER
        if zebra:
            cell.fill = ZEBRA_FILL


def auto_width(ws, widths: dict[int, int]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width


def coletar_dados():
    todos = []
    valor_rec = Counter()
    nome_rec = Counter()
    cnpj_rec = Counter()

    for path, mes, conta in ARQUIVOS:
        txs = ler_ofx(path)
        for t in txs:
            res = classificar(t)
            todos.append((mes, conta, t, res))
            if t.valor < 0:
                valor_rec[round(abs(t.valor), 2)] += 1
                nome_simp = (t.nome or t.memo or "")[:40].upper().strip()
                if nome_simp and len(nome_simp) > 5:
                    nome_rec[nome_simp] += 1
                m = RX_CNPJ.search(t.nome or "") or RX_CNPJ.search(t.memo or "")
                if m:
                    cnpj_rec[re.sub(r"\D", "", m.group(1))] += 1
    return todos, valor_rec, nome_rec, cnpj_rec


def aba_resumo(wb: Workbook, todos, totais_est: Counter) -> None:
    ws = wb.create_sheet("Resumo", 0)
    n = len(todos)
    cred = sum(t.valor for _, _, t, _ in todos if t.valor > 0)
    deb = sum(t.valor for _, _, t, _ in todos if t.valor < 0)

    ws["A1"] = "RELATORIO CONSOLIDADO - EXTRATOS 2026"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")
    ws["A2"] = f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')} - OrgConc/OrgNeural2"
    ws["A2"].font = Font(italic=True, color="64748B", size=9)
    ws.merge_cells("A2:D2")

    ws["A4"] = "INDICADOR"
    ws["B4"] = "VALOR"
    style_header_row(ws, 4, 2)

    rows = [
        ("Total de transacoes", n),
        ("Volume de creditos (R$)", round(cred, 2)),
        ("Volume de debitos (R$)", round(deb, 2)),
        ("Saldo de fluxo (R$)", round(cred + deb, 2)),
        ("Periodo", "Janeiro a Maio/2026"),
        ("Contas", "158083-3, 9695-4, 51785-2 (Sicoob 756)"),
    ]
    for i, (k, v) in enumerate(rows, start=5):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws.cell(row=i, column=2, value=v)
        if isinstance(v, (int, float)) and "R$" in k:
            ws.cell(row=i, column=2).number_format = "#,##0.00"
        style_data_row(ws, i, 2, zebra=(i % 2 == 0))

    # Distribuição por estágio
    ws["A12"] = "DISTRIBUICAO POR ESTAGIO"
    ws["A12"].font = TITLE_FONT
    ws.merge_cells("A12:E12")

    ws["A14"] = "Estagio"
    ws["B14"] = "Tipo"
    ws["C14"] = "Quantidade"
    ws["D14"] = "%"
    ws["E14"] = "Status"
    style_header_row(ws, 14, 5)

    r = 15
    for est in range(7):
        if totais_est.get(est):
            qtd = totais_est[est]
            pct = qtd / n
            ws.cell(row=r, column=1, value=est)
            ws.cell(row=r, column=2, value=ESTAGIOS[est])
            ws.cell(row=r, column=3, value=qtd).number_format = "#,##0"
            cell_pct = ws.cell(row=r, column=4, value=pct)
            cell_pct.number_format = "0.0%"
            ws.cell(row=r, column=5, value=STATUS_MAP[est])
            style_data_row(ws, r, 5, zebra=(r % 2 == 0))
            r += 1

    # Total
    ws.cell(row=r, column=1, value="TOTAL").font = TOTAL_FONT
    ws.cell(row=r, column=1).fill = TOTAL_FILL
    ws.cell(row=r, column=3, value=n).number_format = "#,##0"
    ws.cell(row=r, column=3).font = TOTAL_FONT
    ws.cell(row=r, column=3).fill = TOTAL_FILL
    ws.cell(row=r, column=4, value=1.0).number_format = "0.0%"
    ws.cell(row=r, column=4).font = TOTAL_FONT
    ws.cell(row=r, column=4).fill = TOTAL_FILL
    for c in (2, 5):
        ws.cell(row=r, column=c).fill = TOTAL_FILL

    auto_width(ws, {1: 18, 2: 22, 3: 14, 4: 10, 5: 36})
    ws.freeze_panes = "A4"


def aba_por_mes_conta(wb: Workbook, todos) -> None:
    ws = wb.create_sheet("Por Mes-Conta")
    ws["A1"] = "FLUXO POR MES E CONTA"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:E1")

    headers = ["Mes", "Conta", "Transacoes", "Creditos (R$)", "Debitos (R$)"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=3, column=c, value=h)
    style_header_row(ws, 3, 5)

    por_mc = defaultdict(lambda: {"n": 0, "cred": 0.0, "deb": 0.0})
    for mes, conta, t, _ in todos:
        k = (mes, conta)
        por_mc[k]["n"] += 1
        if t.valor > 0:
            por_mc[k]["cred"] += t.valor
        else:
            por_mc[k]["deb"] += t.valor

    r = 4
    tot_n, tot_c, tot_d = 0, 0.0, 0.0
    for (mes, conta), v in sorted(por_mc.items(), key=lambda x: (x[0][1], x[0][0])):
        ws.cell(row=r, column=1, value=mes)
        ws.cell(row=r, column=2, value=conta)
        ws.cell(row=r, column=3, value=v["n"]).number_format = "#,##0"
        ws.cell(row=r, column=4, value=round(v["cred"], 2)).number_format = "#,##0.00"
        ws.cell(row=r, column=5, value=round(v["deb"], 2)).number_format = "#,##0.00"
        style_data_row(ws, r, 5, zebra=(r % 2 == 0))
        tot_n += v["n"]
        tot_c += v["cred"]
        tot_d += v["deb"]
        r += 1

    # Totais
    ws.cell(row=r, column=1, value="TOTAL")
    ws.cell(row=r, column=3, value=tot_n).number_format = "#,##0"
    ws.cell(row=r, column=4, value=round(tot_c, 2)).number_format = "#,##0.00"
    ws.cell(row=r, column=5, value=round(tot_d, 2)).number_format = "#,##0.00"
    for c in range(1, 6):
        ws.cell(row=r, column=c).fill = TOTAL_FILL
        ws.cell(row=r, column=c).font = TOTAL_FONT

    auto_width(ws, {1: 14, 2: 14, 3: 13, 4: 18, 5: 18})
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:E{r-1}"


def aba_top_recorrentes(wb: Workbook, valor_rec: Counter) -> None:
    ws = wb.create_sheet("Top Valores Recorrentes")
    ws["A1"] = "TOP 30 VALORES RECORRENTES (Candidatos a Contrato)"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:C1")
    ws["A2"] = "Debitos com o mesmo valor exato - investigue para cadastrar contratos recorrentes."
    ws["A2"].font = Font(italic=True, color="64748B", size=9)
    ws.merge_cells("A2:C2")

    headers = ["Valor (R$)", "Ocorrencias", "Sugestao"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=4, column=c, value=h)
    style_header_row(ws, 4, 3)

    r = 5
    for valor, qtd in valor_rec.most_common(30):
        sug = ("Cadastrar como contrato" if qtd >= 10
               else "Investigar" if qtd >= 5 else "Pode ser coincidencia")
        cell_v = ws.cell(row=r, column=1, value=valor)
        cell_v.number_format = "#,##0.00"
        ws.cell(row=r, column=2, value=qtd).number_format = "#,##0"
        cell_s = ws.cell(row=r, column=3, value=sug)
        if qtd >= 10:
            cell_s.fill = PatternFill("solid", fgColor="DCFCE7")
        elif qtd >= 5:
            cell_s.fill = PatternFill("solid", fgColor="FEF3C7")
        style_data_row(ws, r, 3, zebra=(r % 2 == 0))
        r += 1

    auto_width(ws, {1: 16, 2: 14, 3: 32})
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:C{r-1}"


def aba_top_cnpjs(wb: Workbook, cnpj_rec: Counter) -> None:
    ws = wb.create_sheet("Top CNPJs")
    ws["A1"] = "TOP 30 CONTRAPARTES POR CNPJ (Candidatos a Cadastro)"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:B1")

    headers = ["CNPJ", "Ocorrencias"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=3, column=c, value=h)
    style_header_row(ws, 3, 2)

    r = 4
    for cnpj, qtd in [(d, q) for d, q in cnpj_rec.most_common(30) if q >= 3]:
        if len(cnpj) == 14:
            cnpj_fmt = f"{cnpj[:2]}.{cnpj[2:5]}.{cnpj[5:8]}/{cnpj[8:12]}-{cnpj[12:14]}"
        else:
            cnpj_fmt = cnpj
        ws.cell(row=r, column=1, value=cnpj_fmt)
        ws.cell(row=r, column=2, value=qtd).number_format = "#,##0"
        style_data_row(ws, r, 2, zebra=(r % 2 == 0))
        r += 1

    auto_width(ws, {1: 22, 2: 14})
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:B{r-1}"


def aba_top_debitos(wb: Workbook, todos) -> None:
    ws = wb.create_sheet("Top Debitos")
    ws["A1"] = "TOP 30 MAIORES DEBITOS"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:F1")

    headers = ["Data", "Conta", "Valor (R$)", "Memo", "Nome / Favorecido", "Mes"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=3, column=c, value=h)
    style_header_row(ws, 3, 6)

    debitos = [(t.valor, t, mes, conta) for mes, conta, t, _ in todos if t.valor < 0]
    debitos.sort(key=lambda x: x[0])

    r = 4
    for valor, t, mes, conta in debitos[:30]:
        ws.cell(row=r, column=1, value=t.data)
        ws.cell(row=r, column=2, value=conta)
        cell_v = ws.cell(row=r, column=3, value=round(valor, 2))
        cell_v.number_format = "#,##0.00"
        cell_v.font = Font(color="DC2626")
        ws.cell(row=r, column=4, value=t.memo or "")
        ws.cell(row=r, column=5, value=t.nome or "")
        ws.cell(row=r, column=6, value=mes)
        style_data_row(ws, r, 6, zebra=(r % 2 == 0))
        r += 1

    auto_width(ws, {1: 12, 2: 12, 3: 16, 4: 38, 5: 38, 6: 12})
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:F{r-1}"


def aba_top_creditos(wb: Workbook, todos) -> None:
    ws = wb.create_sheet("Top Creditos")
    ws["A1"] = "TOP 30 MAIORES CREDITOS"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:F1")

    headers = ["Data", "Conta", "Valor (R$)", "Memo", "Nome / Remetente", "Mes"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=3, column=c, value=h)
    style_header_row(ws, 3, 6)

    creditos = [(t.valor, t, mes, conta) for mes, conta, t, _ in todos if t.valor > 0]
    creditos.sort(key=lambda x: -x[0])

    r = 4
    for valor, t, mes, conta in creditos[:30]:
        ws.cell(row=r, column=1, value=t.data)
        ws.cell(row=r, column=2, value=conta)
        cell_v = ws.cell(row=r, column=3, value=round(valor, 2))
        cell_v.number_format = "#,##0.00"
        cell_v.font = Font(color="16A34A")
        ws.cell(row=r, column=4, value=t.memo or "")
        ws.cell(row=r, column=5, value=t.nome or "")
        ws.cell(row=r, column=6, value=mes)
        style_data_row(ws, r, 6, zebra=(r % 2 == 0))
        r += 1

    auto_width(ws, {1: 12, 2: 12, 3: 16, 4: 38, 5: 38, 6: 12})
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:F{r-1}"


def aba_transacoes(wb: Workbook, todos) -> None:
    """Todas as 5.880 transações com classificação completa."""
    ws = wb.create_sheet("Transacoes (Todas)")
    ws["A1"] = f"TRANSACOES CLASSIFICADAS - {len(todos):,} LINHAS"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:I1")

    headers = ["Mes", "Conta", "Data", "Tipo", "Valor (R$)", "Estagio", "Metodo", "Memo", "Nome"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=3, column=c, value=h)
    style_header_row(ws, 3, 9)

    r = 4
    for mes, conta, t, res in todos:
        ws.cell(row=r, column=1, value=mes)
        ws.cell(row=r, column=2, value=conta)
        ws.cell(row=r, column=3, value=t.data)
        ws.cell(row=r, column=4, value=t.tipo)
        cell_v = ws.cell(row=r, column=5, value=round(t.valor, 2))
        cell_v.number_format = "#,##0.00"
        if t.valor < 0:
            cell_v.font = Font(color="DC2626")
        else:
            cell_v.font = Font(color="16A34A")
        ws.cell(row=r, column=6, value=res.estagio)
        ws.cell(row=r, column=7, value=res.metodo)
        ws.cell(row=r, column=8, value=t.memo or "")
        ws.cell(row=r, column=9, value=t.nome or "")
        # zebra leve apenas no estilo da borda — sem fill (5880 linhas pesariam)
        for c in range(1, 10):
            ws.cell(row=r, column=c).border = THIN_BORDER
        r += 1

    auto_width(ws, {1: 11, 2: 12, 3: 12, 4: 8, 5: 14, 6: 9, 7: 22, 8: 35, 9: 35})
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:I{r-1}"


def main() -> None:
    print("Coletando dados dos OFXs...")
    todos, valor_rec, nome_rec, cnpj_rec = coletar_dados()
    totais_est = Counter(res.estagio for _, _, _, res in todos)
    print(f"  {len(todos):,} transacoes carregadas.")

    wb = Workbook()
    # Remove sheet padrão
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    print("Gerando abas...")
    aba_resumo(wb, todos, totais_est)
    aba_por_mes_conta(wb, todos)
    aba_top_recorrentes(wb, valor_rec)
    aba_top_cnpjs(wb, cnpj_rec)
    aba_top_debitos(wb, todos)
    aba_top_creditos(wb, todos)
    aba_transacoes(wb, todos)

    print(f"Salvando em {OUT_PATH}...")
    wb.save(str(OUT_PATH))
    print(f"OK - 7 abas, {len(todos):,} transacoes no detalhamento")


if __name__ == "__main__":
    main()
