"""Enriquece os CNPJs detectados nos OFXs via BrasilAPI/RFB e regera XLSX/PDF."""
from __future__ import annotations

import asyncio
import re
import sys
import time
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, r"C:\OrgConc")
from api.matchers.cascata import classificar, ler_ofx
from api.matchers.cnpj_enricher import CnpjInfo, enriquecer_lote, normaliza_cnpj

ARQUIVOS = [
    (r"C:\Users\Veloso\Downloads\extrato-conta-corrente-ofx-unix_202605_20260514110822 (1).ofx", "JAN/2026", "158083-3"),
    (r"C:\Users\Veloso\Downloads\extrato-conta-corrente-ofx-unix_202605_20260514110900 (1).ofx", "MAR/2026", "158083-3"),
    (r"C:\Users\Veloso\Downloads\extrato-conta-corrente-ofx-unix_202605_20260514110917 (1).ofx", "ABR/2026", "158083-3"),
    (r"C:\Users\Veloso\Downloads\extrato-conta-corrente-ofx-unix_202605_20260514110938 (1).ofx", "MAI/2026", "158083-3"),
    (r"C:\Users\Veloso\Downloads\extrato-conta-corrente-ofx-unix_202605_20260504172522.ofx", "ABR/2026", "9695-4"),
    (r"C:\Users\Veloso\Downloads\extrato-conta-corrente-ofx-unix_202605_20260504172614.ofx", "ABR/2026", "51785-2"),
]

ESTAGIOS = {0: "TRANSF.INTERNA", 1: "CNPJ/CPF", 2: "NF-e", 3: "TARIFA",
            4: "TRIBUTO", 5: "CONTRATO", 6: "ALIAS/FUZZY"}

OUT_PATH = Path(r"C:\Users\Veloso\Downloads\RELATORIO_ENRIQUECIDO_v3.xlsx")
RX_CNPJ = re.compile(r"(\d{2})[.](\d{3})[.](\d{3})[ /](\d{4})[-](\d{2})")

NAVY = "0F172A"
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TOTAL_FILL = PatternFill("solid", fgColor="1E3A8A")
TOTAL_FONT = Font(bold=True, color="FFFFFF", size=11)
ZEBRA_FILL = PatternFill("solid", fgColor="F8FAFC")
ALERT_FILL = PatternFill("solid", fgColor="FEE2E2")
TITLE_FONT = Font(bold=True, size=14, color=NAVY)
BORDER_LIGHT = Side(border_style="thin", color="E2E8F0")
THIN_BORDER = Border(top=BORDER_LIGHT, left=BORDER_LIGHT, right=BORDER_LIGHT, bottom=BORDER_LIGHT)


def style_header_row(ws, row: int, n_cols: int) -> None:
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = THIN_BORDER


def auto_width(ws, widths: dict[int, int]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width


def extrair_cnpj(t) -> str | None:
    for fonte in (t.nome or "", t.memo or ""):
        m = RX_CNPJ.search(fonte)
        if m:
            return "".join(m.groups())
    return None


def coletar_transacoes():
    todos = []
    for path, mes, conta in ARQUIVOS:
        for t in ler_ofx(path):
            res = classificar(t)
            cnpj = extrair_cnpj(t)
            todos.append((mes, conta, t, res, cnpj))
    return todos


async def main_async():
    print("Coletando transações dos OFXs...")
    todos = coletar_transacoes()
    cnpjs_unicos = sorted({c for _, _, _, _, c in todos if c})
    print(f"  {len(todos):,} transações | {len(cnpjs_unicos)} CNPJs únicos")

    print(f"\nEnriquecendo {len(cnpjs_unicos)} CNPJs (cache -> RFB local -> BrasilAPI)...")
    t0 = time.time()

    contagem = {"feitos": 0}

    def progresso(feitos, total):
        contagem["feitos"] = feitos
        if feitos % 25 == 0 or feitos == total:
            elapsed = time.time() - t0
            rps = feitos / max(elapsed, 0.1)
            print(f"  {feitos}/{total}  ({rps:.1f} req/s, {elapsed:.0f}s)")

    infos = await enriquecer_lote(cnpjs_unicos, db=None, max_concurrency=2, progress_cb=progresso)
    elapsed = time.time() - t0
    print(f"\n  Concluído em {elapsed:.0f}s")
    fontes = Counter(i.fonte for i in infos.values())
    print(f"  Fontes: {dict(fontes)}")
    achados = sum(1 for i in infos.values() if i.razao_social)
    print(f"  Razão social obtida: {achados}/{len(infos)} ({100*achados/len(infos):.1f}%)")

    # Gera XLSX enriquecido
    print(f"\nGerando XLSX em {OUT_PATH}...")
    gerar_xlsx_enriquecido(todos, infos)
    print("OK")


def gerar_xlsx_enriquecido(todos, infos: dict[str, CnpjInfo]) -> None:
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # ── Aba 1: Resumo executivo ─────────────────────────────────────────
    ws = wb.create_sheet("Resumo")
    ws["A1"] = "RELATORIO CONSOLIDADO ENRIQUECIDO - EXTRATOS 2026"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:E1")
    ws["A2"] = f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')} - OrgConc/OrgNeural2"
    ws["A2"].font = Font(italic=True, color="64748B", size=9)
    ws.merge_cells("A2:E2")

    n = len(todos)
    cred = sum(t.valor for _, _, t, _, _ in todos if t.valor > 0)
    deb = sum(t.valor for _, _, t, _, _ in todos if t.valor < 0)
    com_cnpj = sum(1 for _, _, _, _, c in todos if c)
    cnpj_com_razao = sum(1 for _, _, _, _, c in todos if c and infos.get(c) and infos[c].razao_social)

    indicadores = [
        ("Total de transacoes", n),
        ("Volume creditos (R$)", round(cred, 2)),
        ("Volume debitos (R$)", round(deb, 2)),
        ("Transacoes com CNPJ identificado", com_cnpj),
        ("Transacoes com razao social enriquecida", cnpj_com_razao),
        ("Taxa de enriquecimento (%)",
         round(100 * cnpj_com_razao / max(com_cnpj, 1), 1)),
        ("Periodo", "Janeiro a Maio/2026"),
        ("Contas", "158083-3, 9695-4, 51785-2 (Sicoob 756)"),
    ]
    ws["A4"] = "INDICADOR"
    ws["B4"] = "VALOR"
    style_header_row(ws, 4, 2)
    for i, (k, v) in enumerate(indicadores, start=5):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        c = ws.cell(row=i, column=2, value=v)
        if isinstance(v, float) and "R$" in k:
            c.number_format = "#,##0.00"
        if i % 2 == 0:
            ws.cell(row=i, column=1).fill = ZEBRA_FILL
            ws.cell(row=i, column=2).fill = ZEBRA_FILL

    auto_width(ws, {1: 36, 2: 32})
    ws.freeze_panes = "A4"

    # ── Aba 2: CNPJs enriquecidos ────────────────────────────────────────
    ws = wb.create_sheet("CNPJs Enriquecidos")
    ws["A1"] = f"CONTRAPARTES IDENTIFICADAS - {len(infos)} CNPJS"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:J1")
    ws["A2"] = "Razao social via BrasilAPI / base RFB / cache. Linhas em vermelho = BAIXADA/INAPTA."
    ws["A2"].font = Font(italic=True, color="64748B", size=9)
    ws.merge_cells("A2:J2")

    headers = ["CNPJ", "Razao Social", "Nome Fantasia", "Situacao", "Data Baixa/Situacao",
               "UF", "Municipio", "CNAE", "CNAE Descricao", "Porte", "Aparicoes"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=4, column=c, value=h)
    style_header_row(ws, 4, 11)

    aparicoes = Counter(c for _, _, _, _, c in todos if c)
    items = sorted(infos.items(), key=lambda x: -aparicoes.get(x[0], 0))
    r = 5
    for cnpj, info in items:
        fmt = f"{cnpj[:2]}.{cnpj[2:5]}.{cnpj[5:8]}/{cnpj[8:12]}-{cnpj[12:14]}"
        is_baixada = "BAIXADA" in info.situacao or "INAPTA" in info.situacao
        ws.cell(row=r, column=1, value=fmt).font = Font(name="Consolas", size=10)
        ws.cell(row=r, column=2, value=info.razao_social or "(nao encontrado)")
        ws.cell(row=r, column=3, value=info.nome_fantasia)
        ws.cell(row=r, column=4, value=info.situacao)
        ws.cell(row=r, column=5, value=info.data_situacao or "")
        ws.cell(row=r, column=6, value=info.uf)
        ws.cell(row=r, column=7, value=info.municipio)
        ws.cell(row=r, column=8, value=info.cnae_principal)
        ws.cell(row=r, column=9, value=info.cnae_descricao)
        ws.cell(row=r, column=10, value=info.porte)
        ws.cell(row=r, column=11, value=aparicoes.get(cnpj, 0)).number_format = "#,##0"
        for c in range(1, 12):
            cell = ws.cell(row=r, column=c)
            cell.border = THIN_BORDER
            if is_baixada:
                cell.fill = ALERT_FILL
                if c == 4:
                    cell.font = Font(bold=True, color="DC2626")
                elif c == 5:
                    cell.font = Font(bold=True, color="DC2626")
            elif r % 2 == 0:
                cell.fill = ZEBRA_FILL
        r += 1

    auto_width(ws, {1: 20, 2: 42, 3: 30, 4: 22, 5: 16, 6: 5, 7: 22, 8: 10, 9: 42, 10: 22, 11: 12})
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:K{r-1}"

    # ── Aba 3: Transações com enriquecimento ─────────────────────────────
    ws = wb.create_sheet("Transacoes Enriquecidas")
    ws["A1"] = f"TRANSACOES + RAZAO SOCIAL - {len(todos):,} LINHAS"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:K1")

    headers = ["Mes", "Conta", "Data Trans.", "Tipo", "Valor (R$)", "Estagio", "Memo",
               "Nome (banco)", "CNPJ", "Razao Social (RFB)", "Situacao", "Data Baixa"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=3, column=c, value=h)
    style_header_row(ws, 3, 12)

    r = 4
    for mes, conta, t, res, cnpj in todos:
        ws.cell(row=r, column=1, value=mes)
        ws.cell(row=r, column=2, value=conta)
        ws.cell(row=r, column=3, value=t.data)
        ws.cell(row=r, column=4, value=t.tipo)
        cell_v = ws.cell(row=r, column=5, value=round(t.valor, 2))
        cell_v.number_format = "#,##0.00"
        cell_v.font = Font(color=("DC2626" if t.valor < 0 else "16A34A"))
        ws.cell(row=r, column=6, value=res.estagio)
        ws.cell(row=r, column=7, value=t.memo or "")
        ws.cell(row=r, column=8, value=t.nome or "")
        if cnpj:
            fmt = f"{cnpj[:2]}.{cnpj[2:5]}.{cnpj[5:8]}/{cnpj[8:12]}-{cnpj[12:14]}"
            ws.cell(row=r, column=9, value=fmt).font = Font(name="Consolas", size=10)
            info = infos.get(cnpj)
            if info:
                ws.cell(row=r, column=10, value=info.razao_social or "")
                ws.cell(row=r, column=11, value=info.situacao or "")
                is_baixada = "BAIXADA" in (info.situacao or "") or "INAPTA" in (info.situacao or "")
                # Data da baixa: só exibe quando empresa está baixada/inapta (a data_situacao
                # de empresas ATIVAS é a data de abertura — irrelevante aqui)
                if is_baixada:
                    cell_db = ws.cell(row=r, column=12, value=info.data_situacao or "")
                    cell_db.font = Font(bold=True, color="DC2626")
                    for c in range(1, 13):
                        ws.cell(row=r, column=c).fill = ALERT_FILL
        for c in range(1, 13):
            ws.cell(row=r, column=c).border = THIN_BORDER
        r += 1

    auto_width(ws, {1: 11, 2: 12, 3: 12, 4: 8, 5: 14, 6: 8, 7: 32, 8: 32, 9: 20, 10: 40, 11: 18, 12: 13})
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:L{r-1}"

    # ── Aba 4: Alertas (situação ≠ ATIVA) ────────────────────────────────
    ws = wb.create_sheet("Alertas CNPJ")
    ws["A1"] = "TRANSACOES COM CONTRAPARTE NAO ATIVA"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:F1")

    headers = ["Data", "Valor (R$)", "CNPJ", "Razao Social", "Situacao", "Aparicoes"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=3, column=c, value=h)
    style_header_row(ws, 3, 6)

    alertas = [
        (mes, conta, t, cnpj, infos[cnpj])
        for mes, conta, t, _, cnpj in todos
        if cnpj and cnpj in infos
        and infos[cnpj].situacao
        and infos[cnpj].situacao not in ("ATIVA",)
        and "ATIVA" not in (infos[cnpj].situacao or "")
    ]
    r = 4
    for _, _, t, cnpj, info in alertas[:200]:
        fmt = f"{cnpj[:2]}.{cnpj[2:5]}.{cnpj[5:8]}/{cnpj[8:12]}-{cnpj[12:14]}"
        ws.cell(row=r, column=1, value=t.data)
        cv = ws.cell(row=r, column=2, value=round(t.valor, 2))
        cv.number_format = "#,##0.00"
        cv.font = Font(color="DC2626")
        ws.cell(row=r, column=3, value=fmt).font = Font(name="Consolas", size=10)
        ws.cell(row=r, column=4, value=info.razao_social)
        ws.cell(row=r, column=5, value=info.situacao).font = Font(bold=True, color="DC2626")
        ws.cell(row=r, column=6, value=aparicoes.get(cnpj, 0))
        for c in range(1, 7):
            ws.cell(row=r, column=c).border = THIN_BORDER
        r += 1

    if r == 4:
        ws.cell(row=4, column=1, value="(nenhum alerta)").font = Font(italic=True, color="16A34A")

    auto_width(ws, {1: 12, 2: 14, 3: 20, 4: 42, 5: 18, 6: 12})
    ws.freeze_panes = "A4"
    if r > 4:
        ws.auto_filter.ref = f"A3:F{r-1}"

    # ── Aba 5: Pagamentos Pos-Baixa (CRITICO para auditoria) ────────────
    aba_pos_baixa(wb, todos, infos)

    wb.save(str(OUT_PATH))


def aba_pos_baixa(wb, todos, infos: dict) -> None:
    """Transacoes feitas APOS a data de baixa do CNPJ - achado critico."""
    ws = wb.create_sheet("Pagamentos Pos-Baixa")
    ws["A1"] = "PAGAMENTOS REALIZADOS APOS A BAIXA DO CNPJ - CRITICO"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:I1")
    ws["A2"] = (
        "Transacoes posteriores a data de baixa/inaptidao da contraparte. "
        "Para auditoria forense: pagamentos suspeitos para empresas extintas."
    )
    ws["A2"].font = Font(italic=True, color="64748B", size=9)
    ws.merge_cells("A2:I2")

    headers = ["Data Trans.", "Conta", "Valor (R$)", "Memo", "Nome (banco)",
               "CNPJ", "Razao Social", "Data Baixa", "Dias Apos Baixa"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=4, column=c, value=h)
    style_header_row(ws, 4, 9)

    # Filtra transacoes pos-baixa
    pos_baixa = []
    for mes, conta, t, _, cnpj in todos:
        if not cnpj or cnpj not in infos:
            continue
        info = infos[cnpj]
        sit = info.situacao or ""
        if "BAIXADA" not in sit and "INAPTA" not in sit:
            continue
        if not info.data_situacao:
            continue
        try:
            data_baixa = date.fromisoformat(info.data_situacao[:10])
            data_trans = date.fromisoformat(t.data[:10])
        except (ValueError, TypeError):
            continue
        if data_trans > data_baixa:
            dias = (data_trans - data_baixa).days
            pos_baixa.append((data_trans, conta, t, cnpj, info, dias, data_baixa))

    # Ordenar pelos mais antigos pagamentos pos-baixa primeiro (maior intervalo = mais grave)
    pos_baixa.sort(key=lambda x: -x[5])

    r = 5
    total_volume = 0.0
    for data_trans, conta, t, cnpj, info, dias, data_baixa in pos_baixa:
        fmt = f"{cnpj[:2]}.{cnpj[2:5]}.{cnpj[5:8]}/{cnpj[8:12]}-{cnpj[12:14]}"
        ws.cell(row=r, column=1, value=t.data).font = Font(bold=True, color="DC2626")
        ws.cell(row=r, column=2, value=conta)
        cv = ws.cell(row=r, column=3, value=round(t.valor, 2))
        cv.number_format = "#,##0.00"
        cv.font = Font(color="DC2626", bold=True)
        ws.cell(row=r, column=4, value=(t.memo or "")[:50])
        ws.cell(row=r, column=5, value=(t.nome or "")[:50])
        ws.cell(row=r, column=6, value=fmt).font = Font(name="Consolas", size=10)
        ws.cell(row=r, column=7, value=info.razao_social[:50])
        ws.cell(row=r, column=8, value=info.data_situacao or "")
        ws.cell(row=r, column=9, value=dias).number_format = "#,##0"
        for c in range(1, 10):
            ws.cell(row=r, column=c).border = THIN_BORDER
            ws.cell(row=r, column=c).fill = ALERT_FILL
        total_volume += abs(t.valor)
        r += 1

    if r == 5:
        ws.cell(row=5, column=1, value="(nenhuma transacao pos-baixa detectada)").font = Font(italic=True, color="16A34A")
    else:
        # Linha de total
        ws.cell(row=r, column=1, value=f"TOTAL ({r-5} transacoes)")
        ws.cell(row=r, column=3, value=round(total_volume, 2)).number_format = "#,##0.00"
        for c in range(1, 10):
            ws.cell(row=r, column=c).fill = TOTAL_FILL
            ws.cell(row=r, column=c).font = TOTAL_FONT

    auto_width(ws, {1: 12, 2: 12, 3: 14, 4: 38, 5: 38, 6: 20, 7: 42, 8: 13, 9: 15})
    ws.freeze_panes = "A5"
    if r > 5:
        ws.auto_filter.ref = f"A4:I{r-1}"


if __name__ == "__main__":
    asyncio.run(main_async())
