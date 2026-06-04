"""Testes do endpoint unificado POST /fiscal/laudo (OFX + XMLs -> abas fiscais)."""
from __future__ import annotations

import io
import os
import zipfile
from pathlib import Path

os.environ.setdefault("ORGCONC_DATA_DIR", str(Path(__file__).resolve().parent / "_data_test"))
os.environ.setdefault("ANTHROPIC_API_KEY", "sk-ant-test")

from fastapi.testclient import TestClient

from api.main import app
from api.services.laudo_forense import carregar_docs_xmls

client = TestClient(app)
_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

_OFX = b"""OFXHEADER:100
DATA:OFXSGML
VERSION:102
<OFX><BANKMSGSRSV1><STMTTRNRS><STMTRS><CURDEF>BRL
<BANKACCTFROM><BANKID>001<ACCTID>12345-6<ACCTTYPE>CHECKING</BANKACCTFROM>
<BANKTRANLIST><DTSTART>20260101<DTEND>20260131
<STMTTRN><TRNTYPE>DEBIT<DTPOSTED>20260115<TRNAMT>-1500.00<FITID>T1<MEMO>PIX 12.345.678 0001-90 FORN</STMTTRN>
</BANKTRANLIST><LEDGERBAL><BALAMT>0<DTASOF>20260131</LEDGERBAL>
</STMTRS></STMTTRNRS></BANKMSGSRSV1></OFX>"""


def _nfe_xml(chave="1" * 44, valor="1500.00"):
    return (
        '<?xml version="1.0"?>'
        '<nfeProc xmlns="http://www.portalfiscal.inf.br/nfe"><NFe>'
        f'<infNFe Id="NFe{chave}">'
        "<ide><mod>55</mod><nNF>1</nNF><serie>1</serie><natOp>VENDA</natOp></ide>"
        "<emit><CNPJ>12345678000190</CNPJ><xNome>FORN</xNome>"
        "<enderEmit><UF>GO</UF></enderEmit></emit>"
        "<dest><CNPJ>99888777000166</CNPJ><xNome>DEST</xNome></dest>"
        "<det><prod><CFOP>5102</CFOP></prod></det>"
        f"<total><ICMSTot><vNF>{valor}</vNF></ICMSTot></total>"
        "</infNFe></NFe></nfeProc>"
    ).encode()


# ── carregar_docs_xmls (parse em memória da engine) ─────────────────────────

def test_carregar_docs_xmls_nfe_direto():
    nfes, ctes, n = carregar_docs_xmls([("nota.xml", _nfe_xml())])
    assert n == 1 and len(nfes) == 1 and len(ctes) == 0
    assert nfes[0]["chave"] == "1" * 44


def test_carregar_docs_xmls_zip():
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as zf:
        zf.writestr("a.xml", _nfe_xml(chave="2" * 44))
        zf.writestr("b.xml", _nfe_xml(chave="3" * 44))
    nfes, _ctes, n = carregar_docs_xmls([("lote.zip", buf.getvalue())])
    assert n == 2 and len(nfes) == 2


# ── Endpoint ────────────────────────────────────────────────────────────────

def test_laudo_formato_invalido_400():
    files = [("arquivos", ("e.ofx", _OFX, "application/octet-stream"))]
    r = client.post("/fiscal/laudo?formato=docx", files=files, data={"empresa_cnpj": "12345678000190"})
    assert r.status_code == 400


def test_laudo_xlsx_so_ofx():
    files = [("arquivos", ("e.ofx", _OFX, "application/octet-stream"))]
    r = client.post("/fiscal/laudo?formato=xlsx", files=files, data={"empresa_cnpj": "12345678000190"})
    assert r.status_code == 200, r.text
    assert r.headers["content-type"] == _XLSX
    assert r.content[:2] == b"PK"


def test_laudo_xlsx_com_xml_inclui_abas_fiscais():
    """OFX + NF-e -> mesmo documento ganha as abas 12 e 13 (mesma engine)."""
    import openpyxl
    files = [
        ("arquivos", ("e.ofx", _OFX, "application/octet-stream")),
        ("arquivos", ("nota.xml", _nfe_xml(), "application/xml")),
    ]
    r = client.post("/fiscal/laudo?formato=xlsx", files=files, data={"empresa_cnpj": "12345678000190"})
    assert r.status_code == 200, r.text
    wb = openpyxl.load_workbook(io.BytesIO(r.content))
    assert "12. Documentos Fiscais" in wb.sheetnames
    assert "13. Conformidade Fiscal" in wb.sheetnames


def test_laudo_html():
    files = [("arquivos", ("e.ofx", _OFX, "application/octet-stream"))]
    r = client.post("/fiscal/laudo?formato=html", files=files, data={"empresa_cnpj": "12345678000190"})
    assert r.status_code == 200, r.text
    assert r.headers["content-type"].startswith("text/html")
    assert "<!DOCTYPE html>" in r.text and "capa" in r.text
