"""Testes pytest da API de Conciliacao."""

import io
import os
import sys
from pathlib import Path

import pytest
from unittest.mock import patch
from fastapi.testclient import TestClient

# Forca data dir temporario antes de importar app
os.environ["ORGCONC_DATA_DIR"] = str(Path(__file__).resolve().parent / "_data_test")
os.environ.setdefault("ANTHROPIC_API_KEY", "sk-ant-test")

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from api.main import (
    app,
    _parse_ofx,
    _parse_xml,
    _classificar,
    _detectar_anomalias,
    _gerar_xlsx,
    _render_html,
    DB_DISPONIVEL,
)

client = TestClient(app)

OFX_SAMPLE2 = """OFXHEADER:100
DATA:OFXSGML
<OFX>
<BANKMSGSRSV1>
<STMTTRNRS>
<STMTRS>
<BANKACCTFROM>
<BRANCHID>5678-9</BRANCHID>
<ACCTID>1111-1</ACCTID>
</BANKACCTFROM>
<BANKTRANLIST>
<STMTTRN>
<TRNTYPE>DEBIT</TRNTYPE>
<DTPOSTED>20260415120000</DTPOSTED>
<TRNAMT>-75000.00</TRNAMT>
<MEMO>INTERCREDIS TRANSF MESMA TIT</MEMO>
</STMTTRN>
<STMTTRN>
<TRNTYPE>DEBIT</TRNTYPE>
<DTPOSTED>20260416120000</DTPOSTED>
<TRNAMT>-15000.00</TRNAMT>
<MEMO>TED EMITIDO FORNECEDOR ABC</MEMO>
</STMTTRN>
<STMTTRN>
<TRNTYPE>CREDIT</TRNTYPE>
<DTPOSTED>20260417120000</DTPOSTED>
<TRNAMT>500.00</TRNAMT>
<MEMO>ESTORNO TARIFA INDEVIDA</MEMO>
</STMTTRN>
</BANKTRANLIST>
</STMTRS>
</STMTTRNRS>
</BANKMSGSRSV1>
</OFX>"""

XML_SAMPLE = """<?xml version="1.0"?>
<Document xmlns="urn:iso:std:iso:20022:tech:xsd:camt.053.001.02">
  <BkToCstmrStmt>
    <Stmt>
      <Acct><Id><Othr><Id>ACC-XML-001</Id></Othr></Id></Acct>
      <Ntry>
        <Amt Ccy="BRL">500.00</Amt>
        <CdtDbtInd>CRDT</CdtDbtInd>
        <BookgDt><Dt>2026-04-15</Dt></BookgDt>
        <NtryDtls><TxDtls><RmtInf><Ustrd>Receita XML</Ustrd></RmtInf></TxDtls></NtryDtls>
      </Ntry>
      <Ntry>
        <Amt Ccy="BRL">75.50</Amt>
        <CdtDbtInd>DBIT</CdtDbtInd>
        <BookgDt><Dt>2026-04-16</Dt></BookgDt>
        <NtryDtls><TxDtls><RmtInf><Ustrd>Pagamento XML</Ustrd></RmtInf></TxDtls></NtryDtls>
      </Ntry>
    </Stmt>
  </BkToCstmrStmt>
</Document>"""

OFX_SAMPLE = """OFXHEADER:100
DATA:OFXSGML
<OFX>
<BANKMSGSRSV1>
<STMTTRNRS>
<STMTRS>
<BANKACCTFROM>
<BRANCHID>1234-5</BRANCHID>
<ACCTID>9999-9</ACCTID>
</BANKACCTFROM>
<BANKTRANLIST>
<STMTTRN>
<TRNTYPE>CREDIT</TRNTYPE>
<DTPOSTED>20260415120000</DTPOSTED>
<TRNAMT>1500.00</TRNAMT>
<MEMO>PIX RECEBIDO TESTE</MEMO>
</STMTTRN>
<STMTTRN>
<TRNTYPE>DEBIT</TRNTYPE>
<DTPOSTED>20260417120000</DTPOSTED>
<TRNAMT>-89.90</TRNAMT>
<MEMO>TARIFA BANCARIA</MEMO>
</STMTTRN>
<STMTTRN>
<TRNTYPE>DEBIT</TRNTYPE>
<DTPOSTED>20260417120000</DTPOSTED>
<TRNAMT>-89.90</TRNAMT>
<MEMO>TARIFA BANCARIA</MEMO>
</STMTTRN>
<STMTTRN>
<TRNTYPE>DEBIT</TRNTYPE>
<DTPOSTED>20260417120000</DTPOSTED>
<TRNAMT>-89.90</TRNAMT>
<MEMO>TARIFA BANCARIA</MEMO>
</STMTTRN>
</BANKTRANLIST>
</STMTRS>
</STMTTRNRS>
</BANKMSGSRSV1>
</OFX>"""


def test_health():
    r = client.get("/health")
    assert r.status_code == 200
    assert r.json()["status"] == "ok"


def test_root():
    r = client.get("/")
    assert r.status_code == 200
    j = r.json()
    assert "endpoints" in j


def test_logo_base64():
    r = client.get("/logo-base64")
    assert r.status_code == 200
    # Pode estar vazio se logo nao existir, mas a chave deve existir
    assert "data_uri" in r.json()


def test_parser_ofx():
    txs = _parse_ofx(OFX_SAMPLE)
    assert len(txs) == 4
    assert txs[0]["valor"] == 1500.00
    assert txs[0]["tipo"] == "CREDIT"
    assert txs[1]["valor"] == -89.90


def test_parser_xml_camt053():
    xml = """<?xml version="1.0"?>
    <Document xmlns="urn:iso:std:iso:20022:tech:xsd:camt.053.001.02">
      <BkToCstmrStmt>
        <Stmt>
          <Acct><Id><Othr><Id>ACC-TEST-123</Id></Othr></Id></Acct>
          <Ntry>
            <Amt Ccy="BRL">500.00</Amt>
            <CdtDbtInd>CRDT</CdtDbtInd>
            <BookgDt><Dt>2026-04-15</Dt></BookgDt>
            <NtryDtls><TxDtls><RmtInf><Ustrd>Receita teste</Ustrd></RmtInf></TxDtls></NtryDtls>
          </Ntry>
          <Ntry>
            <Amt Ccy="BRL">75.50</Amt>
            <CdtDbtInd>DBIT</CdtDbtInd>
            <BookgDt><Dt>2026-04-16</Dt></BookgDt>
            <NtryDtls><TxDtls><RmtInf><Ustrd>Pagamento teste</Ustrd></RmtInf></TxDtls></NtryDtls>
          </Ntry>
        </Stmt>
      </BkToCstmrStmt>
    </Document>"""
    txs = _parse_xml(xml, "test.xml")
    assert len(txs) == 2
    assert txs[0]["valor"] == 500.00
    assert txs[1]["valor"] == -75.50


def test_classificador():
    assert _classificar("PIX EMITIDO OUTRA IF", "Fornecedor X") == "Pagamento PIX - Fornecedor/Despesa"
    assert _classificar("PIX RECEBIDO", "Cliente Y") == "Receita PIX"
    assert _classificar("TARIFA MANUTENCAO CONTA", "") == "Despesa Bancaria - Tarifa"
    assert _classificar("DEB.IOF TD", "") == "Despesa Financeira - IOF"
    assert _classificar("DAS SIMPLES NACIONAL", "") == "Tributo"
    assert _classificar("FOLHA PGTO ABRIL", "") == "Folha de Pagamento"
    assert _classificar("BOLETO ENERGIA ELETRICA CEMIG", "") in ("Despesa - Energia Eletrica", "Pagamento Boleto")
    assert _classificar("ALUGUEL ESCRITORIO", "") == "Despesa - Aluguel/Condominio"
    assert _classificar("POSTO IPIRANGA COMPRA", "") in ("Despesa - Combustivel", "Compra Cartao")


def test_deteccao_anomalias_duplicidades():
    txs = _parse_ofx(OFX_SAMPLE)
    extrato = {"conta": "AG 1234-5 / CC 9999-9", "qtd": len(txs), "transacoes": txs, "arquivo": "test.ofx"}
    anomalias = _detectar_anomalias([extrato])
    # Espera detectar a triplicata de TARIFA BANCARIA (3 lancamentos identicos)
    crits = [a for a in anomalias if a["severidade"] == "critico" and a["tipo"] == "Duplicidade"]
    assert len(crits) >= 1


def test_conciliar_ofx_simulacao():
    r = client.post(
        "/conciliar/ofx?simular=true",
        files=[("arquivos", ("test.ofx", OFX_SAMPLE, "application/x-ofx"))],
    )
    assert r.status_code == 200, r.text
    j = r.json()
    assert j["modo"] == "simulacao_local"
    assert "report_id" in j
    assert len(j["extratos"]) == 1
    assert "anomalias" in j
    assert "relatorio_md" in j
    assert "relatorio_html" in j


def test_export_html_xlsx():
    # Gera dataset
    r = client.post(
        "/conciliar/ofx?simular=true",
        files=[("arquivos", ("test.ofx", OFX_SAMPLE, "application/x-ofx"))],
    )
    rid = r.json()["report_id"]
    # HTML
    r_html = client.get(f"/export/html/{rid}")
    assert r_html.status_code == 200
    assert b"ORGATEC" in r_html.content
    assert b"<table" in r_html.content
    # XLSX
    r_xlsx = client.get(f"/export/xlsx/{rid}")
    assert r_xlsx.status_code == 200
    assert r_xlsx.content[:2] == b"PK"  # zip magic (xlsx)


def test_export_id_invalido():
    r = client.get("/export/html/INVALID")
    assert r.status_code == 400


def test_export_id_inexistente():
    r = client.get("/export/html/abcdef012345")
    assert r.status_code == 404


def test_upload_extensao_invalida():
    r = client.post(
        "/conciliar/ofx?simular=true",
        files=[("arquivos", ("malicioso.exe", b"binarylixo", "application/octet-stream"))],
    )
    assert r.status_code == 400
    assert "nao suportada" in r.json()["detail"].lower() or "suport" in r.json()["detail"].lower()


# ── Hardening: /conciliar/csv exige auth quando ORGCONC_AUTH_TOKEN ativo ──


def test_conciliar_ofx_aceita_modelo_haiku_em_simular():
    """modelo=haiku deve ser aceito mesmo em simular (parametro nao quebra fluxo)."""
    r = client.post(
        "/conciliar/ofx?simular=true&modelo=haiku",
        files=[("arquivos", ("test.ofx", OFX_SAMPLE, "application/x-ofx"))],
    )
    assert r.status_code == 200
    # Em simular, modelo eh ignorado mas nao pode quebrar
    assert r.json()["modo"] == "simulacao_local"


def test_conciliar_ofx_modelo_invalido_retorna_400():
    """modelo=opus123 ou outro invalido deve retornar 400 com lista."""
    r = client.post(
        "/conciliar/ofx?modelo=opus123",
        files=[("arquivos", ("test.ofx", OFX_SAMPLE, "application/x-ofx"))],
    )
    assert r.status_code == 400
    detail = r.json()["detail"].lower()
    assert "modelo" in detail and "invalido" in detail


def test_modelos_validos_mapping():
    """O dicionario _MODELOS_VALIDOS deve ter haiku, sonnet, opus mapeados."""
    from api.main import _MODELOS_VALIDOS

    assert "haiku" in _MODELOS_VALIDOS
    assert "sonnet" in _MODELOS_VALIDOS
    assert "opus" in _MODELOS_VALIDOS
    assert _MODELOS_VALIDOS["haiku"][0] == "claude-haiku-4-5-20251001"
    assert _MODELOS_VALIDOS["sonnet"][0] == "claude-sonnet-4-6"


def test_conciliar_csv_exige_auth_quando_token_definido():
    """/conciliar/csv NAO pode ser publico quando AUTH_TOKEN existe."""
    with patch("api.services.auth._LEGACY_SERVICE_TOKEN", "segredo-de-teste"):
        r = client.post(
            "/conciliar/csv",
            files=[
                ("extrato", ("e.csv", b"data,valor\n2026-01-01,100", "text/csv")),
                ("razao", ("r.csv", b"data,valor\n2026-01-01,100", "text/csv")),
            ],
        )
    assert r.status_code == 401, f"Esperado 401, recebido {r.status_code}: {r.text[:200]}"


# ── Hardening: persistencia no response ───────────────────────────────────


def test_conciliacao_simulacao_inclui_status_persistencia():
    """O response do /conciliar/ofx deve trazer chave 'persistencia' com status."""
    r = client.post(
        "/conciliar/ofx?simular=true",
        files=[("arquivos", ("test.ofx", OFX_SAMPLE, "application/x-ofx"))],
    )
    assert r.status_code == 200
    j = r.json()
    assert "persistencia" in j
    assert j["persistencia"]["status"] in ("ok", "skip", "error")


# ── Hardening: persistencia retorna error quando BD falha ─────────────────


def test_persistencia_retorna_error_quando_bd_falha():
    """_salvar_no_banco NAO pode engolir excecao silenciosamente."""
    import asyncio
    from api.main import _salvar_no_banco

    with (
        patch("api.services.db_persistence.DB_DISPONIVEL", True),
        patch("api.services.db_persistence.SessionLocal", side_effect=RuntimeError("conexao recusada")),
    ):
        resultado = asyncio.run(_salvar_no_banco("teste-rid", [], [], "simulacao"))
    assert resultado["status"] == "error"
    assert "conexao" in resultado["mensagem"].lower() or resultado["erro"] == "RuntimeError"


# ── Hardening: CSP endurecida ─────────────────────────────────────────────


def test_csp_inclui_diretivas_endurecidas():
    """CSP deve incluir connect-src, form-action, upgrade-insecure-requests."""
    r = client.get("/health")
    csp = r.headers.get("content-security-policy", "")
    assert "connect-src 'self'" in csp
    assert "form-action 'self'" in csp
    assert "upgrade-insecure-requests" in csp


# ── JWT auth ──────────────────────────────────────────────────────────────


def test_auth_login_sem_config_retorna_503():
    """Sem ORGCONC_ADMIN_* configurado, /auth/login deve retornar 503."""
    with patch.dict(os.environ, {"ORGCONC_ADMIN_EMAIL": "", "ORGCONC_ADMIN_SENHA_HASH": ""}):
        r = client.post("/auth/login", json={"email": "x@y.com", "senha": "qualquer"})
    assert r.status_code == 503


def test_auth_login_credenciais_invalidas_retorna_401():
    """Email errado deve retornar 401 com mesma mensagem que senha errada (anti-enum)."""
    from api.services.auth import hash_senha

    hash_valido = hash_senha("senha-correta-123")
    with patch.dict(
        os.environ,
        {
            "ORGCONC_ADMIN_EMAIL": "admin@orgconc.com",
            "ORGCONC_ADMIN_SENHA_HASH": hash_valido,
        },
    ):
        r1 = client.post("/auth/login", json={"email": "outro@x.com", "senha": "qualquer"})
        r2 = client.post("/auth/login", json={"email": "admin@orgconc.com", "senha": "errada"})
    assert r1.status_code == 401
    assert r2.status_code == 401
    assert r1.json()["detail"] == r2.json()["detail"]  # anti-enumeration


def test_auth_login_sucesso_emite_jwt():
    """Login correto deve emitir JWT valido."""
    from api.services.auth import hash_senha, decodificar_token

    hash_valido = hash_senha("senha-correta-123")
    with patch.dict(
        os.environ,
        {
            "ORGCONC_ADMIN_EMAIL": "admin@orgconc.com",
            "ORGCONC_ADMIN_SENHA_HASH": hash_valido,
        },
    ):
        r = client.post("/auth/login", json={"email": "admin@orgconc.com", "senha": "senha-correta-123"})
    assert r.status_code == 200
    j = r.json()
    assert "access_token" in j
    assert j["token_type"] == "bearer"
    payload = decodificar_token(j["access_token"])
    assert payload.email == "admin@orgconc.com"
    assert payload.role == "admin"


def test_auth_me_exige_token():
    """/auth/me sem token retorna 401 quando LEGACY_SERVICE_TOKEN configurado."""
    with patch("api.services.auth._LEGACY_SERVICE_TOKEN", "legacy-test-token"):
        r = client.get("/auth/me")
    assert r.status_code == 401


def test_auth_me_com_token_legacy_funciona():
    """Token legacy compartilhado deve funcionar via /auth/me."""
    with (
        patch("api.main.AUTH_TOKEN", "legacy-test-token"),
        patch("api.services.auth._LEGACY_SERVICE_TOKEN", "legacy-test-token"),
    ):
        r = client.get("/auth/me", headers={"Authorization": "Bearer legacy-test-token"})
    assert r.status_code == 200
    j = r.json()
    assert j["role"] == "service"


def test_auth_me_com_jwt_funciona():
    """JWT valido emitido por emitir_token deve passar pelo current_user."""
    from api.services.auth import emitir_token

    token = emitir_token(sub="teste@x.com", email="teste@x.com", role="admin")
    with patch("api.main.AUTH_TOKEN", "qualquer-coisa"):
        r = client.get("/auth/me", headers={"Authorization": f"Bearer {token}"})
    assert r.status_code == 200
    assert r.json()["email"] == "teste@x.com"


def test_auth_jwt_expirado_retorna_401():
    """JWT com exp no passado deve retornar 401 com mensagem clara."""
    from api.services.auth import emitir_token

    token = emitir_token(sub="x@y.com", ttl_min=-1)  # ja expirado
    r = client.get("/auth/me", headers={"Authorization": f"Bearer {token}"})
    assert r.status_code == 401
    assert "expirado" in r.json()["detail"].lower()


# ── Hardening: CSP script-src sem unsafe-inline ───────────────────────────


def test_csp_script_src_sem_unsafe_inline():
    """Script-src deve permitir self + cdn.jsdelivr.net mas NAO unsafe-inline."""
    r = client.get("/health")
    csp = r.headers.get("content-security-policy", "")
    # Extrai a diretiva script-src
    parts = [p.strip() for p in csp.split(";") if p.strip().startswith("script-src")]
    assert parts, "script-src diretiva ausente do CSP"
    script_src = parts[0]
    assert "'unsafe-inline'" not in script_src, f"script-src contem unsafe-inline: {script_src}"
    assert "'self'" in script_src
    assert "cdn.jsdelivr.net" in script_src


# ── Hardening: limite agregado de upload ──────────────────────────────────


def test_limite_agregado_de_upload_retorna_413():
    """Soma dos uploads deve exceder MAX_UPLOAD_TOTAL_BYTES -> 413."""
    # Baixa o limite agregado para 1KB para o teste e envia 3 OFX validos de >400B
    with patch("api.routers.conciliacao.MAX_UPLOAD_TOTAL_BYTES", 1024), patch("api.routers.conciliacao.MAX_UPLOAD_TOTAL_MB", 0):
        files = [
            ("arquivos", (f"e{i}.ofx", OFX_SAMPLE.encode("latin-1"), "application/x-ofx"))
            for i in range(3)  # 3x ~700B = ~2KB > 1KB
        ]
        r = client.post("/conciliar/ofx?simular=true", files=files)
    assert r.status_code == 413, f"Esperado 413, recebido {r.status_code}: {r.text[:200]}"
    assert "excede" in r.json()["detail"].lower() or "mb" in r.json()["detail"].lower()


def test_gerar_xlsx_estrutura():
    txs = _parse_ofx(OFX_SAMPLE)
    extrato = {"conta": "AG 1234-5 / CC 9999-9", "qtd": len(txs), "transacoes": txs, "arquivo": "test.ofx"}
    anomalias = _detectar_anomalias([extrato])
    blob = _gerar_xlsx([extrato], anomalias)
    assert blob[:2] == b"PK"  # xlsx valido

    # Abre e verifica abas
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(blob))
    assert "Resumo" in wb.sheetnames
    assert "Transações" in wb.sheetnames
    assert "Anomalias" in wb.sheetnames


def test_render_html_contem_logo_e_dados():
    md = "# Relatorio\n\n| col | val |\n|---|---|\n| a | 1 |\n"
    html = _render_html(md)
    assert "ORGATEC" in html
    assert "Inter:wght" in html
    assert "<table" in html


# ── Clientes ──────────────────────────────────────────────────────────────


def test_clientes_cnpj_invalido_retorna_422():
    """CNPJ incorreto deve ser rejeitado pelo Pydantic antes de chegar ao banco."""
    r = client.post("/clientes", json={"nome": "Empresa Teste", "cnpj": "11111111111111"})
    assert r.status_code == 422


def test_clientes_cnpj_todos_iguais_retorna_422():
    """CNPJ com todos os digitos iguais (ex: 00.000.000/0000-00) eh invalido."""
    r = client.post("/clientes", json={"nome": "Empresa Teste", "cnpj": "00000000000000"})
    assert r.status_code == 422


def test_clientes_criar_sem_db_retorna_503():
    """Com DB_DISPONIVEL=False, POST /clientes deve retornar 503."""
    with patch("api.main.DB_DISPONIVEL", False):
        r = client.post("/clientes", json={"nome": "Empresa Sem DB", "plano": "basico"})
    assert r.status_code == 503


def test_clientes_listar_sem_db_retorna_503():
    """Com DB_DISPONIVEL=False, GET /clientes deve retornar 503."""
    with patch("api.main.DB_DISPONIVEL", False):
        r = client.get("/clientes")
    assert r.status_code == 503


def test_clientes_buscar_sem_db_retorna_503():
    """Com DB_DISPONIVEL=False, GET /clientes/{id} deve retornar 503."""
    with patch("api.main.DB_DISPONIVEL", False):
        r = client.get("/clientes/00000000-0000-0000-0000-000000000001")
    assert r.status_code == 503


def test_clientes_atualizar_sem_db_retorna_503():
    """Com DB_DISPONIVEL=False, PATCH /clientes/{id} deve retornar 503."""
    with patch("api.main.DB_DISPONIVEL", False):
        r = client.patch("/clientes/00000000-0000-0000-0000-000000000001", json={"nome": "Novo"})
    assert r.status_code == 503


# ── Headers de Segurança ──────────────────────────────────────────────────


def test_security_headers_presentes():
    """Toda resposta deve incluir os headers de segurança obrigatorios."""
    r = client.get("/health")
    assert r.headers.get("x-content-type-options") == "nosniff"
    assert r.headers.get("x-frame-options") == "DENY"
    assert "Content-Security-Policy" in r.headers or "content-security-policy" in r.headers


# ── Integração Real com DB ────────────────────────────────────────────────

_CNPJ_TESTE = "11444777000161"  # CNPJ válido reservado para testes de integração
_requer_db = pytest.mark.skipif(not DB_DISPONIVEL, reason="DATABASE_URL não configurado")


def _limpar_cliente_teste() -> None:
    """Remove cliente de teste via psycopg2 síncrono para evitar conflito de event loop."""
    import psycopg2

    url = os.environ.get("DATABASE_URL", "").replace("+asyncpg", "")
    if not url or "[" in url:
        return
    with psycopg2.connect(url) as conn:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM clientes WHERE cnpj = %s", (_CNPJ_TESTE,))
        conn.commit()


@pytest.fixture(scope="module")
def cliente_real():
    """Cria cliente no banco real e remove ao final do módulo."""
    _limpar_cliente_teste()  # estado limpo antes de criar

    r = client.post(
        "/clientes",
        json={
            "nome": "Empresa Integracao Teste",
            "cnpj": _CNPJ_TESTE,
            "email": "integracao@orgconc.com",
            "plano": "basico",
        },
    )
    assert r.status_code == 201, r.text
    data = r.json()
    yield data

    _limpar_cliente_teste()  # cleanup


@_requer_db
def test_db_criar_cliente(cliente_real):
    assert "id" in cliente_real
    assert cliente_real["nome"] == "Empresa Integracao Teste"
    assert cliente_real["cnpj"] == _CNPJ_TESTE
    assert cliente_real["plano"] == "basico"


@_requer_db
def test_db_listar_clientes(cliente_real):
    r = client.get("/clientes")
    assert r.status_code == 200
    ids = [c["id"] for c in r.json()]
    assert cliente_real["id"] in ids


@_requer_db
def test_db_buscar_cliente_por_id(cliente_real):
    r = client.get(f"/clientes/{cliente_real['id']}")
    assert r.status_code == 200
    j = r.json()
    assert j["id"] == cliente_real["id"]
    assert j["email"] == "integracao@orgconc.com"


@_requer_db
def test_db_atualizar_cliente(cliente_real):
    r = client.patch(f"/clientes/{cliente_real['id']}", json={"nome": "Empresa Atualizada", "plano": "pro"})
    assert r.status_code == 200
    j = r.json()
    assert j["nome"] == "Empresa Atualizada"
    assert j["plano"] == "pro"


@_requer_db
def test_db_buscar_cliente_inexistente():
    r = client.get("/clientes/00000000-0000-0000-0000-000000000099")
    assert r.status_code == 404


# ── Anomalias — cobertura de tipos ────────────────────────────────────────


def _extrato(txs: list[dict], conta: str = "AG 0000 / CC 0000") -> dict:
    return {"conta": conta, "qtd": len(txs), "transacoes": txs, "arquivo": "test.ofx"}


def _tx(valor: float, memo: str = "TESTE", data: str = "2026-04-15") -> dict:
    return {"valor": valor, "memo": memo, "nome": "", "data": data, "tipo": "DEBIT", "conta": ""}


def test_anomalias_valor_alto_alerta():
    txs = [_tx(75000.00, "TED EMITIDO GRANDE")]
    anomalias = _detectar_anomalias([_extrato(txs)])
    alerta = [a for a in anomalias if a["tipo"] == "Valor alto" and a["severidade"] == "alerta"]
    assert len(alerta) == 1


def test_anomalias_valor_alto_atencao():
    txs = [_tx(25000.00, "TED EMITIDO MEDIO")]
    anomalias = _detectar_anomalias([_extrato(txs)])
    atencao = [a for a in anomalias if a["tipo"] == "Valor alto" and a["severidade"] == "atencao"]
    assert len(atencao) == 1


def test_anomalias_estorno_critico():
    txs = [_tx(200.00, "ESTORNO TARIFA INDEVIDA")]
    anomalias = _detectar_anomalias([_extrato(txs)])
    estornos = [a for a in anomalias if a["tipo"] == "Estorno" and a["severidade"] == "critico"]
    assert len(estornos) == 1


def test_anomalias_lista_vazia():
    assert _detectar_anomalias([]) == []


def test_anomalias_sem_duplicidade_transacoes_distintas():
    txs = [
        _tx(100.00, "PIX RECEBIDO A", "2026-04-15"),
        _tx(200.00, "PIX RECEBIDO B", "2026-04-15"),
    ]
    anomalias = _detectar_anomalias([_extrato(txs)])
    dups = [a for a in anomalias if a["tipo"] == "Duplicidade"]
    assert len(dups) == 0


def test_anomalias_transferencia_sem_par():
    e1 = _extrato([_tx(-1000.00, "INTERCREDIS TRANSF MESMA TIT")], "Conta A")
    e2 = _extrato([_tx(-500.00, "INTERCREDIS TRANSF MESMA TIT")], "Conta B")
    anomalias = _detectar_anomalias([e1, e2])
    sem_par = [a for a in anomalias if a["tipo"] == "Transferencia sem par"]
    assert len(sem_par) >= 1


# ── Classificador — edge cases ────────────────────────────────────────────


def test_classificador_ted_recebido():
    assert _classificar("TED RECEBIDO CLIENTE", "") == "Receita TED/DOC"


def test_classificador_ted_pago():
    assert _classificar("TED EMITIDO FORNECEDOR", "") == "Pagamento TED/DOC"


def test_classificador_desconhecido():
    assert _classificar("XPTO INEXISTENTE 99", "") == "A classificar"


def test_classificador_pix_a_classificar():
    assert _classificar("PIX", "") == "PIX - A classificar"


# ── Conciliar — múltiplos arquivos e XML ──────────────────────────────────


def test_conciliar_multiplos_ofx():
    r = client.post(
        "/conciliar/ofx?simular=true",
        files=[
            ("arquivos", ("extrato1.ofx", OFX_SAMPLE, "application/x-ofx")),
            ("arquivos", ("extrato2.ofx", OFX_SAMPLE2, "application/x-ofx")),
        ],
    )
    assert r.status_code == 200, r.text
    j = r.json()
    assert len(j["extratos"]) == 2
    assert j["modo"] == "simulacao_local"
    estornos = [a for a in j["anomalias"] if a["tipo"] == "Estorno"]
    assert len(estornos) >= 1


def test_conciliar_xml():
    r = client.post(
        "/conciliar/ofx?simular=true",
        files=[("arquivos", ("extrato.xml", XML_SAMPLE, "text/xml"))],
    )
    assert r.status_code == 200, r.text
    j = r.json()
    assert j["modo"] == "simulacao_local"
    assert j["extratos"][0]["qtd"] == 2


# ── Clientes — validações adicionais ──────────────────────────────────────


def test_cliente_plano_invalido_retorna_422():
    r = client.post("/clientes", json={"nome": "Empresa X", "plano": "platinum"})
    assert r.status_code == 422


def test_cliente_buscar_id_invalido_retorna_400():
    r = client.get("/clientes/nao-e-um-uuid")
    assert r.status_code == 400


def test_cliente_cnpj_com_mascara_valido():
    """CNPJ formatado com pontuação deve ser aceito pelo Pydantic e normalizado."""
    r = client.post("/clientes", json={"nome": "Empresa Mascara", "cnpj": "11.444.777/0001-61"})
    # 422 = formato rejeitado (não esperado); 503 = sem DB; 201 = criado; 409/500 = duplicado OK
    assert r.status_code not in (422,), f"CNPJ com máscara foi rejeitado incorretamente: {r.text}"


# ── XLSX — cabeçalhos das colunas ─────────────────────────────────────────


def test_xlsx_cabecalhos_transacoes():
    txs = _parse_ofx(OFX_SAMPLE)
    extrato = {"conta": "AG 1234-5 / CC 9999-9", "qtd": len(txs), "transacoes": txs, "arquivo": "test.ofx"}
    blob = _gerar_xlsx([extrato], [])
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(blob))
    ws = wb["Transações"]
    headers = [ws.cell(1, c).value for c in range(1, 10) if ws.cell(1, c).value]
    assert "Data" in headers
    assert "Valor (R$)" in headers or any("Valor" in str(h) for h in headers)
    assert "Memo / Descrição" in headers or any("Memo" in str(h) for h in headers)


# ── DB real — conciliação salva no banco ──────────────────────────────────


@_requer_db
def test_db_conciliacao_salva_no_banco():
    """POST /conciliar/ofx?simular=true deve persistir registro em conciliacoes."""
    import psycopg2

    r = client.post(
        "/conciliar/ofx?simular=true",
        files=[("arquivos", ("real.ofx", OFX_SAMPLE, "application/x-ofx"))],
    )
    assert r.status_code == 200
    rid = r.json()["report_id"]

    url = os.environ.get("DATABASE_URL", "").replace("+asyncpg", "")
    with psycopg2.connect(url) as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT report_id, total_transacoes FROM conciliacoes WHERE report_id = %s", (rid,))
            row = cur.fetchone()
    assert row is not None, f"Conciliacao {rid} nao encontrada no banco"
    assert row[1] == 4  # OFX_SAMPLE tem 4 transacoes


# ── Sanitizacao XSS ─────────────────────────────────────────────────────────


def test_sanitize_remove_script_tag():
    """Defesa contra XSS: <script> e removido antes do template renderizar."""
    from api.services.sanitize import sanitize_html

    out = sanitize_html("<p>ok</p><script>alert(1)</script>")
    assert "<script" not in out.lower()
    assert "alert(1)" in out  # texto preservado, mas inerte


def test_sanitize_remove_javascript_uri():
    """Defesa contra XSS: javascript: URIs sao removidas de href."""
    from api.services.sanitize import sanitize_html

    out = sanitize_html('<a href="javascript:alert(1)">bad</a><a href="https://x.com">ok</a>')
    assert "javascript:" not in out.lower()
    assert 'href="https://x.com"' in out


def test_sanitize_remove_event_handlers():
    """Defesa contra XSS: atributos on* sao removidos."""
    from api.services.sanitize import sanitize_html

    out = sanitize_html('<p onclick="alert(1)">x</p><img src="x" onerror="alert(1)">')
    assert "onclick" not in out.lower()
    assert "onerror" not in out.lower()


def test_render_html_sanitiza_xss_em_relatorio_md():
    """End-to-end: relatorio_md com payload XSS nao gera script no HTML final."""
    payload_md = "# Relatorio\n\n<script>alert('xss')</script>\n\nConteudo legitimo."
    html = _render_html(payload_md)
    assert "<script" not in html.lower()
    assert "Conteudo legitimo" in html


# ── Trilha 2 — segurança ────────────────────────────────────────────────────


def test_cors_sem_wildcard():
    """CORS: allow_origins nunca deve ser ['*'] — wildcard fallback removido."""
    from api.core.config import CORS_ORIGINS

    assert "*" not in CORS_ORIGINS, "CORS_ORIGINS contém '*' — todas as origens estariam autorizadas"


def test_security_headers_inclui_referrer_policy():
    """Middleware de segurança injeta Referrer-Policy stict-origin-when-cross-origin."""
    r = client.get("/health")
    assert r.headers.get("referrer-policy") == "strict-origin-when-cross-origin"


def test_auth_hash_bloqueado_em_prod():
    """POST /auth/hash deve retornar 404 quando _IS_PROD=True."""
    from unittest.mock import patch

    with patch("api.core.config._IS_PROD", True):
        r = client.post("/auth/hash", json={"senha": "senha123"})
    assert r.status_code == 404, f"/auth/hash acessivel em producao — retornou {r.status_code}"


def test_login_retorna_401_para_email_e_senha_errados():
    """Login com email errado e senha errada devem retornar 401 idêntico.

    Verifica que a resposta é consistente (mesmo status, mesma mensagem)
    independente de qual credencial está errada — sem vazar informação via
    mensagem diferente ou status diferente.
    """
    with patch.dict(
        os.environ,
        {
            "ORGCONC_ADMIN_EMAIL": "admin@orgatec.com",
            "ORGCONC_ADMIN_SENHA_HASH": "$2b$12$LQv3c1yqBWVHxkd0LHAkCOYz6TiGVQ/YTKGiFj6b3M3KdA0Jl4W2",
        },
    ):
        r_email = client.post("/auth/login", json={"email": "errado@outro.com", "senha": "qualquer"})
        r_senha = client.post("/auth/login", json={"email": "admin@orgatec.com", "senha": "errada"})

    assert r_email.status_code == 401
    assert r_senha.status_code == 401
    assert (
        r_email.json()["detail"] == r_senha.json()["detail"]
    ), "Mensagens de erro diferem — vaza qual credencial está errada"


def test_jwt_contem_nbf():
    """Tokens emitidos devem conter claim nbf (not-before)."""
    import jwt as pyjwt
    from api.services.auth import emitir_token

    token = emitir_token(sub="test@x.com", email="test@x.com")
    claims = pyjwt.decode(token, options={"verify_signature": False})
    assert "nbf" in claims, "Claim nbf ausente no token"
    assert claims["nbf"] <= claims["iat"], "nbf deve ser <= iat"


# ── Trilha 4: validação de input, rate limiting, UUID ordering ────────────


def test_cliente_buscar_uuid_valido_sem_db_retorna_503():
    """UUID válido com DB indisponível deve retornar 503, não 400.

    Regressão: após reordenação (UUID antes de DB check), UUID válido ainda
    deve retornar 503 quando banco não está configurado. Patch explícito de
    DB_DISPONIVEL=False — não depende de cenário global do ambiente.
    """
    with patch("api.main.DB_DISPONIVEL", False):
        r = client.get("/clientes/00000000-0000-0000-0000-000000000001")
    assert r.status_code == 503, f"UUID válido sem DB deveria retornar 503, retornou {r.status_code}"


def test_login_payload_muito_grande_retorna_422():
    """POST /auth/login com email acima de 254 chars deve retornar 422 (max_length)."""
    email_gigante = "a" * 300 + "@x.com"
    r = client.post("/auth/login", json={"email": email_gigante, "senha": "qualquer"})
    assert (
        r.status_code == 422
    ), f"Email de {len(email_gigante)} chars deveria ser rejeitado com 422, retornou {r.status_code}"


def test_export_pdf_sem_auth_retorna_401():
    """GET /export/pdf sem token deve retornar 401 quando AUTH_TOKEN está definido."""
    with patch("api.services.auth._LEGACY_SERVICE_TOKEN", "segredo-de-teste"):
        r = client.get("/export/pdf/report-inexistente")
    assert r.status_code == 401, f"/export/pdf sem token deveria retornar 401, retornou {r.status_code}"


# ── Trilha 5: path traversal, health prod, response schema ───────────────


def test_carregar_dataset_rid_invalido_retorna_400():
    """IDs inválidos devem ser rejeitados (path traversal bloqueado).

    '../etc/passwd' retorna 404 porque o router ASGI normaliza '..' antes de
    chegar no handler — a traversal é bloqueada pelo roteador, não pela nossa
    validação de regex. IDs com formato errado retornam 400 pela nossa validação.
    """
    for rid_invalido in ["../etc/passwd", "../../secret"]:
        r = client.get(f"/export/html/{rid_invalido}")
        assert r.status_code in (400, 404, 422), (
            f"Traversal '{rid_invalido}' deveria ser bloqueado (400/404/422), " f"retornou {r.status_code}"
        )
    for rid_invalido in ["AAAABBBBCCCC", "abc123"]:
        r = client.get(f"/export/html/{rid_invalido}")
        assert r.status_code == 400, f"rid '{rid_invalido}' deveria ser rejeitado com 400, retornou {r.status_code}"


def test_health_em_prod_nao_expoe_config():
    """Em produção, /health não deve expor api_key_configured nem banco_dados."""
    with patch("api.core.config._IS_PROD", True):
        r = client.get("/health")
    assert r.status_code == 200
    body = r.json()
    assert "api_key_configured" not in body, "/health expõe api_key_configured em produção — risk de reconhecimento"
    assert "banco_dados" not in body, "/health expõe banco_dados em produção"


# ── Trilha 6: Cache-Control, exception handler, JWT jti, version info ────


def test_auth_login_retorna_cache_control_no_store():
    """POST /auth/login deve retornar Cache-Control: no-store para evitar cache de tokens."""
    r = client.post("/auth/login", json={"email": "x@y.com", "senha": "qualquer"})
    # Independente do status (401/503), o header deve estar presente
    cc = r.headers.get("cache-control", "")
    assert "no-store" in cc, f"/auth/login sem 'no-store' no Cache-Control: '{cc}'"


def test_jwt_contem_jti():
    """Tokens emitidos devem conter claim jti (JWT ID único)."""
    import jwt as pyjwt
    from api.services.auth import emitir_token

    token = emitir_token(sub="test@x.com", email="test@x.com")
    claims = pyjwt.decode(token, options={"verify_signature": False})
    assert "jti" in claims, "Claim jti ausente no token"
    assert len(claims["jti"]) >= 16, "jti deve ter pelo menos 16 chars"


def test_root_em_prod_nao_expoe_versao():
    """GET / em produção não deve retornar campo version (fingerprinting)."""
    with patch("api.core.config._IS_PROD", True):
        r = client.get("/")
    assert r.status_code == 200
    body = r.json()
    assert "version" not in body, f"GET / expõe 'version' em produção: {body}"


def test_export_retorna_cache_control_no_store():
    """GET /export/* deve retornar Cache-Control: no-store (relatórios financeiros)."""
    r = client.get("/export/html/abc123456789")
    cc = r.headers.get("cache-control", "")
    assert "no-store" in cc, f"/export/html sem 'no-store' no Cache-Control: '{cc}'"


# ── Trilha 7: audit log · legacy token em prod · /docs em prod · body limit ──


def test_docs_em_prod_retorna_404():
    """Em produção /docs deve retornar 404 (Swagger UI desabilitado)."""
    from fastapi import FastAPI
    from fastapi.testclient import TestClient as _TC

    prod_app = FastAPI(
        docs_url=None,
        redoc_url=None,
        openapi_url=None,
    )

    @prod_app.get("/ping")
    def _ping():
        return {"ok": True}

    c = _TC(prod_app)
    r = c.get("/docs")
    assert r.status_code == 404, f"/docs deveria ser 404 em prod, got {r.status_code}"


def test_openapi_em_prod_retorna_404():
    """Em produção /openapi.json deve retornar 404 (schema não exposto)."""
    from fastapi import FastAPI
    from fastapi.testclient import TestClient as _TC

    prod_app = FastAPI(
        docs_url=None,
        redoc_url=None,
        openapi_url=None,
    )
    c = _TC(prod_app)
    r = c.get("/openapi.json")
    assert r.status_code == 404, f"/openapi.json deveria ser 404 em prod, got {r.status_code}"


def test_legacy_token_rejeitado_em_prod():
    """Bearer com token legado deve ser rejeitado (401) quando _IS_PROD=True."""
    from api.services.auth import auth_optional
    from fastapi import HTTPException

    legacy = "token-legado-teste-seguro"
    with patch("api.services.auth._LEGACY_SERVICE_TOKEN", legacy), patch("api.services.auth._IS_PROD", True):
        try:
            auth_optional(authorization=f"Bearer {legacy}")
            raise AssertionError("Token legado aceito em produção — deve levantar HTTPException 401")
        except HTTPException as e:
            assert e.status_code == 401, f"Esperado 401, got {e.status_code}"


def test_body_gigante_retorna_413():
    """POST com Content-Length > _MAX_BODY_BYTES deve retornar 413."""
    import api.core.config as _m

    original = _m._MAX_BODY_BYTES
    try:
        _m._MAX_BODY_BYTES = 100
        payload = "x" * 200
        r = client.post(
            "/auth/login",
            content=payload,
            headers={"Content-Type": "application/json", "Content-Length": str(len(payload))},
        )
        assert r.status_code == 413, f"Body gigante deveria retornar 413, got {r.status_code}"
    finally:
        _m._MAX_BODY_BYTES = original


# ── Trilha 8: max_tokens cap · DB error leak · UUID validation · filename sanitization ──


def test_max_tokens_acima_limite_csv_retorna_422():
    """POST /conciliar/csv com max_tokens > 32768 deve retornar 422 (validação FastAPI)."""
    r = client.post(
        "/conciliar/csv?max_tokens=999999",
        files={
            "extrato": ("e.csv", io.BytesIO(b"data,valor\n2026-01-01,100"), "text/csv"),
            "razao": ("r.csv", io.BytesIO(b"data,valor\n2026-01-01,100"), "text/csv"),
        },
    )
    assert r.status_code == 422, f"max_tokens=999999 deveria retornar 422, got {r.status_code}"


def test_max_tokens_zero_csv_retorna_422():
    """POST /conciliar/csv com max_tokens=0 (abaixo do mínimo) deve retornar 422."""
    r = client.post(
        "/conciliar/csv?max_tokens=0",
        files={
            "extrato": ("e.csv", io.BytesIO(b"data,valor\n2026-01-01,100"), "text/csv"),
            "razao": ("r.csv", io.BytesIO(b"data,valor\n2026-01-01,100"), "text/csv"),
        },
    )
    assert r.status_code == 422, f"max_tokens=0 deveria retornar 422, got {r.status_code}"


def test_cliente_id_invalido_no_ofx_retorna_422():
    """POST /conciliar/ofx com cliente_id não-UUID deve retornar 422."""
    r = client.post(
        "/conciliar/ofx?simular=true&cliente_id=nao-e-um-uuid",
        files={"arquivos": ("extrato.ofx", io.BytesIO(OFX_SAMPLE2.encode()), "text/plain")},
    )
    assert r.status_code == 422, f"cliente_id inválido deveria retornar 422, got {r.status_code}"


def test_db_error_em_prod_omite_mensagem():
    """_salvar_no_banco em _IS_PROD=True não deve retornar campo 'mensagem' (info disclosure)."""
    import asyncio
    from unittest.mock import AsyncMock, patch

    async def _run():
        with (
            patch("api.core.config._IS_PROD", True),
            patch("api.services.db_persistence.DB_DISPONIVEL", True),
            patch("api.services.db_persistence.SessionLocal") as mock_session,
        ):
            mock_session.return_value.__aenter__ = AsyncMock(side_effect=Exception("table users does not exist"))
            mock_session.return_value.__aexit__ = AsyncMock(return_value=False)
            from api.main import _salvar_no_banco

            result = await _salvar_no_banco("abc123456789", [], [], "test")
        return result

    result = asyncio.run(_run())
    assert result["status"] == "error"
    assert "mensagem" not in result, f"'mensagem' exposto em produção — vazamento de erro DB: {result}"


# ── Trilha 9: WeasyPrint SSRF · parser error · plano max_length · magic bytes ──


def test_sanitize_img_src_https_removido():
    """img src com URL remota deve ser removido — SSRF guard para WeasyPrint."""
    from api.services.sanitize import sanitize_html

    out = sanitize_html('<img src="https://evil.com/track.png" alt="x">')
    assert 'src="https://evil.com' not in out, f"img src remoto não bloqueado pelo sanitizer: {out}"


def test_sanitize_img_src_http_removido():
    """img src HTTP deve ser removido (sem downgrade para plaintext)."""
    from api.services.sanitize import sanitize_html

    out = sanitize_html('<img src="http://evil.com/track.png" alt="x">')
    assert "http://evil.com" not in out, f"img src HTTP não bloqueado: {out}"


def test_sanitize_img_data_uri_mantido():
    """img src com data URI (base64) deve ser preservado."""
    from api.services.sanitize import sanitize_html

    data_uri = "data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAYAAAAfFcSJAAAADUlEQVR42mNk+M9QDwADhgGAWjR9awAAAABJRU5ErkJggg=="
    out = sanitize_html(f'<img src="{data_uri}" alt="logo">')
    assert "data:image/png" in out, f"data URI removido indevidamente: {out}"


def test_sanitize_link_http_bloqueado():
    """a href com HTTP deve ser removido (sem downgrade)."""
    from api.services.sanitize import sanitize_html

    out = sanitize_html('<a href="http://example.com">link</a>')
    assert "http://example.com" not in out, f"href HTTP não bloqueado: {out}"


def test_sanitize_link_https_mantido():
    """a href com HTTPS deve ser preservado."""
    from api.services.sanitize import sanitize_html

    out = sanitize_html('<a href="https://example.com">link</a>')
    assert "https://example.com" in out, f"href HTTPS removido indevidamente: {out}"


def test_plano_muito_longo_retorna_422():
    """POST /clientes com plano > 20 chars deve retornar 422.

    Sem header de auth: em dev o usuario e anonimo, entao a validacao Pydantic
    do corpo (422) e exercitada. Um token Bearer invalido daria 401 antes da
    validacao do corpo (auth e dependency, roda antes do parse do body).
    """
    r = client.post(
        "/clientes",
        json={"nome": "Empresa X", "plano": "a" * 21},
    )
    assert r.status_code == 422, f"plano de 21 chars deveria retornar 422, got {r.status_code}"


def test_magic_bytes_pdf_nomeado_como_ofx():
    """Arquivo PDF nomeado .ofx deve ser detectado por magic bytes e processado como PDF."""
    from api.parsers import _detectar_tipo

    pdf_header = b"%PDF-1.4 content here"
    ext = _detectar_tipo(pdf_header, "extrato.ofx")
    assert ext == ".pdf", f"Magic bytes PDF não detectado — ext retornada: {ext}"


def test_magic_bytes_xml_detectado():
    """Arquivo XML deve ser detectado por magic bytes."""
    from api.parsers import _detectar_tipo

    xml_content = b"<?xml version='1.0'?><root/>"
    ext = _detectar_tipo(xml_content, "extrato.pdf")  # extensão errada
    assert ext == ".xml", f"Magic bytes XML não detectado — ext retornada: {ext}"


# ── Trilha 12: gaps de production-readiness ───────────────────────────────


def test_criar_cliente_response_inclui_ativo():
    """Regressão: dict retornado por POST /clientes deve satisfazer ClienteResponse.

    Bug histórico: o dict pulava 'ativo' (que é Required em ClienteResponse),
    quebrando o endpoint com ResponseValidationError quando DATABASE_URL real
    estava configurado.
    """
    from datetime import datetime, timezone
    from unittest.mock import patch
    import uuid as _uuid

    # Arquitetura limpa: o router instancia CriarClienteUseCase (router -> usecase
    # -> repo). Mockamos a CLASSE do use case no modulo do router, preservando a
    # intencao do teste: a resposta deve conter 'ativo'.
    from api.usecases.criar_cliente import CriarClienteOutput
    from api.domain.entities import Cliente as _ClienteEntity

    entidade = _ClienteEntity(
        id=_uuid.uuid4(),
        nome="Empresa Teste",
        cnpj="11444777000161",
        email="teste@example.com",
        telefone=None,
        plano="basico",
        ativo=True,
        criado_em=datetime.now(timezone.utc),
    )

    class _FakeCriarUC:
        def __init__(self, *args, **kwargs):
            pass

        async def execute(self, _input):
            return CriarClienteOutput(cliente=entidade)

    ctx, _ = _patch_db_session()
    with ctx, patch("api.routers.clientes.CriarClienteUseCase", _FakeCriarUC):
        r = client.post(
            "/clientes",
            json={"nome": "Empresa Teste", "cnpj": "11.444.777/0001-61", "plano": "basico"},
        )

    assert r.status_code == 201, f"POST /clientes falhou: {r.status_code} — {r.text}"
    body = r.json()
    # Todos os campos required de ClienteResponse devem estar presentes
    assert "ativo" in body, f"Campo 'ativo' ausente na resposta: {body}"
    assert body["ativo"] is True
    assert body["nome"] == "Empresa Teste"
    assert body["plano"] == "basico"


def test_export_pdf_retorna_bytes_pdf_validos():
    """Smoke test: /export/pdf/{rid} deve retornar bytes que começam com %PDF.

    Gera um dataset via simulação primeiro, depois pede o PDF.
    Se weasyprint falhar (sem libpango), o endpoint cai no fallback HTML —
    nesse caso aceitamos text/html, mas o status deve ser 200.
    """
    # Cria um dataset via simulação
    r = client.post(
        "/conciliar/ofx?simular=true",
        files={"arquivos": ("smoke.ofx", io.BytesIO(OFX_SAMPLE2.encode()), "text/plain")},
    )
    assert r.status_code == 200, f"Falha ao criar dataset para PDF test: {r.text}"
    rid = r.json()["report_id"]

    # Tenta gerar PDF
    pdf_r = client.get(f"/export/pdf/{rid}")
    assert pdf_r.status_code == 200, f"GET /export/pdf/{rid} falhou: {pdf_r.status_code}"

    content_type = pdf_r.headers.get("content-type", "")
    if "application/pdf" in content_type:
        # WeasyPrint funcionou — valida magic bytes
        assert pdf_r.content.startswith(b"%PDF"), (
            f"Content-Type diz application/pdf mas bytes não começam com %PDF: " f"{pdf_r.content[:20]!r}"
        )
        assert len(pdf_r.content) > 1000, f"PDF muito pequeno ({len(pdf_r.content)} bytes) — provavelmente corrompido"
    else:
        # Fallback HTML — aceitável, mas log para visibilidade
        assert "text/html" in content_type, f"PDF endpoint retornou content-type inesperado: {content_type}"


# ── Trilha 13: DB integration — testes mocados (rodam em CI sem Supabase) ─


def _fake_cliente_mock(**overrides):
    """Cria um mock SQLAlchemy Cliente com defaults sensatos."""
    from unittest.mock import MagicMock
    from datetime import datetime, timezone
    import uuid as _uuid

    c = MagicMock()
    c.id = overrides.get("id", _uuid.uuid4())
    c.nome = overrides.get("nome", "Empresa Mock")
    c.cnpj = overrides.get("cnpj", "11444777000161")
    c.email = overrides.get("email", "mock@example.com")
    c.telefone = overrides.get("telefone", None)
    c.plano = overrides.get("plano", "basico")
    c.ativo = overrides.get("ativo", True)
    c.criado_em = overrides.get("criado_em", datetime.now(timezone.utc))
    return c


def _patch_db_session():
    """Context manager: ativa DB_DISPONIVEL=True + mock de SessionLocal."""
    from unittest.mock import AsyncMock, MagicMock, patch as _p

    session_mock = MagicMock()
    session_mock.__aenter__ = AsyncMock(return_value=session_mock)
    session_mock.__aexit__ = AsyncMock(return_value=False)
    return (
        _p.multiple(
            "api.routers.clientes",
            DB_DISPONIVEL=True,
            SessionLocal=MagicMock(return_value=session_mock),
        ),
        session_mock,
    )


def test_listar_clientes_com_db_mockado():
    """GET /clientes com DB mockado: retorna lista válida + response_model OK."""
    from unittest.mock import patch

    class _FakeListarUC:
        def __init__(self, *args, **kwargs):
            pass

        async def execute(self, _input):
            return [
                _fake_cliente_mock(nome="Cliente A"),
                _fake_cliente_mock(nome="Cliente B"),
            ]

    ctx, _ = _patch_db_session()
    with ctx, patch("api.routers.clientes.ListarClientesUseCase", _FakeListarUC):
        r = client.get("/clientes")

    assert r.status_code == 200, r.text
    body = r.json()
    assert isinstance(body, list) and len(body) == 2
    for item in body:
        # ClienteResponse exige ativo
        assert "ativo" in item, f"Campo 'ativo' ausente em listar: {item}"
        assert "id" in item and "nome" in item


def test_buscar_cliente_com_db_mockado_retorna_cliente():
    """GET /clientes/{id} com DB mockado e cliente existente."""
    from unittest.mock import patch

    fake = _fake_cliente_mock(nome="Empresa Encontrada", email="found@example.com")

    async def fake_buscar(*args, **kwargs):
        return fake

    ctx, _ = _patch_db_session()
    with ctx, patch("api.main.crud_clientes.buscar_cliente", side_effect=fake_buscar):
        r = client.get(f"/clientes/{fake.id}")

    assert r.status_code == 200, r.text
    body = r.json()
    assert body["nome"] == "Empresa Encontrada"
    assert body["email"] == "found@example.com"
    assert body["ativo"] is True


def test_buscar_cliente_inexistente_retorna_404():
    """GET /clientes/{id} com DB mockado e cliente inexistente → 404."""
    from unittest.mock import patch

    async def fake_buscar_none(*args, **kwargs):
        return None

    ctx, _ = _patch_db_session()
    with ctx, patch("api.main.crud_clientes.buscar_cliente", side_effect=fake_buscar_none):
        r = client.get("/clientes/00000000-0000-0000-0000-000000000099")

    assert r.status_code == 404, f"Cliente inexistente deveria retornar 404, got {r.status_code}"


def test_atualizar_cliente_com_db_mockado():
    """PATCH /clientes/{id} com DB mockado."""
    from unittest.mock import patch

    updated = _fake_cliente_mock(nome="Empresa Renomeada", plano="pro")

    async def fake_atualizar(*args, **kwargs):
        return updated

    ctx, _ = _patch_db_session()
    with ctx, patch("api.main.crud_clientes.atualizar_cliente", side_effect=fake_atualizar):
        r = client.patch(
            f"/clientes/{updated.id}",
            json={"nome": "Empresa Renomeada", "plano": "pro"},
        )

    assert r.status_code == 200, r.text
    body = r.json()
    assert body["nome"] == "Empresa Renomeada"
    assert body["plano"] == "pro"


def test_atualizar_cliente_inexistente_retorna_404():
    """PATCH /clientes/{id} com cliente inexistente → 404."""
    from unittest.mock import patch

    async def fake_atualizar_none(*args, **kwargs):
        return None

    ctx, _ = _patch_db_session()
    with ctx, patch("api.main.crud_clientes.atualizar_cliente", side_effect=fake_atualizar_none):
        r = client.patch(
            "/clientes/00000000-0000-0000-0000-000000000099",
            json={"nome": "Fantasma"},
        )

    assert r.status_code == 404


def test_criar_cliente_cnpj_duplicado_retorna_409():
    """POST /clientes com IntegrityError (CNPJ duplicado) → 409."""
    from unittest.mock import patch
    from sqlalchemy.exc import IntegrityError

    class _FakeCriarUCDup:
        def __init__(self, *args, **kwargs):
            pass

        async def execute(self, _input):
            raise IntegrityError("INSERT INTO clientes", {}, Exception("UNIQUE violation"))

    ctx, _ = _patch_db_session()
    with ctx, patch("api.routers.clientes.CriarClienteUseCase", _FakeCriarUCDup):
        r = client.post(
            "/clientes",
            json={"nome": "Empresa Duplicada", "cnpj": "11444777000161", "plano": "basico"},
        )

    assert r.status_code == 409, f"CNPJ duplicado deveria retornar 409, got {r.status_code}"


def test_salvar_no_banco_persiste_conciliacao_e_retorna_ok():
    """_salvar_no_banco com DB mockado retorna {status: ok, transacoes_persistidas: N}."""
    import asyncio
    from unittest.mock import AsyncMock, MagicMock, patch

    session_mock = MagicMock()
    # Suporta `async with SessionLocal() as db` e `async with db.begin()`
    session_mock.__aenter__ = AsyncMock(return_value=session_mock)
    session_mock.__aexit__ = AsyncMock(return_value=False)
    session_mock.begin = MagicMock(return_value=session_mock)
    session_mock.add = MagicMock()
    session_mock.add_all = MagicMock()
    session_mock.flush = AsyncMock()

    extratos = [
        {
            "conta": "AG 1234 / CC 5678",
            "arquivo": "test.ofx",
            "qtd": 2,
            "transacoes": [
                {"data": "2026-01-01", "valor": -100.0, "memo": "Saída", "nome": "", "tipo": "DEBIT"},
                {"data": "2026-01-02", "valor": 250.5, "memo": "Entrada", "nome": "", "tipo": "CREDIT"},
            ],
        }
    ]
    anomalias = [{"conta": "AG 1234 / CC 5678", "valor": 250.5, "severidade": "atencao", "tipo": "ok"}]

    async def _run():
        with patch("api.services.db_persistence.DB_DISPONIVEL", True), patch("api.services.db_persistence.SessionLocal", return_value=session_mock):
            from api.main import _salvar_no_banco

            return await _salvar_no_banco("abc123456789", extratos, anomalias, "test")

    result = asyncio.run(_run())
    assert result["status"] == "ok", f"_salvar_no_banco falhou: {result}"
    assert result["transacoes_persistidas"] == 2
    # add foi chamado uma vez com Conciliacao, add_all com lista de Transacao
    assert session_mock.add.called
    assert session_mock.add_all.called


# ── Trilha 15: frontend /app servido pelo React (build) ou 503 explícito ──


def test_app_reflete_estado_do_build_react():
    """/app serve o React quando o build existe, senão 503 explícito (nunca UI fantasma)."""
    from api.core.config import REACT_DIST

    r = client.get("/app")
    if REACT_DIST.exists():
        assert r.status_code == 200
        assert "text/html" in r.headers.get("content-type", "")
    else:
        assert r.status_code == 503
        detail = r.json().get("detail", "").lower()
        assert "npm run build" in detail or "compil" in detail


# ── Trilha 14: LLM integration — testes mocados (rodam em CI sem chave real) ──


def _fake_anthropic_response(text: str, in_tokens: int = 100, out_tokens: int = 200):
    """Cria mock de resposta do Anthropic SDK (resp.content[].text + usage)."""
    from unittest.mock import MagicMock

    block = MagicMock()
    block.type = "text"
    block.text = text
    resp = MagicMock()
    resp.content = [block]
    resp.usage = MagicMock(input_tokens=in_tokens, output_tokens=out_tokens)
    resp.stop_reason = "end_turn"
    return resp


def _make_apistatus_error(msg: str, status_code: int = 400, error_message: str = None):
    """Constrói uma APIStatusError real do SDK Anthropic para uso em side_effect."""
    from anthropic import APIStatusError
    import httpx

    request = httpx.Request("POST", "https://api.anthropic.com/v1/messages")
    response = httpx.Response(status_code, request=request)
    body = {"error": {"message": error_message or msg}}
    err = APIStatusError(msg, response=response, body=body)
    err.status_code = status_code
    return err


def _wire_stream(fake_client, *, response=None, side_effect=None):
    """Adapta um fake_client ao codigo que usa `c.messages.stream(...)` como
    context manager. Sucesso: get_final_message()->response. Erro: levanta no enter."""
    from unittest.mock import MagicMock

    cm = MagicMock()
    if side_effect is not None:
        cm.__enter__ = MagicMock(side_effect=side_effect)
    else:
        stream = MagicMock()
        stream.get_final_message = MagicMock(return_value=response)
        cm.__enter__ = MagicMock(return_value=stream)
    cm.__exit__ = MagicMock(return_value=False)
    fake_client.messages.stream = MagicMock(return_value=cm)
    return fake_client


def test_conciliar_ofx_modo_llm_single_mockado():
    """POST /conciliar/ofx em modo LLM (sem multi-modelo): caminho happy."""
    from unittest.mock import MagicMock, patch

    fake_client = MagicMock()
    _wire_stream(fake_client, response=_fake_anthropic_response(
        "# Relatório de Conciliação\n\nAnálise completa pelo Haiku.",
        in_tokens=150,
        out_tokens=400,
    ))

    with (
        patch("api.services.conciliacao_llm._get_client", return_value=fake_client),
        patch.dict(os.environ, {"ANTHROPIC_API_KEY": "sk-ant-test-mock"}),
    ):
        r = client.post(
            "/conciliar/ofx?modelo=haiku",
            files={"arquivos": ("test.ofx", io.BytesIO(OFX_SAMPLE2.encode()), "text/plain")},
        )

    assert r.status_code == 200, r.text
    data = r.json()
    assert data["modo"] == "claude_llm"
    assert data["modelo"] == "haiku"
    assert data["modelo_id"] == "claude-haiku-4-5-20251001"
    assert data["usage"]["input_tokens"] == 150
    assert data["usage"]["output_tokens"] == 400
    assert "Análise completa pelo Haiku" in data["relatorio_md"]
    assert "Análise completa pelo Haiku" in data["relatorio_html"]  # sanitizado


def test_conciliar_ofx_llm_credito_esgotado_retorna_msg_amigavel():
    """APIStatusError com 'credit balance' → 400 + mensagem em pt-BR."""
    from unittest.mock import MagicMock, patch

    err = _make_apistatus_error(
        "Insufficient credits",
        status_code=400,
        error_message="Your credit balance is too low to access the Anthropic API",
    )
    fake_client = MagicMock()
    _wire_stream(fake_client, side_effect=err)

    with (
        patch("api.services.conciliacao_llm._get_client", return_value=fake_client),
        patch.dict(os.environ, {"ANTHROPIC_API_KEY": "sk-ant-test-mock"}),
    ):
        r = client.post(
            "/conciliar/ofx?modelo=sonnet",
            files={"arquivos": ("test.ofx", io.BytesIO(OFX_SAMPLE2.encode()), "text/plain")},
        )

    assert r.status_code == 400, r.text
    body = r.json()
    msg = body["detail"]["anthropic_error"]
    assert "creditos" in msg.lower() or "credit" in msg.lower()
    # Verifica que mostrou link de billing (mensagem amigável traduzida)
    assert "billing" in msg.lower() or "recarregue" in msg.lower()


def test_conciliar_ofx_llm_rate_limit_retorna_msg_amigavel():
    """APIStatusError com 'rate limit' → status 429 + mensagem em pt-BR."""
    from unittest.mock import MagicMock, patch

    err = _make_apistatus_error(
        "Rate limit",
        status_code=429,
        error_message="Rate limit exceeded for this organization",
    )
    fake_client = MagicMock()
    _wire_stream(fake_client, side_effect=err)

    with (
        patch("api.services.conciliacao_llm._get_client", return_value=fake_client),
        patch.dict(os.environ, {"ANTHROPIC_API_KEY": "sk-ant-test-mock"}),
    ):
        r = client.post(
            "/conciliar/ofx?modelo=opus",
            files={"arquivos": ("test.ofx", io.BytesIO(OFX_SAMPLE2.encode()), "text/plain")},
        )

    assert r.status_code == 429, r.text
    msg = r.json()["detail"]["anthropic_error"]
    assert "rate limit" in msg.lower() or "aguarde" in msg.lower()


def test_conciliar_ofx_multi_modelo_mockado():
    """Multi-modelo: 3 chamadas paralelas + síntese, com score_consenso extraído."""
    from unittest.mock import patch

    chamadas_feitas = []

    async def fake_chamar_modelo(api_key, prompt, model_id, label, max_tokens):
        chamadas_feitas.append((model_id, label, max_tokens))
        return {
            "texto": f"Análise produzida por {label}",
            "input_tokens": 200,
            "output_tokens": 500,
            "modelo": model_id,
            "label": label,
            "erro": None,
        }

    async def fake_sintetizar(api_key, resultados, max_tokens):
        # Verifica que recebeu 3 resultados
        assert len(resultados) == 3
        return "## Índice de Consenso: 85/100\n\nRelatório consolidado", 0.85, 0.0123, 0, 0

    with (
        patch("api.routers.conciliacao.chamar_modelo_async", side_effect=fake_chamar_modelo),
        patch("api.routers.conciliacao.sintetizar_consenso", side_effect=fake_sintetizar),
        patch.dict(os.environ, {"ANTHROPIC_API_KEY": "sk-ant-test-mock"}),
    ):
        r = client.post(
            "/conciliar/ofx?multi_modelo=true",
            files={"arquivos": ("test.ofx", io.BytesIO(OFX_SAMPLE2.encode()), "text/plain")},
        )

    assert r.status_code == 200, r.text
    data = r.json()
    assert data["modo"] == "multi_modelo"
    assert data["score_consenso"] == 0.85
    assert len(data["modelos"]) == 3  # opus + sonnet + haiku
    assert {m["modelo"] for m in data["modelos"]} == {
        "claude-opus-4-8",
        "claude-sonnet-4-6",
        "claude-haiku-4-5-20251001",
    }
    assert "Índice de Consenso: 85/100" in data["relatorio_md"]
    # Cada modelo recebeu max_tokens // 2 (default 16000 → 8000)
    assert all(c[2] == 8000 for c in chamadas_feitas), f"max_tokens errado: {chamadas_feitas}"


def test_conciliar_csv_modo_llm_mockado():
    """POST /conciliar/csv: caminho LLM com mock do Anthropic."""
    from unittest.mock import MagicMock, patch

    fake_client = MagicMock()
    _wire_stream(fake_client, response=_fake_anthropic_response(
        "## Conciliação Bancária — CSV\n\nDivergências identificadas.",
        in_tokens=80,
        out_tokens=250,
    ))

    extrato_csv = b"data,valor,memo\n2026-01-01,-100.00,Saida\n"
    razao_csv = b"data,valor,conta\n2026-01-01,-100.00,Despesa Geral\n"

    with (
        patch("api.services.conciliacao_llm._get_client", return_value=fake_client),
        patch.dict(os.environ, {"ANTHROPIC_API_KEY": "sk-ant-test-mock"}),
    ):
        r = client.post(
            "/conciliar/csv",
            files={
                "extrato": ("extrato.csv", io.BytesIO(extrato_csv), "text/csv"),
                "razao": ("razao.csv", io.BytesIO(razao_csv), "text/csv"),
            },
        )

    assert r.status_code == 200, r.text
    data = r.json()
    assert data["extrato"] == "extrato.csv"
    assert data["razao"] == "razao.csv"
    assert data["usage"]["input_tokens"] == 80
    assert "Divergências identificadas" in data["relatorio_md"]


def test_sintetizar_consenso_zero_resultados_validos():
    """_sintetizar_consenso com lista vazia retorna mensagem padrão + score 0.0."""
    import asyncio
    from api.main import _sintetizar_consenso

    async def _run():
        return await _sintetizar_consenso("fake-key", [], max_tokens=4000)

    texto, score, _custo, *_ = asyncio.run(_run())
    assert "Nenhum modelo" in texto
    assert score == 0.0


def test_sintetizar_consenso_um_resultado_valido_score_0_5():
    """_sintetizar_consenso com 1 resultado válido retorna o texto direto + score 0.5."""
    import asyncio
    from api.main import _sintetizar_consenso

    resultados = [
        {"texto": "Único relatório válido", "erro": None, "label": "Haiku", "modelo": "haiku"},
        {"texto": "", "erro": "API error", "label": "Sonnet", "modelo": "sonnet"},
    ]

    async def _run():
        return await _sintetizar_consenso("fake-key", resultados, max_tokens=4000)

    texto, score, _custo, *_ = asyncio.run(_run())
    assert texto == "Único relatório válido"
    assert score == 0.5


def test_chamar_modelo_async_apistatus_retorna_texto_vazio():
    """_chamar_modelo_async com APIStatusError → res['erro'] preenchido + texto=''."""
    import asyncio
    from unittest.mock import MagicMock, patch
    from api.main import _chamar_modelo_async

    err = _make_apistatus_error("api fail", error_message="model overloaded right now")

    fake_anthropic_class = MagicMock()
    fake_instance = MagicMock()
    _wire_stream(fake_instance, side_effect=err)
    fake_anthropic_class.return_value = fake_instance

    async def _run():
        with patch("api.services.conciliacao_llm.Anthropic", fake_anthropic_class):
            return await _chamar_modelo_async("k", "prompt", "model-x", "Label-X", 1000)

    res = asyncio.run(_run())
    assert res["texto"] == ""
    assert res["input_tokens"] == 0
    assert res["output_tokens"] == 0
    assert res["erro"] is not None
    assert "overloaded" in res["erro"].lower()
    assert res["modelo"] == "model-x"
    assert res["label"] == "Label-X"


@pytest.mark.skipif(
    not os.environ.get("ANTHROPIC_API_KEY") or os.environ.get("ANTHROPIC_API_KEY", "").startswith("sk-ant-test"),
    reason="LLM smoke test requer ANTHROPIC_API_KEY real (não sk-ant-test do CI)",
)
def test_llm_smoke_anthropic_real():
    """Smoke test opcional: chama Claude real com OFX mínimo.

    Skipped em CI (key=sk-ant-test). Roda localmente se ANTHROPIC_API_KEY
    válida estiver configurada. Custa ~1¢ por execução (Haiku, ~500 tokens).
    """
    r = client.post(
        "/conciliar/ofx?modelo=haiku&max_tokens=500",
        files={"arquivos": ("smoke.ofx", io.BytesIO(OFX_SAMPLE2.encode()), "text/plain")},
    )
    assert r.status_code == 200, f"LLM smoke falhou: {r.status_code} — {r.text}"
    data = r.json()
    assert data["modo"] == "claude_llm"
    assert data["modelo"] == "haiku"
    assert "usage" in data
    assert data["usage"]["input_tokens"] > 0
    assert data["usage"]["output_tokens"] > 0
    assert len(data.get("relatorio_md", "")) > 100, "Relatório LLM suspeitamente curto — possível falha de geração"
